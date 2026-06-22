# tests/test_cockpit_ingenieur.py
"""Tests TDD pour cockpit_ingenieur_generator."""
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.models import UOInstance, UOType, Activity, System, Project, StatutUO


def _make_uo(uid: str, engineer: str, hours: float, end: date,
             uo_type_id: str = "spec_fonctionnelle") -> UOInstance:
    activities = [
        Activity(id="ACT-1", name="Analyse", default_hours=hours * 0.4),
        Activity(id="ACT-2", name="Rédaction", default_hours=hours * 0.6),
    ]
    uo_type = UOType(id=uo_type_id, name=f"Type {uo_type_id}", activities=activities)
    system = System(id="clim", name="Climatisation")
    project = Project(id="MI20", name="MI20 RATP")
    return UOInstance(
        id=uid, uo_type_id=uo_type_id, system_id="clim", project_id="MI20",
        engineer_name=engineer, total_hours=hours,
        start_date=date(2026, 4, 1), end_date=end,
        statut=StatutUO.EN_COURS,
        uo_type=uo_type, system=system, project=project,
    )


ALL_UOS = [
    _make_uo("UO-001", "Alice Dubois",  32, date(2026, 6, 30)),
    _make_uo("UO-002", "Alice Dubois",  48, date(2026, 7, 15)),
    _make_uo("UO-003", "Bruno Lecomte", 40, date(2026, 6, 20)),
]


class TestCockpitIngenieurFichier:
    def test_fichier_cree(self, tmp_path):
        from src.generators.cockpit_ingenieur_generator import generate_cockpit_ingenieur
        path = generate_cockpit_ingenieur("Alice Dubois", ALL_UOS, output_dir=tmp_path)
        assert path.exists()
        assert path.name == "Cockpit_Alice_Dubois.xlsx"

    def test_trois_onglets_presents(self, tmp_path):
        from src.generators.cockpit_ingenieur_generator import generate_cockpit_ingenieur
        path = generate_cockpit_ingenieur("Alice Dubois", ALL_UOS, output_dir=tmp_path)
        wb = load_workbook(path)
        assert "Agenda" in wb.sheetnames
        assert "Mes UOs" in wb.sheetnames
        assert "_Manifeste" in wb.sheetnames


class TestCockpitMesUOs:
    def test_seules_les_uo_de_alice(self, tmp_path):
        from src.generators.cockpit_ingenieur_generator import generate_cockpit_ingenieur
        path = generate_cockpit_ingenieur("Alice Dubois", ALL_UOS, output_dir=tmp_path)
        wb = load_workbook(path)
        ws = wb["Mes UOs"]
        ids = [ws.cell(row=r, column=2).value for r in range(6, 20) if ws.cell(row=r, column=2).value]
        assert "UO-001" in ids
        assert "UO-002" in ids
        assert "UO-003" not in ids  # Bruno, pas Alice

    def test_en_tetes_onglet_mes_uo(self, tmp_path):
        from src.generators.cockpit_ingenieur_generator import generate_cockpit_ingenieur
        path = generate_cockpit_ingenieur("Alice Dubois", ALL_UOS, output_dir=tmp_path)
        wb = load_workbook(path)
        ws = wb["Mes UOs"]
        headers = [ws.cell(row=5, column=c).value for c in range(2, 11)]
        assert "UO ID" in headers
        assert "% Avancement" in headers
        assert "H réalisées" in headers
        assert "Alerte" in headers

    def test_zone_saisie_avancement_col_f(self, tmp_path):
        from src.generators.cockpit_ingenieur_generator import generate_cockpit_ingenieur
        path = generate_cockpit_ingenieur("Alice Dubois", ALL_UOS, output_dir=tmp_path)
        wb = load_workbook(path)
        ws = wb["Mes UOs"]
        assert ws.cell(row=5, column=7).value == "% Avancement"
        assert ws.cell(row=5, column=8).value == "H réalisées"

    def test_formule_alerte_presente(self, tmp_path):
        from src.generators.cockpit_ingenieur_generator import generate_cockpit_ingenieur
        path = generate_cockpit_ingenieur("Alice Dubois", ALL_UOS, output_dir=tmp_path)
        wb = load_workbook(path)
        ws = wb["Mes UOs"]
        alerte_cell = ws.cell(row=6, column=10).value
        assert alerte_cell is not None
        assert str(alerte_cell).startswith("=IF(")

    def test_table_nommee_tbl_mes_uos_presente(self, tmp_path):
        from src.generators.cockpit_ingenieur_generator import generate_cockpit_ingenieur
        path = generate_cockpit_ingenieur("Alice Dubois", ALL_UOS, output_dir=tmp_path)
        wb = load_workbook(path)
        ws = wb["Mes UOs"]
        table_names = list(ws.tables.keys())
        assert "tbl_mes_uos" in table_names


class TestCockpitAgenda:
    def test_en_tetes_onglet_agenda(self, tmp_path):
        from src.generators.cockpit_ingenieur_generator import generate_cockpit_ingenieur
        path = generate_cockpit_ingenieur("Alice Dubois", ALL_UOS, output_dir=tmp_path)
        wb = load_workbook(path)
        ws = wb["Agenda"]
        all_values = [ws.cell(row=r, column=c).value for r in range(1, 15) for c in range(1, 7)]
        assert "UO ID" in all_values
        assert "Activité" in all_values

    def test_section_semaine_presente(self, tmp_path):
        from src.generators.cockpit_ingenieur_generator import generate_cockpit_ingenieur
        path = generate_cockpit_ingenieur("Alice Dubois", ALL_UOS, output_dir=tmp_path)
        wb = load_workbook(path)
        ws = wb["Agenda"]
        all_values = [ws.cell(row=r, column=c).value for r in range(1, 30) for c in range(1, 4)]
        assert any("Semaine" in str(v) for v in all_values if v)

    def test_section_prochaines_echeances_presente(self, tmp_path):
        from src.generators.cockpit_ingenieur_generator import generate_cockpit_ingenieur
        path = generate_cockpit_ingenieur("Alice Dubois", ALL_UOS, output_dir=tmp_path)
        wb = load_workbook(path)
        ws = wb["Agenda"]
        all_values = [ws.cell(row=r, column=c).value for r in range(1, 50) for c in range(1, 4)]
        assert any("Prochaines" in str(v) or "échéance" in str(v).lower() for v in all_values if v)


class TestCockpitManifeste:
    def test_version_manifeste_a1(self, tmp_path):
        from src.generators.cockpit_ingenieur_generator import generate_cockpit_ingenieur
        path = generate_cockpit_ingenieur("Alice Dubois", ALL_UOS, output_dir=tmp_path)
        wb = load_workbook(path)
        ws = wb["_Manifeste"]
        assert str(ws["A1"].value).startswith("MANIFESTE_V=")

    def test_ligne2_vide(self, tmp_path):
        """Ligne 2 doit être vide — le parser MXL la skippe."""
        from src.generators.cockpit_ingenieur_generator import generate_cockpit_ingenieur
        path = generate_cockpit_ingenieur("Alice Dubois", ALL_UOS, output_dir=tmp_path)
        wb = load_workbook(path)
        ws = wb["_Manifeste"]
        assert ws["A2"].value is None

    def test_file_type_en_a3(self, tmp_path):
        """A3 doit contenir FILE_TYPE: cockpit_ingenieur."""
        from src.generators.cockpit_ingenieur_generator import generate_cockpit_ingenieur
        path = generate_cockpit_ingenieur("Alice Dubois", ALL_UOS, output_dir=tmp_path)
        wb = load_workbook(path)
        ws = wb["_Manifeste"]
        assert ws["A3"].value == "FILE_TYPE: cockpit_ingenieur"

    def test_commentaires_en_colonne_c(self, tmp_path):
        """Chaque instruction MXL doit avoir un commentaire non vide en colonne C."""
        from src.generators.cockpit_ingenieur_generator import generate_cockpit_ingenieur
        path = generate_cockpit_ingenieur("Alice Dubois", ALL_UOS, output_dir=tmp_path)
        wb = load_workbook(path)
        ws = wb["_Manifeste"]
        # Lignes avec une instruction en col A doivent avoir un commentaire en col C
        for r in range(3, 15):
            instr = ws.cell(row=r, column=1).value
            if instr and str(instr).strip():
                comment = ws.cell(row=r, column=3).value
                assert comment and len(str(comment)) > 5, \
                    f"Commentaire manquant ou trop court en ligne {r}: '{comment}'"

    def test_push_instruction_presente(self, tmp_path):
        """Une instruction PUSH $mes_uos -> ... doit être présente."""
        from src.generators.cockpit_ingenieur_generator import generate_cockpit_ingenieur
        path = generate_cockpit_ingenieur("Alice Dubois", ALL_UOS, output_dir=tmp_path)
        wb = load_workbook(path)
        ws = wb["_Manifeste"]
        instrs = [ws.cell(row=r, column=1).value for r in range(1, 20)]
        assert any(
            str(v).startswith("PUSH $mes_uos") for v in instrs if v
        )

    def test_def_get_table_presente(self, tmp_path):
        """Une instruction DEF $mes_uos = GET_TABLE(...) doit être présente."""
        from src.generators.cockpit_ingenieur_generator import generate_cockpit_ingenieur
        path = generate_cockpit_ingenieur("Alice Dubois", ALL_UOS, output_dir=tmp_path)
        wb = load_workbook(path)
        ws = wb["_Manifeste"]
        instrs = [ws.cell(row=r, column=1).value for r in range(1, 20)]
        assert any(
            str(v).startswith("DEF $mes_uos = GET_TABLE") for v in instrs if v
        )

    def test_colonne_b_non_polluee(self, tmp_path):
        """Col B = ancres uniquement. Les commentaires ne doivent PAS être en col B."""
        from src.generators.cockpit_ingenieur_generator import generate_cockpit_ingenieur
        path = generate_cockpit_ingenieur("Alice Dubois", ALL_UOS, output_dir=tmp_path)
        wb = load_workbook(path)
        ws = wb["_Manifeste"]
        for r in range(3, 15):
            b_val = ws.cell(row=r, column=2).value
            # Col B est vide ou contient une ancre (pas un commentaire long)
            if b_val:
                assert len(str(b_val)) < 60, \
                    f"Col B ligne {r} semble contenir un commentaire : '{b_val}'"

    def test_mxl_parseable_zero_erreurs(self, tmp_path):
        """Le _Manifeste généré doit être parseable par parser.py sans erreur."""
        from src.generators.cockpit_ingenieur_generator import generate_cockpit_ingenieur
        from src.parser import parse_file
        path = generate_cockpit_ingenieur("Alice Dubois", ALL_UOS, output_dir=tmp_path)
        ast = parse_file(path)
        assert ast is not None, "parse_file() a retourné None — pas de feuille _Manifeste"
        errors = [f"L{e.line_num}: {e.message}" for e in ast.errors]
        assert not ast.errors, f"Erreurs de parse MXL : {errors}"
