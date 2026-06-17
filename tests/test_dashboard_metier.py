# tests/test_dashboard_metier.py
"""Tests TDD pour dashboard_metier_generator."""
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.models import (
    UOInstance, UOType, Activity, System, Project, StatutUO,
    ProfilActeur, Role, TypeFiltre, NiveauAcces,
)
from src.store import JsonStore


def _make_uo(uid: str, engineer: str, hours: float, end: date) -> UOInstance:
    uo_type = UOType(id="spec_fonctionnelle", name="Spec Fonctionnelle", activities=[
        Activity(id="A1", name="Analyse", default_hours=hours * 0.5),
    ])
    return UOInstance(
        id=uid, uo_type_id="spec_fonctionnelle", system_id="clim", project_id="MI20",
        engineer_name=engineer, total_hours=hours,
        start_date=date(2026, 4, 1), end_date=end,
        statut=StatutUO.EN_COURS,
        uo_type=uo_type,
        system=System(id="clim", name="Climatisation"),
        project=Project(id="MI20", name="MI20 RATP"),
    )


def _make_pilote_metier() -> ProfilActeur:
    return ProfilActeur(
        id="USR004", nom="Jean-Luc Bernard",
        role=Role.PILOTE_METIER,
        filtre_type=TypeFiltre.INGENIEUR,
        filtre_valeur="Alice Dubois,Bruno Lecomte",
        acces=NiveauAcces.READ,
    )


ALL_UOS = [
    _make_uo("UO-001", "Alice Dubois",  32, date(2026, 6, 30)),
    _make_uo("UO-002", "Alice Dubois",  48, date(2026, 7, 15)),
    _make_uo("UO-003", "Bruno Lecomte", 40, date(2026, 6, 20)),
    _make_uo("UO-004", "Denis Renard",  24, date(2026, 8, 1)),   # hors périmètre
]


class TestDashboardFichier:
    def test_fichier_cree(self, tmp_path):
        from src.generators.dashboard_metier_generator import generate_dashboard_metier
        acteur = _make_pilote_metier()
        store = JsonStore(tmp_path / "store.json")
        path = generate_dashboard_metier(acteur, ALL_UOS, store, output_dir=tmp_path)
        assert path.exists()
        assert path.name == "Dashboard_Jean-Luc_Bernard.xlsx"

    def test_quatre_onglets_presents(self, tmp_path):
        from src.generators.dashboard_metier_generator import generate_dashboard_metier
        acteur = _make_pilote_metier()
        store = JsonStore(tmp_path / "store.json")
        path = generate_dashboard_metier(acteur, ALL_UOS, store, output_dir=tmp_path)
        wb = load_workbook(path)
        assert "Synthèse" in wb.sheetnames
        assert "Par Ingénieur" in wb.sheetnames
        assert "Alertes" in wb.sheetnames
        assert "_Manifeste" in wb.sheetnames


class TestDashboardFiltrage:
    def test_filtre_ingenieur_respecte(self, tmp_path):
        """Denis Renard ne doit pas apparaître dans le dashboard de Jean-Luc."""
        from src.generators.dashboard_metier_generator import generate_dashboard_metier
        acteur = _make_pilote_metier()
        store = JsonStore(tmp_path / "store.json")
        path = generate_dashboard_metier(acteur, ALL_UOS, store, output_dir=tmp_path)
        wb = load_workbook(path)
        ws = wb["Synthèse"]
        all_values = [ws.cell(row=r, column=c).value
                      for r in range(1, 50) for c in range(1, 12)]
        assert "UO-004" not in all_values
        assert "Denis Renard" not in all_values

    def test_uo_alice_et_bruno_presents(self, tmp_path):
        from src.generators.dashboard_metier_generator import generate_dashboard_metier
        acteur = _make_pilote_metier()
        store = JsonStore(tmp_path / "store.json")
        path = generate_dashboard_metier(acteur, ALL_UOS, store, output_dir=tmp_path)
        wb = load_workbook(path)
        ws = wb["Synthèse"]
        all_values = [ws.cell(row=r, column=c).value
                      for r in range(1, 50) for c in range(1, 12)]
        assert "UO-001" in all_values
        assert "UO-002" in all_values
        assert "UO-003" in all_values


class TestDashboardKPIs:
    def test_charge_totale_correcte(self, tmp_path):
        """32 + 48 + 40 = 120h pour Alice + Bruno (Denis exclu)."""
        from src.generators.dashboard_metier_generator import generate_dashboard_metier
        acteur = _make_pilote_metier()
        store = JsonStore(tmp_path / "store.json")
        path = generate_dashboard_metier(acteur, ALL_UOS, store, output_dir=tmp_path)
        wb = load_workbook(path)
        ws = wb["Synthèse"]
        all_values = [ws.cell(row=r, column=c).value
                      for r in range(1, 10) for c in range(1, 15)]
        assert 120 in all_values or "120h" in [str(v) for v in all_values if v]

    def test_nb_uo_kpi_correct(self, tmp_path):
        """3 UOs dans le périmètre (Alice×2 + Bruno×1)."""
        from src.generators.dashboard_metier_generator import generate_dashboard_metier
        acteur = _make_pilote_metier()
        store = JsonStore(tmp_path / "store.json")
        path = generate_dashboard_metier(acteur, ALL_UOS, store, output_dir=tmp_path)
        wb = load_workbook(path)
        ws = wb["Synthèse"]
        all_values = [ws.cell(row=r, column=c).value
                      for r in range(1, 10) for c in range(1, 15)]
        assert 3 in all_values or "3 UOs" in [str(v) for v in all_values if v]


class TestDashboardAlertes:
    def test_alerte_depassement_heures(self, tmp_path):
        """UO avec heures_realisees > charge doit apparaître dans Alertes."""
        from src.generators.dashboard_metier_generator import generate_dashboard_metier
        acteur = _make_pilote_metier()
        store = JsonStore(tmp_path / "store.json")
        store.set("uo.UO-001.heures_realisees", 50.0)  # > 32h allouées
        path = generate_dashboard_metier(acteur, ALL_UOS, store, output_dir=tmp_path)
        wb = load_workbook(path)
        ws = wb["Alertes"]
        all_values = [ws.cell(row=r, column=c).value
                      for r in range(1, 30) for c in range(1, 8)]
        assert "UO-001" in all_values

    def test_pas_alerte_si_heures_ok(self, tmp_path):
        """UO sans dépassement ne doit pas apparaître comme alerte dérive."""
        from src.generators.dashboard_metier_generator import generate_dashboard_metier
        acteur = _make_pilote_metier()
        store = JsonStore(tmp_path / "store.json")
        store.set("uo.UO-001.heures_realisees", 10.0)  # < 32h — OK
        path = generate_dashboard_metier(acteur, ALL_UOS, store, output_dir=tmp_path)
        wb = load_workbook(path, data_only=True)
        ws = wb["Alertes"]
        found = False
        for r in range(3, 30):
            uid = ws.cell(row=r, column=2).value
            type_alerte = ws.cell(row=r, column=3).value
            if uid == "UO-001" and type_alerte and "Dépassement" in str(type_alerte):
                found = True
        assert not found


class TestDashboardManifeste:
    def test_version_manifeste(self, tmp_path):
        from src.generators.dashboard_metier_generator import generate_dashboard_metier
        acteur = _make_pilote_metier()
        store = JsonStore(tmp_path / "store.json")
        path = generate_dashboard_metier(acteur, ALL_UOS, store, output_dir=tmp_path)
        wb = load_workbook(path)
        ws = wb["_Manifeste"]
        assert str(ws["A1"].value).startswith("MANIFESTE_V=")

    def test_colonne_commentaire_presente(self, tmp_path):
        from src.generators.dashboard_metier_generator import generate_dashboard_metier
        acteur = _make_pilote_metier()
        store = JsonStore(tmp_path / "store.json")
        path = generate_dashboard_metier(acteur, ALL_UOS, store, output_dir=tmp_path)
        wb = load_workbook(path)
        ws = wb["_Manifeste"]
        headers = [ws.cell(row=2, column=c).value for c in range(1, 15)]
        assert "COMMENTAIRE" in headers

    def test_regles_pull_avancement_presentes(self, tmp_path):
        from src.generators.dashboard_metier_generator import generate_dashboard_metier
        acteur = _make_pilote_metier()
        store = JsonStore(tmp_path / "store.json")
        path = generate_dashboard_metier(acteur, ALL_UOS, store, output_dir=tmp_path)
        wb = load_workbook(path)
        ws = wb["_Manifeste"]
        nom_globals = [ws.cell(row=r, column=3).value for r in range(3, 30)
                       if ws.cell(row=r, column=3).value]
        assert "uo.UO-001.avancement" in nom_globals
        assert "uo.UO-003.avancement" in nom_globals
        assert "uo.UO-004.avancement" not in nom_globals


class TestPushPullCycle:
    """Vérifie le cycle complet : store → dashboard (simulation du push ingénieur)."""

    def test_avancement_store_visible_dans_synthese(self, tmp_path):
        """Simule un push de 80% d'avancement → vérifie que Synthèse affiche 0.8."""
        from src.generators.dashboard_metier_generator import generate_dashboard_metier
        acteur = _make_pilote_metier()
        store = JsonStore(tmp_path / "store.json")

        store.set_many({
            "uo.UO-001.avancement": 0.8,
            "uo.UO-001.heures_realisees": 25.0,
            "uo.UO-002.avancement": 0.5,
            "uo.UO-002.heures_realisees": 24.0,
            "uo.UO-003.avancement": 0.3,
            "uo.UO-003.heures_realisees": 12.0,
        })

        path = generate_dashboard_metier(acteur, ALL_UOS, store, output_dir=tmp_path)
        wb = load_workbook(path, data_only=True)
        ws = wb["Synthèse"]

        for row in range(6, 20):
            if ws.cell(row=row, column=1).value == "UO-001":
                avanc = ws.cell(row=row, column=7).value
                assert avanc == pytest.approx(0.8, abs=0.01), \
                    f"Attendu 0.8, obtenu {avanc}"
                break
        else:
            pytest.fail("UO-001 non trouvé dans Synthèse")

    def test_alerte_generee_apres_depassement(self, tmp_path):
        """Simule un dépassement : heures_realisees > charge → alerte dans Alertes."""
        from src.generators.dashboard_metier_generator import generate_dashboard_metier
        acteur = _make_pilote_metier()
        store = JsonStore(tmp_path / "store.json")

        store.set("uo.UO-003.heures_realisees", 55.0)

        path = generate_dashboard_metier(acteur, ALL_UOS, store, output_dir=tmp_path)
        wb = load_workbook(path, data_only=True)
        ws = wb["Alertes"]

        found = any(
            ws.cell(row=r, column=2).value == "UO-003"
            for r in range(3, 20)
        )
        assert found, "UO-003 devrait apparaître dans les Alertes après dépassement"

    def test_store_vide_ne_plante_pas(self, tmp_path):
        """Store vide → dashboard généré avec des 0 (pas d'exception)."""
        from src.generators.dashboard_metier_generator import generate_dashboard_metier
        acteur = _make_pilote_metier()
        store = JsonStore(tmp_path / "store.json")

        path = generate_dashboard_metier(acteur, ALL_UOS, store, output_dir=tmp_path)
        assert path.exists()
        wb = load_workbook(path, data_only=True)
        ws = wb["Synthèse"]
        for row in range(6, 20):
            if ws.cell(row=row, column=1).value == "UO-001":
                avanc = ws.cell(row=row, column=7).value
                assert avanc == pytest.approx(0.0, abs=0.01)
                break
