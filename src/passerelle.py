"""
Parser et exécuteur de la feuille _Manifeste (legacy : _Passerelle).

Responsabilités :
- Lire la feuille _Manifeste d'un fichier Excel
- Parser les règles (ReglePasserelle)
- Exécuter les pulls (store → fichier Excel)
- Exécuter les pushes (fichier Excel → store)
- Résoudre les tableaux (nommé natif ou scan de headers)
- Évaluer les formules COMPUTED
- Détecter et gérer le versioning (-MOD)
"""
import re
import statistics
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src import store as Store
from src.models import (
    DirectionPasserelle, Passerelle, ReglePasserelle,
    ScopePasserelle, TypePasserelle,
)

MANIFESTE_SHEET  = "_Manifeste"
PASSERELLE_SHEET = MANIFESTE_SHEET   # alias rétro-compat — ne pas supprimer
VERSION_CELL = "A1"
HEADER_ROW = 2
DATA_START_ROW = 3

COLONNES_PASSERELLE = [
    "TYPE", "SCOPE", "NOM_GLOBAL", "NOM_LOCAL",
    "FEUILLE", "TABLEAU", "CLE", "COLONNES",
    "CELLULE", "DIRECTION", "FORMULE",
]


# ─── Lecture de la passerelle ─────────────────────────────────────────────────

def lire_passerelle(ws: Worksheet) -> Passerelle:
    """Parse la feuille _Passerelle et retourne un objet Passerelle."""
    version_raw = str(ws[VERSION_CELL].value or "").strip()
    if not (version_raw.startswith("MANIFESTE_V=") or version_raw.startswith("PASSERELLE_V=")):
        raise ValueError(f"En-tête de version manquant ou invalide : '{version_raw}'")
    version = version_raw.replace("MANIFESTE_V=", "").replace("PASSERELLE_V=", "").strip()

    # Lire les en-têtes (ligne 2) pour mapper col → champ
    col_index: Dict[str, int] = {}
    for col in range(1, 20):
        val = str(ws.cell(row=HEADER_ROW, column=col).value or "").strip().upper()
        if val in COLONNES_PASSERELLE:
            col_index[val] = col

    regles: List[ReglePasserelle] = []
    row = DATA_START_ROW
    while True:
        type_val = ws.cell(row=row, column=col_index.get("TYPE", 1)).value
        if type_val is None:
            break
        type_val = str(type_val).strip().upper()
        if not type_val:
            break

        def _get(col_name: str) -> str:
            idx = col_index.get(col_name)
            if idx is None:
                return ""
            v = ws.cell(row=row, column=idx).value
            return str(v).strip() if v is not None else ""

        try:
            regle = ReglePasserelle(
                type=TypePasserelle(type_val),
                scope=ScopePasserelle(_get("SCOPE").upper() or "GLOBAL"),
                nom_global=_get("NOM_GLOBAL"),
                nom_local=_get("NOM_LOCAL"),
                feuille=_get("FEUILLE"),
                tableau=_get("TABLEAU"),
                cle=_get("CLE"),
                colonnes=_get("COLONNES"),
                cellule=_get("CELLULE"),
                direction=DirectionPasserelle(_get("DIRECTION").lower()) if _get("DIRECTION") else DirectionPasserelle.PULL,
                formule=_get("FORMULE"),
            )
            regles.append(regle)
        except (ValueError, KeyError) as e:
            # Règle invalide → loguée mais on continue
            print(f"  [WARN] Passerelle ligne {row} ignorée : {e}")
        row += 1

    return Passerelle(version=version, regles=regles)


def ecrire_version(ws: Worksheet, version: str) -> None:
    ws[VERSION_CELL] = f"MANIFESTE_V={version}"


# ─── Résolution de tableaux ───────────────────────────────────────────────────

def _trouver_tableau_nomme(wb, sheet_name: str, table_name: str) -> Optional[Tuple[Worksheet, int, int]]:
    """Cherche un tableau Excel nommé. Retourne (ws, header_row, first_data_row) ou None."""
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    for tbl in ws.tables.values():
        if tbl.name.lower() == table_name.lower():
            ref = tbl.ref  # ex. "A2:F10"
            from openpyxl.utils.cell import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(ref)
            return ws, min_row, min_row + 1
    return None


def _scanner_headers(ws: Worksheet, colonnes_attendues: List[str]) -> Optional[Tuple[int, Dict[str, int]]]:
    """
    Scan la feuille pour trouver une ligne dont les cellules correspondent
    aux colonnes attendues. Retourne (header_row, {col_name: col_index}).
    """
    colonnes_lower = [c.lower().strip() for c in colonnes_attendues]
    for row in range(1, min(20, ws.max_row + 1)):
        row_vals = [str(ws.cell(row=row, column=c).value or "").lower().strip()
                    for c in range(1, ws.max_column + 1)]
        matches = sum(1 for c in colonnes_lower if c in row_vals)
        if matches >= max(1, len(colonnes_lower) // 2):
            col_map = {}
            for col_name in colonnes_attendues:
                for idx, val in enumerate(row_vals, 1):
                    if val == col_name.lower().strip():
                        col_map[col_name] = idx
                        break
            return row, col_map
    return None


def _lire_tableau(wb, sheet_name: str, table_name: str, colonnes: List[str]) -> List[Dict[str, Any]]:
    """
    Lit un tableau depuis un fichier Excel.
    Essaie d'abord le tableau nommé natif, puis scan de headers.
    """
    if sheet_name not in wb.sheetnames:
        raise LookupError(f"Feuille '{sheet_name}' introuvable")

    ws = wb[sheet_name]
    header_row = None
    col_map: Dict[str, int] = {}

    # Tentative 1 : tableau nommé natif
    if table_name:
        result = _trouver_tableau_nomme(wb, sheet_name, table_name)
        if result:
            ws_t, h_row, _ = result
            header_row = h_row
            for col_name in colonnes:
                for col_idx in range(1, ws_t.max_column + 1):
                    cell_val = str(ws_t.cell(row=h_row, column=col_idx).value or "").strip()
                    if cell_val.lower() == col_name.lower():
                        col_map[col_name] = col_idx
                        break

    # Tentative 2 : scan des headers
    if not col_map:
        scan = _scanner_headers(ws, colonnes)
        if scan:
            header_row, col_map = scan

    if header_row is None:
        raise LookupError(
            f"Tableau '{table_name}' et headers {colonnes} introuvables dans '{sheet_name}'"
        )

    # Lire les données
    rows = []
    for row in range(header_row + 1, ws.max_row + 1):
        # Arrêter sur ligne entièrement vide
        if all(ws.cell(row=row, column=c).value is None for c in col_map.values()):
            break
        record = {}
        for col_name, col_idx in col_map.items():
            record[col_name] = ws.cell(row=row, column=col_idx).value
        rows.append(record)
    return rows


def _lire_cellule(wb, sheet_name: str, cellule: str) -> Any:
    if sheet_name not in wb.sheetnames:
        raise LookupError(f"Feuille '{sheet_name}' introuvable")
    return wb[sheet_name][cellule].value


def _ecrire_cellule(wb, sheet_name: str, cellule: str, valeur: Any) -> None:
    if sheet_name not in wb.sheetnames:
        raise LookupError(f"Feuille '{sheet_name}' introuvable")
    ws = wb[sheet_name]
    cell = ws[cellule]
    # Si c'est une cellule fusionnée non-maître, trouver la cellule maître
    from openpyxl.cell.cell import MergedCell
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
            col_letter, row_num = coordinate_from_string(cellule)
            col_idx = column_index_from_string(col_letter)
            if (merged_range.min_row <= row_num <= merged_range.max_row and
                    merged_range.min_col <= col_idx <= merged_range.max_col):
                ws.cell(row=merged_range.min_row, column=merged_range.min_col, value=valeur)
                return
        return  # cellule fusionnée non trouvée dans les ranges → skip
    cell.value = valeur


# ─── Évaluation des formules COMPUTED ────────────────────────────────────────

def _evaluer_formule(formule: str, store_snapshot: Dict[str, Any]) -> Any:
    """
    Évalue une formule COMPUTED Python.
    Fonctions supportées :
      MEAN(var)
      MEAN_WEIGHTED(var_values, var_weights)
      SUM(var)
      COUNT_IF(var, "valeur")
    """
    formule = formule.strip()

    def _resolve_list(var_name: str) -> List[Any]:
        """Résout une variable de store en liste de valeurs scalaires."""
        val = store_snapshot.get(var_name)
        if val is None:
            return []
        if isinstance(val, list):
            # Si c'est une liste de dicts, on ne peut pas agréger directement
            # → la var doit être une colonne extraite, ex: uo.UO-001.activites.avancement
            return [v for v in val if v is not None]
        return [val]

    # MEAN(var)
    m = re.fullmatch(r"MEAN\(([^,)]+)\)", formule)
    if m:
        vals = [v for v in _resolve_list(m.group(1).strip()) if isinstance(v, (int, float))]
        return statistics.mean(vals) if vals else 0.0

    # MEAN_WEIGHTED(var_values, var_weights)
    m = re.fullmatch(r"MEAN_WEIGHTED\(([^,)]+),\s*([^)]+)\)", formule)
    if m:
        vals = [v for v in _resolve_list(m.group(1).strip()) if isinstance(v, (int, float))]
        weights = [w for w in _resolve_list(m.group(2).strip()) if isinstance(w, (int, float))]
        if not vals or not weights or len(vals) != len(weights):
            return 0.0
        total_w = sum(weights)
        return sum(v * w for v, w in zip(vals, weights)) / total_w if total_w else 0.0

    # SUM(var)
    m = re.fullmatch(r"SUM\(([^)]+)\)", formule)
    if m:
        vals = [v for v in _resolve_list(m.group(1).strip()) if isinstance(v, (int, float))]
        return sum(vals)

    # COUNT_IF(var, "valeur")
    m = re.fullmatch(r'COUNT_IF\(([^,)]+),\s*"([^"]*)"\)', formule)
    if m:
        vals = _resolve_list(m.group(1).strip())
        cible = m.group(2)
        return sum(1 for v in vals if str(v) == cible)

    raise ValueError(f"Formule COMPUTED non reconnue : '{formule}'")


# ─── Exécution de la passerelle ───────────────────────────────────────────────

def executer_passerelle(
    fichier_path: Path,
    uo_id: str,
    log: List[str],
) -> Dict[str, Any]:
    """
    Exécute la passerelle d'un fichier Excel :
    1. Pull (store → fichier)
    2. Push (fichier → store)
    3. Computed (calcul → store)
    4. Gestion du versioning (-MOD)

    Retourne un dict des variables pushées (pour debug/rapport).
    """
    if not fichier_path.exists():
        log.append(f"[ERREUR] Fichier introuvable : {fichier_path}")
        return {}

    try:
        wb = load_workbook(fichier_path)
    except Exception as e:
        log.append(f"[ERREUR] Impossible d'ouvrir {fichier_path.name} : {e}")
        return {}

    if PASSERELLE_SHEET not in wb.sheetnames:
        log.append(f"[INFO] Pas de feuille {PASSERELLE_SHEET} dans {fichier_path.name} — skipped")
        return {}

    ws_pass = wb[PASSERELLE_SHEET]

    try:
        passerelle = lire_passerelle(ws_pass)
    except ValueError as e:
        log.append(f"[ERREUR] {fichier_path.name} — passerelle invalide : {e}")
        return {}

    # Gestion du versioning
    if passerelle.est_modifiee:
        nouvelle_version = str(int(passerelle.version_num) + 1)
        log.append(
            f"[VERSION] {fichier_path.name} — passerelle modifiée "
            f"v{passerelle.version_num} → v{nouvelle_version}"
        )
        ecrire_version(ws_pass, nouvelle_version)

    # Variables locales (REF)
    locals_: Dict[str, Any] = {}

    store_snapshot = Store.get_all()
    pushed: Dict[str, Any] = {}

    # ── Phase 1 : PULL (store → cellules Excel) ───────────────────────────────
    for regle in passerelle.regles:
        if regle.direction != DirectionPasserelle.PULL:
            continue
        if regle.scope == ScopePasserelle.LOCAL:
            # LOCAL : mémoriser la valeur de la cellule pour les REF
            try:
                val = _lire_cellule(wb, regle.feuille, regle.cellule)
                locals_[regle.nom_local] = val
            except LookupError as e:
                log.append(f"  [WARN] LOCAL '{regle.nom_local}' : {e}")
            continue

        # Construire la clé globale avec le prefixe uo_id
        cle_globale = f"{regle.nom_global}.{uo_id}" if uo_id else regle.nom_global
        valeur = store_snapshot.get(cle_globale) or store_snapshot.get(regle.nom_global)
        if valeur is None:
            log.append(f"  [INFO] Variable '{regle.nom_global}' absente du store — pull ignoré")
            continue

        try:
            if regle.type in (TypePasserelle.CELL, TypePasserelle.CELL_NUM,
                               TypePasserelle.CELL_DATE, TypePasserelle.CELL_PCT):
                _ecrire_cellule(wb, regle.feuille, regle.cellule, valeur)
            elif regle.type in (TypePasserelle.TABLE_FULL, TypePasserelle.TABLE_COL):
                # Pour l'instant : log uniquement (écriture de tableau = phase ultérieure)
                log.append(f"  [INFO] Pull TABLE vers '{regle.feuille}' — non implémenté en v1")
        except LookupError as e:
            log.append(f"  [WARN] Pull '{regle.nom_global}' → {regle.feuille}!{regle.cellule} : {e}")

    # ── Résolution des REF (variables locales) ────────────────────────────────
    for regle in passerelle.regles:
        if regle.type == TypePasserelle.REF and regle.nom_local in locals_:
            try:
                _ecrire_cellule(wb, regle.feuille, regle.cellule, locals_[regle.nom_local])
            except LookupError as e:
                log.append(f"  [WARN] REF '{regle.nom_local}' → {regle.feuille}!{regle.cellule} : {e}")

    # ── Phase 2 : PUSH (cellules Excel → store) ───────────────────────────────
    for regle in passerelle.regles:
        if regle.direction != DirectionPasserelle.PUSH:
            continue
        if regle.type == TypePasserelle.COMPUTED:
            continue  # traité en phase 3

        cle = f"{regle.nom_global}.{uo_id}" if uo_id else regle.nom_global

        try:
            if regle.type in (TypePasserelle.CELL, TypePasserelle.CELL_NUM,
                               TypePasserelle.CELL_DATE, TypePasserelle.CELL_PCT):
                val = _lire_cellule(wb, regle.feuille, regle.cellule)
                pushed[cle] = val

            elif regle.type in (TypePasserelle.TABLE_FULL, TypePasserelle.TABLE_COL):
                colonnes = [c.strip() for c in regle.colonnes.split(",") if c.strip()]
                rows = _lire_tableau(wb, regle.feuille, regle.tableau, colonnes)
                # Si TABLE_COL, extraire uniquement la colonne
                if regle.type == TypePasserelle.TABLE_COL and colonnes:
                    pushed[cle] = [r.get(colonnes[0]) for r in rows]
                else:
                    pushed[cle] = rows
                # Indexer aussi les colonnes individuelles pour COMPUTED
                for col in colonnes:
                    pushed[f"{cle}.{col}"] = [r.get(col) for r in rows]

        except LookupError as e:
            log.append(f"  [WARN] Push '{regle.nom_global}' : {e}")

    # ── Phase 3 : COMPUTED ────────────────────────────────────────────────────
    # Mise à jour du snapshot avec ce qu'on vient de pusher
    store_snapshot.update(pushed)

    for regle in passerelle.regles:
        if regle.type != TypePasserelle.COMPUTED:
            continue
        cle = f"{regle.nom_global}.{uo_id}" if uo_id else regle.nom_global
        try:
            val = _evaluer_formule(regle.formule, store_snapshot)
            pushed[cle] = val
            store_snapshot[cle] = val
            log.append(f"  [COMPUTED] {cle} = {val}")
        except ValueError as e:
            log.append(f"  [WARN] COMPUTED '{regle.nom_global}' : {e}")

    # ── Sauvegardes ───────────────────────────────────────────────────────────
    if pushed:
        Store.set_many(pushed)

    wb.save(fichier_path)
    return pushed
