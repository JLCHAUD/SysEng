"""
Tests pour les nouveaux modules (M04, M07 étendu, M08, M09, M10, M11, M13)
"""
import json
import shutil
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _dummy_resultats():
    return [
        {"id": "UO-001", "chemin": "UOs/UO-001.xlsx", "type": "uo_instance",
         "statut": "ok", "log": ["[OK] sync OK"], "timestamp": "2026-04-24T10:00:00"},
        {"id": "UO-002", "chemin": "UOs/UO-002.xlsx", "type": "uo_instance",
         "statut": "erreur", "log": ["[ERR] echec"], "timestamp": "2026-04-24T10:01:00"},
    ]


# ─── M04 — Historique ─────────────────────────────────────────────────────────

class TestHistory:
    def test_save_run_history(self, tmp_path):
        from src.history import save_run_history
        debut = datetime(2026, 4, 24, 10, 0, 0)
        fin   = datetime(2026, 4, 24, 10, 0, 5)
        # Patch HISTORY_DIR temporairement
        import src.history as H
        orig = H.HISTORY_DIR
        H.HISTORY_DIR = tmp_path / "history"
        try:
            p = save_run_history(_dummy_resultats(), debut, fin)
            assert p.exists()
            data = json.loads(p.read_text(encoding="utf-8"))
            assert data["nb_total"] == 2
            assert data["nb_ok"] == 1
            assert data["nb_erreur"] == 1
            assert data["duree_s"] == 5.0
        finally:
            H.HISTORY_DIR = orig

    def test_save_store_snapshot(self, tmp_path):
        from src.history import save_store_snapshot
        import src.history as H
        orig = H.SNAPSHOT_DIR
        H.SNAPSHOT_DIR = tmp_path / "snaps"
        # Store source bidon
        store_path = tmp_path / "store.json"
        store_path.write_text(
            json.dumps({"variables": {"x": 1}, "derniere_maj": "2026-04-24"}),
            encoding="utf-8"
        )
        try:
            p = save_store_snapshot(store_path=store_path)
            assert p.exists()
            data = json.loads(p.read_text(encoding="utf-8"))
            assert data["variables"]["x"] == 1
        finally:
            H.SNAPSHOT_DIR = orig

    def test_list_runs(self, tmp_path):
        from src.history import save_run_history, list_runs
        import src.history as H
        orig = H.HISTORY_DIR
        H.HISTORY_DIR = tmp_path / "history"
        try:
            d = datetime(2026, 4, 24, 10, 0, 0)
            save_run_history(_dummy_resultats(), d, d)
            runs = list_runs(tmp_path / "history")
            assert len(runs) == 1
        finally:
            H.HISTORY_DIR = orig

    def test_compare_snapshots(self, tmp_path):
        from src.history import compare_snapshots
        a = tmp_path / "store_a.json"
        b = tmp_path / "store_b.json"
        a.write_text(json.dumps({"variables": {"x": 1, "y": 2}}), encoding="utf-8")
        b.write_text(json.dumps({"variables": {"x": 10, "z": 3}}), encoding="utf-8")
        diff = compare_snapshots(a, b)
        assert "x" in diff["modifications"]
        assert diff["modifications"]["x"] == {"avant": 1, "apres": 10}
        assert "y" in diff["suppressions"]
        assert "z" in diff["ajouts"]

    def test_purge_old_files(self, tmp_path):
        from src.history import save_run_history, purge_old_files, list_runs
        import src.history as H
        orig = H.HISTORY_DIR
        hist_dir = tmp_path / "history"
        H.HISTORY_DIR = hist_dir
        try:
            # Créer 5 runs
            for i in range(5):
                d = datetime(2026, 4, 24, 10, i, 0)
                save_run_history(_dummy_resultats(), d, d)
            assert len(list_runs(hist_dir)) == 5
            # Purger pour en garder max 3
            supprime = purge_old_files(max_runs=3, max_snapshots=99, history_dir=hist_dir)
            assert supprime == 2
            assert len(list_runs(hist_dir)) == 3
        finally:
            H.HISTORY_DIR = orig

    def test_history_of_key(self, tmp_path):
        from src.history import save_store_snapshot, history_of_key
        import src.history as H
        orig = H.SNAPSHOT_DIR
        snap_dir = tmp_path / "snaps"
        H.SNAPSHOT_DIR = snap_dir
        store_path = tmp_path / "store.json"
        try:
            store_path.write_text(
                json.dumps({"variables": {"uo.x": 10}, "derniere_maj": "2026-04-24T09:00:00"}),
                encoding="utf-8"
            )
            save_store_snapshot(store_path=store_path)
            store_path.write_text(
                json.dumps({"variables": {"uo.x": 20}, "derniere_maj": "2026-04-24T10:00:00"}),
                encoding="utf-8"
            )
            save_store_snapshot(store_path=store_path)
            hist = history_of_key("uo.x", snapshot_dir=snap_dir)
            vals = [v for _, v in hist]
            assert 10 in vals
            assert 20 in vals
        finally:
            H.SNAPSHOT_DIR = orig


# ─── M07 — Nouvelles règles VALIDATE ─────────────────────────────────────────

class TestValidateReglesEtendues:
    """Tests des règles ajoutées : NOT_EMPTY, MAX_LENGTH, MIN_LENGTH, MATCHES, MAX, MIN."""

    def _run(self, rule, values):
        from src.executor import _validate_rule
        return _validate_rule(rule, values)

    def test_not_empty_ok(self):
        assert self._run("NOT_EMPTY", ["hello", "world"]) == []

    def test_not_empty_vide(self):
        violations = self._run("NOT_EMPTY", ["", None, "ok"])
        assert len(violations) == 1
        assert "2" in violations[0]   # 2 valeurs vides

    def test_max_length_ok(self):
        assert self._run("MAX_LENGTH(5)", ["abc", "de"]) == []

    def test_max_length_violation(self):
        v = self._run("MAX_LENGTH(3)", ["toolong", "ok"])
        assert len(v) == 1

    def test_min_length_ok(self):
        assert self._run("MIN_LENGTH(2)", ["ab", "abc"]) == []

    def test_min_length_violation(self):
        v = self._run("MIN_LENGTH(3)", ["ab", "a"])
        assert len(v) == 1

    def test_matches_ok(self):
        assert self._run("MATCHES(^[A-Z]{2}-\\d+$)", ["UO-001", "AB-123"]) == []

    def test_matches_violation(self):
        v = self._run("MATCHES(^\\d+$)", ["123", "abc"])
        assert len(v) == 1

    def test_matches_regex_invalide(self):
        v = self._run("MATCHES([invalid)", ["x"])
        assert "invalide" in v[0]

    def test_max_ok(self):
        assert self._run("MAX(100)", [50, 99, 100]) == []

    def test_max_violation(self):
        v = self._run("MAX(10)", [5, 15, 8])
        assert len(v) == 1

    def test_min_ok(self):
        assert self._run("MIN(0)", [0, 1, 5]) == []

    def test_min_violation(self):
        v = self._run("MIN(5)", [3, 10])
        assert len(v) == 1


# ─── M08 — Résilience ─────────────────────────────────────────────────────────

class TestAtomicStore:
    def test_ecriture_atomique_cree_pas_de_tmp(self, tmp_path):
        """Après un set(), aucun fichier .tmp ne doit traîner."""
        from src.store import JsonStore
        store = JsonStore(tmp_path / "store.json")
        store.set("x.y", 42)
        tmp_files = list(tmp_path.glob("*.tmp"))
        assert tmp_files == []

    def test_ecriture_atomique_contenu_correct(self, tmp_path):
        from src.store import JsonStore
        store = JsonStore(tmp_path / "store.json")
        store.set("a.b", "hello")
        store2 = JsonStore(tmp_path / "store.json")
        assert store2.get("a.b") == "hello"


class TestLogSheet:
    def test_log_sheet_cree(self, tmp_path):
        """_write_log_sheet crée la feuille _Log si absente."""
        from src.executor import _write_log_sheet, ExecutionResult, LOG_SHEET_NAME
        wb = Workbook()
        result = ExecutionResult()
        result.errors.append("ERREUR test")
        _write_log_sheet(wb, result)
        assert LOG_SHEET_NAME in wb.sheetnames

    def test_log_sheet_contient_erreurs(self, tmp_path):
        from src.executor import _write_log_sheet, ExecutionResult, LOG_SHEET_NAME
        wb = Workbook()
        result = ExecutionResult()
        result.errors = ["err1", "err2"]
        _write_log_sheet(wb, result)
        ws = wb[LOG_SHEET_NAME]
        # Ligne 1 = header, lignes 2+ = données
        assert ws.max_row >= 3   # header + 2 erreurs + 1 ligne INFO

    def test_safe_save_cree_backup(self, tmp_path):
        from src.executor import _safe_save
        xlsx = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.save(str(xlsx))
        wb2 = Workbook()
        _safe_save(wb2, xlsx)
        bak = xlsx.with_suffix(".bak")
        assert bak.exists()


# ─── M10 — Sécurité ───────────────────────────────────────────────────────────

class TestSecurity:
    def test_cle_valide(self):
        from src.security import is_valid_store_key
        assert is_valid_store_key("uo.UO-001.avancement") is True
        assert is_valid_store_key("ref.types_uo") is True
        assert is_valid_store_key("projet.acteurs") is True

    def test_cle_invalide_espaces(self):
        from src.security import is_valid_store_key
        assert is_valid_store_key("uo. UO-001") is False

    def test_cle_invalide_traversal(self):
        from src.security import is_valid_store_key
        assert is_valid_store_key("uo/../secret") is False

    def test_cle_invalide_vide(self):
        from src.security import is_valid_store_key
        assert is_valid_store_key("") is False

    def test_cle_invalide_trop_longue(self):
        from src.security import is_valid_store_key
        assert is_valid_store_key("x" * 201) is False

    def test_cle_invalide_slash(self):
        from src.security import is_valid_store_key
        assert is_valid_store_key("uo/UO-001") is False

    def test_validate_store_key_leve_exception(self):
        from src.security import validate_store_key, StoreKeyError
        with pytest.raises(StoreKeyError):
            validate_store_key("uo/../secret")

    def test_namespace_ok(self):
        from src.security import validate_namespace
        assert validate_namespace("uo.UO-001.av", ["uo.", "ref."]) is True

    def test_namespace_interdit(self):
        from src.security import validate_namespace
        assert validate_namespace("secret.key", ["uo.", "ref."]) is False

    def test_hash_fichier(self, tmp_path):
        from src.security import compute_manifest_hash, check_manifest_integrity
        f = tmp_path / "test.bin"
        f.write_bytes(b"contenu test")
        h = compute_manifest_hash(f)
        assert len(h) == 64
        assert check_manifest_integrity(f, h) is True
        assert check_manifest_integrity(f, "deadbeef" * 8) is False

    def test_store_rejette_cle_invalide(self, tmp_path):
        from src.store import JsonStore
        from src.security import StoreKeyError
        store = JsonStore(tmp_path / "store.json")
        with pytest.raises(StoreKeyError):
            store.set("../etc/passwd", "hack")


# ─── M09 — NOTIFY (parsing) ───────────────────────────────────────────────────

class TestNotifyParsing:
    def _parse(self, line):
        from src.parser import _parse_notify
        return _parse_notify(line)

    def test_notify_log(self):
        n = self._parse('NOTIFY log "Avancement faible"')
        assert n is not None
        assert n.channel == "log"
        assert n.message == "Avancement faible"
        assert n.condition == ""

    def test_notify_log_avec_condition(self):
        n = self._parse('NOTIFY log "Alerte" IF $av < 10')
        assert n is not None
        assert n.condition == "$av < 10"

    def test_notify_email_avec_to(self):
        n = self._parse('NOTIFY email "Message" TO admin@example.com')
        assert n is not None
        assert n.channel == "email"
        assert n.target == "admin@example.com"

    def test_notify_webhook(self):
        n = self._parse('NOTIFY webhook "alert" TO https://hooks.example.com/xyz')
        assert n is not None
        assert n.channel == "webhook"
        assert "hooks.example.com" in n.target

    def test_notify_invalide(self):
        n = self._parse("NOTIFY invalide")
        assert n is None


class TestNotifyExecution:
    def test_notify_log_ajoute_warning(self):
        from src.executor import execute_notifies, ExecutionResult
        from src.parser import NotifyNode, PasserelleAST
        ast = PasserelleAST()
        ast.notifies.append(NotifyNode(channel="log", message="Test notification"))
        result = ExecutionResult()
        execute_notifies(ast, {}, result)
        assert any("NOTIFY" in w for w in result.warnings)

    def test_notify_condition_non_satisfaite(self):
        from src.executor import execute_notifies, ExecutionResult
        from src.parser import NotifyNode, PasserelleAST
        ast = PasserelleAST()
        ast.notifies.append(NotifyNode(
            channel="log", message="Alerte", condition="1 > 100"
        ))
        result = ExecutionResult()
        execute_notifies(ast, {}, result)
        assert len(result.warnings) == 0
        assert any("condition" in s for s in result.skipped)

    def test_notify_email_sans_smtp(self):
        """Sans SMTP_HOST, l'email est loggué comme warning non bloquant."""
        import os
        from src.executor import execute_notifies, ExecutionResult
        from src.parser import NotifyNode, PasserelleAST
        os.environ.pop("SMTP_HOST", None)
        ast = PasserelleAST()
        ast.notifies.append(NotifyNode(
            channel="email", message="Hello", target="x@y.com"
        ))
        result = ExecutionResult()
        execute_notifies(ast, {}, result)
        assert any("non configure" in w for w in result.warnings)


# ─── M13 — Documentation HTML ─────────────────────────────────────────────────

class TestDocGenerator:
    def test_generate_html_cree_fichier(self, tmp_path):
        from src.doc_generator import generate_html_doc
        out = generate_html_doc(output_dir=tmp_path / "doc")
        assert out.exists()
        content = out.read_text(encoding="utf-8")
        assert "ExoSync" in content
        assert "<html" in content

    def test_html_contient_stats(self, tmp_path):
        from src.doc_generator import generate_html_doc
        out = generate_html_doc(output_dir=tmp_path / "doc")
        content = out.read_text(encoding="utf-8")
        assert "Fichiers" in content
        assert "Arcs" in content

    def test_html_avec_ecosystem_custom(self, tmp_path):
        """Génère la doc depuis un ecosystem vide dans tmp_path."""
        from src.doc_generator import generate_html_doc
        eco_path = tmp_path / "ecosystem.json"
        eco_path.write_text(
            json.dumps({
                "version": "2.0",
                "files": {},
                "edges": [],
                "tables": {},
                "variables": {},
                "last_scan": "2026-04-24",
            }),
            encoding="utf-8"
        )
        out = generate_html_doc(
            output_dir=tmp_path / "doc",
            ecosystem_path=eco_path,
        )
        assert out.exists()
        content = out.read_text(encoding="utf-8")
        assert "Aucun fichier" in content


# ─── M11 — CLI doctor/history ─────────────────────────────────────────────────

class TestCLIDoctor:
    def test_doctor_retourne_code(self):
        from src.cli import main
        code = main(["doctor"])
        assert isinstance(code, int)
        assert code in (0, 1, 2)

    def test_history_retourne_code(self):
        from src.cli import main
        code = main(["history"])
        assert code == 0

    def test_history_compare(self):
        from src.cli import main
        # Pas de snapshots → doit retourner 1 avec message d'erreur
        import src.history as H
        from pathlib import Path
        orig = H.SNAPSHOT_DIR
        H.SNAPSHOT_DIR = Path("/nonexistent_snapshot_dir_xyz")
        try:
            code = main(["history", "--compare"])
            assert code == 1
        finally:
            H.SNAPSHOT_DIR = orig

    def test_doc_command(self, tmp_path):
        from src.cli import main
        code = main(["doc", "--output", str(tmp_path)])
        assert code == 0
        assert len(list(tmp_path.glob("*.html"))) == 1
