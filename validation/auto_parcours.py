"""Execution automatique du Parcours de Validation (marches 0 a 5).
Reproduit fidelement ce que ferait le testeur humain, y compris les
etats intermediaires du recepteur (marche 2 : PULL seul ; marche 3 :
+ COMPUTE/BIND ; marche 4 : + 2e PULL APPEND_NEW)."""
import io
import subprocess
import sys
from contextlib import redirect_stdout
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo

VDIR = ROOT / "validation"
RESULTS = []


def check(marche, label, ok, detail=""):
    RESULTS.append((marche, label, ok, detail))
    print("  [%s] M%s %s %s" % ("OK " if ok else "KO!", marche, label,
                                ("- " + detail) if detail else ""))


def run(cmd):
    return subprocess.run(cmd, cwd=str(ROOT), capture_output=True, text=True)


def store_clear():
    run([sys.executable, "-m", "src", "store", "clear"])


def valider(relpath):
    return run([sys.executable, "scripts/valider_un.py", relpath])


def add_table(ws, name, headers, rows, r0=1, c0=1):
    for ci, h in enumerate(headers, c0):
        ws.cell(row=r0, column=ci, value=h)
    for ri, rd in enumerate(rows, r0 + 1):
        for ci, h in enumerate(headers, c0):
            ws.cell(row=ri, column=ci, value=rd.get(h, ""))
    ref = f"{get_column_letter(c0)}{r0}:{get_column_letter(c0+len(headers)-1)}{r0+len(rows)}"
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(t)


def manifeste(ws, lines):
    ws["A1"] = "MANIFESTE_V=1"
    ws["A2"] = "instruction"
    for i, l in enumerate(lines, 3):
        ws.cell(row=i, column=1, value=l)


def build_source(path, table, file_id, key, rows):
    wb = Workbook(); ws = wb.active; ws.title = "Activites"
    add_table(ws, table, ["id", "libelle", "avancement"], rows)
    manifeste(wb.create_sheet("_Manifeste"), [
        "FILE_TYPE: uo_instance", f"FILE_ID: {file_id}", "",
        f"DEF $act = GET_TABLE(Activites, {table})",
        f"PUSH $act -> {key}"])
    wb.save(path)


def build_recepteur(path, etape):
    """etape 2 : PULL seul / etape 3 : + dashboard / etape 4 : + 2e PULL."""
    wb = Workbook(); ws = wb.active; ws.title = "Donnees"
    add_table(ws, "tbl_recue", ["id", "libelle", "avancement"],
              [{"id": "-", "libelle": "(en attente de synchro)", "avancement": 0}])
    mxl = ["FILE_TYPE: cockpit", "FILE_ID: TEST-RECEPTEUR", "",
           "PULL test.source.activites  -> FILL_TABLE(Donnees, tbl_recue) MODE=OVERWRITE"]
    if etape >= 4:
        mxl.append("PULL test.source2.activites -> FILL_TABLE(Donnees, tbl_recue) MODE=APPEND_NEW KEY=id")
    if etape >= 3:
        dash = wb.create_sheet("Dashboard")
        dash["E3"] = "Avancement moyen"
        wb.defined_names["avancement"] = DefinedName("avancement", attr_text="'Dashboard'!$F$3")
        mxl += ["", "DEF $recu = GET_TABLE(Donnees, tbl_recue)",
                "DEF $moy = COMPUTE(AVG($recu.avancement))", "",
                "BIND $moy -> Dashboard.avancement"]
    manifeste(wb.create_sheet("_Manifeste"), mxl)
    wb.save(path)


def donnees_rows(path):
    ws = load_workbook(path)["Donnees"]
    return [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0] is not None]


def f3(path):
    return load_workbook(path)["Dashboard"]["F3"].value


# ════════ MARCHE 0 — outillage + demo ════════
print("\n== MARCHE 0 ==")
r = run([sys.executable, "scripts/generate_demo.py"])
check(0, "generate_demo.py", r.returncode == 0 and "7 fichiers generes" in r.stdout)
r = run([sys.executable, "scripts/sync_demo.py"])
check(0, "sync_demo.py sans erreur", r.returncode == 0 and "SANS erreur" in r.stdout)
v = load_workbook(ROOT / "output/demo/COCKPIT-ALICE.xlsx")["Dashboard"]["F5"].value
check(0, "COCKPIT-ALICE Dashboard F5 = 55 %", abs(v - 0.55) < 1e-9, f"lu : {v}")

# ════════ MARCHE 1 — un fichier pousse ════════
print("\n== MARCHE 1 ==")
store_clear()
build_source(VDIR / "m1_source.xlsx", "tbl_act", "TEST-SOURCE", "test.source.activites",
             [{"id": "A1", "libelle": "Analyse fonctionnelle", "avancement": 80},
              {"id": "A2", "libelle": "Redaction specification", "avancement": 60}])
r = valider("validation/m1_source.xlsx")
check(1, "PUSH = 1, zero erreur", r.returncode == 0 and "-> test.source.activites" in r.stdout)
check(1, "store : cle avec 2 lignes", "[2 lignes]" in r.stdout)

# ════════ MARCHE 2 — deux fichiers communiquent ════════
print("\n== MARCHE 2 ==")
store_clear()
valider("validation/m1_source.xlsx")
build_recepteur(VDIR / "m2_recepteur.xlsx", etape=2)
r = valider("validation/m2_recepteur.xlsx")
rows = donnees_rows(VDIR / "m2_recepteur.xlsx")
check(2, "PULL OK, zero erreur", r.returncode == 0)
check(2, "les 2 lignes de A sont dans B", len(rows) == 2 and rows[0][0] == "A1" and rows[1][0] == "A2",
      f"lignes : {[x[0] for x in rows]}")
check(2, "ligne 'en attente' disparue", all("attente" not in str(x[1]) for x in rows))

# ════════ MARCHE 3 — calcul + affichage ════════
print("\n== MARCHE 3 ==")
store_clear()
valider("validation/m1_source.xlsx")
build_recepteur(VDIR / "m2_recepteur.xlsx", etape=3)
r = valider("validation/m2_recepteur.xlsx")
val = f3(VDIR / "m2_recepteur.xlsx")
check(3, "BIND = 1, zero erreur", r.returncode == 0 and "= avancement" in r.stdout)
check(3, "Dashboard F3 = 70", val == 70, f"lu : {val}")

# ════════ MARCHE 4 — accumulation sans doublon ════════
print("\n== MARCHE 4 ==")
store_clear()
build_source(VDIR / "m4_source2.xlsx", "tbl_act2", "TEST-SOURCE-2", "test.source2.activites",
             [{"id": "A3", "libelle": "Plan de test", "avancement": 50},
              {"id": "A4", "libelle": "Execution recette", "avancement": 30}])
valider("validation/m1_source.xlsx")
valider("validation/m4_source2.xlsx")
build_recepteur(VDIR / "m2_recepteur.xlsx", etape=4)
r = valider("validation/m2_recepteur.xlsx")
rows = donnees_rows(VDIR / "m2_recepteur.xlsx")
val = f3(VDIR / "m2_recepteur.xlsx")
check(4, "4 lignes (A1..A4)", [x[0] for x in rows] == ["A1", "A2", "A3", "A4"],
      f"lignes : {[x[0] for x in rows]}")
check(4, "Dashboard F3 = 55", val == 55, f"lu : {val}")
valider("validation/m2_recepteur.xlsx")  # relance SANS vider le store
rows2 = donnees_rows(VDIR / "m2_recepteur.xlsx")
check(4, "idempotence : toujours 4 lignes", len(rows2) == 4, f"lignes : {len(rows2)}")

# ════════ MARCHE 5 — demo complete 3 niveaux ════════
print("\n== MARCHE 5 ==")
store_clear()
run([sys.executable, "scripts/generate_demo.py"])
r = run([sys.executable, "scripts/sync_demo.py"])
check(5, "sync demo sans erreur", r.returncode == 0 and "SANS erreur" in r.stdout)
ca = load_workbook(ROOT / "output/demo/COCKPIT-ALICE.xlsx")["Dashboard"]
cb = load_workbook(ROOT / "output/demo/COCKPIT-BRUNO.xlsx")["Dashboard"]
pi = load_workbook(ROOT / "output/demo/PILOTE.xlsx")["Dashboard"]
check(5, "ALICE : 70 / 40 / 55 %",
      (ca["F3"].value, ca["F4"].value, ca["F5"].value) == (0.7, 0.4, 0.55),
      f"lu : {ca['F3'].value}, {ca['F4'].value}, {ca['F5'].value}")
check(5, "BRUNO : 80 / 15 / 47.5 %",
      (cb["F3"].value, cb["F4"].value, cb["F5"].value) == (0.8, 0.15, 0.475),
      f"lu : {cb['F3'].value}, {cb['F4'].value}, {cb['F5'].value}")
check(5, "PILOTE : Alice 55 %, Bruno 47.5 %",
      (pi["F3"].value, pi["F4"].value) == (0.55, 0.475),
      f"lu : {pi['F3'].value}, {pi['F4'].value}")

# ════════ BILAN ════════
print("\n" + "=" * 60)
nok = sum(1 for *_, ok, _ in [(m, l, ok, d) for m, l, ok, d in RESULTS] if ok)
print("BILAN : %d/%d verifications OK" % (nok, len(RESULTS)))
for m, l, ok, d in RESULTS:
    if not ok:
        print("  ECHEC -> M%s %s (%s)" % (m, l, d))
print("=" * 60)
sys.exit(0 if nok == len(RESULTS) else 1)
