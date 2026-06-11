"""
build_uo_type.py — Construit le fichier UO type v1 du projet Train System.
===========================================================================

Genere L09U1-CFL2400-CLIM.xlsx : la premiere UO type ExoSync du projet
Alstom Train System, conforme aux decisions de conception du 2026-06-11
(vault Obsidian : 20-Projets/ExoSync/07-Roadmap-v2 + conversation C1).

Architecture du classeur :
  General            identite + heures vendues (saisi a l'instanciation)
  Description_Besoin cahier des charges (copie du Catalogue UO)
  Donnees_Entree     tbl_donnees_entree — suivi des inputs (DOORS...)
  Activites          tbl_activites — coeur du pilotage (10 colonnes)
  Livrables          tbl_livrables — livrables et maturites
  OIL                tbl_oil — Open Items List (points ouverts, 9 colonnes)
  KPI                sortie du moteur (BIND) — plages nommees
  Dashboard          presentation libre (lit l'onglet KPI)
  Planning           reserve a la visualisation planning (deplacee ici)
  Orga               organisation projet (statique v1)
  _Manifeste         MXL — genere ici

Contrainte moteur : les colonnes consommees par le Manifeste contiennent des
VALEURS (le moteur lit les formules comme texte). Les colonnes calculees
(heures_allouees, reste_a_faire) sont des formules Excel pour l'humain
uniquement — le Manifeste pondere par `poids`.

Usage :
    python projet_TrainSystem/build_uo_type.py
"""
import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = Path(__file__).parent / "L09U1-CFL2400-CLIM.xlsx"
NAVY = "1F4E79"
FONT = "Arial"


def style_header(ws, row, c1, c2):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def add_table(ws, name, headers, rows, widths=None, start_row=1):
    for ci, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=ci, value=h)
    for ri, rd in enumerate(rows, start_row + 1):
        for ci, h in enumerate(headers, 1):
            ws.cell(row=ri, column=ci, value=rd.get(h))
    style_header(ws, start_row, 1, len(headers))
    end = f"{get_column_letter(len(headers))}{start_row + len(rows)}"
    t = Table(displayName=name, ref=f"A{start_row}:{end}")
    t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(t)
    if widths:
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
    return t


def named(wb, name, sheet, cell):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${cell[0]}${cell[1:]}")


wb = Workbook()

# ── General ───────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "General"
ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 40
rows = [
    ("Pôle", "Train Système"),
    ("Lot", "LOT 9"),
    ("UO", "UO1"),
    ("Libellé", "Préparation et passage des IDR (Maturité M4)"),
    ("Projet", "CFL 2400"),
    ("Système", "Climatisation"),
    ("Ingénieur Système (SE)", "Jean Dujardin"),
    ("Heures vendues", 200),
    ("Date de création", "2026-06-11"),
]
ws["A1"] = "UO — L09U1-CFL2400-CLIM"
ws["A1"].font = Font(name=FONT, bold=True, size=14)
for i, (k, v) in enumerate(rows, 3):
    ws.cell(row=i, column=1, value=k).font = Font(name=FONT, bold=True)
    ws.cell(row=i, column=2, value=v)
named(wb, "heures_vendues", "General", "B10")

# ── Description_Besoin (copie du Catalogue UO) ────────────────────────────────
ws = wb.create_sheet("Description_Besoin")
ws.column_dimensions["A"].width = 110
ws["A1"] = "Cahier des charges — extrait du Catalogue UO (L09U1) — copié à la création"
ws["A1"].font = Font(name=FONT, bold=True, size=12)
besoin = [
    "SD.L1.03 : SAA requirement Specification",
    "  · Définir les exigences techniques applicables au SAA",
    "SD.L1.04 : MoP Type / Planned demonstration identification",
    "  · Sélectionner les types pertinents de Moyens de Preuve",
    "SD.L1.06 : SAA Requirement Review",
    "  · Garantir la qualité de chaque exigence technique SAA (critères SMART)",
    "SA.L1.05 : System architecture top level definition",
    "  · Sélectionner ou définir une architecture / solution",
    "SA.L1.06 : Calculation / Simulations and subsystem scope",
    "  · Dimensionner les sous-systèmes et éléments de conception",
]
for i, line in enumerate(besoin, 3):
    ws.cell(row=i, column=1, value=line)

# ── Donnees_Entree ────────────────────────────────────────────────────────────
ws = wb.create_sheet("Donnees_Entree")
add_table(ws, "tbl_donnees_entree",
    ["id", "designation", "type", "origine", "pic", "statut",
     "date_reception", "maturite", "date_update", "commentaire"],
    [
        {"id": "DE-001", "designation": ".CUS — Customer requirements", "origine": "DOORS", "statut": "RECUE", "maturite": "M3"},
        {"id": "DE-002", "designation": ".STD — Standards applicables", "origine": "DOORS", "statut": "EN_ATTENTE"},
        {"id": "DE-003", "designation": ".RSL — RefSol selectionnees", "origine": "DOORS", "statut": "RECUE", "maturite": "M2"},
    ],
    widths={"A": 9, "B": 38, "C": 10, "D": 10, "F": 12, "G": 13, "H": 9, "I": 12, "J": 28})
dv = DataValidation(type="list", formula1='"NA,EN_ATTENTE,RECUE"', allow_blank=True)
ws.add_data_validation(dv); dv.add("F2:F40")

# ── Activites — le coeur ──────────────────────────────────────────────────────
ws = wb.create_sheet("Activites")
acts = [
    {"id": "SD.L1.03", "designation": "SAA requirement Specification", "applicable": "OUI",
     "poids": 25, "statut": "EN_COURS", "avancement": 60, "heures_consommees": 30,
     "commentaire": ""},
    {"id": "SD.L1.04", "designation": "MoP Type / Planned demonstration", "applicable": "OUI",
     "poids": 25, "statut": "TERMINEE", "avancement": 100, "heures_consommees": 24,
     "commentaire": ""},
    {"id": "SD.L1.06", "designation": "SAA Requirement Review", "applicable": "OUI",
     "poids": 25, "statut": "EN_COURS", "avancement": 20, "heures_consommees": 5,
     "commentaire": ""},
    {"id": "SA.L1.05", "designation": "System architecture top level", "applicable": "NON",
     "poids": 15, "statut": "A_FAIRE", "avancement": 0, "heures_consommees": 0,
     "commentaire": "Hors perimetre — couvert par le donneur"},
    {"id": "SA.L1.06", "designation": "Calculations / Simulations", "applicable": "OUI",
     "poids": 25, "statut": "A_FAIRE", "avancement": 0, "heures_consommees": 0,
     "commentaire": ""},
]
headers = ["id", "designation", "applicable", "poids", "heures_allouees",
           "statut", "avancement", "heures_consommees", "reste_a_faire", "commentaire"]
rows = []
for a in acts:
    rows.append({h: a.get(h) for h in headers})
add_table(ws, "tbl_activites", headers, rows,
          widths={"A": 10, "B": 36, "C": 11, "D": 8, "E": 15, "F": 11,
                  "G": 12, "H": 17, "I": 13, "J": 32})
n = len(acts)
for r in range(2, 2 + n):
    # heures_allouees : poids renormalise sur les applicables x heures vendues
    ws.cell(row=r, column=5).value = (
        f'=IF(C{r}="OUI",D{r}/SUMIF($C$2:$C${1+n},"OUI",$D$2:$D${1+n})*heures_vendues,0)'
    )
    # reste_a_faire : defaut arithmetique, surchargeable par l'ingenieur
    ws.cell(row=r, column=9).value = f"=(1-G{r}/100)*E{r}"
    ws.cell(row=r, column=5).number_format = "0.00"
    ws.cell(row=r, column=9).number_format = "0.00"
dv = DataValidation(type="list", formula1='"OUI,NON"', allow_blank=True)
ws.add_data_validation(dv); dv.add(f"C2:C40")
dv = DataValidation(type="list", formula1='"A_FAIRE,EN_COURS,TERMINEE,STAND_BY"', allow_blank=True)
ws.add_data_validation(dv); dv.add(f"F2:F40")

# ── Livrables ─────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Livrables")
add_table(ws, "tbl_livrables",
    ["id", "designation", "type", "maturite_attendue", "date_attendue",
     "pic", "statut", "date_revisee", "date_livraison", "commentaire"],
    [
        {"id": "LIV-001", "designation": "SAA.RS — Requirement Spec", "type": "MaJ",
         "maturite_attendue": "M3", "statut": "EN_COURS"},
        {"id": "LIV-002", "designation": "RSS.RS — Subsystem Spec", "type": "MaJ",
         "maturite_attendue": "M3", "statut": "A_FAIRE"},
        {"id": "LIV-003", "designation": "DR.RS — Design Review file", "type": "Spec",
         "statut": "A_FAIRE"},
    ],
    widths={"A": 9, "B": 36, "C": 8, "D": 16, "E": 14, "F": 10, "G": 11, "H": 13, "I": 13, "J": 28})
dv = DataValidation(type="list", formula1='"A_FAIRE,EN_COURS,LIVRE,VALIDE"', allow_blank=True)
ws.add_data_validation(dv); dv.add("G2:G40")

# ── OIL — Open Items List ─────────────────────────────────────────────────────
ws = wb.create_sheet("OIL")
add_table(ws, "tbl_oil",
    ["id", "titre", "description", "en_action", "domaine", "criticite",
     "statut", "date_ouverture", "date_besoin", "journal"],
    [
        {"id": "PO-001", "titre": "Relecture spec SAA par fournisseur",
         "description": "Spec SAA v0.3 envoyee au fournisseur pour relecture avant IDR.",
         "en_action": "FOURNISSEUR", "domaine": "", "criticite": "MOYENNE", "statut": "OUVERT",
         "date_ouverture": "2026-06-02", "date_besoin": "2026-06-20",
         "journal": "02/06 : envoi v0.3 au fournisseur"},
        {"id": "PO-002", "titre": "Exigence feu-fumee EN 45545 a clarifier",
         "description": "Classification HL2 vs HL3 a confirmer pour le compartiment clim.",
         "en_action": "EXPERT", "domaine": "feu-fumee", "criticite": "HAUTE", "statut": "OUVERT",
         "date_ouverture": "2026-06-05", "date_besoin": "2026-06-30",
         "journal": "05/06 : question posee a l'expert FS"},
        {"id": "PO-003", "titre": "Acces DOORS module .CUS",
         "description": "Droits d'acces au module client demandes.",
         "en_action": "SE", "domaine": "", "criticite": "BASSE", "statut": "CLOS",
         "date_ouverture": "2026-05-28", "date_besoin": "",
         "journal": "28/05 : demande IT — 30/05 : acces obtenu, point clos"},
    ],
    widths={"A": 8, "B": 34, "C": 44, "D": 13, "E": 12, "F": 11, "G": 9, "H": 13, "I": 12, "J": 50})
dv = DataValidation(type="list", formula1='"SE,FOURNISSEUR,EXPERT,AT,CLIENT,AUTRE"', allow_blank=True)
ws.add_data_validation(dv); dv.add("D2:D60")
dv = DataValidation(type="list", formula1='"BASSE,MOYENNE,HAUTE"', allow_blank=True)
ws.add_data_validation(dv); dv.add("F2:F60")
dv = DataValidation(type="list", formula1='"OUVERT,CLOS"', allow_blank=True)
ws.add_data_validation(dv); dv.add("G2:G60")
for r in range(2, 5):
    ws.cell(row=r, column=10).alignment = Alignment(wrap_text=True, vertical="top")

# ── KPI — sortie du moteur ────────────────────────────────────────────────────
ws = wb.create_sheet("KPI")
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 64
ws["A1"] = "KPI — calculés par ExoSync à chaque synchronisation (ne pas saisir)"
ws["A1"].font = Font(name=FONT, bold=True, size=12)
kpis = [
    ("Avancement UO (%)",            "kpi_avancement",     "0.00", "Moyenne des avancements pondérée par les poids (activités applicables)"),
    ("Heures consommées",            "kpi_h_conso",        "0.00", "Somme des heures consommées (activités applicables)"),
    ("Points ouverts",               "kpi_po_ouverts",     "0",    "Nombre de points OIL au statut OUVERT"),
    ("Points fermés",                "kpi_po_fermes",      "0",    "Nombre de points OIL au statut CLOS"),
    ("Points critiques ouverts",     "kpi_po_critiques",   "0",    "Points OUVERTS avec criticite = HAUTE"),
    ("Dont balle chez fournisseur",  "kpi_po_fournisseur", "0",    "Points OUVERTS avec en_action = FOURNISSEUR"),
    ("Dont balle chez expert",       "kpi_po_expert",      "0",    "Points OUVERTS avec en_action = EXPERT"),
]
r = 3
for label, name, fmt, desc in kpis:
    ws.cell(row=r, column=1, value=label).font = Font(name=FONT, bold=True)
    named(wb, name, "KPI", f"B{r}")
    ws.cell(row=r, column=2).number_format = fmt
    ws.cell(row=r, column=3, value=desc).font = Font(name=FONT, size=9, italic=True)
    r += 1
# Lignes calculees en Excel local (formules — le moteur ne les lit pas en v1)
excel_kpis = [
    ("Heures vendues",            "kpi_h_vendues", "=heures_vendues", "0",
     "Lu depuis General (formule Excel)"),
    ("Reste à faire total (h)",   "kpi_raf",       "=SUM(Activites!I2:I40)", "0.00",
     "Somme colonne reste_a_faire (formule Excel)"),
    ("Heures estimées à terminaison (EAC)", "kpi_eac", "=kpi_h_conso+kpi_raf", "0.00",
     "EAC = consommé + reste à faire (formule Excel)"),
    ("Dérive à terminaison (h)",  "kpi_derive",    "=kpi_eac-kpi_h_vendues", "0.00",
     "EAC − heures vendues : >0 = dépassement prévu (formule Excel)"),
    ("Santé",                     "kpi_sante",
     '=IF(kpi_po_critiques>0,"ROUGE",IF(kpi_po_ouverts>0,"ORANGE","VERT"))', "General",
     "ROUGE si point critique ouvert · ORANGE si point ouvert · VERT sinon (formule Excel)"),
]
for label, name, formula, fmt, desc in excel_kpis:
    ws.cell(row=r, column=1, value=label).font = Font(name=FONT, bold=True)
    ws.cell(row=r, column=2, value=formula)
    ws.cell(row=r, column=2).number_format = fmt
    named(wb, name, "KPI", f"B{r}")
    ws.cell(row=r, column=3, value=desc).font = Font(name=FONT, size=9, italic=True)
    r += 1

# ── Dashboard — presentation libre ────────────────────────────────────────────
ws = wb.create_sheet("Dashboard")
ws["A1"] = "Dashboard — L09U1 · CFL 2400 · Climatisation · Jean Dujardin"
ws["A1"].font = Font(name=FONT, bold=True, size=14)
ws.column_dimensions["A"].width = 26
for col in "BCDEFG":
    ws.column_dimensions[col].width = 16
cards = [
    ("A4", "AVANCEMENT UO",   "A6", '=ROUND(kpi_avancement,2)&" %"'),
    ("C4", "HEURES",          "C6", '=ROUND(kpi_h_conso,2)&" / "&heures_vendues&" h"'),
    ("E4", "POINTS OUVERTS",  "E6", "=kpi_po_ouverts"),
    ("G4", "ATTENTE FOURN.",  "G6", "=kpi_po_fournisseur"),
    ("A8", "SANTÉ",           "A10", "=kpi_sante"),
    ("C8", "EAC (h)",         "C10", "=ROUND(kpi_eac,2)"),
    ("E8", "DÉRIVE (h)",      "E10", "=ROUND(kpi_derive,2)"),
    ("G8", "PTS CRITIQUES",   "G10", "=kpi_po_critiques"),
]
for lc, label, vc, formula in cards:
    ws[lc] = label
    ws[lc].font = Font(name=FONT, bold=True, size=10, color=NAVY)
    ws[vc] = formula
    ws[vc].font = Font(name=FONT, bold=True, size=18)
ws["A13"] = "Cette feuille est libre : l'ingénieur la customise (elle ne fait que LIRE l'onglet KPI)."
ws["A13"].font = Font(name=FONT, italic=True, size=9)

# ── Planning / Orga (reserves) ────────────────────────────────────────────────
ws = wb.create_sheet("Planning")
ws["A1"] = "Réservé : visualisation planning (formules à migrer depuis le template d'origine)."
ws["A1"].font = Font(name=FONT, italic=True)
ws = wb.create_sheet("Orga")
ws["A1"] = "Organisation projet (statique v1 — renseignée à la création depuis le Catalogue Projets)."
ws["A1"].font = Font(name=FONT, italic=True)

# ── _Manifeste ────────────────────────────────────────────────────────────────
ws = wb.create_sheet("_Manifeste")
ws.column_dimensions["A"].width = 95
ws["A1"] = "MANIFESTE_V=1"
ws["A2"] = "instruction"
ws["A2"].font = Font(bold=True)
UO = "L09U1-CFL2400-CLIM"
manifeste = [
    "FILE_TYPE: uo_instance",
    f"FILE_ID: {UO}",
    "",
    "# ── Lecture des donnees locales ─────────────────────────────",
    "DEF $act = GET_TABLE(Activites, tbl_activites)",
    "DEF $oil = GET_TABLE(OIL, tbl_oil)",
    "DEF $liv = GET_TABLE(Livrables, tbl_livrables)",
    "",
    "# ── Sous-ensembles ──────────────────────────────────────────",
    'DEF $actifs = COMPUTE(FILTER($act, applicable = "OUI"))',
    'DEF $po_ouv = COMPUTE(FILTER($oil, statut = "OUVERT"))',
    "",
    "# ── KPI (pondere par poids : valeurs saisies, pas formules) ─",
    "DEF $avancement = COMPUTE(MEAN_WEIGHTED($actifs.avancement, $actifs.poids))",
    "DEF $h_conso = COMPUTE(SUM($actifs.heures_consommees))",
    'DEF $po_ouverts = COMPUTE(COUNT_IF($oil.statut, "OUVERT"))',
    'DEF $po_fermes = COMPUTE(COUNT_IF($oil.statut, "CLOS"))',
    'DEF $po_critiques = COMPUTE(COUNT_IF($po_ouv.criticite, "HAUTE"))',
    'DEF $po_fournisseur = COMPUTE(COUNT_IF($po_ouv.en_action, "FOURNISSEUR"))',
    'DEF $po_expert = COMPUTE(COUNT_IF($po_ouv.en_action, "EXPERT"))',
    "",
    "# ── Regles de qualite ───────────────────────────────────────",
    "VALIDATE $actifs.avancement : RANGE(0, 100)",
    'VALIDATE $act.applicable : IN("OUI", "NON")',
    'VALIDATE $oil.statut : IN("OUVERT", "CLOS")',
    "",
    "# ── Affichage local (onglet KPI) ────────────────────────────",
    "BIND $avancement -> KPI.kpi_avancement",
    "BIND $h_conso -> KPI.kpi_h_conso",
    "BIND $po_ouverts -> KPI.kpi_po_ouverts",
    "BIND $po_fermes -> KPI.kpi_po_fermes",
    "BIND $po_critiques -> KPI.kpi_po_critiques",
    "BIND $po_fournisseur -> KPI.kpi_po_fournisseur",
    "BIND $po_expert -> KPI.kpi_po_expert",
    "",
    "# ── Publication vers l'ecosysteme ───────────────────────────",
    f"PUSH $avancement -> uo.{UO}.avancement",
    f"PUSH $h_conso -> uo.{UO}.heures_consommees",
    f"PUSH $po_ouverts -> uo.{UO}.po_ouverts",
    f"PUSH $po_fermes -> uo.{UO}.po_fermes",
    f"PUSH $po_critiques -> uo.{UO}.po_critiques",
    f"PUSH $actifs -> uo.{UO}.activites",
    f"PUSH $oil -> uo.{UO}.points_ouverts",
    f"PUSH $liv -> uo.{UO}.livrables",
]
for i, line in enumerate(manifeste, 3):
    ws.cell(row=i, column=1, value=line)

wb.save(OUT)
print(f"[OK] {OUT.name} genere ({len(wb.sheetnames)} feuilles)")
