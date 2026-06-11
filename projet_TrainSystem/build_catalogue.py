"""
build_catalogue.py — Genere Catalogue_UO_TrainSystem.xlsx depuis catalogue_uo.json.
====================================================================================
Le catalogue est LE referentiel editable dans Excel : l'assembleur (creer_uo.py)
le lit pour instancier les UO.

Feuilles :
  Index                    1 ligne par UO type (code, lot, libelles, criteres...)
  Catalogue_Activites      table centrale : uo_type | id | designation | poids | description
  Catalogue_Livrables      uo_type | designation
  Catalogue_DonneesEntree  uo_type | designation

⚠ CONFIDENTIEL : contenu issu d'un document Alstom — ne pas versionner sur GitHub.

Usage : python projet_TrainSystem/build_catalogue.py
"""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = Path(__file__).parent
CAT = json.loads((HERE / "catalogue_uo.json").read_text(encoding="utf-8"))
OUT = HERE / "Catalogue_UO_TrainSystem.xlsx"
NAVY = "1F4E79"
FONT = "Arial"


def add_table(ws, name, headers, rows, widths=None):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.font = Font(name=FONT, bold=True, color="FFFFFF")
    for ri, rd in enumerate(rows, 2):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=ri, column=ci, value=rd.get(h))
            c.alignment = Alignment(wrap_text=True, vertical="top")
    end = f"{get_column_letter(len(headers))}{1 + max(len(rows), 1)}"
    t = Table(displayName=name, ref=f"A1:{end}")
    t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(t)
    if widths:
        for col, w in widths.items():
            ws.column_dimensions[col].width = w


wb = Workbook()

# ── Index ─────────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Index"
rows = []
for code, u in sorted(CAT.items()):
    rows.append({
        "uo_type": code,
        "lot": u["lot"],
        "lot_libelle": u["lot_libelle"],
        "uo_libelle": u["uo_libelle"],
        "nb_activites": len(u["activites"]),
        "heures_reference": None,  # a remplir par JLC (heures vendues typiques)
        "criteres_acceptation": "\n".join(u["criteres"]),
        "niveaux_complexite": "\n".join(u["complexite"]),
    })
add_table(ws, "tbl_index",
          ["uo_type", "lot", "lot_libelle", "uo_libelle", "nb_activites",
           "heures_reference", "criteres_acceptation", "niveaux_complexite"],
          rows,
          widths={"A": 9, "B": 6, "C": 38, "D": 30, "E": 11, "F": 14, "G": 45, "H": 40})

# ── Catalogue_Activites ───────────────────────────────────────────────────────
ws = wb.create_sheet("Catalogue_Activites")
rows = []
for code, u in sorted(CAT.items()):
    for act in u["activites"]:
        rows.append({"uo_type": code, "id": act["id"],
                     "designation": act["designation"], "poids": 1,
                     "description": act["description"]})
add_table(ws, "tbl_cat_activites",
          ["uo_type", "id", "designation", "poids", "description"],
          rows, widths={"A": 9, "B": 10, "C": 50, "D": 7, "E": 90})

# ── Catalogue_Livrables ───────────────────────────────────────────────────────
ws = wb.create_sheet("Catalogue_Livrables")
rows = []
for code, u in sorted(CAT.items()):
    for item in u["livrables"]:
        rows.append({"uo_type": code, "designation": item})
add_table(ws, "tbl_cat_livrables", ["uo_type", "designation"], rows,
          widths={"A": 9, "B": 90})

# ── Catalogue_DonneesEntree ───────────────────────────────────────────────────
ws = wb.create_sheet("Catalogue_DonneesEntree")
rows = []
for code, u in sorted(CAT.items()):
    for item in u["donnees_entree"]:
        rows.append({"uo_type": code, "designation": item})
add_table(ws, "tbl_cat_donnees", ["uo_type", "designation"], rows,
          widths={"A": 9, "B": 90})

wb.save(OUT)
print(f"[OK] {OUT.name} — {len(CAT)} UO types, "
      f"{sum(len(u['activites']) for u in CAT.values())} activites, "
      f"{sum(len(u['livrables']) for u in CAT.values())} livrables, "
      f"{sum(len(u['donnees_entree']) for u in CAT.values())} donnees d'entree")
