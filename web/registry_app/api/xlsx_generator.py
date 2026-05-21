"""Génère un fichier Excel squelette pour un Post du registre.

Crée la feuille _Manifeste avec un MXL valide pré-rempli (syntaxe MXL v1 :
MANIFESTE_V=1 en A1, FILE_TYPE:/FILE_ID:/VERSION: avec colon,
DEF $var = GET_TABLE(...), COL $var.col : KEY|DATA, PUSH $var -> ns.key),
plus les feuilles de données avec les headers de colonnes issus des std_tables.

La génération du contenu MXL est déléguée à web.mxl_service (partagé avec N1).
"""
import io
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from web.registry_app.services.config_service import load_registre, save_registre, load_file_types, load_tables
from web.mxl_service import build_class_mxl_lines

router = APIRouter()

# ── Styles ────────────────────────────────────────────────────────────────────

_BLUE_DARK  = "1E3A5F"
_BLUE_MID   = "2563EB"
_GREY_LIGHT = "F1F5F9"
_WHITE      = "FFFFFF"

def _hfont(bold=True, size=11, color="FFFFFF"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def _bfont(bold=False, size=10, color="1E293B"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def _fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def _border():
    s = Side(border_style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_row(ws, row, headers, col_widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font      = _hfont()
        c.fill      = _fill(_BLUE_MID)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border()
    ws.row_dimensions[row].height = 20
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


# ── Générateur MXL ────────────────────────────────────────────────────────────
# Délégué à web.mxl_service — voir build_class_mxl_lines pour la logique complète.

def _get_std_tables(type_fichier: str) -> list[dict]:
    tables_data = load_tables().get("tables", {})
    return [
        t for t in tables_data.values()
        if t.get("file_id") == f"__class__{type_fichier}"
    ]


# ── Feuille _Manifeste ────────────────────────────────────────────────────────

def _sheet_manifeste(wb: Workbook, file_id: str, file_rec: dict, ft: dict):
    ws = wb.create_sheet("_Manifeste")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 80

    # A1 DOIT être "MANIFESTE_V=1" — parse_sheet() lit spécifiquement ws["A1"]
    # pour extraire le numéro de version avant de lire les instructions depuis min_row=3.
    ws["A1"] = "MANIFESTE_V=1"
    ws["A1"].font      = Font(name="Courier New", bold=True, size=11, color="4ADE80")
    ws["A1"].fill      = _fill(_BLUE_DARK)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 22

    # Row 2 : titre décoratif (ignoré par le parser — parse_sheet lit depuis min_row=3)
    ws["A2"] = f"# {file_id} — {file_rec['type_fichier']}"
    ws["A2"].font      = Font(name="Courier New", size=9.5, color="6B7280")
    ws["A2"].fill      = _fill("0F172A")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=0)
    ws.row_dimensions[2].height = 18

    # Rows 3+ : instructions MXL (délégué à web.mxl_service)
    lines = build_class_mxl_lines(
        class_id   = file_rec["type_fichier"],
        file_id    = file_id,
        ft         = ft,
        std_tables = _get_std_tables(file_rec["type_fichier"]),
    )
    for i, line in enumerate(lines, start=3):
        c          = ws.cell(row=i, column=1, value=line)
        is_comment = line.startswith("#")
        c.font     = Font(
            name="Courier New", size=9.5,
            color="6B7280" if is_comment else "4ADE80",
            bold=bool(line and not line.startswith(" ") and not is_comment),
        )
        c.fill      = _fill("0F172A")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=0)

    ws.freeze_panes = "A3"


# ── Feuille de données générique ──────────────────────────────────────────────

def _sheet_data(wb: Workbook, sheet_name: str, columns: list[dict]):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    title_end = max(len(columns), 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=title_end)
    ws["A1"]            = sheet_name
    ws["A1"].font       = _hfont(size=11)
    ws["A1"].fill       = _fill(_BLUE_DARK)
    ws["A1"].alignment  = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 22

    if not columns:
        ws.cell(row=3, column=1, value="(feuille vide — définir les colonnes dans Tables & colonnes)")
        ws.cell(row=3, column=1).font = Font(name="Calibri", size=9, color="94A3B8", italic=True)
        ws.column_dimensions["A"].width = 60
        return

    headers  = [c.get("header") or c.get("name", "") for c in columns]
    col_wids = [max(len(h) + 4, 14) for h in headers]
    _header_row(ws, 2, headers, col_wids)

    for r in range(3, 6):
        for i in range(1, len(columns) + 1):
            c        = ws.cell(row=r, column=i, value="")
            c.fill   = _fill(_GREY_LIGHT if r % 2 == 1 else _WHITE)
            c.border = _border()
            c.font   = _bfont()

    ws.freeze_panes = "A3"


# ── Feuille vide nommée ───────────────────────────────────────────────────────

def _sheet_empty(wb: Workbook, sheet_name: str):
    ws              = wb.create_sheet(sheet_name)
    ws["A1"]        = sheet_name
    ws["A1"].font   = _hfont(size=11)
    ws["A1"].fill   = _fill(_BLUE_DARK)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.column_dimensions["A"].width = 50
    ws.row_dimensions[1].height = 22


# ── Endpoint ──────────────────────────────────────────────────────────────────

@router.get("/{file_id}")
def generate_xlsx(file_id: str):
    registre = {f["id"]: f for f in load_registre()}
    if file_id not in registre:
        raise HTTPException(404, f"Post '{file_id}' non trouvé dans le registre")

    file_rec = registre[file_id]
    if not file_rec.get("type_fichier"):
        raise HTTPException(422, "Ce Post est un Post libre (sans Classe) — la génération Excel requiert une Classe")

    ft_all = load_file_types()
    ft     = ft_all.get(file_rec["type_fichier"])
    if not ft:
        raise HTTPException(422, f"Type de fichier '{file_rec['type_fichier']}' inconnu")

    # Colonnes par feuille depuis std_tables de la Classe
    tables_data    = load_tables().get("tables", {})
    cols_by_sheet: dict[str, list[dict]] = {}
    for t in tables_data.values():
        if t.get("file_id") == f"__class__{file_rec['type_fichier']}":
            sheet = t.get("sheet", t["table_name"])
            cols_by_sheet[sheet] = t.get("columns", [])

    wb = Workbook()
    wb.remove(wb.active)

    _sheet_manifeste(wb, file_id, file_rec, ft)

    for sheet in ft.get("required_sheets", []):
        if sheet == "_Manifeste":
            continue
        _sheet_data(wb, sheet, cols_by_sheet.get(sheet, []))

    for sheet in ft.get("optional_sheets", []):
        if sheet in ("_Manifeste", "_Log"):
            _sheet_empty(wb, sheet)
        else:
            _sheet_data(wb, sheet, cols_by_sheet.get(sheet, []))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    class_version = ft.get("schema_version", 1)
    all_fichiers = load_registre()
    for i, f in enumerate(all_fichiers):
        if f["id"] == file_id:
            all_fichiers[i] = {**f, "schema_version": class_version}
            break
    save_registre(all_fichiers)

    safe_id  = file_id.replace("/", "-").replace("\\", "-")
    filename = f"{safe_id}_{file_rec['type_fichier']}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
