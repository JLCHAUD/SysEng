"""
Tests d'intégration — execute_ast() sur de vrais fichiers Excel
===============================================================
Génère 3 fichiers Excel en mémoire avec openpyxl, lance execute_ast(),
vérifie que le store et le dashboard sont correctement mis à jour.

Scénario :
  1. referentiel.xlsx  — pousse des activités types dans le store
  2. UO-001.xlsx       — PULL depuis store, COMPUTE, PUSH, BIND sur Dashboard
  3. cockpit.xlsx      — lit les résultats du store publiés par UO-001
"""
import json
import tempfile
from pathlib import Path
from typing import Any, Dict

import pytest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table

from src.executor import execute_ast, ExecutionResult
from src.parser import PasserelleAST, FileHeader, DefNode, PullNode, PushNode, BindNode


# ─── Store isolé pour les tests ───────────────────────────────────────────────

class MemStore:
    """Store en mémoire (n'écrit pas sur disque) pour isoler les tests."""

    def __init__(self, initial: Dict[str, Any] = None):
        self._data = dict(initial or {})

    def get(self, key: str) -> Any:
        return self._data.get(key)

    def set(self, key: str, value: Any) -> None:
        self._data[key] = value

    def set_many(self, variables: Dict[str, Any]) -> None:
        self._data.update(variables)

    def get_all(self) -> Dict[str, Any]:
        return dict(self._data)


# ─── Helpers de construction Excel ────────────────────────────────────────────

def _make_table(ws, table_name: str, headers: list, rows: list, start_row: int = 1):
    """Crée une Table Excel nommée dans une feuille openpyxl."""
    # En-têtes
    for ci, h in enumerate(headers, 1):
        ws.cell(start_row, ci).value = h
    # Données
    for ri, row in enumerate(rows, start_row + 1):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci).value = val
    # Objet Table
    end_row = start_row + len(rows)
    end_col = get_column_letter(len(headers))
    ref = f"A{start_row}:{end_col}{end_row}"
    tbl = Table(displayName=table_name, ref=ref)
    ws.add_table(tbl)
    return tbl


def _add_named_range(wb, sheet_name: str, cell: str, range_name: str):
    """Crée une plage nommée dans le workbook."""
    attr = f"'{sheet_name}'!${cell.replace('A', 'A').replace('B', 'B')}"
    # Construire l'attribut correct
    col = ''.join(c for c in cell if c.isalpha())
    row = ''.join(c for c in cell if c.isdigit())
    attr_text = f"'{sheet_name}'!${col}${row}"
    defn = DefinedName(range_name, attr_text=attr_text)
    wb.defined_names.add(defn)


# ─── Fixture : fichier UO-001.xlsx ────────────────────────────────────────────

def _build_uo001_workbook() -> Workbook:
    """
    Construit UO-001.xlsx en mémoire :
      - Feuille 'Activites' avec Table 'TabActivites' (données pré-remplies)
      - Feuille 'Dashboard' avec plages nommées pour BIND
    """
    wb = Workbook()

    # ── Feuille Activites ──────────────────────────────────────────────────
    ws_act = wb.active
    ws_act.title = "Activites"

    headers = ["id", "libelle", "statut", "avancement", "heures"]
    rows = [
        ("ACT-001", "Pose câbles signaux",  "EN_COURS",  60,  100),
        ("ACT-002", "Tests fonctionnels",   "EN_COURS",  30,  200),
        ("ACT-003", "Mise en service",      "CLOTUREE",  100, 50),
        ("ACT-004", "Documentation",        "SUSPENDUE", 10,  80),
    ]
    _make_table(ws_act, "TabActivites", headers, rows)

    # ── Feuille Dashboard ─────────────────────────────────────────────────
    ws_dash = wb.create_sheet("Dashboard")
    ws_dash["A1"] = "Indicateur"
    ws_dash["B1"] = "Valeur"
    ws_dash["A2"] = "Avancement global"
    ws_dash["A3"] = "Total heures"
    ws_dash["A4"] = "Nb activités"
    ws_dash["A5"] = "Statut global"

    _add_named_range(wb, "Dashboard", "B2", "avancement_global")
    _add_named_range(wb, "Dashboard", "B3", "total_heures")
    _add_named_range(wb, "Dashboard", "B4", "nb_activites")
    _add_named_range(wb, "Dashboard", "B5", "statut_global")

    return wb


def _build_uo001_ast() -> PasserelleAST:
    """
    Construit l'AST correspondant à ce Manifeste MXL :

      FILE_TYPE: uo_instance
      FILE_ID:   UO-001

      DEF $activites = GET_TABLE(Activites, TabActivites)
      DEF $actives   = COMPUTE(FILTER($activites, statut != "CLOTUREE"))
      DEF $avancement_global = COMPUTE(MEAN_WEIGHTED($actives.avancement, $actives.heures))
      DEF $total_heures      = COMPUTE(SUM($activites.heures))
      DEF $nb_total          = COMPUTE(COUNT($activites.id))
      DEF $statut_global     = COMPUTE(TRAFFIC_LIGHT($avancement_global, warn=30, ok=70))

      BIND $avancement_global -> Dashboard.avancement_global
      BIND $total_heures      -> Dashboard.total_heures
      BIND $nb_total          -> Dashboard.nb_activites
      BIND $statut_global     -> Dashboard.statut_global

      PUSH $avancement_global -> uo.UO-001.avancement_global
      PUSH $total_heures      -> uo.UO-001.total_heures
      PUSH $nb_total          -> uo.UO-001.nb_activites
      PUSH $statut_global     -> uo.UO-001.statut
    """
    ast = PasserelleAST()
    ast.header = FileHeader(file_type="uo_instance", file_id="UO-001")

    ast.defs = [
        DefNode("$activites",        "GET_TABLE", sheet="Activites", table_name="TabActivites"),
        DefNode("$actives",          "COMPUTE",   formula='FILTER($activites, statut != "CLOTUREE")'),
        DefNode("$avancement_global","COMPUTE",   formula="MEAN_WEIGHTED($actives.avancement, $actives.heures)"),
        DefNode("$total_heures",     "COMPUTE",   formula="SUM($activites.heures)"),
        DefNode("$nb_total",         "COMPUTE",   formula="COUNT($activites.id)"),
        DefNode("$statut_global",    "COMPUTE",   formula="TRAFFIC_LIGHT($avancement_global, warn=30, ok=70)"),
    ]
    for d in ast.defs:
        ast._defs_index[d.var_name] = d

    ast.binds = [
        BindNode("$avancement_global", "Dashboard", "avancement_global"),
        BindNode("$total_heures",      "Dashboard", "total_heures"),
        BindNode("$nb_total",          "Dashboard", "nb_activites"),
        BindNode("$statut_global",     "Dashboard", "statut_global"),
    ]

    ast.pushes = [
        PushNode("$avancement_global", "uo.UO-001.avancement_global"),
        PushNode("$total_heures",      "uo.UO-001.total_heures"),
        PushNode("$nb_total",          "uo.UO-001.nb_activites"),
        PushNode("$statut_global",     "uo.UO-001.statut"),
    ]

    return ast


# ─── Tests ────────────────────────────────────────────────────────────────────

class TestExecuteAstUO001:
    """Tests d'intégration sur UO-001.xlsx (COMPUTE + BIND + PUSH, pas de PULL)."""

    @pytest.fixture
    def setup(self, tmp_path):
        """Crée le fichier Excel et le store en mémoire."""
        filepath = tmp_path / "UO-001.xlsx"
        wb = _build_uo001_workbook()
        wb.save(str(filepath))

        ast   = _build_uo001_ast()
        store = MemStore()
        return filepath, ast, store

    def test_execution_sans_erreur(self, setup):
        filepath, ast, store = setup
        result = execute_ast(ast, filepath, store)
        assert result.errors == [], f"Erreurs inattendues : {result.errors}"

    def test_push_avancement_dans_store(self, setup):
        filepath, ast, store = setup
        execute_ast(ast, filepath, store)

        # Actives : ACT-001(60%, 100h) + ACT-002(30%, 200h) + ACT-004(10%, 80h)
        # MEAN_WEIGHTED = (60*100 + 30*200 + 10*80) / (100+200+80)
        expected = (60 * 100 + 30 * 200 + 10 * 80) / (100 + 200 + 80)
        val = store.get("uo.UO-001.avancement_global")
        assert val == pytest.approx(expected, rel=1e-3)

    def test_push_total_heures_dans_store(self, setup):
        filepath, ast, store = setup
        execute_ast(ast, filepath, store)
        assert store.get("uo.UO-001.total_heures") == 430  # 100+200+50+80

    def test_push_nb_activites_dans_store(self, setup):
        filepath, ast, store = setup
        execute_ast(ast, filepath, store)
        assert store.get("uo.UO-001.nb_activites") == 4

    def test_push_statut_dans_store(self, setup):
        filepath, ast, store = setup
        execute_ast(ast, filepath, store)
        statut = store.get("uo.UO-001.statut")
        assert statut in ("ROUGE", "ORANGE", "VERT")

    def test_bind_ecrit_dans_dashboard(self, setup):
        """Vérifie que les valeurs sont bien écrites dans le fichier Excel."""
        filepath, ast, store = setup
        execute_ast(ast, filepath, store)

        # Re-ouvrir le fichier et vérifier les cellules
        from openpyxl import load_workbook
        wb2 = load_workbook(str(filepath), data_only=True)
        ws_dash = wb2["Dashboard"]

        assert ws_dash["B2"].value is not None, "avancement_global non écrit"
        assert ws_dash["B3"].value == 430,       "total_heures incorrect"
        assert ws_dash["B4"].value == 4,         "nb_activites incorrect"
        assert ws_dash["B5"].value in ("ROUGE", "ORANGE", "VERT"), "statut invalide"

    def test_result_pushed_count(self, setup):
        filepath, ast, store = setup
        result = execute_ast(ast, filepath, store)
        assert len(result.pushed) == 4

    def test_result_bound_count(self, setup):
        filepath, ast, store = setup
        result = execute_ast(ast, filepath, store)
        assert len(result.bound) == 4


class TestExecuteAstPull:
    """Tests d'intégration pour la phase PULL (store → Excel)."""

    @pytest.fixture
    def setup(self, tmp_path):
        """Prépare un store avec des données et un fichier UO avec table vide."""
        # Store pré-rempli avec des données référentiel
        store = MemStore({
            "ref.activites_type": [
                {"id": "TYPE-A", "libelle": "Pose", "heures_allouees": 100},
                {"id": "TYPE-B", "libelle": "Test", "heures_allouees": 50},
                {"id": "TYPE-C", "libelle": "Doc",  "heures_allouees": 30},
            ]
        })

        # Fichier Excel avec une table Activites vide (1 ligne = en-têtes seulement)
        wb = Workbook()
        ws = wb.active
        ws.title = "Activites"
        _make_table(ws, "TabActivites",
                    ["id", "libelle", "heures_allouees"],
                    [("TYPE-EXISTANT", "Existant", 80)])  # 1 ligne existante

        filepath = tmp_path / "UO-PULL.xlsx"
        wb.save(str(filepath))

        # AST avec PULL APPEND_NEW + GET_TABLE + PUSH COUNT
        ast = PasserelleAST()
        ast.header = FileHeader(file_type="uo_instance", file_id="UO-PULL")
        ast.pulls = [
            PullNode(
                global_name="ref.activites_type",
                operation="FILL_TABLE",
                sheet="Activites",
                table="TabActivites",
                mode="APPEND_NEW",
                key="id",
            )
        ]
        ast.defs = [
            DefNode("$activites", "GET_TABLE", sheet="Activites", table_name="TabActivites"),
            DefNode("$nb_total",  "COMPUTE",   formula="COUNT($activites.id)"),
        ]
        for d in ast.defs:
            ast._defs_index[d.var_name] = d
        ast.pushes = [PushNode("$nb_total", "uo.UO-PULL.nb_total")]

        return filepath, ast, store

    def test_pull_append_new_ajoute_lignes(self, setup):
        filepath, ast, store = setup
        result = execute_ast(ast, filepath, store)
        assert result.errors == [], f"Erreurs : {result.errors}"
        # Store : 3 nouvelles lignes + 1 existante = 4 total
        assert store.get("uo.UO-PULL.nb_total") == 4

    def test_pull_overwrite_ecrase_tout(self, tmp_path):
        store = MemStore({
            "ref.acteurs": [
                {"id": "P1", "nom": "Alice"},
                {"id": "P2", "nom": "Bob"},
            ]
        })

        wb = Workbook()
        ws = wb.active
        ws.title = "Acteurs"
        # Table avec données initiales différentes
        _make_table(ws, "TabActeurs",
                    ["id", "nom"],
                    [("OLD-1", "Ancien"), ("OLD-2", "Vieux")])
        filepath = tmp_path / "UO-OW.xlsx"
        wb.save(str(filepath))

        ast = PasserelleAST()
        ast.header = FileHeader(file_type="uo_instance", file_id="UO-OW")
        ast.pulls = [
            PullNode("ref.acteurs", "FILL_TABLE", "Acteurs", "TabActeurs", "OVERWRITE")
        ]
        ast.defs = [
            DefNode("$acteurs", "GET_TABLE", sheet="Acteurs", table_name="TabActeurs"),
            DefNode("$nb",      "COMPUTE",   formula="COUNT($acteurs.id)"),
        ]
        for d in ast.defs:
            ast._defs_index[d.var_name] = d
        ast.pushes = [PushNode("$nb", "uo.UO-OW.nb")]

        execute_ast(ast, filepath, store)
        assert store.get("uo.UO-OW.nb") == 2  # exactement les 2 du store, pas 4


class TestExecuteAstStatusFormatage:
    """Vérifie que BIND applique le formatage couleur sur les STATUS."""

    def test_bind_statut_formate(self, tmp_path):
        wb = Workbook()
        ws_dash = wb.create_sheet("Dashboard")
        wb.remove(wb.active)     # supprimer la feuille par défaut
        ws_dash["A1"] = "Statut"
        _add_named_range(wb, "Dashboard", "A1", "statut_global")

        filepath = tmp_path / "UO-FORMAT.xlsx"
        wb.save(str(filepath))

        ast = PasserelleAST()
        ast.header = FileHeader(file_type="uo_instance", file_id="UO-FORMAT")
        ast.defs = [DefNode("$statut", "COMPUTE", formula='"VERT"')]
        ast._defs_index["$statut"] = ast.defs[0]
        ast.binds = [BindNode("$statut", "Dashboard", "statut_global")]

        store = MemStore()
        execute_ast(ast, filepath, store)

        from openpyxl import load_workbook
        wb2 = load_workbook(str(filepath), data_only=True)
        cell = wb2["Dashboard"]["A1"]
        assert cell.value == "VERT"
        # Vérifier que la couleur de fond a été appliquée
        assert cell.fill is not None
        assert cell.fill.fgColor.rgb.endswith("00B050")  # vert ExoSync


class TestExecuteAstErreurs:
    """Vérifie la robustesse face aux erreurs."""

    def test_fichier_inexistant(self, tmp_path):
        ast = PasserelleAST()
        store = MemStore()
        result = execute_ast(ast, tmp_path / "INEXISTANT.xlsx", store)
        assert len(result.errors) == 1
        assert "introuvable" in result.errors[0].lower()

    def test_feuille_manquante_dans_get_table(self, tmp_path):
        wb = Workbook()
        wb.active.title = "AutreFeuille"
        filepath = tmp_path / "UO-ERR.xlsx"
        wb.save(str(filepath))

        ast = PasserelleAST()
        ast.header = FileHeader(file_type="uo_instance", file_id="UO-ERR")
        ast.defs = [DefNode("$t", "GET_TABLE", sheet="Activites", table_name="TabActivites")]
        ast._defs_index["$t"] = ast.defs[0]

        store = MemStore()
        result = execute_ast(ast, filepath, store)
        # Doit signaler l'erreur sans planter
        assert len(result.errors) >= 1
        assert any("Activites" in e for e in result.errors)

    def test_store_vide_pull_skippe(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Activites"
        _make_table(ws, "TabActivites", ["id"], [("A1",)])
        filepath = tmp_path / "UO-SKIP.xlsx"
        wb.save(str(filepath))

        ast = PasserelleAST()
        ast.pulls = [PullNode("ref.absent", "FILL_TABLE", "Activites", "TabActivites", "OVERWRITE")]

        store = MemStore()  # store vide
        result = execute_ast(ast, filepath, store)
        assert result.errors == []          # pas d'erreur
        assert len(result.skipped) == 1     # mais bien skippé


class TestExecutionResult:
    """Tests sur le résultat lui-même."""

    def test_summary_format(self):
        r = ExecutionResult(
            pushed=["k1", "k2"],
            pulled=["t1"],
            bound=["b1", "b2", "b3"],
            errors=["oops"],
            skipped=["s1"],
        )
        summary = r.summary()
        assert "PULL" in summary
        assert "PUSH" in summary
        assert "BIND" in summary
        assert "oops" in summary
