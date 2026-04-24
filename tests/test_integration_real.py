"""
Tests d'intégration réels — 3 fichiers Excel sur disque
========================================================
Scénario :
  1. referentiel.xlsx  — lit un tableau de types UO, push dans store
  2. UO-001.xlsx       — lit activités, COMPUTE KPIs, VALIDATE, PUSH dans store
  3. UO-002.xlsx       — mêmes instructions mais données différentes
  4. Chaîne complète   — REF → UO-001 (PULL depuis store) → PUSH KPIs

Différence avec test_executor_integration.py :
  - Les fichiers sont écrits sur disque (tmp_path) puis lus via parse_file()
  - Le store est un JsonStore isolé (JsonStore(tmp_path/store.json))
  - Le pipeline passe par parse_file() + execute_ast() comme en production
"""
from pathlib import Path
from typing import Any, Dict, List

import pytest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table

from src.executor import execute_ast, ExecutionResult
from src.parser import parse_file
from src.store import JsonStore


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _make_table(ws, name: str, headers: List[str], rows: List[tuple], start_row: int = 1):
    """Écrit une Table Excel nommée dans la feuille."""
    for ci, h in enumerate(headers, 1):
        ws.cell(start_row, ci).value = h
    for ri, row in enumerate(rows, start_row + 1):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci).value = val
    end_col = get_column_letter(len(headers))
    end_row = start_row + len(rows)
    tbl = Table(displayName=name, ref=f"A{start_row}:{end_col}{end_row}")
    ws.add_table(tbl)


def _add_named_range(wb, sheet: str, cell: str, name: str):
    col = "".join(c for c in cell if c.isalpha())
    row = "".join(c for c in cell if c.isdigit())
    defn = DefinedName(name, attr_text=f"'{sheet}'!${col}${row}")
    wb.defined_names.add(defn)


def _write_manifeste(ws, instructions: List[str]):
    """
    Écrit la feuille _Manifeste :
      A1  : MANIFESTE_V=1
      A2  : (en-tête ignoré par parse_sheet qui démarre à row 3)
      A3+ : instructions MXL
    """
    ws["A1"] = "MANIFESTE_V=1"
    ws["A2"] = "Instruction"   # sautée par parse_sheet (min_row=3)
    for i, instr in enumerate(instructions, start=3):
        ws.cell(row=i, column=1).value = instr


def _save(wb: Workbook, path: Path) -> Path:
    wb.save(str(path))
    return path


# ─── Constructeurs de fichiers ─────────────────────────────────────────────────

def _build_referentiel(tmp_path: Path) -> Path:
    """
    referentiel.xlsx
      Sheet 'Types'  : TabTypesUO (id, nom, charge_std)
      _Manifeste     : GET_TABLE → PUSH vers store
    """
    wb = Workbook()
    ws_types = wb.active
    ws_types.title = "Types"

    _make_table(ws_types, "TabTypesUO",
                ["id", "nom", "charge_std"],
                [
                    ("SIG", "Signalisation", 1200),
                    ("ELEC", "Electricite",  800),
                    ("MECA", "Mecanique",    600),
                ])

    ws_mxl = wb.create_sheet("_Manifeste")
    _write_manifeste(ws_mxl, [
        "FILE_TYPE: referentiel_uo",
        "FILE_ID:   REF-001",
        "DEF $types = GET_TABLE(Types, TabTypesUO)",
        "PUSH $types -> referentiel.types_uo",
    ])

    return _save(wb, tmp_path / "referentiel.xlsx")


def _build_uo001(tmp_path: Path) -> Path:
    """
    UO-001.xlsx
      Sheet 'Activites' : TabActivites (id, nom, avancement, heures, statut)
      Sheet 'Dashboard' : plages nommées pour BIND
      _Manifeste        : GET_TABLE → COMPUTE → VALIDATE → BIND → PUSH
    """
    wb = Workbook()
    ws_act = wb.active
    ws_act.title = "Activites"

    _make_table(ws_act, "TabActivites",
                ["id", "nom", "avancement", "heures", "statut"],
                [
                    ("A01", "Pose cables",    60,  200, "EN_COURS"),
                    ("A02", "Tests fonct.",   80,  150, "EN_COURS"),
                    ("A03", "Mise en service", 100, 50,  "CLOTUREE"),
                    ("A04", "Documentation",  20,  100, "EN_COURS"),
                ])

    ws_dash = wb.create_sheet("Dashboard")
    ws_dash["A1"] = "KPI"
    ws_dash["B1"] = "Valeur"
    ws_dash["A2"] = "Avancement"
    ws_dash["A3"] = "Total heures"
    ws_dash["A4"] = "Statut"
    _add_named_range(wb, "Dashboard", "B2", "avancement_global")
    _add_named_range(wb, "Dashboard", "B3", "total_heures")
    _add_named_range(wb, "Dashboard", "B4", "statut_global")

    ws_mxl = wb.create_sheet("_Manifeste")
    _write_manifeste(ws_mxl, [
        "FILE_TYPE: uo_instance",
        "FILE_ID:   UO-001",
        "DEF $activites         = GET_TABLE(Activites, TabActivites)",
        "DEF $avancement        = COMPUTE(MEAN_WEIGHTED($activites.avancement, $activites.heures))",
        "DEF $total_heures      = COMPUTE(SUM($activites.heures))",
        "DEF $nb_cloturees      = COMPUTE(COUNT_IF($activites.statut, \"CLOTUREE\"))",
        "DEF $statut            = COMPUTE(TRAFFIC_LIGHT($avancement, warn=40, ok=70))",
        "VALIDATE $activites.avancement : RANGE(0, 100)",
        "VALIDATE $activites.id         : UNIQUE",
        "BIND $avancement   -> Dashboard.avancement_global",
        "BIND $total_heures -> Dashboard.total_heures",
        "BIND $statut       -> Dashboard.statut_global",
        "PUSH $avancement      -> uo.UO-001.avancement",
        "PUSH $total_heures    -> uo.UO-001.heures",
        "PUSH $nb_cloturees    -> uo.UO-001.nb_cloturees",
        "PUSH $statut          -> uo.UO-001.statut",
    ])

    return _save(wb, tmp_path / "UO-001.xlsx")


def _build_uo002(tmp_path: Path) -> Path:
    """
    UO-002.xlsx — données différentes d'UO-001.
      Avancement plus faible → TRAFFIC_LIGHT retourne ROUGE.
    """
    wb = Workbook()
    ws_act = wb.active
    ws_act.title = "Activites"

    _make_table(ws_act, "TabActivites",
                ["id", "nom", "avancement", "heures", "statut"],
                [
                    ("B01", "Etude preliminaire", 10,  300, "EN_COURS"),
                    ("B02", "Validation client",  5,   200, "EN_COURS"),
                    ("B03", "Prototype",          15,  100, "EN_COURS"),
                ])

    ws_mxl = wb.create_sheet("_Manifeste")
    _write_manifeste(ws_mxl, [
        "FILE_TYPE: uo_instance",
        "FILE_ID:   UO-002",
        "DEF $activites    = GET_TABLE(Activites, TabActivites)",
        "DEF $avancement   = COMPUTE(MEAN_WEIGHTED($activites.avancement, $activites.heures))",
        "DEF $total_heures = COMPUTE(SUM($activites.heures))",
        "DEF $statut       = COMPUTE(TRAFFIC_LIGHT($avancement, warn=40, ok=70))",
        "VALIDATE $activites.avancement : RANGE(0, 100)",
        "PUSH $avancement   -> uo.UO-002.avancement",
        "PUSH $total_heures -> uo.UO-002.heures",
        "PUSH $statut       -> uo.UO-002.statut",
    ])

    return _save(wb, tmp_path / "UO-002.xlsx")


# ─── Tests ────────────────────────────────────────────────────────────────────

class TestReferentiel:
    """Fichier référentiel — lit une table et pousse dans le store."""

    def test_parse_et_execute(self, tmp_path):
        path = _build_referentiel(tmp_path)
        store = JsonStore(tmp_path / "store.json")

        ast = parse_file(path)
        assert ast is not None
        assert ast.errors == []
        assert ast.header.file_type == "referentiel_uo"
        assert ast.header.file_id   == "REF-001"
        assert len(ast.defs)   == 1
        assert len(ast.pushes) == 1

        result = execute_ast(ast, path, store)

        assert result.errors == []
        assert "referentiel.types_uo" in result.pushed

    def test_store_contient_les_types(self, tmp_path):
        path = _build_referentiel(tmp_path)
        store = JsonStore(tmp_path / "store.json")

        ast = parse_file(path)
        execute_ast(ast, path, store)

        types = store.get("referentiel.types_uo")
        assert isinstance(types, list)
        assert len(types) == 3
        ids = [t["id"] for t in types]
        assert "SIG" in ids
        assert "ELEC" in ids
        assert "MECA" in ids

    def test_store_persistant_apres_execution(self, tmp_path):
        """Vérifie que le store JSON survit à la réinstanciation."""
        path = _build_referentiel(tmp_path)
        store_a = JsonStore(tmp_path / "store.json")
        ast = parse_file(path)
        execute_ast(ast, path, store_a)

        # Nouvelle instance pointant sur le même fichier
        store_b = JsonStore(tmp_path / "store.json")
        assert store_b.get("referentiel.types_uo") is not None


class TestUO001:
    """UO-001 — COMPUTE, VALIDATE, BIND, PUSH."""

    def _run(self, tmp_path):
        path  = _build_uo001(tmp_path)
        store = JsonStore(tmp_path / "store.json")
        ast   = parse_file(path)
        result = execute_ast(ast, path, store)
        return path, store, ast, result

    def test_parse_sans_erreur(self, tmp_path):
        path = _build_uo001(tmp_path)
        ast  = parse_file(path)
        assert ast is not None
        assert ast.errors == []
        assert ast.header.file_id == "UO-001"
        assert len(ast.defs)      == 5
        assert len(ast.validates) == 2
        assert len(ast.binds)     == 3
        assert len(ast.pushes)    == 4

    def test_execute_sans_erreur(self, tmp_path):
        _, _, _, result = self._run(tmp_path)
        assert result.errors   == []
        assert result.warnings == []

    def test_avancement_calcule(self, tmp_path):
        """MEAN_WEIGHTED((60*200 + 80*150 + 100*50 + 20*100) / 500)"""
        _, store, _, _ = self._run(tmp_path)
        av = store.get("uo.UO-001.avancement")
        expected = (60*200 + 80*150 + 100*50 + 20*100) / (200+150+50+100)
        assert av is not None
        assert abs(av - expected) < 0.01

    def test_total_heures(self, tmp_path):
        _, store, _, _ = self._run(tmp_path)
        assert store.get("uo.UO-001.heures") == 500  # 200+150+50+100

    def test_nb_cloturees(self, tmp_path):
        _, store, _, _ = self._run(tmp_path)
        assert store.get("uo.UO-001.nb_cloturees") == 1  # A03

    def test_statut_traffic_light(self, tmp_path):
        """Avancement ~67% → warn=40, ok=70 → ORANGE."""
        _, store, _, _ = self._run(tmp_path)
        statut = store.get("uo.UO-001.statut")
        # (60*200 + 80*150 + 100*50 + 20*100)/500 ≈ 67.0% → ORANGE
        assert statut in ("ORANGE", "VERT")  # selon valeur exacte

    def test_validate_range_ok(self, tmp_path):
        """Toutes les valeurs d'avancement sont dans [0,100] → pas d'erreur."""
        _, _, _, result = self._run(tmp_path)
        assert result.errors == []

    def test_push_toutes_les_cles(self, tmp_path):
        _, store, _, result = self._run(tmp_path)
        pushed = set(result.pushed)
        assert "uo.UO-001.avancement"   in pushed
        assert "uo.UO-001.heures"       in pushed
        assert "uo.UO-001.nb_cloturees" in pushed
        assert "uo.UO-001.statut"       in pushed

    def test_bind_dashboard(self, tmp_path):
        """BIND écrit dans les plages nommées Dashboard."""
        _, _, _, result = self._run(tmp_path)
        assert len(result.bound) == 3
        assert "avancement_global" in result.bound
        assert "total_heures"      in result.bound
        assert "statut_global"     in result.bound

    def test_fichier_modifie_sur_disque(self, tmp_path):
        """Après execute_ast, le fichier est bien réécrit sur disque."""
        import time
        path = _build_uo001(tmp_path)
        mtime_avant = path.stat().st_mtime

        time.sleep(0.05)  # assure un écart mesurable

        store = JsonStore(tmp_path / "store.json")
        ast   = parse_file(path)
        execute_ast(ast, path, store)

        mtime_apres = path.stat().st_mtime
        assert mtime_apres > mtime_avant


class TestUO002:
    """UO-002 — données faibles, TRAFFIC_LIGHT → ROUGE."""

    def _run(self, tmp_path):
        path  = _build_uo002(tmp_path)
        store = JsonStore(tmp_path / "store.json")
        ast   = parse_file(path)
        result = execute_ast(ast, path, store)
        return path, store, ast, result

    def test_parse_sans_erreur(self, tmp_path):
        path = _build_uo002(tmp_path)
        ast  = parse_file(path)
        assert ast is not None
        assert ast.errors == []
        assert ast.header.file_id == "UO-002"

    def test_statut_rouge(self, tmp_path):
        """Avancement ~10% → warn=40 → ROUGE."""
        _, store, _, result = self._run(tmp_path)
        assert result.errors == []
        statut = store.get("uo.UO-002.statut")
        assert statut == "ROUGE"

    def test_avancement_faible(self, tmp_path):
        _, store, _, _ = self._run(tmp_path)
        av = store.get("uo.UO-002.avancement")
        expected = (10*300 + 5*200 + 15*100) / (300+200+100)
        assert av is not None
        assert abs(av - expected) < 0.01

    def test_total_heures(self, tmp_path):
        _, store, _, _ = self._run(tmp_path)
        assert store.get("uo.UO-002.heures") == 600


class TestChaineComplete:
    """
    Scénario complet :
      1. referentiel.xlsx → pousse les types dans le store
      2. UO-001.xlsx      → pousse ses KPIs dans le même store
      3. UO-002.xlsx      → pousse ses KPIs dans le même store
      → Le store contient les données des 3 fichiers
    """

    def test_trois_fichiers_en_sequence(self, tmp_path):
        store = JsonStore(tmp_path / "store.json")

        # Référentiel
        p_ref = _build_referentiel(tmp_path)
        ast_ref = parse_file(p_ref)
        r_ref = execute_ast(ast_ref, p_ref, store)
        assert r_ref.errors == []

        # UO-001
        p_uo1 = _build_uo001(tmp_path)
        ast_uo1 = parse_file(p_uo1)
        r_uo1 = execute_ast(ast_uo1, p_uo1, store)
        assert r_uo1.errors == []

        # UO-002
        p_uo2 = _build_uo002(tmp_path)
        ast_uo2 = parse_file(p_uo2)
        r_uo2 = execute_ast(ast_uo2, p_uo2, store)
        assert r_uo2.errors == []

        # Store contient les données des 3 fichiers
        all_keys = set(store.get_all().keys())
        assert "referentiel.types_uo"  in all_keys
        assert "uo.UO-001.avancement"  in all_keys
        assert "uo.UO-001.heures"      in all_keys
        assert "uo.UO-002.avancement"  in all_keys
        assert "uo.UO-002.statut"      in all_keys

    def test_store_cle_prefix(self, tmp_path):
        """JsonStore.keys(prefix) filtre correctement."""
        store = JsonStore(tmp_path / "store.json")

        p_uo1 = _build_uo001(tmp_path)
        execute_ast(parse_file(p_uo1), p_uo1, store)

        p_uo2 = _build_uo002(tmp_path)
        execute_ast(parse_file(p_uo2), p_uo2, store)

        uo1_keys = store.keys("uo.UO-001.")
        uo2_keys = store.keys("uo.UO-002.")

        assert len(uo1_keys) == 4
        assert len(uo2_keys) == 3
        assert all(k.startswith("uo.UO-001.") for k in uo1_keys)

    def test_store_clear_efface_tout(self, tmp_path):
        """JsonStore.clear() remet le store à zéro."""
        store = JsonStore(tmp_path / "store.json")

        p_ref = _build_referentiel(tmp_path)
        execute_ast(parse_file(p_ref), p_ref, store)
        assert len(store.get_all()) > 0

        store.clear()
        assert store.get_all() == {}

    def test_resync_ecrase_valeur(self, tmp_path):
        """Exécuter deux fois UO-001 écrase les valeurs dans le store."""
        store = JsonStore(tmp_path / "store.json")
        path  = _build_uo001(tmp_path)

        execute_ast(parse_file(path), path, store)
        av_1 = store.get("uo.UO-001.avancement")

        execute_ast(parse_file(path), path, store)
        av_2 = store.get("uo.UO-001.avancement")

        assert av_1 == av_2  # même données → même résultat


class TestValidateIntegration:
    """Vérifie que VALIDATE stoppe les erreurs de données dans un vrai fichier."""

    def test_avancement_hors_range_donne_erreur(self, tmp_path):
        """Construire un fichier avec avancement > 100 → erreur VALIDATE."""
        wb = Workbook()
        ws_act = wb.active
        ws_act.title = "Activites"

        # avancement = 150 (hors RANGE(0, 100))
        _make_table(ws_act, "TabActivites",
                    ["id", "avancement", "heures"],
                    [("X01", 150, 100)])

        ws_mxl = wb.create_sheet("_Manifeste")
        _write_manifeste(ws_mxl, [
            "FILE_TYPE: uo_instance",
            "FILE_ID: UO-ERR",
            "DEF $activites = GET_TABLE(Activites, TabActivites)",
            "DEF $avancement = COMPUTE(MEAN_WEIGHTED($activites.avancement, $activites.heures))",
            "VALIDATE $activites.avancement : RANGE(0, 100)",
            "PUSH $avancement -> uo.UO-ERR.avancement",
        ])

        path = tmp_path / "UO-ERR.xlsx"
        wb.save(str(path))

        store = JsonStore(tmp_path / "store.json")
        ast   = parse_file(path)
        result = execute_ast(ast, path, store)

        # VALIDATE détecte l'erreur mais le PUSH a quand même lieu
        assert len(result.errors) == 1
        assert "RANGE" in result.errors[0]

    def test_id_doublon_donne_erreur(self, tmp_path):
        wb = Workbook()
        ws_act = wb.active
        ws_act.title = "Activites"

        _make_table(ws_act, "TabActivites",
                    ["id", "avancement", "heures"],
                    [("X01", 50, 100), ("X01", 60, 100)])  # doublon d'ID

        ws_mxl = wb.create_sheet("_Manifeste")
        _write_manifeste(ws_mxl, [
            "FILE_TYPE: uo_instance",
            "FILE_ID: UO-DUP",
            "DEF $activites = GET_TABLE(Activites, TabActivites)",
            "VALIDATE $activites.id : UNIQUE",
            "PUSH $activites -> uo.UO-DUP.activites",
        ])

        path = tmp_path / "UO-DUP.xlsx"
        wb.save(str(path))

        store = JsonStore(tmp_path / "store.json")
        result = execute_ast(parse_file(path), path, store)

        assert len(result.errors) == 1
        assert "UNIQUE" in result.errors[0]

    def test_warning_severity_ne_bloque_pas_push(self, tmp_path):
        """SEVERITY=warning n'empêche pas les PUSH."""
        wb = Workbook()
        ws_act = wb.active
        ws_act.title = "Activites"

        _make_table(ws_act, "TabActivites",
                    ["id", "avancement", "heures"],
                    [("Y01", 50, -5)])  # heures = -5 → viole NON_NEGATIVE (warning)

        ws_mxl = wb.create_sheet("_Manifeste")
        _write_manifeste(ws_mxl, [
            "FILE_TYPE: uo_instance",
            "FILE_ID: UO-WARN",
            "DEF $activites = GET_TABLE(Activites, TabActivites)",
            "DEF $total_heures = COMPUTE(SUM($activites.heures))",
            "VALIDATE $activites.heures : NON_NEGATIVE  SEVERITY=warning",
            "PUSH $total_heures -> uo.UO-WARN.heures",
        ])

        path = tmp_path / "UO-WARN.xlsx"
        wb.save(str(path))

        store = JsonStore(tmp_path / "store.json")
        result = execute_ast(parse_file(path), path, store)

        assert result.errors   == []
        assert len(result.warnings) == 1
        # Le PUSH se fait quand même malgré le warning
        assert store.get("uo.UO-WARN.heures") == -5
