"""
generate_demo.py — Génère l'écosystème de démonstration ExoSync.
================================================================

Crée 7 fichiers Excel dans output/demo/ :

  UO-A1  Alice — Analyse système        (avancement moyen cible : 70 %)
  UO-A2  Alice — Validation fonct.      (avancement moyen cible : 40 %)
  UO-B1  Bruno — Architecture système   (avancement moyen cible : 80 %)
  UO-B2  Bruno — Essais terrain         (avancement moyen cible : 15 %)

  COCKPIT-ALICE  → PULL UO-A1 + UO-A2 → av par UO + global Alice
  COCKPIT-BRUNO  → PULL UO-B1 + UO-B2 → av par UO + global Bruno
  PILOTE         → PULL toutes activités → av Alice + av Bruno

Chaîne de données (store keys) :
  demo.alice.uo1.activites   ← poussé par UO-A1
  demo.alice.uo2.activites   ← poussé par UO-A2
  demo.bruno.uo1.activites   ← poussé par UO-B1
  demo.bruno.uo2.activites   ← poussé par UO-B2

Usage :
    python scripts/generate_demo.py
"""

import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo

DEMO_DIR = ROOT / "output" / "demo"
DEMO_DIR.mkdir(parents=True, exist_ok=True)


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _add_excel_table(ws, table_name: str, headers: list, rows: list,
                     start_row: int = 1, start_col: int = 1) -> Table:
    """
    Crée un tableau Excel openpyxl avec en-têtes + données.
    Retourne l'objet Table créé.
    """
    # En-têtes
    for ci, h in enumerate(headers, start_col):
        ws.cell(row=start_row, column=ci, value=h)

    # Données
    for ri, row_data in enumerate(rows, start_row + 1):
        for ci, h in enumerate(headers, start_col):
            ws.cell(row=ri, column=ci, value=row_data.get(h, ""))

    end_row = start_row + len(rows)
    end_col = start_col + len(headers) - 1
    ref = (f"{get_column_letter(start_col)}{start_row}:"
           f"{get_column_letter(end_col)}{end_row}")

    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)
    return tbl


def _add_named_range(wb, name: str, sheet: str, row: int, col: int) -> None:
    """Crée une plage nommée (DefinedName) dans le classeur."""
    col_letter = get_column_letter(col)
    attr_text = f"'{sheet}'!${col_letter}${row}"
    wb.defined_names[name] = DefinedName(name, attr_text=attr_text)


def _style_header(ws, row: int, col_start: int, col_end: int,
                  bg="1F4E79", fg="FFFFFF") -> None:
    """Colorie une ligne d'en-tête."""
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.font = Font(color=fg, bold=True)
        cell.alignment = Alignment(horizontal="center")


def _write_manifeste(ws, instructions: list) -> None:
    """
    Écrit les instructions MXL dans la feuille _Manifeste.
      Ligne 1  : MANIFESTE_V=1
      Ligne 2  : en-têtes (instruction | ancre)   ← sautée par le parser
      Lignes 3+: instructions MXL
    """
    ws["A1"] = "MANIFESTE_V=1"
    ws["A2"] = "instruction"
    ws["B2"] = "ancre"
    ws["A2"].font = Font(bold=True)
    ws["B2"].font = Font(bold=True)

    for i, line in enumerate(instructions, 3):
        ws.cell(row=i, column=1, value=line)

    # Largeur colonne A
    ws.column_dimensions["A"].width = 80


def _dashboard_kpi(ws, row: int, label: str, init_value=0) -> None:
    """Écrit un KPI label (col E) + valeur (col F) dans le Dashboard."""
    ws.cell(row=row, column=5, value=label)
    ws.cell(row=row, column=5).font = Font(bold=True)
    ws.cell(row=row, column=6, value=init_value)
    ws.cell(row=row, column=6).number_format = "0%"


# ─── Générateur UO ────────────────────────────────────────────────────────────

def generate_uo(file_id: str, store_key: str, activites: list, filename: str) -> Path:
    """
    Génère un fichier UO de démo.

    Structure :
      Activites   → table tbl_activites (uo_id, activite, avancement)
      Dashboard   → plage avancement_global (F3)
      _Manifeste  → MXL : GET_TABLE + AVG + PUSH + BIND
    """
    wb = Workbook()

    # ── Feuille Activites ──────────────────────────────────────────────────────
    ws_act = wb.active
    ws_act.title = "Activites"
    ws_act.column_dimensions["A"].width = 12
    ws_act.column_dimensions["B"].width = 35
    ws_act.column_dimensions["C"].width = 14

    headers = ["uo_id", "activite", "avancement"]
    _add_excel_table(ws_act, "tbl_activites", headers, activites)
    _style_header(ws_act, 1, 1, 3)

    # Formater la colonne avancement en %
    for ri in range(2, len(activites) + 2):
        ws_act.cell(row=ri, column=3).number_format = "0%"

    # ── Feuille Dashboard ──────────────────────────────────────────────────────
    ws_dash = wb.create_sheet("Dashboard")
    ws_dash["A1"] = f"Dashboard — {file_id}"
    ws_dash["A1"].font = Font(bold=True, size=14)
    _dashboard_kpi(ws_dash, 3, "Avancement global", 0)
    _add_named_range(wb, "avancement_global", "Dashboard", 3, 6)

    # ── Feuille _Manifeste ─────────────────────────────────────────────────────
    ws_man = wb.create_sheet("_Manifeste")
    manifeste = [
        f"FILE_TYPE: uo_instance",
        f"FILE_ID: {file_id}",
        "",
        "# Lecture des activités depuis la feuille Excel",
        "DEF $activites = GET_TABLE(Activites, tbl_activites)",
        "DEF $av_global = COMPUTE(AVG($activites.avancement))",
        "",
        "# Publication dans le store central",
        f"PUSH $activites -> {store_key}",
        "",
        "# Affichage dans le Dashboard",
        "BIND $av_global -> Dashboard.avancement_global",
    ]
    _write_manifeste(ws_man, manifeste)

    path = DEMO_DIR / filename
    wb.save(str(path))
    print(f"  [OK] {filename}")
    return path


# ─── Générateur Cockpit ───────────────────────────────────────────────────────

def _init_detail_sheet(ws, tbl_name: str, placeholder: list) -> None:
    """Initialise une feuille de détail avec une table placeholder."""
    headers = ["uo_id", "activite", "avancement"]
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 14
    _add_excel_table(ws, tbl_name, headers, placeholder)
    _style_header(ws, 1, 1, 3)
    for ri in range(2, len(placeholder) + 2):
        ws.cell(row=ri, column=3).number_format = "0%"


def generate_cockpit(file_id: str, engineer: str,
                     uo1_id: str, uo1_key: str, uo1_label: str,
                     uo2_id: str, uo2_key: str, uo2_label: str,
                     filename: str) -> Path:
    """
    Génère un fichier Cockpit de démo.

    Structure :
      Detail_UO1     → table tbl_uo1  (PULL OVERWRITE depuis store)
      Detail_UO2     → table tbl_uo2  (PULL OVERWRITE depuis store)
      Detail_Global  → table tbl_global (PULL OVERWRITE UO1 + APPEND_NEW UO2)
      Dashboard      → plages av_uo1, av_uo2, av_global
      _Manifeste     → MXL : 4 PULL + 6 DEF + 3 BIND
    """
    wb = Workbook()

    placeholder_1 = [{"uo_id": "-", "activite": "(en attente de synchro)", "avancement": 0}]
    placeholder_2 = [
        {"uo_id": "-", "activite": "(en attente UO1)", "avancement": 0},
        {"uo_id": "-", "activite": "(en attente UO2)", "avancement": 0},
    ]

    # ── Feuille Detail_UO1 (renommer la feuille active par défaut) ─────────────
    ws_d1 = wb.active
    ws_d1.title = "Detail_UO1"
    _init_detail_sheet(ws_d1, "tbl_uo1", placeholder_1)

    # ── Feuille Detail_UO2 ─────────────────────────────────────────────────────
    ws_d2 = wb.create_sheet("Detail_UO2")
    _init_detail_sheet(ws_d2, "tbl_uo2", placeholder_1)

    # ── Feuille Detail_Global ──────────────────────────────────────────────────
    ws_dg = wb.create_sheet("Detail_Global")
    _init_detail_sheet(ws_dg, "tbl_global", placeholder_2)

    # ── Feuille Dashboard ──────────────────────────────────────────────────────
    ws_dash = wb.create_sheet("Dashboard")
    ws_dash["A1"] = "Cockpit - %s" % engineer
    ws_dash["A1"].font = Font(bold=True, size=14)

    _dashboard_kpi(ws_dash, 3, "Avancement %s" % uo1_label, 0)
    _dashboard_kpi(ws_dash, 4, "Avancement %s" % uo2_label, 0)
    _dashboard_kpi(ws_dash, 5, "Avancement global", 0)

    _add_named_range(wb, "av_uo1", "Dashboard", 3, 6)
    _add_named_range(wb, "av_uo2", "Dashboard", 4, 6)
    _add_named_range(wb, "av_global", "Dashboard", 5, 6)

    # ── Feuille _Manifeste ─────────────────────────────────────────────────────
    ws_man = wb.create_sheet("_Manifeste")
    manifeste = [
        f"FILE_TYPE: cockpit",
        f"FILE_ID: {file_id}",
        "",
        f"# Récupérer les activités de chaque UO depuis le store",
        f"PULL {uo1_key} -> FILL_TABLE(Detail_UO1, tbl_uo1) MODE=OVERWRITE",
        f"PULL {uo2_key} -> FILL_TABLE(Detail_UO2, tbl_uo2) MODE=OVERWRITE",
        "",
        f"# Vue globale : UO1 (OVERWRITE) puis UO2 (APPEND)",
        f"PULL {uo1_key} -> FILL_TABLE(Detail_Global, tbl_global) MODE=OVERWRITE",
        f"PULL {uo2_key} -> FILL_TABLE(Detail_Global, tbl_global) MODE=APPEND_NEW KEY=activite",
        "",
        f"# Calcul des avancements",
        f"DEF $act_uo1 = GET_TABLE(Detail_UO1, tbl_uo1)",
        f"DEF $act_uo2 = GET_TABLE(Detail_UO2, tbl_uo2)",
        f"DEF $all = GET_TABLE(Detail_Global, tbl_global)",
        f"DEF $av_uo1 = COMPUTE(AVG($act_uo1.avancement))",
        f"DEF $av_uo2 = COMPUTE(AVG($act_uo2.avancement))",
        f"DEF $av_global = COMPUTE(AVG($all.avancement))",
        "",
        f"# Affichage dans le Dashboard",
        f"BIND $av_uo1 -> Dashboard.av_uo1",
        f"BIND $av_uo2 -> Dashboard.av_uo2",
        f"BIND $av_global -> Dashboard.av_global",
    ]
    _write_manifeste(ws_man, manifeste)

    path = DEMO_DIR / filename
    wb.save(str(path))
    print(f"  [OK] {filename}")
    return path


# ─── Générateur Pilote ────────────────────────────────────────────────────────

def generate_pilote(filename: str,
                    alice_keys: list, bruno_keys: list) -> Path:
    """
    Génère le tableau de bord Pilote.

    Structure :
      Donnees_Alice  → table tbl_alice  (PULL OVERWRITE UO-A1 + APPEND UO-A2)
      Donnees_Bruno  → table tbl_bruno  (PULL OVERWRITE UO-B1 + APPEND UO-B2)
      Dashboard      → plages av_alice, av_bruno
      _Manifeste     → MXL : 4 PULL + 4 DEF + 2 BIND
    """
    wb = Workbook()

    placeholder_2 = [
        {"uo_id": "-", "activite": "(en attente UO1)", "avancement": 0},
        {"uo_id": "-", "activite": "(en attente UO2)", "avancement": 0},
    ]

    # ── Feuille Donnees_Alice (renommer la feuille active par défaut) ──────────
    ws_alice = wb.active
    ws_alice.title = "Donnees_Alice"
    _init_detail_sheet(ws_alice, "tbl_alice", placeholder_2)

    # ── Feuille Donnees_Bruno ─────────────────────────────────────────────────
    ws_bruno = wb.create_sheet("Donnees_Bruno")
    _init_detail_sheet(ws_bruno, "tbl_bruno", placeholder_2)

    # ── Feuille Dashboard ──────────────────────────────────────────────────────
    ws_dash = wb.create_sheet("Dashboard")
    ws_dash["A1"] = "Tableau de bord Pilote"
    ws_dash["A1"].font = Font(bold=True, size=16)

    _dashboard_kpi(ws_dash, 3, "Avancement Alice Dupont", 0)
    _dashboard_kpi(ws_dash, 4, "Avancement Bruno Martin", 0)

    _add_named_range(wb, "av_alice", "Dashboard", 3, 6)
    _add_named_range(wb, "av_bruno", "Dashboard", 4, 6)

    # ── Feuille _Manifeste ─────────────────────────────────────────────────────
    ws_man = wb.create_sheet("_Manifeste")
    a1_key, a2_key = alice_keys
    b1_key, b2_key = bruno_keys
    manifeste = [
        "FILE_TYPE: pilote",
        "FILE_ID: DEMO-PILOTE",
        "",
        "# Charger toutes les activités Alice",
        f"PULL {a1_key} -> FILL_TABLE(Donnees_Alice, tbl_alice) MODE=OVERWRITE",
        f"PULL {a2_key} -> FILL_TABLE(Donnees_Alice, tbl_alice) MODE=APPEND_NEW KEY=activite",
        "",
        "# Charger toutes les activités Bruno",
        f"PULL {b1_key} -> FILL_TABLE(Donnees_Bruno, tbl_bruno) MODE=OVERWRITE",
        f"PULL {b2_key} -> FILL_TABLE(Donnees_Bruno, tbl_bruno) MODE=APPEND_NEW KEY=activite",
        "",
        "# Calcul des avancements par ingénieur",
        "DEF $data_alice = GET_TABLE(Donnees_Alice, tbl_alice)",
        "DEF $data_bruno = GET_TABLE(Donnees_Bruno, tbl_bruno)",
        "DEF $av_alice = COMPUTE(AVG($data_alice.avancement))",
        "DEF $av_bruno = COMPUTE(AVG($data_bruno.avancement))",
        "",
        "# Affichage dans le Dashboard",
        "BIND $av_alice -> Dashboard.av_alice",
        "BIND $av_bruno -> Dashboard.av_bruno",
    ]
    _write_manifeste(ws_man, manifeste)

    path = DEMO_DIR / filename
    wb.save(str(path))
    print(f"  [OK] {filename}")
    return path


# ─── Données de la démo ───────────────────────────────────────────────────────

ACTIVITES_A1 = [
    {"uo_id": "UO-A1", "activite": "A1-Analyse fonctionnelle", "avancement": 0.80},
    {"uo_id": "UO-A1", "activite": "A1-Redaction specification", "avancement": 0.60},
]
# Moyenne A1 = 70 %

ACTIVITES_A2 = [
    {"uo_id": "UO-A2", "activite": "A2-Plan de test", "avancement": 0.50},
    {"uo_id": "UO-A2", "activite": "A2-Execution tests", "avancement": 0.30},
]
# Moyenne A2 = 40 %
# Moyenne Alice (4 activités) = 55 %

ACTIVITES_B1 = [
    {"uo_id": "UO-B1", "activite": "B1-Architecture systeme", "avancement": 0.90},
    {"uo_id": "UO-B1", "activite": "B1-Dimensionnement", "avancement": 0.70},
]
# Moyenne B1 = 80 %

ACTIVITES_B2 = [
    {"uo_id": "UO-B2", "activite": "B2-Essais terrain", "avancement": 0.20},
    {"uo_id": "UO-B2", "activite": "B2-Rapport final", "avancement": 0.10},
]
# Moyenne B2 = 15 %
# Moyenne Bruno (4 activités) = 47.5 %


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("Generation de l'ecosysteme de demo dans : %s\n" % DEMO_DIR)

    print("-- UOs Alice --")
    generate_uo(
        file_id="DEMO-UO-A1",
        store_key="demo.alice.uo1.activites",
        activites=ACTIVITES_A1,
        filename="UO-A1_analyse_systeme.xlsx",
    )
    generate_uo(
        file_id="DEMO-UO-A2",
        store_key="demo.alice.uo2.activites",
        activites=ACTIVITES_A2,
        filename="UO-A2_validation_fonct.xlsx",
    )

    print("\n-- UOs Bruno --")
    generate_uo(
        file_id="DEMO-UO-B1",
        store_key="demo.bruno.uo1.activites",
        activites=ACTIVITES_B1,
        filename="UO-B1_architecture_sys.xlsx",
    )
    generate_uo(
        file_id="DEMO-UO-B2",
        store_key="demo.bruno.uo2.activites",
        activites=ACTIVITES_B2,
        filename="UO-B2_essais_terrain.xlsx",
    )

    print("\n-- Cockpits --")
    generate_cockpit(
        file_id="DEMO-COCKPIT-ALICE",
        engineer="Alice Dupont",
        uo1_id="DEMO-UO-A1",
        uo1_key="demo.alice.uo1.activites",
        uo1_label="UO-A1 (Analyse)",
        uo2_id="DEMO-UO-A2",
        uo2_key="demo.alice.uo2.activites",
        uo2_label="UO-A2 (Validation)",
        filename="COCKPIT-ALICE.xlsx",
    )
    generate_cockpit(
        file_id="DEMO-COCKPIT-BRUNO",
        engineer="Bruno Martin",
        uo1_id="DEMO-UO-B1",
        uo1_key="demo.bruno.uo1.activites",
        uo1_label="UO-B1 (Architecture)",
        uo2_id="DEMO-UO-B2",
        uo2_key="demo.bruno.uo2.activites",
        uo2_label="UO-B2 (Essais)",
        filename="COCKPIT-BRUNO.xlsx",
    )

    print("\n-- Pilote --")
    generate_pilote(
        filename="PILOTE.xlsx",
        alice_keys=["demo.alice.uo1.activites", "demo.alice.uo2.activites"],
        bruno_keys=["demo.bruno.uo1.activites", "demo.bruno.uo2.activites"],
    )

    print("\n[OK] 7 fichiers generes dans %s" % DEMO_DIR)
    print("\nResultats attendus apres sync :")
    print("  COCKPIT-ALICE  : UO-A1=70%  UO-A2=40%  Global=55%")
    print("  COCKPIT-BRUNO  : UO-B1=80%  UO-B2=15%  Global=47.5%")
    print("  PILOTE         : Alice=55%  Bruno=47.5%")
    print("\nLancer la synchronisation :")
    print("  python scripts/sync_demo.py")


if __name__ == "__main__":
    main()
