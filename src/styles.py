from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, Rule

# ─── Colours ──────────────────────────────────────────────────────────────────
BLUE_DARK = "1F3864"
BLUE_MID = "2F5496"
BLUE_LIGHT = "D6E4F0"
GREEN = "70AD47"
GREEN_LIGHT = "E2EFDA"
ORANGE = "ED7D31"
ORANGE_LIGHT = "FCE4D6"
RED = "C00000"
RED_LIGHT = "FFDCDC"
GREY_LIGHT = "F2F2F2"
GREY_MID = "D9D9D9"
WHITE = "FFFFFF"
YELLOW_LIGHT = "FFFF99"

# ─── Thin border helper ────────────────────────────────────────────────────────
_THIN = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def header_fill(color: str = BLUE_DARK) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def solid_fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def header_font(size: int = 11, color: str = WHITE) -> Font:
    return Font(bold=True, color=color, size=size)


def body_font(size: int = 11, bold: bool = False, color: str = "000000") -> Font:
    return Font(bold=bold, size=size, color=color)


def center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ─── Apply helpers ─────────────────────────────────────────────────────────────

def style_header_row(ws, row: int, col_start: int, col_end: int, color: str = BLUE_DARK):
    """Apply header styling to a range of cells in a single row."""
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill(color)
        cell.font = header_font()
        cell.alignment = center()
        cell.border = THIN_BORDER


def style_data_row(ws, row: int, col_start: int, col_end: int, alternate: bool = False):
    """Apply alternating row styling."""
    fill_color = GREY_LIGHT if alternate else WHITE
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = solid_fill(fill_color)
        cell.font = body_font()
        cell.alignment = left()
        cell.border = THIN_BORDER


def set_column_widths(ws, widths: dict):
    """widths = {col_letter: width} e.g. {'A': 15, 'B': 30}"""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def freeze_top_row(ws):
    ws.freeze_panes = "A2"
