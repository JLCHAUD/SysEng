"""
Génère les fichiers Excel de test pour la feature LIST + COLLECT.

Usage :
    python tests/generate_test_fixtures.py [dossier_sortie]

Par défaut, les fichiers sont créés dans tests/fixtures/list_collect/.
"""

import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableColumn
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl est requis : pip install openpyxl")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _write_table(ws, table_name: str, headers: list[str], rows: list[list], start_row: int = 1):
    """
    Écrit headers + rows dans ws à partir de start_row,
    puis déclare une Table Excel nommée table_name.
    """
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col_idx, value=header)
    for row_idx, row in enumerate(rows, start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    end_row   = start_row + len(rows)
    end_col   = get_column_letter(len(headers))
    ref       = f"A{start_row}:{end_col}{end_row}"
    tbl       = Table(displayName=table_name, ref=ref)
    ws.add_table(tbl)


def _write_empty_table(ws, table_name: str, headers: list[str], start_row: int = 1):
    """Écrit uniquement la ligne d'en-tête + une ligne vide (nécessaire pour une Table valide)."""
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col_idx, value=header)
    # ligne vide obligatoire pour que la Table soit valide en openpyxl
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=start_row + 1, column=col_idx, value=None)

    end_col = get_column_letter(len(headers))
    ref     = f"A{start_row}:{end_col}{start_row + 1}"
    tbl     = Table(displayName=table_name, ref=ref)
    ws.add_table(tbl)


def _manifeste_lines(file_type: str, file_id: str, **extra_meta) -> str:
    """Génère le contenu MXL minimal pour la feuille _Manifeste."""
    lines = [
        f"FILE_TYPE: {file_type}",
        f"FILE_ID:   {file_id}",
    ]
    for k, v in extra_meta.items():
        lines.append(f"{k}: {v}")
    return "\n".join(lines)


def _write_manifeste_sheet(ws, content: str):
    for row_idx, line in enumerate(content.splitlines(), start=1):
        ws.cell(row=row_idx, column=1, value=line)


# ---------------------------------------------------------------------------
# Fichiers fils
# ---------------------------------------------------------------------------

FILS = [
    {
        "file_id":       "fils_uo1",
        "nom_ingenieur": "Alice Dubois",
        "projet":        "MI20",
        "planning_rows": [
            ["T01", "Etude préliminaire",    "2026-06-01", "2026-06-15", "en-cours"],
            ["T02", "Rédaction note calcul", "2026-06-16", "2026-07-01", "à-faire"],
            ["T03", "Revue interne",         "2026-07-02", "2026-07-05", "à-faire"],
        ],
        "risques_rows": [
            ["R01", "Retard fournisseur",  2, "ouvert"],
            ["R02", "Ressource critique",  4, "ouvert"],
        ],
        "livrables_rows": [
            ["L01", "Note de calcul v1", "en-cours"],
            ["L02", "Plan d'ensemble",   "à-faire"],
        ],
    },
    {
        "file_id":       "fils_uo2",
        "nom_ingenieur": "Bruno Lecomte",
        "projet":        "MI20",
        "planning_rows": [
            ["T04", "Conception détaillée", "2026-06-01", "2026-06-30", "en-cours"],
            ["T05", "Vérification",         "2026-07-01", "2026-07-10", "à-faire"],
        ],
        "risques_rows": [
            ["R03", "Non-conformité interface", 3, "ouvert"],
            ["R04", "Dérive planning",          5, "critique"],
        ],
        "livrables_rows": [
            ["L03", "Rapport de vérification", "à-faire"],
        ],
    },
    {
        "file_id":       "fils_uo3",
        "nom_ingenieur": "Camille Vidal",
        "projet":        "MI22",
        "planning_rows": [
            ["T06", "Analyse des besoins", "2026-07-01", "2026-07-20", "à-faire"],
        ],
        "risques_rows": [
            ["R05", "Risque faible test", 1, "fermé"],
        ],
        "livrables_rows": [
            ["L04", "Cahier des charges", "à-faire"],
        ],
    },
]


def create_fils(output_dir: Path):
    for fils in FILS:
        wb = Workbook()

        # --- Feuille _Manifeste ---
        ws_man = wb.active
        ws_man.title = "_Manifeste"
        _write_manifeste_sheet(ws_man, _manifeste_lines(
            file_type     = "uo_instance",
            file_id       = fils["file_id"],
            owner         = fils["nom_ingenieur"].split()[0],
            projet        = fils["projet"],
        ))

        # --- Feuille Planning ---
        ws_pl = wb.create_sheet("Planning")
        _write_table(
            ws_pl, "Planning",
            headers = ["id", "tache", "date_debut", "date_fin", "statut"],
            rows    = fils["planning_rows"],
        )

        # --- Feuille Risques ---
        ws_ri = wb.create_sheet("Risques")
        _write_table(
            ws_ri, "Risques",
            headers = ["id", "libelle", "criticite", "statut"],
            rows    = fils["risques_rows"],
        )

        # --- Feuille Livrables ---
        ws_li = wb.create_sheet("Livrables")
        _write_table(
            ws_li, "Livrables",
            headers = ["id", "libelle", "statut"],
            rows    = fils["livrables_rows"],
        )

        path = output_dir / f"{fils['file_id']}.xlsx"
        wb.save(path)
        print(f"  créé : {path}")


# ---------------------------------------------------------------------------
# Fichier père (synthese_mi20)
# ---------------------------------------------------------------------------

def create_pere(output_dir: Path):
    wb = Workbook()

    # --- Feuille _Manifeste (père) ---
    ws_man = wb.active
    ws_man.title = "_Manifeste"
    manifeste_content = "\n".join([
        "FILE_TYPE: synthese",
        "FILE_ID:   synthese_mi20",
        "owner:     Jean",
        "projet:    MI20",
        "",
        "# Liste depuis la table Excel du père",
        "LIST UOs_actifs FROM TABLE liste_uo",
        "",
        "# Agrégations",
        "COLLECT Planning  FROM UOs_actifs INTO vue_planning",
        "COLLECT Risques   FROM UOs_actifs INTO vue_risques  WHERE criticite >= 3",
        "COLLECT Livrables FROM UOs_actifs INTO vue_livrables COLS=[id, libelle, statut]",
    ])
    _write_manifeste_sheet(ws_man, manifeste_content)

    # --- Feuille listes : table liste_uo (source de LIST FROM TABLE) ---
    # Colonnes : FILE_ID (obligatoire) + colonnes contextuelles
    ws_listes = wb.create_sheet("listes")
    _write_table(
        ws_listes, "liste_uo",
        headers = ["FILE_ID", "Nom_ingenieur", "Projet"],
        rows    = [
            ["fils_uo1", "Alice Dubois",   "MI20"],
            ["fils_uo2", "Bruno Lecomte",  "MI20"],
            ["fils_uo3", "Camille Vidal",  "MI22"],
        ],
    )

    # --- Feuilles cibles COLLECT (tables vides prêtes à recevoir les données) ---
    for table_name, headers in [
        ("vue_planning",   ["_source_file_id", "Nom_ingenieur", "Projet", "id", "tache", "date_debut", "date_fin", "statut"]),
        ("vue_risques",    ["_source_file_id", "Nom_ingenieur", "Projet", "id", "libelle", "criticite", "statut"]),
        ("vue_livrables",  ["_source_file_id", "Nom_ingenieur", "Projet", "id", "libelle", "statut"]),
    ]:
        ws = wb.create_sheet(table_name)
        _write_empty_table(ws, table_name, headers)

    path = output_dir / "synthese_mi20.xlsx"
    wb.save(path)
    print(f"  créé : {path}")


# ---------------------------------------------------------------------------
# Fichier père alternatif : synthese_mi20_partial
# (fils_uo3 absent → test COLLECT_FILE_NOT_FOUND)
# ---------------------------------------------------------------------------

def create_pere_partial(output_dir: Path):
    """Père dont la liste_uo référence fils_uo_absent (fichier inexistant)."""
    wb = Workbook()

    ws_man = wb.active
    ws_man.title = "_Manifeste"
    manifeste_content = "\n".join([
        "FILE_TYPE: synthese",
        "FILE_ID:   synthese_partial",
        "owner:     Jean",
        "projet:    MI20",
        "",
        "LIST UOs_partiel FROM TABLE liste_uo_partiel",
        "COLLECT Planning FROM UOs_partiel INTO vue_planning_partiel",
    ])
    _write_manifeste_sheet(ws_man, manifeste_content)

    ws_listes = wb.create_sheet("listes")
    _write_table(
        ws_listes, "liste_uo_partiel",
        headers = ["FILE_ID", "Nom_ingenieur", "Projet"],
        rows    = [
            ["fils_uo1",      "Alice Dubois",   "MI20"],
            ["fils_uo_absent","Inconnu",         "MI99"],   # ← ce fichier n'existera pas
        ],
    )

    ws = wb.create_sheet("vue_planning_partiel")
    _write_empty_table(
        ws, "vue_planning_partiel",
        ["_source_file_id", "Nom_ingenieur", "Projet", "id", "tache", "date_debut", "date_fin", "statut"],
    )

    path = output_dir / "synthese_partial.xlsx"
    wb.save(path)
    print(f"  créé : {path}")


# ---------------------------------------------------------------------------
# Fichier père : test COLLECT sans LIST déclarée (COLLECT_LIST_UNKNOWN)
# ---------------------------------------------------------------------------

def create_pere_orphan_collect(output_dir: Path):
    """Père dont le COLLECT référence une liste non déclarée."""
    wb = Workbook()

    ws_man = wb.active
    ws_man.title = "_Manifeste"
    manifeste_content = "\n".join([
        "FILE_TYPE: synthese",
        "FILE_ID:   synthese_orphan",
        "",
        "# Pas de LIST déclarée !",
        "COLLECT Planning FROM liste_inconnue INTO vue_planning_orphan",
    ])
    _write_manifeste_sheet(ws_man, manifeste_content)

    ws = wb.create_sheet("vue_planning_orphan")
    _write_empty_table(ws, "vue_planning_orphan", ["_source_file_id", "id", "tache"])

    path = output_dir / "synthese_orphan.xlsx"
    wb.save(path)
    print(f"  créé : {path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    output_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).parent / "fixtures" / "list_collect"
    output_dir.mkdir(parents=True, exist_ok=True)
    print(f"\nGénération des fixtures LIST/COLLECT dans : {output_dir}\n")

    print("── Fichiers fils ──")
    create_fils(output_dir)

    print("\n── Fichier père principal ──")
    create_pere(output_dir)

    print("\n── Fichier père partiel (test fichier manquant) ──")
    create_pere_partial(output_dir)

    print("\n── Fichier père orphan (test COLLECT sans LIST) ──")
    create_pere_orphan_collect(output_dir)

    print(f"\n✓ {len(list(output_dir.glob('*.xlsx')))} fichiers générés dans {output_dir}")
    print("\nStructure attendue :")
    print("  synthese_mi20.xlsx      ← père principal")
    print("    liste_uo : fils_uo1, fils_uo2, fils_uo3")
    print("    COLLECT Planning  → vue_planning   (6 lignes attendues)")
    print("    COLLECT Risques   → vue_risques    WHERE criticite>=3 (3 lignes : R02, R03, R04)")
    print("    COLLECT Livrables → vue_livrables  COLS=[id,libelle,statut]")
    print("  fils_uo1.xlsx           ← Planning(3), Risques(2), Livrables(2)")
    print("  fils_uo2.xlsx           ← Planning(2), Risques(2), Livrables(1)")
    print("  fils_uo3.xlsx           ← Planning(1), Risques(1), Livrables(1)")
    print("  synthese_partial.xlsx   ← père avec 1 fils manquant")
    print("  synthese_orphan.xlsx    ← père avec COLLECT sans LIST")


if __name__ == "__main__":
    main()
