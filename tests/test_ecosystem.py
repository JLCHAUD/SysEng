"""
Tests — M03 Écosystème / Exomap v2
====================================
Couvre :
  1. FileRecord / EdgeRecord / sérialisation round-trip
  2. record_file_sync()
  3. record_edges_from_ast()
  4. check_consistency() — 3 règles
  5. lineage_text() / lineage_dict()
  6. Migration v1 → v2
  7. Intégration avec un vrai fichier Excel (via test_integration_real helpers)
"""
import json
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table

from src.ecosystem import (
    EcosystemSchema, FileRecord, EdgeRecord, ColumnSchema, TableSchema, VariableSchema,
    ConsistencyWarning,
    load, save,
    register_many, record_file_sync, record_edges_from_ast,
    check_consistency, lineage_text, lineage_dict, summary,
)
from src.parser import PasserelleAST, FileHeader, PullNode, PushNode, DefNode


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _eco(tmp_path: Path) -> Path:
    """Retourne le chemin d'un ecosystem.json isolé dans tmp_path."""
    return tmp_path / "ecosystem.json"


def _ast_with_edges(file_id="UO-001", pulls=None, pushes=None) -> PasserelleAST:
    """Construit un AST minimal avec des PULL et PUSH."""
    ast = PasserelleAST()
    ast.header = FileHeader(file_type="uo_instance", file_id=file_id)
    ast.pulls  = pulls  or []
    ast.pushes = pushes or []
    ast.defs   = []
    return ast


# ─── 1. Sérialisation round-trip ──────────────────────────────────────────────

class TestSerialisation:

    def test_schema_vide_round_trip(self, tmp_path):
        eco = _eco(tmp_path)
        schema = EcosystemSchema()
        save(schema, eco)
        loaded = load(eco)
        assert loaded.version == "2.0"
        assert loaded.files == {}
        assert loaded.edges == []

    def test_file_record_round_trip(self, tmp_path):
        eco = _eco(tmp_path)
        schema = EcosystemSchema()
        schema.files["UO-001"] = FileRecord(
            file_id="UO-001",
            path="UOs/UO-001.xlsx",
            file_type="uo_instance",
            status="ok",
        )
        save(schema, eco)
        loaded = load(eco)
        assert "UO-001" in loaded.files
        assert loaded.files["UO-001"].file_type == "uo_instance"
        assert loaded.files["UO-001"].status == "ok"

    def test_edge_record_round_trip(self, tmp_path):
        eco = _eco(tmp_path)
        schema = EcosystemSchema()
        schema.edges.append(EdgeRecord(
            edge_type="PUSH",
            from_node="UO-001::$avancement",
            to_node="store::uo.UO-001.avancement",
        ))
        schema.edges.append(EdgeRecord(
            edge_type="PULL",
            from_node="store::projet.acteurs",
            to_node="UO-001::TabActeurs",
            mode="OVERWRITE",
        ))
        save(schema, eco)
        loaded = load(eco)
        assert len(loaded.edges) == 2
        types = {e.edge_type for e in loaded.edges}
        assert types == {"PUSH", "PULL"}

    def test_migration_v1_vers_v2(self, tmp_path):
        """Un ecosystem.json v1 (sans 'files' ni 'edges') est migré automatiquement."""
        eco = _eco(tmp_path)
        v1 = {
            "version": "1",
            "tables": {},
            "variables": {},
        }
        eco.write_text(json.dumps(v1), encoding="utf-8")
        loaded = load(eco)
        assert loaded.version == "2.0"
        assert loaded.files == {}
        assert loaded.edges == []


# ─── 2. record_file_sync ──────────────────────────────────────────────────────

class TestRecordFileSync:

    def test_enregistre_nouveau_fichier(self, tmp_path):
        eco = _eco(tmp_path)
        record_file_sync("UO-001", "UOs/UO-001.xlsx", "uo_instance", "ok",
                         ecosystem_path=eco)
        schema = load(eco)
        assert "UO-001" in schema.files
        frec = schema.files["UO-001"]
        assert frec.file_type == "uo_instance"
        assert frec.status == "ok"
        assert frec.last_sync is not None

    def test_met_a_jour_fichier_existant(self, tmp_path):
        eco = _eco(tmp_path)
        record_file_sync("UO-001", "UOs/UO-001.xlsx", "uo_instance", "error",
                         ecosystem_path=eco)
        record_file_sync("UO-001", "UOs/UO-001.xlsx", "uo_instance", "ok",
                         ecosystem_path=eco)
        schema = load(eco)
        assert schema.files["UO-001"].status == "ok"

    def test_plusieurs_fichiers(self, tmp_path):
        eco = _eco(tmp_path)
        for fid, ftype in [("REF-001", "referentiel_uo"),
                            ("UO-001", "uo_instance"),
                            ("UO-002", "uo_instance")]:
            record_file_sync(fid, f"{fid}.xlsx", ftype, "ok", ecosystem_path=eco)
        schema = load(eco)
        assert len(schema.files) == 3


# ─── 3. record_edges_from_ast ─────────────────────────────────────────────────

class TestRecordEdgesFromAst:

    def test_enregistre_push(self, tmp_path):
        eco = _eco(tmp_path)
        ast = _ast_with_edges(
            file_id="UO-001",
            pushes=[PushNode("$avancement", "uo.UO-001.avancement")],
        )
        record_edges_from_ast(ast, "UO-001", ecosystem_path=eco)
        schema = load(eco)
        push_edges = [e for e in schema.edges if e.edge_type == "PUSH"]
        assert len(push_edges) == 1
        assert push_edges[0].from_node == "UO-001::$avancement"
        assert push_edges[0].to_node   == "store::uo.UO-001.avancement"

    def test_enregistre_pull(self, tmp_path):
        eco = _eco(tmp_path)
        ast = _ast_with_edges(
            file_id="UO-001",
            pulls=[PullNode("projet.acteurs", "FILL_TABLE", "Org", "TabActeurs", "OVERWRITE")],
        )
        record_edges_from_ast(ast, "UO-001", ecosystem_path=eco)
        schema = load(eco)
        pull_edges = [e for e in schema.edges if e.edge_type == "PULL"]
        assert len(pull_edges) == 1
        assert pull_edges[0].from_node == "store::projet.acteurs"
        assert pull_edges[0].to_node   == "UO-001::TabActeurs"
        assert pull_edges[0].mode      == "OVERWRITE"

    def test_plusieurs_push_et_pull(self, tmp_path):
        eco = _eco(tmp_path)
        ast = _ast_with_edges(
            file_id="UO-001",
            pulls=[
                PullNode("projet.acteurs", "FILL_TABLE", "Org", "TabActeurs", "OVERWRITE"),
                PullNode("ref.types",      "FILL_TABLE", "Types", "TabTypes", "APPEND_NEW"),
            ],
            pushes=[
                PushNode("$avancement", "uo.UO-001.avancement"),
                PushNode("$heures",     "uo.UO-001.heures"),
            ],
        )
        record_edges_from_ast(ast, "UO-001", ecosystem_path=eco)
        schema = load(eco)
        assert len(schema.edges) == 4

    def test_re_sync_remplace_anciens_arcs(self, tmp_path):
        """Resynchroniser un fichier remplace ses arcs précédents."""
        eco = _eco(tmp_path)
        ast1 = _ast_with_edges(
            file_id="UO-001",
            pushes=[PushNode("$av", "uo.UO-001.av")],
        )
        record_edges_from_ast(ast1, "UO-001", ecosystem_path=eco)

        # Deuxième sync avec un push différent
        ast2 = _ast_with_edges(
            file_id="UO-001",
            pushes=[PushNode("$heures", "uo.UO-001.heures")],
        )
        record_edges_from_ast(ast2, "UO-001", ecosystem_path=eco)

        schema = load(eco)
        to_nodes = {e.to_node for e in schema.edges}
        assert "store::uo.UO-001.heures" in to_nodes
        assert "store::uo.UO-001.av" not in to_nodes

    def test_deux_fichiers_arcs_independants(self, tmp_path):
        """Les arcs de deux fichiers coexistent sans collision."""
        eco = _eco(tmp_path)
        for fid in ("UO-001", "UO-002"):
            ast = _ast_with_edges(
                file_id=fid,
                pushes=[PushNode("$av", f"uo.{fid}.avancement")],
            )
            record_edges_from_ast(ast, fid, ecosystem_path=eco)
        schema = load(eco)
        assert len(schema.edges) == 2


# ─── 4. check_consistency ─────────────────────────────────────────────────────

class TestCheckConsistency:

    def _setup_eco(self, tmp_path, edges, files=None) -> Path:
        eco = _eco(tmp_path)
        schema = EcosystemSchema()
        schema.edges = [EdgeRecord(**e) for e in edges]
        if files:
            for fid, frec in files.items():
                schema.files[fid] = FileRecord(**frec)
        save(schema, eco)
        return eco

    def test_pas_de_warning_si_coherent(self, tmp_path):
        eco = self._setup_eco(tmp_path, edges=[
            {"edge_type": "PUSH", "from_node": "REF::$types",  "to_node": "store::ref.types"},
            {"edge_type": "PULL", "from_node": "store::ref.types", "to_node": "UO-001::TabTypes"},
        ])
        warnings = check_consistency(eco)
        assert warnings == []

    def test_pull_jamais_pushed(self, tmp_path):
        """UO-001 PULL depuis store::ref.types mais personne ne PUSH ref.types."""
        eco = self._setup_eco(tmp_path, edges=[
            {"edge_type": "PULL", "from_node": "store::ref.types", "to_node": "UO-001::TabTypes"},
        ])
        warnings = check_consistency(eco)
        codes = [w.code for w in warnings]
        assert "PULL_NEVER_PUSHED" in codes

    def test_conflit_push(self, tmp_path):
        """UO-001 et UO-002 pushent tous les deux vers store::uo.avancement."""
        eco = self._setup_eco(tmp_path, edges=[
            {"edge_type": "PUSH", "from_node": "UO-001::$av", "to_node": "store::uo.avancement"},
            {"edge_type": "PUSH", "from_node": "UO-002::$av", "to_node": "store::uo.avancement"},
        ])
        warnings = check_consistency(eco)
        codes = [w.code for w in warnings]
        assert "PUSH_CONFLICT" in codes

    def test_fichier_jamais_sync(self, tmp_path):
        eco = self._setup_eco(tmp_path, edges=[], files={
            "UO-001": {"file_id": "UO-001", "path": "UO-001.xlsx",
                       "file_type": "uo_instance", "last_sync": None, "status": "unknown"},
        })
        warnings = check_consistency(eco)
        codes = [w.code for w in warnings]
        assert "STALE_FILE" in codes

    def test_fichier_sync_ok_pas_de_warning_stale(self, tmp_path):
        eco = self._setup_eco(tmp_path, edges=[], files={
            "UO-001": {"file_id": "UO-001", "path": "UO-001.xlsx",
                       "file_type": "uo_instance", "last_sync": "2026-04-24T10:00:00", "status": "ok"},
        })
        warnings = check_consistency(eco)
        codes = [w.code for w in warnings]
        assert "STALE_FILE" not in codes

    def test_plusieurs_warnings_independants(self, tmp_path):
        """Plusieurs règles peuvent se déclencher en même temps."""
        eco = self._setup_eco(tmp_path, edges=[
            # PULL jamais pushé
            {"edge_type": "PULL", "from_node": "store::ref.types", "to_node": "UO-001::Tab"},
            # Conflit PUSH
            {"edge_type": "PUSH", "from_node": "UO-001::$x", "to_node": "store::shared.x"},
            {"edge_type": "PUSH", "from_node": "UO-002::$x", "to_node": "store::shared.x"},
        ], files={
            "STALE": {"file_id": "STALE", "path": "stale.xlsx",
                      "file_type": "uo_instance", "last_sync": None, "status": "unknown"},
        })
        warnings = check_consistency(eco)
        codes = {w.code for w in warnings}
        assert "PULL_NEVER_PUSHED" in codes
        assert "PUSH_CONFLICT"    in codes
        assert "STALE_FILE"       in codes


# ─── 5. lineage_text / lineage_dict ──────────────────────────────────────────

class TestLineage:

    def test_lineage_vide(self, tmp_path):
        eco = _eco(tmp_path)
        text = lineage_text(path=eco)
        assert "vide" in text.lower() or "aucun" in text.lower()

    def test_lineage_affiche_fichier(self, tmp_path):
        eco = _eco(tmp_path)
        record_file_sync("UO-001", "UO-001.xlsx", "uo_instance", "ok",
                         ecosystem_path=eco)
        ast = _ast_with_edges(
            file_id="UO-001",
            pushes=[PushNode("$av", "uo.UO-001.avancement")],
        )
        record_edges_from_ast(ast, "UO-001", ecosystem_path=eco)
        text = lineage_text(path=eco)
        assert "UO-001" in text
        assert "PUSH" in text
        assert "uo.UO-001.avancement" in text

    def test_lineage_filtre_par_fichier(self, tmp_path):
        eco = _eco(tmp_path)
        for fid in ("UO-001", "UO-002"):
            record_file_sync(fid, f"{fid}.xlsx", "uo_instance", "ok", ecosystem_path=eco)
            ast = _ast_with_edges(fid, pushes=[PushNode("$av", f"uo.{fid}.av")])
            record_edges_from_ast(ast, fid, ecosystem_path=eco)

        text = lineage_text(file_id="UO-001", path=eco)
        assert "UO-001" in text
        assert "UO-002" not in text

    def test_lineage_dict_stats(self, tmp_path):
        eco = _eco(tmp_path)
        record_file_sync("UO-001", "UO-001.xlsx", "uo_instance", "ok", ecosystem_path=eco)
        ast = _ast_with_edges(
            file_id="UO-001",
            pulls=[PullNode("ref.types", "FILL_TABLE", "Types", "TabT", "OVERWRITE")],
            pushes=[PushNode("$av", "uo.UO-001.av")],
        )
        record_edges_from_ast(ast, "UO-001", ecosystem_path=eco)

        d = lineage_dict(eco)
        assert d["stats"]["nb_files"]      == 1
        assert d["stats"]["nb_push_edges"] == 1
        assert d["stats"]["nb_pull_edges"] == 1
        assert d["stats"]["nb_edges"]      == 2

    def test_lineage_dict_json_serializable(self, tmp_path):
        eco = _eco(tmp_path)
        record_file_sync("UO-001", "UO-001.xlsx", "uo_instance", "ok", ecosystem_path=eco)
        d = lineage_dict(eco)
        # Ne doit pas lever
        json_str = json.dumps(d)
        assert "UO-001" in json_str


# ─── 6. summary ───────────────────────────────────────────────────────────────

class TestSummary:

    def test_summary_vide(self, tmp_path):
        eco = _eco(tmp_path)
        schema = EcosystemSchema()
        save(schema, eco)

        # patch ECOSYSTEM_PATH pour ce test
        import src.ecosystem as eco_mod
        original = eco_mod.ECOSYSTEM_PATH
        eco_mod.ECOSYSTEM_PATH = eco
        try:
            s = summary()
            assert s["nb_files"] == 0
            assert s["nb_edges"] == 0
        finally:
            eco_mod.ECOSYSTEM_PATH = original

    def test_summary_apres_enregistrement(self, tmp_path):
        eco = _eco(tmp_path)
        record_file_sync("UO-001", "UO-001.xlsx", "uo_instance", "ok", ecosystem_path=eco)

        s = summary(ecosystem_path=eco)
        assert s["nb_files"] == 1


# ─── 7. Intégration réelle avec execute_ast ──────────────────────────────────

class TestIntegrationEcosystemAvecExcel:
    """
    Vérifie que record_edges_from_ast() reflète correctement
    un AST produit par parse_file() sur un vrai xlsx.
    """

    def test_ast_reel_donne_arcs_corrects(self, tmp_path):
        from openpyxl.workbook.defined_name import DefinedName
        from src.executor import execute_ast
        from src.parser import parse_file
        from src.store import JsonStore

        # Construire un xlsx simple avec _Manifeste
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        headers = ["id", "val", "heures"]
        for ci, h in enumerate(headers, 1):
            ws.cell(1, ci).value = h
        for ri, row in enumerate([(1, 50, 100), (2, 80, 200)], 2):
            for ci, v in enumerate(row, 1):
                ws.cell(ri, ci).value = v
        tbl = Table(displayName="TabData", ref="A1:C3")
        ws.add_table(tbl)

        ws_mxl = wb.create_sheet("_Manifeste")
        ws_mxl["A1"] = "MANIFESTE_V=1"
        ws_mxl["A2"] = "Instruction"
        instructions = [
            "FILE_TYPE: uo_instance",
            "FILE_ID: UO-ECO",
            "DEF $data = GET_TABLE(Data, TabData)",
            "DEF $moy = COMPUTE(MEAN_WEIGHTED($data.val, $data.heures))",
            "PUSH $moy -> uo.UO-ECO.avancement",
        ]
        for i, instr in enumerate(instructions, 3):
            ws_mxl.cell(i, 1).value = instr

        path = tmp_path / "UO-ECO.xlsx"
        wb.save(str(path))

        # Exécuter le pipeline
        store = JsonStore(tmp_path / "store.json")
        ast = parse_file(path)
        result = execute_ast(ast, path, store)
        assert result.errors == []

        # Enregistrer dans l'Exomap
        eco = _eco(tmp_path)
        record_file_sync("UO-ECO", str(path), "uo_instance", "ok",
                         ecosystem_path=eco)
        record_edges_from_ast(ast, "UO-ECO", ecosystem_path=eco)

        schema = load(eco)
        push_edges = [e for e in schema.edges if e.edge_type == "PUSH"]
        assert len(push_edges) == 1
        assert push_edges[0].to_node == "store::uo.UO-ECO.avancement"
        assert "UO-ECO" in schema.files
