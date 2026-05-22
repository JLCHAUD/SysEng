"""
Test d'intégration : COLLECT injecte les colonnes IDENT du child en tête des lignes.

Crée deux vrais fichiers Excel (child + parent) dans tmp_path,
exécute le parent via execute_ast, et vérifie que les colonnes identitaires
déclarées par IDENT dans le child apparaissent en tête de la table collectée.
"""
import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter

from src.parser import parse_file
from src.executor import execute_ast
from src import store as Store


def _make_named_table(ws, table_name: str, rows: list) -> None:
    """
    Crée une table Excel nommée dans la feuille ws.
    rows[0] = headers, rows[1:] = données.
    """
    for row in rows:
        ws.append(row)
    n_cols = len(rows[0])
    n_rows = len(rows)
    ref = f"A1:{get_column_letter(n_cols)}{n_rows}"
    ws.add_table(Table(displayName=table_name, ref=ref))


def test_collect_uses_ident_columns(tmp_path):
    """
    COLLECT injecte les colonnes identitaires (IDENT) du child en tête.

    Setup :
    - child  ENFANT-001.xlsx : IDENT nom="UO-Alpha", IDENT site="Paris"
                               table TabTaches : [id, libelle]
    - parent parent.xlsx     : LIST MesUOs FROM TABLE ListeUOs
                               COLLECT TabTaches FROM MesUOs INTO VueTaches

    Attendu : VueTaches contient _source_file_id + nom + site + id + libelle,
              avec "UO-Alpha" et "Paris" correctement injectés.
    """
    # ── Child : ENFANT-001.xlsx ───────────────────────────────────────────────
    child_path = tmp_path / "ENFANT-001.xlsx"
    wbc = Workbook()
    wbc.remove(wbc.active)

    # _Manifeste du child avec deux IDENT
    ws_mc = wbc.create_sheet("_Manifeste")
    ws_mc["A1"] = "MANIFESTE_V=1"
    # row 2 = titre décoratif, ignoré par parse_sheet (min_row=3)
    ws_mc["A3"] = "FILE_TYPE: uo_test"
    ws_mc["A4"] = "FILE_ID:   ENFANT-001"
    ws_mc["A5"] = "VERSION:   1"
    ws_mc["A6"] = 'IDENT nom : LABEL="Nom"'
    ws_mc["B6"] = "UO-Alpha"           # col B = valeur saisie par l'utilisateur
    ws_mc["A7"] = 'IDENT site : LABEL="Site"'
    ws_mc["B7"] = "Paris"

    # Feuille Taches avec table nommée TabTaches
    ws_tc = wbc.create_sheet("Taches")
    _make_named_table(ws_tc, "TabTaches", [
        ["id",  "libelle"],
        ["T1",  "Analyse"],
        ["T2",  "Conception"],
    ])
    wbc.save(str(child_path))

    # ── Parent : parent.xlsx ──────────────────────────────────────────────────
    parent_path = tmp_path / "parent.xlsx"
    wbp = Workbook()
    wbp.remove(wbp.active)

    # _Manifeste du parent : LIST + COLLECT
    ws_mp = wbp.create_sheet("_Manifeste")
    ws_mp["A1"] = "MANIFESTE_V=1"
    ws_mp["A3"] = "FILE_TYPE: projet"
    ws_mp["A4"] = "FILE_ID:   PROJ-001"
    ws_mp["A5"] = "VERSION:   1"
    ws_mp["A6"] = "LIST MesUOs FROM TABLE ListeUOs"
    ws_mp["A7"] = "COLLECT TabTaches FROM MesUOs INTO VueTaches"

    # Table de liste des UOs (colonne FILE_ID obligatoire)
    ws_l = wbp.create_sheet("Liste")
    _make_named_table(ws_l, "ListeUOs", [
        ["FILE_ID"],
        ["ENFANT-001"],
    ])

    # Table cible (pré-existante, sera écrasée par COLLECT)
    ws_v = wbp.create_sheet("Vue")
    _make_named_table(ws_v, "VueTaches", [
        ["nom", "site", "id", "libelle"],
        ["",    "",     "",   ""],        # ligne vide initiale
    ])
    wbp.save(str(parent_path))

    # ── Exécution ─────────────────────────────────────────────────────────────
    ast = parse_file(parent_path)
    result = execute_ast(ast, parent_path, Store)

    assert not result.errors, f"Erreurs executor : {result.errors}"
    assert len(result.collected) == 1, f"COLLECT attendu, got : {result.collected}"

    # ── Vérification de VueTaches ─────────────────────────────────────────────
    wb_r = load_workbook(str(parent_path), data_only=True)
    # Trouver la feuille contenant VueTaches
    ws_result = None
    for sn in wb_r.sheetnames:
        if "VueTaches" in wb_r[sn].tables:
            ws_result = wb_r[sn]
            break
    assert ws_result is not None, "Table VueTaches introuvable dans le parent après COLLECT"

    tbl_ref = ws_result.tables["VueTaches"].ref
    cells   = list(ws_result[tbl_ref])
    headers = [c.value for c in cells[0]]
    data_rows = [
        [c.value for c in row]
        for row in cells[1:]
        if any(c.value is not None for c in row)
    ]
    wb_r.close()

    # Les colonnes identitaires doivent précéder les colonnes de données du child
    assert "nom"  in headers, f"'nom' absent des headers : {headers}"
    assert "site" in headers, f"'site' absent des headers : {headers}"
    assert "id"   in headers, f"'id' absent des headers : {headers}"
    assert headers.index("nom")  < headers.index("id"),  "nom doit précéder id"
    assert headers.index("site") < headers.index("id"),  "site doit précéder id"

    # Les valeurs IDENT sont correctement injectées sur chaque ligne
    nom_idx  = headers.index("nom")
    site_idx = headers.index("site")
    id_idx   = headers.index("id")

    assert len(data_rows) == 2, f"Attendu 2 lignes collectées, got {len(data_rows)}"
    assert data_rows[0][nom_idx]  == "UO-Alpha"
    assert data_rows[0][site_idx] == "Paris"
    assert data_rows[0][id_idx]   == "T1"
    assert data_rows[1][nom_idx]  == "UO-Alpha"
    assert data_rows[1][site_idx] == "Paris"
    assert data_rows[1][id_idx]   == "T2"
