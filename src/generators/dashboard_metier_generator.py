# src/generators/dashboard_metier_generator.py
"""Dashboard métier — Vue consolidée équipe : Synthèse + Par Ingénieur + Alertes + _Manifeste."""
from datetime import date, timedelta
from pathlib import Path
from typing import List, Tuple

from openpyxl import Workbook

from src.models import ProfilActeur, TypeFiltre, UOInstance
from src.store import JsonStore
from src.xl_design import XD

OUTPUT_DIR = Path(__file__).parent.parent.parent / "output" / "cockpits"


def _filter_instances(acteur: ProfilActeur, instances: List[UOInstance]) -> List[UOInstance]:
    """Filtre les UOs selon le filtre_type et filtre_valeur de l'acteur."""
    if acteur.filtre_valeur == "ALL":
        return list(instances)
    if acteur.filtre_type == TypeFiltre.INGENIEUR:
        noms = {n.strip() for n in acteur.filtre_valeur.split(",")}
        return [u for u in instances if u.engineer_name in noms]
    if acteur.filtre_type == TypeFiltre.PROJET:
        projets = {p.strip() for p in acteur.filtre_valeur.split(",")}
        return [u for u in instances if u.project_id in projets]
    return list(instances)


def generate_dashboard_metier(
    acteur: ProfilActeur,
    all_instances: List[UOInstance],
    store: JsonStore,
    output_dir: Path = OUTPUT_DIR,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    uo_list = _filter_instances(acteur, all_instances)

    wb = Workbook()
    wb.remove(wb.active)

    _sheet_synthese(wb, acteur, uo_list, store)
    _sheet_vue_synthese(wb)
    _sheet_par_ingenieur(wb, acteur, uo_list, store)
    _sheet_alertes(wb, uo_list, store)
    _sheet_manifeste_dashboard(wb, acteur, uo_list)

    safe = acteur.nom.replace(" ", "_")
    filepath = output_dir / f"Dashboard_{safe}.xlsx"
    wb.save(filepath)
    return filepath


def _get_store_float(store: JsonStore, key: str, default: float = 0.0) -> float:
    """Lit une valeur float du store, retourne default si absente."""
    val = store.get(key)
    if val is None:
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _sheet_synthese(wb: Workbook, acteur: ProfilActeur,
                    uo_list: List[UOInstance], store: JsonStore):
    ws = wb.create_sheet("Synthèse")
    ws.sheet_view.showGridLines = False

    XD.banner(ws, "dashboard", f"Dashboard Métier — {acteur.nom}",
              subtitle=date.today().strftime('%d/%m/%Y'), n_cols=10)

    total_h = int(sum(u.total_hours for u in uo_list))
    avancement_vals = [_get_store_float(store, f"uo.{u.id}.avancement") for u in uo_list]
    avg_avanc = sum(avancement_vals) / len(avancement_vals) if avancement_vals else 0

    kpis = [
        ("A2:B2", len(uo_list), "Périmètre équipe"),
        ("C2:D2", total_h, "Charge totale"),
        ("E2:F2", avg_avanc, "Avancement moyen"),
    ]
    for cell_range, value, label in kpis:
        ws.merge_cells(cell_range)
        first_cell = cell_range.split(":")[0]
        c = ws[first_cell]
        c.value = value
        c.fill = XD.fill(XD.sheet("dashboard").accent)
        c.font = XD.fnt(11, bold=True, color="1F3864")
        c.alignment = XD.center()
        c.border = XD.HAIR
    # Label row under KPIs
    ws.merge_cells("A3:B3")
    ws["A3"].value = "Périmètre équipe"
    ws.merge_cells("C3:D3")
    ws["C3"].value = "Charge totale (h)"
    ws.merge_cells("E3:F3")
    ws["E3"].value = "Avancement moyen"
    for col_label in ("A3", "C3", "E3"):
        ws[col_label].font = XD.fnt(11, color="555555")
        ws[col_label].alignment = XD.center()
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 18

    ws.merge_cells("A4:J4")
    sec = ws["A4"]
    sec.value = "Toutes les UOs de l'équipe"
    sec.fill = XD.fill(XD.sheet("dashboard").header)
    sec.font = XD.fnt(11, bold=True, color=XD.WHITE)
    sec.alignment = XD.center()

    headers = ["UO ID", "Ingénieur", "Type UO", "Système", "Projet",
               "Charge (h)", "% Avancement", "H réalisées", "Date fin", "Alerte"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=h)
    XD.table_header(ws, 5, headers, "dashboard")

    for i, uo in enumerate(uo_list):
        row = 6 + i
        avancement = _get_store_float(store, f"uo.{uo.id}.avancement")
        h_real = _get_store_float(store, f"uo.{uo.id}.heures_realisees")
        type_name = uo.uo_type.name if uo.uo_type else uo.uo_type_id
        sys_name = uo.system.name if uo.system else uo.system_id
        proj_name = uo.project.name if uo.project else uo.project_id

        ws.cell(row=row, column=1, value=uo.id)
        ws.cell(row=row, column=2, value=uo.engineer_name)
        ws.cell(row=row, column=3, value=type_name)
        ws.cell(row=row, column=4, value=sys_name)
        ws.cell(row=row, column=5, value=proj_name)
        ws.cell(row=row, column=6, value=uo.total_hours)

        avanc_cell = ws.cell(row=row, column=7, value=avancement)
        avanc_cell.number_format = "0%"

        ws.cell(row=row, column=8, value=h_real)

        date_cell = ws.cell(row=row, column=9, value=uo.end_date)
        date_cell.number_format = "DD/MM/YYYY"

        today = date.today()
        if h_real > uo.total_hours:
            alerte = "⚠ Dérive heures"
        elif uo.end_date and uo.end_date <= today + timedelta(days=7):
            alerte = "⏰ Échéance proche"
        else:
            alerte = "✅ OK"
        ws.cell(row=row, column=10, value=alerte)

        alternate = (i % 2 == 1)
        bg = XD.sheet("dashboard").accent if alternate else XD.WHITE
        for c in range(1, 11):
            cell = ws.cell(row=row, column=c)
            cell.fill = XD.fill(bg)
            cell.font = XD.fnt(11)
            cell.alignment = XD.left()
            cell.border = XD.HAIR
        # Restore number formats after style_data_row equivalent
        ws.cell(row=row, column=7).number_format = "0%"
        ws.cell(row=row, column=9).number_format = "DD/MM/YYYY"

    for col, w in {"A": 12, "B": 22, "C": 28, "D": 18, "E": 22,
                   "F": 13, "G": 16, "H": 14, "I": 14, "J": 22}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"


def _sheet_vue_synthese(wb: Workbook):
    """Feuille réceptrice du COLLECT — tbl_vue_synthese remplie par ExoSync."""
    ws = wb.create_sheet("Vue Synthèse")
    ws.sheet_properties.tabColor = XD.sheet("dashboard").header
    ws.sheet_view.showGridLines = False

    # En-têtes alignés sur tbl_mes_uos pour que COLLECT fusionne correctement
    headers = ["_source_file_id", "ingenieur", "UO ID", "Type UO", "Système",
               "Projet", "Charge (h)", "% Avancement", "H réalisées", "Date fin", "Alerte"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    XD.table_header(ws, 1, headers, "dashboard")
    # Ligne vide d'initialisation (table doit avoir au moins 1 ligne de données)
    ws.cell(row=2, column=1, value="(vide — rempli par ExoSync)")

    ref = f"A1:{chr(64+len(headers))}2"
    XD.named_table(ws, "tbl_vue_synthese", ref, "dashboard")
    for col, w in {"A": 22, "B": 22, "C": 18, "D": 28, "E": 18,
                   "F": 20, "G": 13, "H": 16, "I": 14, "J": 14, "K": 22}.items():
        ws.column_dimensions[col].width = w


def _sheet_par_ingenieur(wb: Workbook, acteur: ProfilActeur,
                         uo_list: List[UOInstance], store: JsonStore):
    ws = wb.create_sheet("Par Ingénieur")
    ws.sheet_view.showGridLines = False

    XD.banner(ws, "activites", f"Détail par Ingénieur — {acteur.nom}", n_cols=8)

    engineers = sorted(set(u.engineer_name for u in uo_list))
    current_row = 3

    for eng in engineers:
        eng_uo = [u for u in uo_list if u.engineer_name == eng]
        total_h = sum(u.total_hours for u in eng_uo)
        avancs = [_get_store_float(store, f"uo.{u.id}.avancement") for u in eng_uo]
        avg_avanc = sum(avancs) / len(avancs) if avancs else 0.0

        ws.merge_cells(f"A{current_row}:H{current_row}")
        eng_cell = ws[f"A{current_row}"]
        eng_cell.value = (f"▶  {eng}   |   {len(eng_uo)} UO   |   "
                          f"{total_h}h   |   {avg_avanc:.0%} avancement moyen")
        eng_cell.fill = XD.fill(XD.sheet("activites").header)
        eng_cell.font = XD.fnt(11, bold=True, color=XD.WHITE)
        eng_cell.alignment = XD.left()
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        sub_headers = ["UO ID", "Type", "Système", "Projet",
                       "Charge (h)", "% Avancement", "H réalisées", "Date fin"]
        XD.table_header(ws, current_row, sub_headers, "activites")
        current_row += 1

        for i, uo in enumerate(eng_uo):
            avancement = _get_store_float(store, f"uo.{uo.id}.avancement")
            h_real = _get_store_float(store, f"uo.{uo.id}.heures_realisees")
            type_name = uo.uo_type.name if uo.uo_type else uo.uo_type_id
            sys_name = uo.system.name if uo.system else uo.system_id
            proj_name = uo.project.name if uo.project else uo.project_id

            ws.cell(row=current_row, column=1, value=uo.id)
            ws.cell(row=current_row, column=2, value=type_name)
            ws.cell(row=current_row, column=3, value=sys_name)
            ws.cell(row=current_row, column=4, value=proj_name)
            ws.cell(row=current_row, column=5, value=uo.total_hours)
            avc = ws.cell(row=current_row, column=6, value=avancement)
            avc.number_format = "0%"
            ws.cell(row=current_row, column=7, value=h_real)
            dc = ws.cell(row=current_row, column=8, value=uo.end_date)
            dc.number_format = "DD/MM/YYYY"

            alternate = (i % 2 == 1)
            bg = XD.sheet("activites").accent if alternate else XD.WHITE
            for c in range(1, 9):
                cell = ws.cell(row=current_row, column=c)
                cell.fill = XD.fill(bg)
                cell.font = XD.fnt(11)
                cell.alignment = XD.left()
                cell.border = XD.HAIR
            # Restore number formats
            ws.cell(row=current_row, column=6).number_format = "0%"
            ws.cell(row=current_row, column=8).number_format = "DD/MM/YYYY"
            current_row += 1

        cockpit_path = OUTPUT_DIR / f"Cockpit_{eng.replace(' ', '_')}.xlsx"
        ws.merge_cells(f"A{current_row}:H{current_row}")
        link_cell = ws[f"A{current_row}"]
        link_cell.value = f"→ Ouvrir cockpit {eng}"
        link_cell.hyperlink = str(cockpit_path)
        link_cell.font = XD.fnt(11, color="0563C1")
        link_cell.alignment = XD.left()
        link_cell.fill = XD.fill(XD.GREY_L)
        link_cell.border = XD.HAIR
        current_row += 2

    for col, w in {"A": 12, "B": 28, "C": 18, "D": 22, "E": 13, "F": 16, "G": 14, "H": 14}.items():
        ws.column_dimensions[col].width = w


def _compute_alerts(uo_list: List[UOInstance], store: JsonStore) -> List[Tuple]:
    """Calcule les alertes à partir des données du store. Retourne liste triée par criticité."""
    today = date.today()
    alerts = []

    for uo in uo_list:
        h_real = _get_store_float(store, f"uo.{uo.id}.heures_realisees")

        if h_real > uo.total_hours:
            alerts.append((
                uo.engineer_name, uo.id,
                "Dépassement H",
                f"{h_real:.0f}h réalisées / {uo.total_hours}h allouées",
                "🔴 Critique", 0,
            ))

        if uo.end_date and uo.end_date <= today + timedelta(days=7):
            jours = (uo.end_date - today).days
            label = f"Échéance dans {jours}j" if jours >= 0 else f"Dépassée de {-jours}j"
            criticite = "🔴 Critique" if jours < 0 else "🟠 Élevée"
            prio = 0 if jours < 0 else 1
            alerts.append((
                uo.engineer_name, uo.id,
                "Échéance critique",
                label,
                criticite, prio,
            ))

    alerts.sort(key=lambda x: x[5])
    return alerts


def _sheet_alertes(wb: Workbook, uo_list: List[UOInstance], store: JsonStore):
    ws = wb.create_sheet("Alertes")
    ws.sheet_view.showGridLines = False

    XD.banner(ws, "oil", f"Alertes & Risques — {date.today().strftime('%d/%m/%Y')}", n_cols=5)

    headers = ["Ingénieur", "UO ID", "Type alerte", "Détail", "Criticité"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    XD.table_header(ws, 2, headers, "oil")

    alerts = _compute_alerts(uo_list, store)

    if not alerts:
        ws.merge_cells("A3:E3")
        c = ws["A3"]
        c.value = "✅ Aucune alerte active"
        c.fill = XD.fill(XD.GREEN_L)
        c.font = XD.fnt(11, bold=True, color="1F3864")
        c.alignment = XD.center()
        c.border = XD.HAIR
    else:
        for i, (eng, uid, type_a, detail, criticite, _) in enumerate(alerts):
            row = 3 + i
            ws.cell(row=row, column=1, value=eng)
            ws.cell(row=row, column=2, value=uid)
            ws.cell(row=row, column=3, value=type_a)
            ws.cell(row=row, column=4, value=detail)
            ws.cell(row=row, column=5, value=criticite)

            fill_color = XD.RED_L if "Critique" in criticite else XD.AMBER_L
            for col in range(1, 6):
                c = ws.cell(row=row, column=col)
                c.fill = XD.fill(fill_color)
                c.border = XD.HAIR
                c.alignment = XD.left()
                c.font = XD.fnt(11)

    for col, w in {"A": 22, "B": 12, "C": 20, "D": 40, "E": 15}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"


def _sheet_manifeste_dashboard(wb: Workbook, acteur, uo_list: List[UOInstance]):
    """Génère l'onglet _Manifeste dashboard au format MXL mono-colonne.
    Col A = instruction MXL, col B = ancre, col C = commentaire français.
    Le pilote_id en métadonnée permet à LIST DYNAMIC de découvrir ce dashboard.
    """
    ws = wb.create_sheet("_Manifeste")
    ws.sheet_view.showGridLines = False

    def _mxl_row(row: int, instr: str, ancre: str = "", comment: str = ""):
        cell_a = ws.cell(row=row, column=1, value=instr)
        cell_a.font = XD.fnt(9.5, bold=instr.startswith(("DEF ", "PUSH ", "PULL ", "LIST ", "COLLECT ")))
        cell_a.alignment = XD.left()
        if ancre:
            ws.cell(row=row, column=2, value=ancre).font = XD.fnt(9, color="888888")
        if comment:
            c = ws.cell(row=row, column=3, value=comment)
            c.font = XD.fnt(9, color="666666")
            c.fill = XD.fill("F9F9F9")

    # Ligne 1 : version
    ws["A1"] = "MANIFESTE_V=1"
    ws["A1"].font = XD.fnt(11, bold=True, color="1F3864")
    ws.sheet_properties.tabColor = XD.sheet("manifeste").header
    # Ligne 2 intentionnellement vide

    # Métadonnées
    _mxl_row(3, "FILE_TYPE: dashboard_pilote",      comment="Type de fichier ExoSync")
    _mxl_row(4, f"FILE_ID: Dashboard_{acteur.id}",  comment="Identifiant unique du dashboard")
    _mxl_row(5, f"pilote_id: {acteur.id}",           comment="Identifiant du pilote propriétaire")

    # Découverte des cockpits de l'équipe via pilote_id
    _mxl_row(7, f"LIST mes_cockpits TYPE=cockpit_ingenieur WHERE pilote_id={acteur.id}",
             comment="Découverte des cockpits ingénieur de l'équipe de ce pilote")
    _mxl_row(8, "COLLECT tbl_mes_uos FROM mes_cockpits INTO tbl_vue_synthese",
             comment="Agrégation de toutes les UOs de tous les cockpits de l'équipe")

    # Lecture + publication de la synthèse agrégée
    _mxl_row(10, "DEF $synthese = GET_TABLE(Vue Synthèse, tbl_vue_synthese)",
              comment="Relit la table agrégée après COLLECT")
    _mxl_row(11, f"PUSH $synthese -> dashboard.{acteur.id}.synthese",
              comment="Publie la vue consolidée de l'équipe dans le store central")

    for col, w in {"A": 70, "B": 18, "C": 60}.items():
        ws.column_dimensions[col].width = w
