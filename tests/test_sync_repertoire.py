# tests/test_sync_repertoire.py
"""Tests TDD pour la synchronisation par scan de répertoire.

Principe : le moteur scrute un dossier, découvre tous les .xlsx via leur
_Manifeste (FILE_TYPE), synchronise dans l'ordre de dépendance, sans
aucun registre JSON externe. Store et ecosystem vivent dans le dossier.

Les classeurs de test sont fabriqués localement (openpyxl + _Manifeste MXL
minimal) — autonomes, sans dépendre d'aucun générateur de production.
"""
import json
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


def _make_uo_xlsx(path: Path, file_id: str):
    """Classeur UO minimal : table tbl_act + _Manifeste qui la PUSH (namespacé)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Activites"
    headers = ["id", "designation", "avancement"]
    rows = [["ACT-1", "Analyse", 50], ["ACT-2", "Rédaction", 20]]
    ws.append(headers)
    for r in rows:
        ws.append(r)
    tbl = Table(displayName="tbl_act", ref="A1:C3")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)

    ws_m = wb.create_sheet("_Manifeste")
    ws_m["A1"] = "MANIFESTE_V=1"
    manifeste = [
        "FILE_TYPE: uo_instance",
        f"FILE_ID: {file_id}",
        "DEF $act = GET_TABLE(Activites, tbl_act)",
        f"PUSH $act -> uo.{file_id}.activites",
    ]
    for i, line in enumerate(manifeste, 3):
        ws_m.cell(row=i, column=1, value=line)
    wb.save(path)


def _make_cockpit_xlsx(path: Path, file_id: str):
    """Classeur cockpit minimal : PUSH d'une table locale (namespacé cockpit.*)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Mes UOs"
    ws.append(["uo_id", "avancement"])
    ws.append(["UO-001", 50])
    tbl = Table(displayName="tbl_mes_uos", ref="A1:B2")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)

    ws_m = wb.create_sheet("_Manifeste")
    ws_m["A1"] = "MANIFESTE_V=1"
    manifeste = [
        "FILE_TYPE: cockpit_ingenieur",
        f"FILE_ID: {file_id}",
        "DEF $mes_uos = GET_TABLE(Mes UOs, tbl_mes_uos)",
        f"PUSH $mes_uos -> cockpit.{file_id}.mes_uos",
    ]
    for i, line in enumerate(manifeste, 3):
        ws_m.cell(row=i, column=1, value=line)
    wb.save(path)


@pytest.fixture
def projet_dir(tmp_path):
    """Répertoire projet de test : 2 UOs + 1 cockpit (3 .xlsx avec _Manifeste)."""
    _make_uo_xlsx(tmp_path / "UO-001.xlsx", "UO-001")
    _make_uo_xlsx(tmp_path / "UO-002.xlsx", "UO-002")
    _make_cockpit_xlsx(tmp_path / "Cockpit_Alice.xlsx", "Cockpit_Alice")
    return tmp_path


class TestSyncRepertoire:
    def test_decouvre_tous_les_xlsx(self, projet_dir):
        """Le scan doit traiter les 3 fichiers .xlsx du dossier."""
        from src.sync import synchroniser_repertoire
        rapport_path = synchroniser_repertoire(projet_dir)
        rapport = json.loads(rapport_path.read_text(encoding="utf-8"))
        assert rapport["nb_total"] == 3

    def test_cree_store_local_dans_repertoire(self, projet_dir):
        """Le store.json doit être créé DANS le répertoire scruté."""
        from src.sync import synchroniser_repertoire
        synchroniser_repertoire(projet_dir)
        assert (projet_dir / "store.json").exists()

    def test_push_alimente_store_local(self, projet_dir):
        """Après sync, les PUSH (namespacés par FILE_ID) alimentent le store local."""
        from src.sync import synchroniser_repertoire
        from src.store import JsonStore
        synchroniser_repertoire(projet_dir)
        store = JsonStore(projet_dir / "store.json")
        keys = store.get_all().keys()
        assert "uo.UO-001.activites" in keys, f"PUSH UO absent : {list(keys)}"
        assert "cockpit.Cockpit_Alice.mes_uos" in keys, f"PUSH cockpit absent : {list(keys)}"

    def test_nouveau_fichier_pris_au_scan_suivant(self, projet_dir):
        """Un .xlsx ajouté après coup est découvert au scan suivant."""
        from src.sync import synchroniser_repertoire

        r1 = synchroniser_repertoire(projet_dir)
        assert json.loads(r1.read_text(encoding="utf-8"))["nb_total"] == 3

        _make_uo_xlsx(projet_dir / "UO-003.xlsx", "UO-003")
        r2 = synchroniser_repertoire(projet_dir)
        assert json.loads(r2.read_text(encoding="utf-8"))["nb_total"] == 4

    def test_fichier_supprime_ignore(self, projet_dir):
        """Un .xlsx retiré n'est plus synchronisé — aucune erreur, juste ignoré."""
        from src.sync import synchroniser_repertoire

        synchroniser_repertoire(projet_dir)
        (projet_dir / "UO-002.xlsx").unlink()
        r2 = synchroniser_repertoire(projet_dir)
        data = json.loads(r2.read_text(encoding="utf-8"))
        assert data["nb_total"] == 2
        assert data["nb_erreur"] == 0

    def test_fichier_temp_excel_ignore(self, projet_dir):
        """Les fichiers temporaires Excel (~$...) sont ignorés."""
        from src.sync import synchroniser_repertoire
        (projet_dir / "~$Cockpit_Alice.xlsx").write_bytes(b"garbage")
        r = synchroniser_repertoire(projet_dir)
        data = json.loads(r.read_text(encoding="utf-8"))
        assert data["nb_total"] == 3

    def test_xlsx_sans_manifeste_ignore_sans_erreur(self, projet_dir):
        """Un .xlsx quelconque sans _Manifeste est ignoré, pas compté en erreur."""
        from src.sync import synchroniser_repertoire
        wb = Workbook()
        wb.active["A1"] = "fichier sans manifeste"
        wb.save(projet_dir / "notes_diverses.xlsx")
        r = synchroniser_repertoire(projet_dir)
        data = json.loads(r.read_text(encoding="utf-8"))
        assert data["nb_erreur"] == 0
        assert data["nb_total"] == 3  # le 4e (sans manifeste) est ignoré

    def test_zero_erreur_sur_projet_valide(self, projet_dir):
        """Un projet valide se synchronise sans erreur."""
        from src.sync import synchroniser_repertoire
        r = synchroniser_repertoire(projet_dir)
        data = json.loads(r.read_text(encoding="utf-8"))
        assert data["nb_erreur"] == 0, f"Erreurs: {data['fichiers']}"
