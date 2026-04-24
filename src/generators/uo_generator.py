"""Generates one Excel file per UO instance (7 sheets + _Manifeste)."""
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.models import UOInstance, StatutUO
from src.styles import (
    BLUE_DARK, BLUE_MID, BLUE_LIGHT, GREEN, GREEN_LIGHT, ORANGE_LIGHT,
    RED_LIGHT, GREY_LIGHT, GREY_MID, WHITE, YELLOW_LIGHT,
    THIN_BORDER, solid_fill, header_font, body_font, center, left,
    style_header_row, style_data_row, set_column_widths, freeze_top_row,
)

OUTPUT_DIR = Path(__file__).parent.parent.parent / "output" / "UOs"


def _sheet_organisation(wb: Workbook, uo: UOInstance):
    ws = wb.create_sheet("Organisation Projet")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value = f"Organisation Projet — {uo.project.name if uo.project else uo.project_id}"
    title.fill = solid_fill(BLUE_DARK)
    title.font = header_font(size=13)
    title.alignment = center()
    ws.row_dimensions[1].height = 28

    # Headers
    headers = ["Nom", "Rôle", "Email", "Remarques"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, 1, 4, color=BLUE_MID)

    # Data
    actors = uo.project.actors if uo.project else []
    for i, actor in enumerate(actors):
        row = 3 + i
        ws.cell(row=row, column=1, value=actor.name)
        ws.cell(row=row, column=2, value=actor.role)
        ws.cell(row=row, column=3, value=actor.email)
        ws.cell(row=row, column=4, value="")
        style_data_row(ws, row, 1, 4, alternate=(i % 2 == 1))

    set_column_widths(ws, {"A": 25, "B": 30, "C": 35, "D": 30})
    freeze_top_row(ws)


def _sheet_livrables(wb: Workbook, uo: UOInstance):
    ws = wb.create_sheet("Livrables")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    title = ws["A1"]
    title.value = "Livrables"
    title.fill = solid_fill(BLUE_DARK)
    title.font = header_font(size=13)
    title.alignment = center()
    ws.row_dimensions[1].height = 28

    headers = ["ID", "Nom du livrable", "Date prévue", "Date réelle", "Statut"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, 1, 5, color=BLUE_MID)

    deliverables = uo.uo_type.deliverables if uo.uo_type else []
    statuses = ["À faire", "En cours", "Livré", "Validé"]
    for i, deliv in enumerate(deliverables):
        row = 3 + i
        ws.cell(row=row, column=1, value=deliv.id)
        ws.cell(row=row, column=2, value=deliv.name)
        ws.cell(row=row, column=3, value=deliv.due_date or "")
        ws.cell(row=row, column=4, value="")
        ws.cell(row=row, column=5, value=deliv.status)
        style_data_row(ws, row, 1, 5, alternate=(i % 2 == 1))
        if deliv.due_date:
            ws.cell(row=row, column=3).number_format = "DD/MM/YYYY"

    # Conditional formatting on Status column
    from openpyxl.formatting.rule import CellIsRule
    last_row = 2 + len(deliverables)
    ws.conditional_formatting.add(
        f"E3:E{last_row}",
        CellIsRule(operator="equal", formula=['"Validé"'], fill=solid_fill(GREEN_LIGHT)),
    )
    ws.conditional_formatting.add(
        f"E3:E{last_row}",
        CellIsRule(operator="equal", formula=['"En cours"'], fill=solid_fill(YELLOW_LIGHT)),
    )

    set_column_widths(ws, {"A": 20, "B": 40, "C": 18, "D": 18, "E": 15})
    freeze_top_row(ws)


def _sheet_planning(wb: Workbook, uo: UOInstance):
    ws = wb.create_sheet("Planning")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "Planning des Livrables"
    title.fill = solid_fill(BLUE_DARK)
    title.font = header_font(size=13)
    title.alignment = center()
    ws.row_dimensions[1].height = 28

    headers = ["ID Livrable", "Nom", "Date prévue", "Date réelle", "Écart (j)", "Commentaire"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, 1, 6, color=BLUE_MID)

    deliverables = uo.uo_type.deliverables if uo.uo_type else []
    for i, deliv in enumerate(deliverables):
        row = 3 + i
        ws.cell(row=row, column=1, value=deliv.id)
        ws.cell(row=row, column=2, value=deliv.name)
        ws.cell(row=row, column=3, value="")
        ws.cell(row=row, column=4, value="")
        # Écart formula: réelle - prévue
        ws.cell(row=row, column=5, value=f'=IF(AND(D{row}<>"",C{row}<>""),D{row}-C{row},"")')
        ws.cell(row=row, column=6, value="")
        style_data_row(ws, row, 1, 6, alternate=(i % 2 == 1))
        ws.cell(row=row, column=3).number_format = "DD/MM/YYYY"
        ws.cell(row=row, column=4).number_format = "DD/MM/YYYY"

    set_column_widths(ws, {"A": 20, "B": 40, "C": 18, "D": 18, "E": 12, "F": 30})
    freeze_top_row(ws)


def _sheet_activites(wb: Workbook, uo: UOInstance):
    ws = wb.create_sheet("Activités")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = f"Activités — Charge totale : {uo.total_hours}h"
    title.fill = solid_fill(BLUE_DARK)
    title.font = header_font(size=13)
    title.alignment = center()
    ws.row_dimensions[1].height = 28

    headers = ["ID", "Activité", "Heures allouées", "Date début", "Date fin", "% Avancement", "Heures réalisées", "Commentaire"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, 1, 8, color=BLUE_MID)

    activities = uo.uo_type.activities if uo.uo_type else []
    total_default = sum(a.default_hours for a in activities) or 1
    data_rows = []

    for i, act in enumerate(activities):
        row = 3 + i
        data_rows.append(row)
        # Proportional allocation from total_hours
        allocated = round(act.default_hours / total_default * uo.total_hours, 1)
        ws.cell(row=row, column=1, value=act.id)
        ws.cell(row=row, column=2, value=act.name)
        ws.cell(row=row, column=3, value=allocated)
        ws.cell(row=row, column=4, value=uo.start_date)
        ws.cell(row=row, column=5, value=uo.end_date)
        ws.cell(row=row, column=6, value=0.0)
        ws.cell(row=row, column=7, value=0.0)
        ws.cell(row=row, column=8, value="")
        style_data_row(ws, row, 1, 8, alternate=(i % 2 == 1))
        ws.cell(row=row, column=4).number_format = "DD/MM/YYYY"
        ws.cell(row=row, column=5).number_format = "DD/MM/YYYY"
        ws.cell(row=row, column=6).number_format = "0%"

    # Footer totals
    if data_rows:
        footer_row = data_rows[-1] + 2
        ws.merge_cells(f"A{footer_row}:B{footer_row}")
        total_cell = ws.cell(row=footer_row, column=1, value="TOTAL / AVANCEMENT GLOBAL")
        total_cell.fill = solid_fill(BLUE_LIGHT)
        total_cell.font = header_font(color="000000")
        total_cell.alignment = left()
        total_cell.border = THIN_BORDER

        first_data = data_rows[0]
        last_data = data_rows[-1]

        # Sum hours
        sum_cell = ws.cell(row=footer_row, column=3, value=f"=SUM(C{first_data}:C{last_data})")
        sum_cell.fill = solid_fill(BLUE_LIGHT)
        sum_cell.font = header_font(color="000000")
        sum_cell.border = THIN_BORDER
        sum_cell.alignment = center()

        # Weighted average progress
        avg_formula = (
            f"=SUMPRODUCT(C{first_data}:C{last_data},F{first_data}:F{last_data})"
            f"/SUM(C{first_data}:C{last_data})"
        )
        avg_cell = ws.cell(row=footer_row, column=6, value=avg_formula)
        avg_cell.fill = solid_fill(BLUE_LIGHT)
        avg_cell.font = header_font(color="000000")
        avg_cell.number_format = "0%"
        avg_cell.border = THIN_BORDER
        avg_cell.alignment = center()

        for col in [4, 5, 7, 8]:
            c = ws.cell(row=footer_row, column=col)
            c.fill = solid_fill(BLUE_LIGHT)
            c.border = THIN_BORDER

    set_column_widths(ws, {"A": 22, "B": 35, "C": 18, "D": 18, "E": 18, "F": 16, "G": 18, "H": 30})
    freeze_top_row(ws)


def _sheet_rex(wb: Workbook, uo: UOInstance):
    ws = wb.create_sheet("REX")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    title = ws["A1"]
    sys_name = uo.system.name if uo.system else uo.system_id
    title.value = f"Retour d'Expérience — {sys_name}"
    title.fill = solid_fill(BLUE_DARK)
    title.font = header_font(size=13)
    title.alignment = center()
    ws.row_dimensions[1].height = 28

    headers = ["Catégorie", "Description / Observation", "Lien / Référence", "Date"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, 1, 4, color=BLUE_MID)

    rex_items = uo.system.rex_prefill if uo.system else []
    for i, item in enumerate(rex_items):
        row = 3 + i
        ws.cell(row=row, column=1, value="Bonne pratique")
        ws.cell(row=row, column=2, value=item)
        ws.cell(row=row, column=3, value="")
        ws.cell(row=row, column=4, value="")
        style_data_row(ws, row, 1, 4, alternate=(i % 2 == 1))

    # Extra empty rows for engineer to fill
    start_extra = 3 + len(rex_items)
    categories = ["Problème rencontré", "Solution appliquée", "À surveiller", "Bonne pratique"]
    for i in range(5):
        row = start_extra + i
        ws.cell(row=row, column=1, value="")
        ws.cell(row=row, column=2, value="")
        ws.cell(row=row, column=3, value="")
        ws.cell(row=row, column=4, value="")
        style_data_row(ws, row, 1, 4, alternate=(i % 2 == 1))

    set_column_widths(ws, {"A": 22, "B": 60, "C": 35, "D": 15})
    freeze_top_row(ws)


def _sheet_points_ouverts(wb: Workbook, uo: UOInstance):
    ws = wb.create_sheet("Points Ouverts")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = "Suivi des Points Ouverts"
    title.fill = solid_fill(BLUE_DARK)
    title.font = header_font(size=13)
    title.alignment = center()
    ws.row_dimensions[1].height = 28

    headers = ["ID", "Date ouv.", "Description", "Nature", "Responsable", "Statut", "Date clôt."]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, 1, 7, color=BLUE_MID)

    # Empty rows ready to fill
    natures = ["Question expert", "Problème technique", "Demande fournisseur", "Action interne"]
    statuses = ["À faire", "En cours", "Fermé"]
    for i in range(10):
        row = 3 + i
        ws.cell(row=row, column=1, value=f"PO-{i+1:03d}")
        for col in range(2, 8):
            ws.cell(row=row, column=col, value="")
        ws.cell(row=row, column=6, value="À faire")
        style_data_row(ws, row, 1, 7, alternate=(i % 2 == 1))

    # Conditional formatting on Statut
    ws.conditional_formatting.add(
        "F3:F12",
        CellIsRule(operator="equal", formula=['"Fermé"'], fill=solid_fill(GREEN_LIGHT)),
    )
    ws.conditional_formatting.add(
        "F3:F12",
        CellIsRule(operator="equal", formula=['"En cours"'], fill=solid_fill(YELLOW_LIGHT)),
    )
    ws.conditional_formatting.add(
        "F3:F12",
        CellIsRule(operator="equal", formula=['"À faire"'], fill=solid_fill(ORANGE_LIGHT)),
    )

    set_column_widths(ws, {"A": 12, "B": 14, "C": 45, "D": 25, "E": 25, "F": 14, "G": 14})
    freeze_top_row(ws)


def _sheet_dashboard(wb: Workbook, uo: UOInstance):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = f"Dashboard — {uo.id} | {uo.uo_type.name if uo.uo_type else ''} | {uo.system.name if uo.system else ''}"
    t.fill = solid_fill(BLUE_DARK)
    t.font = header_font(size=14)
    t.alignment = center()
    ws.row_dimensions[1].height = 32

    # ── KPI block ──────────────────────────────────────────────────────────────
    kpis = [
        ("Projet", uo.project.name if uo.project else uo.project_id),
        ("Ingénieur", uo.engineer_name),
        ("Charge totale (h)", uo.total_hours),
        ("Date début", uo.start_date),
        ("Date fin prévue", uo.end_date),
    ]
    for i, (label, value) in enumerate(kpis):
        row = 3 + i
        lbl = ws.cell(row=row, column=1, value=label)
        lbl.fill = solid_fill(BLUE_LIGHT)
        lbl.font = body_font(bold=True)
        lbl.alignment = left()
        lbl.border = THIN_BORDER

        val = ws.cell(row=row, column=2, value=value)
        val.font = body_font()
        val.alignment = left()
        val.border = THIN_BORDER
        if isinstance(value, date):
            val.number_format = "DD/MM/YYYY"

    # ── Avancement global (linked from Activités sheet) ────────────────────────
    ws.merge_cells("D3:E3")
    label_av = ws["D3"]
    label_av.value = "Avancement global"
    label_av.fill = solid_fill(BLUE_LIGHT)
    label_av.font = body_font(bold=True)
    label_av.alignment = center()
    label_av.border = THIN_BORDER

    ws.merge_cells("F3:H3")
    av_val = ws["F3"]
    # Reference weighted average from Activités sheet footer (row 3+len(activities)+2, col 6)
    activities = uo.uo_type.activities if uo.uo_type else []
    footer_row = 3 + len(activities) + 2
    av_val.value = f"=Activités!F{footer_row}"
    av_val.number_format = "0%"
    av_val.font = header_font(size=16, color="000000")
    av_val.alignment = center()
    av_val.border = THIN_BORDER

    # ── Progress bar (row 10) ──────────────────────────────────────────────────
    ws.merge_cells("A9:H9")
    bar_label = ws["A9"]
    bar_label.value = "Barre de progression"
    bar_label.fill = solid_fill(BLUE_MID)
    bar_label.font = header_font()
    bar_label.alignment = center()

    # 8 cells = 100%, fill with green for done%, grey for remaining
    num_cells = 8
    for col in range(1, num_cells + 1):
        cell = ws.cell(row=10, column=col, value="")
        threshold = col / num_cells
        cell.fill = solid_fill("70AD47")  # all green by default; CF will handle graying
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = 8

    # Conditional formatting: grey out cells beyond avancement%
    for col in range(1, num_cells + 1):
        col_letter = get_column_letter(col)
        threshold_pct = col / num_cells
        ws.conditional_formatting.add(
            f"{col_letter}10",
            CellIsRule(
                operator="lessThan",
                formula=[f"F3-{threshold_pct}+0.001"],
                fill=solid_fill(GREY_MID),
            ),
        )

    # ── Livrables à venir (simple table) ──────────────────────────────────────
    ws.merge_cells("A12:H12")
    deliv_title = ws["A12"]
    deliv_title.value = "Prochains livrables"
    deliv_title.fill = solid_fill(BLUE_MID)
    deliv_title.font = header_font()
    deliv_title.alignment = center()

    headers = ["ID", "Livrable", "Date prévue", "Statut"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=13, column=col, value=h)
        c.fill = solid_fill(BLUE_LIGHT)
        c.font = body_font(bold=True)
        c.alignment = center()
        c.border = THIN_BORDER

    deliverables = uo.uo_type.deliverables if uo.uo_type else []
    for i, deliv in enumerate(deliverables):
        row = 14 + i
        ws.cell(row=row, column=1, value=f"=Livrables!A{3+i}")
        ws.cell(row=row, column=2, value=f"=Livrables!B{3+i}")
        ws.cell(row=row, column=3, value=f"=Livrables!C{3+i}")
        ws.cell(row=row, column=4, value=f"=Livrables!E{3+i}")
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).fill = solid_fill(GREY_LIGHT if i % 2 else WHITE)

    set_column_widths(ws, {"A": 18, "B": 30, "C": 18, "D": 15, "E": 15, "F": 15, "G": 15, "H": 15})


def _add_named_table(ws, ref: str, name: str, style: str = "TableStyleMedium2"):
    """Ajoute un tableau Excel nommé à une plage donnée."""
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True)
    ws.add_table(tbl)


def _sheet_manifeste(wb: Workbook, uo: UOInstance):
    """
    Génère la feuille _Manifeste avec le méta-langage MXL.
    Col A = instruction, Col B = ancre (plage nommée ou cellule réelle).
    """
    ws = wb.create_sheet("_Manifeste")
    ws.sheet_view.showGridLines = False

    # ── Styles locaux ──────────────────────────────────────────────────────────
    from src.styles import (
        solid_fill, header_font, body_font, left, center,
        BLUE_DARK, BLUE_MID, BLUE_LIGHT, GREY_LIGHT, WHITE, THIN_BORDER,
    )

    def _hdr(ws, row, col, val, color=BLUE_LIGHT):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = solid_fill(color)
        c.font = body_font(bold=True)
        c.border = THIN_BORDER
        c.alignment = center()
        return c

    def _row(ws, row, instr, anchor="", even=True):
        ci = ws.cell(row=row, column=1, value=instr)
        ca = ws.cell(row=row, column=2, value=anchor)
        bg = WHITE if even else GREY_LIGHT
        for c in (ci, ca):
            c.fill = solid_fill(bg)
            c.font = body_font(size=10)
            c.border = THIN_BORDER
            c.alignment = left()

    def _comment(ws, row, text):
        ci = ws.cell(row=row, column=1, value=f"# {text}")
        ca = ws.cell(row=row, column=2, value="")
        for c in (ci, ca):
            c.fill = solid_fill(BLUE_LIGHT)
            c.font = body_font(size=9, color="4472C4")
            c.border = THIN_BORDER

    # ── Ligne 1 : version ─────────────────────────────────────────────────────
    ws["A1"] = "MANIFESTE_V=1"
    ws["A1"].fill = solid_fill(BLUE_DARK)
    ws["A1"].font = header_font()
    ws["B1"] = ""

    # ── Ligne 2 : en-têtes ────────────────────────────────────────────────────
    _hdr(ws, 2, 1, "Instruction", BLUE_MID)
    _hdr(ws, 2, 2, "Ancre / Plage", BLUE_MID)

    # ── Contenu méta-langage ──────────────────────────────────────────────────
    rows: list[tuple[str, str, bool]] = []  # (instr, anchor, is_comment)

    # En-tête du fichier
    rows.append(("# --- EN-TETE FICHIER ---", "", True))
    rows.append((f"FILE_TYPE: uo_instance", "", False))
    rows.append((f"FILE_ID:   {uo.id}", "", False))
    rows.append(("VERSION:   1", "", False))

    # Imports depuis l'écosystème
    rows.append(("# --- IMPORTS (PULL) ---", "", True))
    rows.append((
        "PULL projet.acteurs -> FILL_TABLE(Organisation Projet, TabActeurs)  MODE=OVERWRITE",
        "",
        False,
    ))
    rows.append((
        "PULL referentiel.uo_types -> FILL_TABLE(Activites, TabActivites)  MODE=APPEND_NEW  KEY=id",
        "",
        False,
    ))

    # Définitions des tables locales
    rows.append(("# --- DEFINITIONS LOCALES (DEF) ---", "", True))
    rows.append((
        "DEF $activites = GET_TABLE(Activites, TabActivites)",
        "",
        False,
    ))
    rows.append((
        "DEF $livrables = GET_TABLE(Livrables, TabLivrables)",
        "",
        False,
    ))
    rows.append((
        "DEF $points_ouverts = GET_TABLE(Points Ouverts, TabPO)",
        "",
        False,
    ))
    rows.append((
        "DEF $avancement_global = COMPUTE(MEAN_WEIGHTED($activites.avancement, $activites.heures))",
        "",
        False,
    ))
    rows.append((
        "DEF $heures_realisees = COMPUTE(SUM($activites.heures_realisees))",
        "",
        False,
    ))
    rows.append((
        "DEF $nb_po_ouverts = COMPUTE(COUNT_IF($points_ouverts.statut, \"En cours\"))",
        "",
        False,
    ))

    # Colonnes de TabActivites
    rows.append(("# Colonnes : TabActivites", "", True))
    rows.append(('COL $activites.id          : KEY  HEADER="ID"', "", False))
    rows.append(('COL $activites.nom         : WRITE=creation  HEADER="Activite"', "", False))
    rows.append(('COL $activites.heures      : WRITE=creation  HEADER="Heures allouees"', "", False))
    rows.append(('COL $activites.date_debut  : WRITE=engineer  HEADER="Date debut"', "", False))
    rows.append(('COL $activites.date_fin    : WRITE=engineer  HEADER="Date fin"', "", False))
    rows.append(('COL $activites.avancement  : WRITE=engineer  HEADER="% Avancement"', "", False))
    rows.append(('COL $activites.heures_realisees : WRITE=engineer  HEADER="Heures realisees"', "", False))
    rows.append(('COL $activites.commentaire : WRITE=engineer  HEADER="Commentaire"', "", False))

    # Colonnes de TabLivrables
    rows.append(("# Colonnes : TabLivrables", "", True))
    rows.append(('COL $livrables.id          : KEY  HEADER="ID"', "", False))
    rows.append(('COL $livrables.nom         : WRITE=creation  HEADER="Nom du livrable"', "", False))
    rows.append(('COL $livrables.date_prevue : WRITE=creation  HEADER="Date prevue"', "", False))
    rows.append(('COL $livrables.date_reelle : WRITE=engineer  HEADER="Date reelle"', "", False))
    rows.append(('COL $livrables.statut      : WRITE=engineer  HEADER="Statut"', "", False))

    # Colonnes de TabPO
    rows.append(("# Colonnes : TabPO (Points Ouverts)", "", True))
    rows.append(('COL $points_ouverts.id           : KEY  HEADER="ID"', "", False))
    rows.append(('COL $points_ouverts.date_ouv     : WRITE=engineer  HEADER="Date ouv."', "", False))
    rows.append(('COL $points_ouverts.description  : WRITE=engineer  HEADER="Description"', "", False))
    rows.append(('COL $points_ouverts.nature       : WRITE=engineer  HEADER="Nature"', "", False))
    rows.append(('COL $points_ouverts.responsable  : WRITE=engineer  HEADER="Responsable"', "", False))
    rows.append(('COL $points_ouverts.statut       : WRITE=engineer  HEADER="Statut"', "", False))
    rows.append(('COL $points_ouverts.date_clot    : WRITE=engineer  HEADER="Date clot."', "", False))

    # Liaisons Dashboard (BIND)
    rows.append(("# --- LIAISONS DASHBOARD (BIND) ---", "", True))
    rows.append(("BIND $avancement_global -> Dashboard.avancement_global", "", False))
    rows.append(("BIND $heures_realisees  -> Dashboard.heures_realisees", "", False))
    rows.append(("BIND $nb_po_ouverts     -> Dashboard.nb_po_ouverts", "", False))

    # Exports vers le store (PUSH)
    rows.append(("# --- EXPORTS STORE (PUSH) ---", "", True))
    rows.append(("PUSH $activites       -> uo.activites", "", False))
    rows.append(("PUSH $livrables       -> uo.livrables", "", False))
    rows.append(("PUSH $points_ouverts  -> uo.points_ouverts", "", False))
    rows.append(("PUSH $avancement_global -> uo.avancement_global", "", False))
    rows.append(("PUSH $heures_realisees  -> uo.heures_realisees", "", False))
    rows.append(("PUSH $nb_po_ouverts     -> uo.nb_po_ouverts", "", False))

    # ── Écrire les lignes ──────────────────────────────────────────────────────
    data_idx = 0
    for r, (instr, anchor, is_cmt) in enumerate(rows, start=3):
        if is_cmt:
            _comment(ws, r, instr[2:])  # enlève le "# " du préfixe
        else:
            _row(ws, r, instr, anchor, even=(data_idx % 2 == 0))
            data_idx += 1

    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 25


def generate_uo_file(uo: UOInstance, output_dir: Path = OUTPUT_DIR) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    _sheet_organisation(wb, uo)
    _sheet_livrables(wb, uo)
    _sheet_planning(wb, uo)
    _sheet_activites(wb, uo)
    _sheet_rex(wb, uo)
    _sheet_points_ouverts(wb, uo)
    _sheet_dashboard(wb, uo)
    _sheet_manifeste(wb, uo)

    # Ajouter les tableaux nommés pour la résolution manifeste
    activities = uo.uo_type.activities if uo.uo_type else []
    deliverables = uo.uo_type.deliverables if uo.uo_type else []
    nb_act = len(activities)
    nb_del = len(deliverables)

    if nb_act > 0:
        last_act = 2 + nb_act
        _add_named_table(wb["Activités"], f"A2:H{last_act}", "TabActivites")
    if nb_del > 0:
        last_del = 2 + nb_del
        _add_named_table(wb["Livrables"], f"A2:E{last_del}", "TabLivrables")
    nb_po = 10  # lignes vides pré-remplies dans Points Ouverts
    _add_named_table(wb["Points Ouverts"], f"A2:G{2 + nb_po}", "TabPO")

    actors = uo.project.actors if uo.project else []
    nb_act_proj = max(len(actors), 1)
    _add_named_table(wb["Organisation Projet"], f"A2:D{2 + nb_act_proj}", "TabActeurs")

    # Statut UO sur le Dashboard
    degrade_label = f" [DEGRADE{': ' + uo.degrade_note if uo.degrade_note else ''}]" if uo.degrade else ""
    wb["Dashboard"]["B3"] = f"{uo.statut.value}{degrade_label}"

    type_name = uo.uo_type.id if uo.uo_type else uo.uo_type_id
    sys_name = uo.system.id if uo.system else uo.system_id
    filename = f"{uo.id}_{type_name}_{sys_name}.xlsx"
    filepath = output_dir / filename
    wb.save(filepath)
    return filepath
