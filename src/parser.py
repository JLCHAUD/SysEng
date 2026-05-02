"""
Parser du méta-langage MXL.

Lit la feuille _Manifeste (colonne A = instruction, colonne B = ancre)
et produit un AST structuré + alimente l'ecosystem schema.

Syntaxe supportée :
  FILE_TYPE: <type>
  FILE_ID:   <id>
  VERSION:   <n>
  # commentaire

  DEF $var = GET_CELL(sheet, named_range)
  DEF $var = GET_TABLE(sheet, table_name)
  DEF $var = COMPUTE(formula)

  COL $table.col : KEY [HEADER="..."]
  COL $table.col : WRITE=<who> [HEADER="..."] [LOCKED]

  BIND $var -> sheet.named_range

  PUSH $var -> global.variable.name

  PULL global.var -> FILL_TABLE(sheet, table)  MODE=<mode> [KEY=col] [COLS=c1;c2]
  PULL global.var -> UPDATE_CELLS(sheet, table, KEY=col, COLS=c1;c2)

Note : la feuille s'appelait auparavant _Passerelle (ADR-001).
"""
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from src import ecosystem as Ecosystem
from src.ecosystem import ColumnSchema, TableSchema, VariableSchema


# ─── Nœuds de l'AST ──────────────────────────────────────────────────────────

@dataclass
class FileHeader:
    file_type: str = ""
    file_id: str = ""
    version: str = "1"
    doc: str = ""
    manifest_metadata: Dict[str, str] = field(default_factory=dict)  # champs libres (owner, projet…)


@dataclass
class DefNode:
    var_name: str                   # "$activites"
    source_type: str                # "GET_CELL" | "GET_TABLE" | "COMPUTE"
    sheet: str = ""                 # pour GET_CELL / GET_TABLE
    table_name: str = ""            # pour GET_TABLE
    named_range: str = ""           # pour GET_CELL
    formula: str = ""               # pour COMPUTE
    anchor: str = ""                # colonne B


@dataclass
class ColNode:
    full_var: str                   # "$activites.avancement"
    table_var: str                  # "$activites"
    col_name: str                   # "avancement"
    is_key: bool = False
    write: str = ""                 # "engineer" | "creation" | "uo_generique" | ...
    header: str = ""                # "% Avancement"
    locked: bool = False


@dataclass
class BindNode:
    var_name: str                   # "$avancement"
    target_sheet: str               # "Dashboard"
    target_range: str               # "avancement_global"
    anchor: str = ""                # colonne B  ex: "Dashboard.F3"


@dataclass
class PushNode:
    var_name: str                   # "$activites"
    global_name: str                # "uo.activites"
    only_if: str = ""               # condition MXL — vide = toujours pousser
                                    # ex: "$total_heures > 0"


@dataclass
class PullNode:
    global_name: str                # "projet.acteurs"
    operation: str                  # "FILL_TABLE" | "UPDATE_CELLS"
    sheet: str = ""
    table: str = ""
    mode: str = "READ_ONLY"         # READ_ONLY | APPEND_NEW | UPDATE | OVERWRITE
    key: str = ""                   # colonne KEY pour APPEND_NEW / UPDATE
    cols: str = ""                  # colonnes pour UPDATE_CELLS  "col1;col2"
    anchor: str = ""


@dataclass
class ExtendsNode:
    template_name: str   # "uo_generique"  (sans extension .mxl)


@dataclass
class ValidateNode:
    var_ref: str            # "$activites.avancement"  ou  "$total_heures"
    rule: str               # "RANGE(0, 100)" | "NOT_NULL" | "IN(...)" | ...
    severity: str = "error" # "error" (bloquant) | "warning" (non bloquant)


@dataclass
class NotifyNode:
    """Instruction NOTIFY — envoie une alerte conditionnelle."""
    channel: str         # "log" | "email" | "webhook"
    message: str         # message MXL (peut contenir $variables)
    condition: str = ""  # condition MXL — vide = toujours notifier
    target: str = ""     # adresse email ou URL webhook


@dataclass
class ListNode:
    """
    Déclaration d'une liste de fichiers fils.

    Forme TABLE   : LIST UOs_actifs FROM TABLE liste_uo
    Forme DYNAMIC : LIST uo_mi20 TYPE=uo_instance WHERE projet=MI20
    """
    name: str                                    # "UOs_actifs"
    form: str                                    # "TABLE" | "DYNAMIC"
    source_table: str = ""                       # TABLE → nom de la table Excel
    filter_type: str = ""                        # DYNAMIC → FILE_TYPE filtre
    filter_where: List[Tuple[str, str, str]] = field(default_factory=list)
    # DYNAMIC → [(champ, op, valeur), ...]  ex: [("projet", "=", "MI20")]


@dataclass
class CollectNode:
    """
    Agrégation cross-fichiers depuis une liste vers une table locale.

    COLLECT Planning FROM UOs_actifs INTO vue_planning
    COLLECT Risques  FROM UOs_actifs INTO vue_risques WHERE criticite >= 3
    COLLECT Livrables FROM UOs_actifs INTO vue_livrables COLS=[id, libelle, statut]
    COLLECT Planning FROM uo_mi20 INTO vue_planning WITH owner, projet
    """
    source_table: str                            # "Planning"
    list_name: str                               # "UOs_actifs"
    target_table: str                            # "vue_planning"
    where_clause: str = ""                       # "criticite >= 3"
    cols_filter: List[str] = field(default_factory=list)   # ["id", "libelle", "statut"]
    with_fields: List[str] = field(default_factory=list)   # ["owner", "projet"]


@dataclass
class ParseError:
    line_num: int
    raw: str
    message: str


@dataclass
class PasserelleAST:
    header: FileHeader = field(default_factory=FileHeader)
    extends: Optional["ExtendsNode"] = None
    defs: List[DefNode] = field(default_factory=list)
    cols: List[ColNode] = field(default_factory=list)
    binds: List[BindNode] = field(default_factory=list)
    pushes: List[PushNode] = field(default_factory=list)
    pulls: List[PullNode] = field(default_factory=list)
    validates: List[ValidateNode] = field(default_factory=list)
    notifies: List[NotifyNode] = field(default_factory=list)
    lists: List[ListNode] = field(default_factory=list)
    collects: List[CollectNode] = field(default_factory=list)
    errors: List[ParseError] = field(default_factory=list)

    # Index rapide : nom de variable → DefNode
    _defs_index: Dict[str, DefNode] = field(default_factory=dict)
    # Index : table_var → liste de ColNode
    _cols_index: Dict[str, List[ColNode]] = field(default_factory=dict)

    def def_for(self, var_name: str) -> Optional[DefNode]:
        return self._defs_index.get(var_name)

    def cols_for(self, table_var: str) -> List[ColNode]:
        return self._cols_index.get(table_var, [])


# ─── Parsers de sous-expressions ─────────────────────────────────────────────

def _extract_args(text: str) -> List[str]:
    """Extrait les arguments d'une expression func(arg1, arg2, ...)."""
    m = re.search(r'\((.+)\)$', text.strip(), re.DOTALL)
    if not m:
        return []
    inner = m.group(1)
    # Split sur virgule en dehors des parenthèses imbriquées
    args, depth, current = [], 0, []
    for ch in inner:
        if ch == '(':
            depth += 1
            current.append(ch)
        elif ch == ')':
            depth -= 1
            current.append(ch)
        elif ch == ',' and depth == 0:
            args.append(''.join(current).strip())
            current = []
        else:
            current.append(ch)
    if current:
        args.append(''.join(current).strip())
    return args


def _parse_kv_attrs(text: str) -> Dict[str, str]:
    """
    Parse les attributs clé=valeur et flags dans une chaîne.
    Ex: 'WRITE=engineer HEADER="% Avancement" LOCKED'
    → {'WRITE': 'engineer', 'HEADER': '% Avancement', 'LOCKED': 'true'}
    """
    attrs: Dict[str, str] = {}
    # KEY=VALUE avec valeur entre guillemets
    for m in re.finditer(r'(\w+)="([^"]*)"', text):
        attrs[m.group(1).upper()] = m.group(2)
    # KEY=VALUE sans guillemets
    for m in re.finditer(r'(\w+)=([^\s"]+)', text):
        key = m.group(1).upper()
        if key not in attrs:
            attrs[key] = m.group(2)
    # Flags seuls (mots isolés en majuscules)
    for m in re.finditer(r'(?<![=\w])([A-Z_]{2,})(?![=\w"])', text):
        flag = m.group(1)
        if flag not in attrs and flag not in ('MODE', 'KEY', 'COLS', 'WRITE', 'HEADER'):
            attrs[flag] = 'true'
    return attrs


# ─── Parsers par instruction ─────────────────────────────────────────────────

_KNOWN_HEADER_KEYS = {"FILE_TYPE", "FILE_ID", "VERSION", "DOC"}


def _parse_header_line(line: str, ast: PasserelleAST) -> bool:
    """
    Tente de parser une ligne d'en-tête. Retourne True si consommée.
    Les clés connues (FILE_TYPE, FILE_ID, VERSION, DOC) alimentent FileHeader.
    Toute autre ligne 'KEY: value' est stockée dans manifest_metadata
    (ex : owner, projet, has_risques) pour les listes DYNAMIC.
    """
    m = re.match(r'^([\w_]+)\s*:\s*(.+)$', line, re.IGNORECASE)
    if not m:
        return False

    key = m.group(1).upper()
    val = m.group(2).strip().strip('"').strip("'")

    if key == "FILE_TYPE":
        ast.header.file_type = val
    elif key == "FILE_ID":
        ast.header.file_id = val
    elif key == "VERSION":
        ast.header.version = val
    elif key == "DOC":
        ast.header.doc = val
    elif key not in {"DEF", "COL", "BIND", "PUSH", "PULL",
                     "VALIDATE", "EXTENDS", "NOTIFY", "LIST", "COLLECT"}:
        # Champ de métadonnée libre (owner, projet, has_risques…)
        ast.header.manifest_metadata[key.lower()] = val
    else:
        return False  # mot-clé MXL, pas une ligne d'en-tête

    return True


def _parse_def(line: str, anchor: str) -> Optional[DefNode]:
    """
    DEF $var = GET_CELL(sheet, named_range)
    DEF $var = GET_TABLE(sheet, table_name)
    DEF $var = COMPUTE(formula...)
    """
    m = re.match(r'^DEF\s+(\$[\w]+)\s*=\s*(\w+)\((.+)\)$', line.strip(), re.DOTALL)
    if not m:
        return None

    var_name = m.group(1)
    func = m.group(2).upper()
    inner = m.group(3).strip()

    node = DefNode(var_name=var_name, source_type=func, anchor=anchor)

    if func == "GET_CELL":
        # GET_CELL(sheet name, named_range)  — la virgule sépare les deux
        parts = [p.strip() for p in inner.split(',', 1)]
        if len(parts) == 2:
            node.sheet = parts[0]
            node.named_range = parts[1]
        else:
            node.sheet = inner

    elif func == "GET_TABLE":
        parts = [p.strip() for p in inner.split(',', 1)]
        node.sheet = parts[0]
        if len(parts) == 2:
            node.table_name = parts[1]

    elif func == "COMPUTE":
        node.formula = inner

    else:
        return None  # fonction inconnue

    return node


def _parse_col(line: str) -> Optional[ColNode]:
    """
    COL $table.col : KEY [HEADER="..."]
    COL $table.col : WRITE=who [HEADER="..."] [LOCKED]
    """
    m = re.match(r'^COL\s+(\$[\w.]+)\s*:\s*(.+)$', line.strip())
    if not m:
        return None

    full_var = m.group(1)
    attrs_str = m.group(2)

    # Décomposer $table.col
    parts = full_var.lstrip('$').split('.', 1)
    table_var = '$' + parts[0]
    col_name = parts[1] if len(parts) > 1 else ""

    attrs = _parse_kv_attrs(attrs_str)
    is_key = 'KEY' in attrs_str.upper().split()

    return ColNode(
        full_var=full_var,
        table_var=table_var,
        col_name=col_name,
        is_key=is_key,
        write=attrs.get('WRITE', ''),
        header=attrs.get('HEADER', col_name),
        locked='LOCKED' in attrs,
    )


def _parse_bind(line: str, anchor: str) -> Optional[BindNode]:
    """BIND $var -> sheet.named_range"""
    m = re.match(r'^BIND\s+(\$[\w]+)\s*->\s*([\w\s]+)\.([\w_]+)$', line.strip())
    if not m:
        return None
    return BindNode(
        var_name=m.group(1),
        target_sheet=m.group(2).strip(),
        target_range=m.group(3).strip(),
        anchor=anchor,
    )


def _parse_push(line: str) -> Optional[PushNode]:
    """
    PUSH $var -> global.variable.name
    PUSH $var -> global.variable.name  ONLY_IF $var > 0
    PUSH $var -> global.variable.name  ONLY_IF $var != NULL
    PUSH $var -> global.variable.name  ONLY_IF $var = "VERT"
    """
    m = re.match(
        r'^PUSH\s+(\$[\w]+)\s*->\s*([\w.\-]+)'
        r'(?:\s+ONLY_IF\s+(.+))?$',
        line.strip(), re.IGNORECASE,
    )
    if not m:
        return None
    return PushNode(
        var_name=m.group(1),
        global_name=m.group(2),
        only_if=(m.group(3) or "").strip(),
    )


def _parse_validate(line: str) -> Optional[ValidateNode]:
    """
    VALIDATE $table.col  : RULE
    VALIDATE $table.col  : RULE  SEVERITY=warning

    Règles supportées :
      NOT_NULL, POSITIVE, NON_NEGATIVE, UNIQUE
      RANGE(min, max)
      IN("a", "b", ...)
    """
    m = re.match(
        r'^VALIDATE\s+(\$[\w.]+)\s*:\s*(.+)$',
        line.strip(), re.IGNORECASE,
    )
    if not m:
        return None

    var_ref  = m.group(1).strip()
    rest     = m.group(2).strip()

    # Extraire SEVERITY= en fin de ligne
    severity = "error"
    sev_m = re.search(r'\bSEVERITY\s*=\s*(\w+)', rest, re.IGNORECASE)
    if sev_m:
        severity = sev_m.group(1).lower()
        rest = rest[:sev_m.start()].strip()

    rule = rest.strip()
    if not rule:
        return None

    return ValidateNode(var_ref=var_ref, rule=rule, severity=severity)


def _parse_notify(line: str) -> Optional["NotifyNode"]:
    """
    NOTIFY  log   "message"
    NOTIFY  log   "message"  IF $var > seuil
    NOTIFY  email "message"  TO addr@example.com  IF $var > seuil
    NOTIFY  webhook "message"  TO https://...  IF $var > seuil
    """
    m = re.match(
        r'^NOTIFY\s+(log|email|webhook)\s+"([^"]+)"(.*)$',
        line.strip(), re.IGNORECASE,
    )
    if not m:
        return None

    channel = m.group(1).lower()
    message = m.group(2)
    rest    = m.group(3).strip()

    condition = ""
    target    = ""

    # Extraire TO ...
    to_m = re.search(r'\bTO\s+(\S+)', rest, re.IGNORECASE)
    if to_m:
        target = to_m.group(1)
        rest   = rest[:to_m.start()].strip() + rest[to_m.end():].strip()

    # Extraire IF ...
    if_m = re.search(r'\bIF\s+(.+)$', rest, re.IGNORECASE)
    if if_m:
        condition = if_m.group(1).strip()

    return NotifyNode(channel=channel, message=message,
                      condition=condition, target=target)


def _parse_pull(line: str, anchor: str) -> Optional[PullNode]:
    """
    PULL global.var -> FILL_TABLE(sheet, table)  MODE=x [KEY=col]
    PULL global.var -> UPDATE_CELLS(sheet, table, KEY=col, COLS=c1;c2)
    """
    m = re.match(
        r'^PULL\s+([\w.*\-]+)\s*->\s*(FILL_TABLE|UPDATE_CELLS)\(([^)]+)\)\s*(.*)$',
        line.strip(), re.IGNORECASE,
    )
    if not m:
        return None

    global_name = m.group(1)
    operation = m.group(2).upper()
    func_args_str = m.group(3)
    trailing = m.group(4)

    # Arguments de la fonction : "sheet, table" ou "sheet, table, KEY=col, COLS=..."
    func_args = [a.strip() for a in func_args_str.split(',')]
    sheet = func_args[0] if len(func_args) > 0 else ""
    table = func_args[1] if len(func_args) > 1 else ""

    # Attributs dans les args de la fonction (pour UPDATE_CELLS)
    func_attrs = _parse_kv_attrs(func_args_str)
    # Attributs dans le trailing  (MODE=x KEY=y COLS=z)
    trail_attrs = _parse_kv_attrs(trailing)
    all_attrs = {**func_attrs, **trail_attrs}

    return PullNode(
        global_name=global_name,
        operation=operation,
        sheet=sheet,
        table=table,
        mode=all_attrs.get('MODE', 'READ_ONLY'),
        key=all_attrs.get('KEY', ''),
        cols=all_attrs.get('COLS', ''),
        anchor=anchor,
    )


# ─── Parsers LIST / COLLECT ───────────────────────────────────────────────────

def _parse_list_where(where_str: str) -> List[Tuple[str, str, str]]:
    """
    Parse 'field=value AND field2=value2' en [(champ, op, valeur), ...].
    Opérateurs supportés : = et !=
    """
    result = []
    for part in re.split(r'\bAND\b', where_str, flags=re.IGNORECASE):
        part = part.strip()
        m = re.match(r'^([\w_]+)\s*(!=|=)\s*(.+)$', part)
        if m:
            result.append((m.group(1).strip(), m.group(2), m.group(3).strip()))
    return result


def _parse_list(line: str) -> Optional[ListNode]:
    """
    LIST UOs_actifs FROM TABLE liste_uo
    LIST uo_mi20 TYPE=uo_instance WHERE projet=MI20
    LIST equipe_Jean TYPE=uo_instance WHERE owner=Jean AND has_risques=oui
    """
    line = line.strip()

    # Forme TABLE
    m = re.match(
        r'^LIST\s+([\w_]+)\s+FROM\s+TABLE\s+([\w_]+)$',
        line, re.IGNORECASE,
    )
    if m:
        return ListNode(name=m.group(1), form="TABLE", source_table=m.group(2))

    # Forme DYNAMIC
    m = re.match(
        r'^LIST\s+([\w_]+)\s+TYPE=([\w_]+)(?:\s+WHERE\s+(.+))?$',
        line, re.IGNORECASE,
    )
    if m:
        where_str = (m.group(3) or "").strip()
        return ListNode(
            name=m.group(1),
            form="DYNAMIC",
            filter_type=m.group(2),
            filter_where=_parse_list_where(where_str) if where_str else [],
        )

    return None


def _parse_collect(line: str) -> Optional[CollectNode]:
    """
    COLLECT <source_table> FROM <list_name> INTO <target_table>
            [WHERE <condition>] [COLS=[col1, col2, ...]] [WITH field1, field2]

    Tous les modificateurs sont inline sur la même ligne Excel.
    Exemples :
      COLLECT Planning FROM UOs_actifs INTO vue_planning
      COLLECT Risques  FROM UOs_actifs INTO vue_risques WHERE criticite >= 3
      COLLECT Livrables FROM UOs_actifs INTO vue_livrables COLS=[id, libelle, statut]
      COLLECT Planning FROM uo_mi20 INTO vue_planning WITH owner, projet
    """
    m = re.match(
        r'^COLLECT\s+([\w_]+)\s+FROM\s+([\w_]+)\s+INTO\s+([\w_]+)(.*)?$',
        line.strip(), re.IGNORECASE,
    )
    if not m:
        return None

    source_table = m.group(1)
    list_name    = m.group(2)
    target_table = m.group(3)
    trailing     = (m.group(4) or "").strip()

    where_clause = ""
    cols_filter: List[str] = []
    with_fields: List[str] = []

    # COLS=[col1, col2, ...] — extraire avant WHERE pour éviter le conflit
    cols_m = re.search(r'\bCOLS=\[([^\]]+)\]', trailing, re.IGNORECASE)
    if cols_m:
        cols_filter = [c.strip() for c in cols_m.group(1).split(',') if c.strip()]
        trailing = trailing[:cols_m.start()] + trailing[cols_m.end():]

    # WITH field1, field2  (doit être en fin de ligne)
    with_m = re.search(r'\bWITH\s+([\w_,\s]+)$', trailing, re.IGNORECASE)
    if with_m:
        with_fields = [f.strip() for f in with_m.group(1).split(',') if f.strip()]
        trailing = trailing[:with_m.start()].strip()

    # WHERE <condition>
    where_m = re.search(r'\bWHERE\s+(.+)$', trailing, re.IGNORECASE)
    if where_m:
        where_clause = where_m.group(1).strip()

    return CollectNode(
        source_table=source_table,
        list_name=list_name,
        target_table=target_table,
        where_clause=where_clause,
        cols_filter=cols_filter,
        with_fields=with_fields,
    )


# ─── Parser principal ─────────────────────────────────────────────────────────

def parse_lines(lines: List[Tuple[str, str]]) -> PasserelleAST:
    """
    Parse une liste de (instruction, ancre) et retourne un PasserelleAST.
    lines = [(col_A_value, col_B_value), ...]
    """
    ast = PasserelleAST()

    for line_num, (raw_instr, anchor) in enumerate(lines, start=1):
        instr = (raw_instr or "").strip()

        # Sauter vides et commentaires
        if not instr or instr.startswith('#'):
            continue

        # En-tête
        if _parse_header_line(instr, ast):
            continue

        # Identifier le mot-clé
        keyword = instr.split()[0].upper() if instr.split() else ""

        if keyword == "DEF":
            node = _parse_def(instr, anchor or "")
            if node:
                ast.defs.append(node)
                ast._defs_index[node.var_name] = node
            else:
                ast.errors.append(ParseError(line_num, instr, "Syntaxe DEF invalide"))

        elif keyword == "COL":
            node = _parse_col(instr)
            if node:
                ast.cols.append(node)
                if node.table_var not in ast._cols_index:
                    ast._cols_index[node.table_var] = []
                ast._cols_index[node.table_var].append(node)
            else:
                ast.errors.append(ParseError(line_num, instr, "Syntaxe COL invalide"))

        elif keyword == "BIND":
            node = _parse_bind(instr, anchor or "")
            if node:
                ast.binds.append(node)
            else:
                ast.errors.append(ParseError(line_num, instr, "Syntaxe BIND invalide"))

        elif keyword == "PUSH":
            node = _parse_push(instr)
            if node:
                ast.pushes.append(node)
            else:
                ast.errors.append(ParseError(line_num, instr, "Syntaxe PUSH invalide"))

        elif keyword == "PULL":
            node = _parse_pull(instr, anchor or "")
            if node:
                ast.pulls.append(node)
            else:
                ast.errors.append(ParseError(line_num, instr, "Syntaxe PULL invalide"))

        elif keyword == "VALIDATE":
            node = _parse_validate(instr)
            if node:
                ast.validates.append(node)
            else:
                ast.errors.append(ParseError(line_num, instr, "Syntaxe VALIDATE invalide"))

        elif keyword == "EXTENDS":
            # EXTENDS <nom_template>
            parts = instr.split()
            if len(parts) == 2:
                ast.extends = ExtendsNode(template_name=parts[1].strip())
            else:
                ast.errors.append(ParseError(line_num, instr, "Syntaxe EXTENDS invalide"))

        elif keyword == "NOTIFY":
            node = _parse_notify(instr)
            if node:
                ast.notifies.append(node)
            else:
                ast.errors.append(ParseError(line_num, instr, "Syntaxe NOTIFY invalide"))

        elif keyword == "LIST":
            node = _parse_list(instr)
            if node:
                ast.lists.append(node)
            else:
                ast.errors.append(ParseError(line_num, instr, "Syntaxe LIST invalide"))

        elif keyword == "COLLECT":
            node = _parse_collect(instr)
            if node:
                ast.collects.append(node)
            else:
                ast.errors.append(ParseError(line_num, instr, "Syntaxe COLLECT invalide"))

        else:
            ast.errors.append(ParseError(line_num, instr,
                                          f"Mot-clé inconnu : '{keyword}'"))

    return ast


def parse_sheet(ws) -> PasserelleAST:
    """
    Parse une feuille openpyxl _Manifeste.
    Colonne A = instruction, Colonne B = ancre.
    Ligne 1 = VERSION (traitée par parse_lines via FILE_TYPE/FILE_ID/VERSION).
    """
    lines: List[Tuple[str, str]] = []

    # Ligne 1 : version (ex: "MANIFESTE_V=1" → on la convertit en "VERSION: 1")
    # Rétro-compat : accepte aussi l'ancien format "PASSERELLE_V="
    v_raw = str(ws["A1"].value or "").strip()
    if v_raw.startswith("MANIFESTE_V="):
        version = v_raw.replace("MANIFESTE_V=", "").replace("-MOD", "").strip()
        lines.append((f"VERSION: {version}", ""))
    elif v_raw.startswith("PASSERELLE_V="):
        version = v_raw.replace("PASSERELLE_V=", "").replace("-MOD", "").strip()
        lines.append((f"VERSION: {version}", ""))

    for row in ws.iter_rows(min_row=3, values_only=False):
        col_a = row[0].value if row[0] else None
        col_b = row[1].value if len(row) > 1 and row[1] else None
        instr = str(col_a).strip() if col_a is not None else ""
        anchor = str(col_b).strip() if col_b is not None else ""
        lines.append((instr, anchor))

    return parse_lines(lines)


MANIFESTE_SHEET  = "_Manifeste"
_LEGACY_SHEET    = "_Passerelle"   # rétro-compat ADR-001
TEMPLATES_DIR    = Path(__file__).parent.parent / "config" / "templates"


# ─── EXTENDS — résolution et fusion ──────────────────────────────────────────

def parse_mxl_file(filepath: Path, substitutions: Dict[str, str] = None) -> PasserelleAST:
    """
    Parse un fichier .mxl texte (template) et retourne un AST.
    Chaque ligne non vide et non commentaire est une instruction (pas d'ancre).
    Substitue les variables {{FILE_ID}}, {{DOC}} si substitutions est fourni.
    """
    substitutions = substitutions or {}
    lines: List[Tuple[str, str]] = []
    with open(filepath, encoding="utf-8") as f:
        for raw in f:
            instr = raw.strip()
            # Substituer les variables template
            for placeholder, value in substitutions.items():
                instr = instr.replace(f"{{{{{placeholder}}}}}", value)
            lines.append((instr, ""))  # ancre vide pour les templates
    return parse_lines(lines)


def merge_asts(child: PasserelleAST, template: PasserelleAST) -> PasserelleAST:
    """
    Fusionne un AST enfant avec un AST template selon les règles de fusion MXL :

      Header (FILE_TYPE, FILE_ID, VERSION, DOC) : l'enfant remplace le template
      DEF $var                                  : l'enfant remplace par même nom
      PULL, COL, VALIDATE, BIND, PUSH           : additifs (template + enfant)
      errors                                    : additifs
    """
    merged = PasserelleAST()

    # ── Header : l'enfant prime ────────────────────────────────────────────────
    merged.header = FileHeader(
        file_type = child.header.file_type or template.header.file_type,
        file_id   = child.header.file_id   or template.header.file_id,
        version   = child.header.version   or template.header.version,
        doc       = child.header.doc       or template.header.doc,
    )

    # ── DEF : l'enfant remplace le template si même var_name ─────────────────
    template_defs: Dict[str, DefNode] = {d.var_name: d for d in template.defs}
    child_defs:    Dict[str, DefNode] = {d.var_name: d for d in child.defs}
    merged_defs = {**template_defs, **child_defs}   # child gagne

    # Reconstruire en préservant l'ordre : template d'abord, puis nouveaux enfant
    seen: set = set()
    ordered: List[DefNode] = []
    for d in template.defs:
        node = merged_defs[d.var_name]
        ordered.append(node)
        seen.add(d.var_name)
    for d in child.defs:
        if d.var_name not in seen:
            ordered.append(d)
    merged.defs = ordered
    merged._defs_index = {d.var_name: d for d in ordered}

    # ── Listes additives ──────────────────────────────────────────────────────
    merged.pulls     = template.pulls     + child.pulls
    merged.cols      = template.cols      + child.cols
    merged.validates = template.validates + child.validates
    merged.binds     = template.binds     + child.binds
    merged.pushes    = template.pushes    + child.pushes
    merged.lists     = template.lists     + child.lists
    merged.collects  = template.collects  + child.collects

    # ── COL index ────────────────────────────────────────────────────────────
    for col in merged.cols:
        if col.table_var not in merged._cols_index:
            merged._cols_index[col.table_var] = []
        merged._cols_index[col.table_var].append(col)

    # ── Erreurs ───────────────────────────────────────────────────────────────
    merged.errors = template.errors + child.errors

    return merged


def resolve_extends(
    ast: PasserelleAST,
    templates_dir: Path = TEMPLATES_DIR,
) -> PasserelleAST:
    """
    Si l'AST contient une instruction EXTENDS, charge le template .mxl,
    substitue les variables et retourne l'AST fusionné.
    Sinon, retourne l'AST inchangé.
    """
    if ast.extends is None:
        return ast

    template_path = templates_dir / f"{ast.extends.template_name}.mxl"
    if not template_path.exists():
        ast.errors.append(ParseError(
            line_num=0,
            raw=f"EXTENDS {ast.extends.template_name}",
            message=f"Template introuvable : {template_path}",
        ))
        return ast

    substitutions = {
        "FILE_ID": ast.header.file_id,
        "DOC":     ast.header.doc,
    }
    try:
        template_ast = parse_mxl_file(template_path, substitutions)
    except Exception as exc:
        ast.errors.append(ParseError(
            line_num=0,
            raw=f"EXTENDS {ast.extends.template_name}",
            message=f"Erreur lecture template : {exc}",
        ))
        return ast

    return merge_asts(child=ast, template=template_ast)


def parse_file(filepath: Path) -> Optional[PasserelleAST]:
    """
    Ouvre un fichier Excel, cherche la feuille _Manifeste et la parse.
    Accepte aussi l'ancien nom _Passerelle (rétro-compat ADR-001).
    Retourne None si aucune des deux feuilles n'est présente.
    """
    from openpyxl import load_workbook
    if not filepath.exists():
        return None
    wb = load_workbook(filepath, read_only=True, data_only=True)
    sheet_name = None
    if MANIFESTE_SHEET in wb.sheetnames:
        sheet_name = MANIFESTE_SHEET
    elif _LEGACY_SHEET in wb.sheetnames:
        sheet_name = _LEGACY_SHEET
    if sheet_name is None:
        wb.close()
        return None
    ast = parse_sheet(wb[sheet_name])
    ast.header.file_id = ast.header.file_id or filepath.stem
    wb.close()
    return resolve_extends(ast)


# ─── Enrichissement de l'écosystème ──────────────────────────────────────────

def enrich_ecosystem(ast: PasserelleAST) -> Tuple[int, int]:
    """
    Extrait les tables et variables déclarées dans l'AST
    et les enregistre dans l'ecosystem schema.

    Retourne (nb_tables_nouvelles, nb_variables_nouvelles).
    """
    new_tables: List[TableSchema] = []
    new_variables: List[VariableSchema] = []

    file_id = ast.header.file_id

    for push in ast.pushes:
        def_node = ast.def_for(push.var_name)
        if def_node is None:
            continue

        global_name = push.global_name

        if def_node.source_type == "GET_TABLE":
            # Construire le TableSchema depuis DEF + COL
            cols_nodes = ast.cols_for(push.var_name)
            columns: Dict[str, ColumnSchema] = {}
            for col in cols_nodes:
                col_type = "KEY" if col.is_key else _infer_col_type(col.col_name)
                columns[col.col_name] = ColumnSchema(
                    name=col.col_name,
                    col_type=col_type,
                    header=col.header or col.col_name,
                    write=col.write,
                )

            new_tables.append(TableSchema(
                id=global_name,
                source_file_id=file_id,
                source_sheet=def_node.sheet,
                table_name=def_node.table_name,
                columns=columns,
                discovered_from=f"{file_id}/_Manifeste",
            ))

        elif def_node.source_type in ("GET_CELL", "COMPUTE"):
            var_type = "COMPUTED" if def_node.source_type == "COMPUTE" else "CELL"
            new_variables.append(VariableSchema(
                id=global_name,
                var_type=var_type,
                source_file_id=file_id,
                formula=def_node.formula if var_type == "COMPUTED" else "",
                discovered_from=f"{file_id}/_Manifeste",
            ))

    Ecosystem.register_many(new_tables, new_variables)
    return len(new_tables), len(new_variables)


def _infer_col_type(col_name: str) -> str:
    """Infère le type d'une colonne depuis son nom."""
    name = col_name.lower()
    if any(k in name for k in ("date", "debut", "fin", "cloture")):
        return "date"
    if any(k in name for k in ("avancement", "pct", "pourcent")):
        return "pct"
    if any(k in name for k in ("heures", "charge", "budget", "nb_", "nombre")):
        return "float"
    return "string"


# ─── Rapport lisible ──────────────────────────────────────────────────────────

def ast_summary(ast: PasserelleAST) -> str:
    """Retourne un résumé lisible de l'AST parsé."""
    lines = [
        f"FILE_TYPE : {ast.header.file_type}",
        f"FILE_ID   : {ast.header.file_id}",
        f"VERSION   : {ast.header.version}",
        f"",
        f"DEF       : {len(ast.defs)} variable(s)",
        f"COL       : {len(ast.cols)} colonne(s)",
        f"BIND      : {len(ast.binds)} lien(s)",
        f"PUSH      : {len(ast.pushes)} export(s)",
        f"PULL      : {len(ast.pulls)} import(s)",
        f"LIST      : {len(ast.lists)} liste(s)",
        f"COLLECT   : {len(ast.collects)} agrégation(s)",
        f"Metadata  : {ast.header.manifest_metadata or '—'}",
        f"Erreurs   : {len(ast.errors)}",
    ]
    if ast.errors:
        lines.append("")
        for e in ast.errors:
            lines.append(f"  [L{e.line_num}] {e.message} — '{e.raw}'")
    return "\n".join(lines)
