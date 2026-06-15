"""
design_demo.py — Demos de design Excel : option A (Studio clair) et B (Cockpit bandeau).
=========================================================================================
Genere deux classeurs de demonstration poussant les capacites natives d'Excel
(sans VBA), pour choisir la direction visuelle des fichiers UO :

  Design_A_Studio.xlsx   minimalisme clair, filets d'accent, data bars vivantes
  Design_B_Cockpit.xlsx  bandeau marine, navigation hyperliens, jauge anneau

Techniques utilisees : quadrillage masque, onglets colores, cellules fusionnees
(cartes KPI), bordures d'accent, mise en forme conditionnelle (data bars, jeux
d'icones feux tricolores, badges colores par valeur), liens HYPERLINK internes,
volets figes, graphique anneau (jauge), formats numeriques personnalises.

Usage : python projet_TrainSystem/design_demo.py
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import DoughnutChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule, DataBarRule, IconSetRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = Path(__file__).parent

# ── Palette (3 couleurs + etats) ──────────────────────────────────────────────
NAVY_D = "0C447C"   # marine fonce — titres, bandeaux
NAVY = "185FA5"     # marine — accents
BLUE = "378ADD"     # bleu — data bars
BLUE_L = "E6F1FB"   # bleu pale — fonds doux
GREY_D = "5F5E5A"   # gris fonce — texte secondaire
GREY_L = "F5F4F0"   # gris chaud pale — fonds de carte
GREY_B = "D3D1C7"   # gris — bordures
GREEN = "639922"; GREEN_L = "EAF3DE"; GREEN_D = "27500A"
AMBER = "EF9F27"; AMBER_L = "FAEEDA"; AMBER_D = "854F0B"
RED = "E24B4A"; RED_L = "FCEBEB"; RED_D = "791F1F"
WHITE = "FFFFFF"

F = "Segoe UI"
THIN_G = Side(style="thin", color=GREY_B)
HAIR = Border(left=THIN_G, right=THIN_G, top=THIN_G, bottom=THIN_G)

ACTIVITES = [
    ("ACT-01", "Analyse fonctionnelle du besoin", "TERMINEE", 100, 24, 22.0),
    ("ACT-02", "Specification des exigences", "EN_COURS", 60, 30, 28.5),
    ("ACT-03", "Definition d'architecture", "EN_COURS", 35, 40, 12.0),
    ("ACT-04", "Calculs et simulations", "A_FAIRE", 0, 35, 0.0),
    ("ACT-05", "Revue de conception", "A_FAIRE", 0, 25, 0.0),
    ("ACT-06", "Preparation gate review", "STAND_BY", 10, 46, 4.0),
]
OIL = [
    ("PO-001", "Question expert normes EN 45545", "EXPERT", "HAUTE", "OUVERT"),
    ("PO-002", "Attente retour relecture fournisseur", "FOURNISSEUR", "MOYENNE", "OUVERT"),
    ("PO-003", "Acces outil documentaire obtenu", "SE", "BASSE", "CLOS"),
]


def fnt(size=11, bold=False, color="2C2C2A", italic=False):
    return Font(name=F, size=size, bold=bold, color=color, italic=italic)


def fill(color):
    return PatternFill("solid", fgColor=color)


def card_border(ws, r1, c1, r2, c2, side=None, color=GREY_B):
    """Encadre une zone (carte) d'une bordure fine, option accent gauche."""
    thin = Side(style="thin", color=color)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            b = {}
            if r == r1: b["top"] = thin
            if r == r2: b["bottom"] = thin
            if c == c1: b["left"] = side or thin
            if c == c2: b["right"] = thin
            old = cell.border
            cell.border = Border(
                left=b.get("left", old.left), right=b.get("right", old.right),
                top=b.get("top", old.top), bottom=b.get("bottom", old.bottom))


def add_table(ws, name, ref):
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name="TableStyleLight15", showRowStripes=True)
    ws.add_table(t)


def statut_cf(ws, rng):
    """Badges colores par valeur de statut."""
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"TERMINEE"'], fill=fill(GREEN_L),
        font=Font(name=F, size=10, bold=True, color=GREEN_D)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"EN_COURS"'], fill=fill(BLUE_L),
        font=Font(name=F, size=10, bold=True, color=NAVY_D)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"STAND_BY"'], fill=fill(AMBER_L),
        font=Font(name=F, size=10, bold=True, color=AMBER_D)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"A_FAIRE"'], fill=fill(GREY_L),
        font=Font(name=F, size=10, color=GREY_D)))


def criticite_cf(ws, rng):
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"HAUTE"'], fill=fill(RED_L),
        font=Font(name=F, size=10, bold=True, color=RED_D)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"MOYENNE"'], fill=fill(AMBER_L),
        font=Font(name=F, size=10, bold=True, color=AMBER_D)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"BASSE"'], fill=fill(GREY_L),
        font=Font(name=F, size=10, color=GREY_D)))


def activites_sheet(wb, banner=None, header_color=NAVY_D):
    """Feuille de saisie Activites — style Studio (commun aux 2 designs)."""
    ws = wb.create_sheet("Activites")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = BLUE
    widths = {"A": 9, "B": 38, "C": 13, "D": 13, "E": 11, "F": 14, "G": 13}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    top = 1
    if banner:
        banner(ws, "Saisie — Activités", 7)
        top = 5

    headers = ["id", "designation", "statut", "avancement", "load (h)",
               "consommé (h)", "reste (h)"]
    hr = top + 1
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=ci, value=h)
        c.font = fnt(10, bold=True, color=WHITE)
        c.fill = fill(header_color)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[hr].height = 22

    for ri, (aid, des, st, av, load, conso) in enumerate(ACTIVITES, hr + 1):
        vals = [aid, des, st, av, load, conso, round((1 - av / 100) * load, 1)]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = fnt(10)
            c.border = HAIR
            c.alignment = Alignment(
                horizontal="center" if ci in (1, 3, 4) else
                ("right" if ci >= 5 else "left"),
                vertical="center")
        ws.cell(row=ri, column=5).number_format = "0.0"
        ws.cell(row=ri, column=6).number_format = "0.0"
        ws.cell(row=ri, column=7).number_format = "0.0"
        ws.row_dimensions[ri].height = 20

    last = hr + len(ACTIVITES)
    rng_av = f"D{hr+1}:D{last}"
    ws.conditional_formatting.add(rng_av, DataBarRule(
        start_type="num", start_value=0, end_type="num", end_value=100,
        color=BLUE, showValue=True))
    statut_cf(ws, f"C{hr+1}:C{last}")
    ws.freeze_panes = f"A{hr+1}"
    return ws


def oil_sheet(wb, banner=None, header_color=NAVY_D):
    ws = wb.create_sheet("OIL")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = AMBER
    widths = {"A": 9, "B": 42, "C": 15, "D": 12, "E": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    top = 1
    if banner:
        banner(ws, "Points ouverts — OIL", 5)
        top = 5
    headers = ["id", "titre", "en_action", "criticite", "statut"]
    hr = top + 1
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=ci, value=h)
        c.font = fnt(10, bold=True, color=WHITE)
        c.fill = fill(header_color)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[hr].height = 22
    for ri, row in enumerate(OIL, hr + 1):
        for ci, v in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = fnt(10)
            c.border = HAIR
            c.alignment = Alignment(
                horizontal="center" if ci != 2 else "left", vertical="center")
        ws.row_dimensions[ri].height = 20
    last = hr + len(OIL)
    criticite_cf(ws, f"D{hr+1}:D{last}")
    ws.conditional_formatting.add(f"E{hr+1}:E{last}", CellIsRule(
        operator="equal", formula=['"OUVERT"'], fill=fill(RED_L),
        font=Font(name=F, size=10, bold=True, color=RED_D)))
    ws.conditional_formatting.add(f"E{hr+1}:E{last}", CellIsRule(
        operator="equal", formula=['"CLOS"'], fill=fill(GREEN_L),
        font=Font(name=F, size=10, color=GREEN_D)))
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# DESIGN A — STUDIO CLAIR
# ══════════════════════════════════════════════════════════════════════════════

def kpi_card_A(ws, col, label, value, sub, accent, value_color, bar_pct=None):
    """Carte KPI : filet d'accent a gauche, grande valeur, sous-texte."""
    c1 = col
    r = 4
    for rr in range(r, r + 4):
        for cc in range(c1, c1 + 3):
            ws.cell(row=rr, column=cc).fill = fill(GREY_L)
    accent_side = Side(style="thick", color=accent)
    card_border(ws, r, c1, r + 3, c1 + 2, side=accent_side)
    lab = ws.cell(row=r, column=c1, value=label)
    lab.font = fnt(9, bold=True, color=GREY_D)
    lab.alignment = Alignment(horizontal="left", vertical="bottom", indent=1)
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c1 + 2)
    val = ws.cell(row=r + 1, column=c1, value=value)
    val.font = Font(name=F, size=22, bold=True, color=value_color)
    val.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=r + 1, start_column=c1, end_row=r + 2, end_column=c1 + 2)
    s = ws.cell(row=r + 3, column=c1, value=sub)
    s.font = fnt(9, color=GREY_D)
    s.alignment = Alignment(horizontal="left", vertical="top", indent=1)
    ws.merge_cells(start_row=r + 3, start_column=c1, end_row=r + 3, end_column=c1 + 2)


def build_design_A():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = NAVY_D
    for i in range(1, 16):
        ws.column_dimensions[get_column_letter(i)].width = 8.5
    ws.column_dimensions["A"].width = 2

    # Titre
    t = ws.cell(row=2, column=2, value="L09U1 — Préparation et passage des IDR")
    t.font = Font(name=F, size=16, bold=True, color=NAVY_D)
    sub = ws.cell(row=3, column=2,
                  value="Projet Demo · Climatisation · Jean Dujardin · sync 12/06 08:12")
    sub.font = fnt(9, color=GREY_D, italic=True)
    ws.row_dimensions[2].height = 26

    # 4 cartes KPI (lignes 4-7)
    for rr in range(4, 8):
        ws.row_dimensions[rr].height = 18
    kpi_card_A(ws, 2, "AVANCEMENT UO", "45,0 %", "moyenne pondérée par poids",
               NAVY, NAVY_D)
    kpi_card_A(ws, 5, "HEURES", "59 / 200", "consommées / vendues", GREY_D, "2C2C2A")
    kpi_card_A(ws, 8, "EAC — DÉRIVE", "169 h", "▼ −31 h vs vendu", AMBER, AMBER_D)
    kpi_card_A(ws, 11, "SANTÉ", "● ROUGE", "1 point critique ouvert", RED, RED_D)

    # Mini-table avancement par activite avec data bars + feux
    hr = 10
    ws.cell(row=9, column=2, value="Avancement par activité").font = \
        fnt(11, bold=True, color=NAVY_D)
    heads = ["", "activité", "", "", "statut", "avancement", "", "heures"]
    for ci, h in [(2, "activité"), (5, "statut"), (6, "avancement"), (8, "heures")]:
        c = ws.cell(row=hr, column=ci, value=h)
        c.font = fnt(9, bold=True, color=GREY_D)
    for ci in range(2, 9):
        ws.cell(row=hr, column=ci).border = Border(bottom=Side(style="thin", color=GREY_B))
    for ri, (aid, des, st, av, load, conso) in enumerate(ACTIVITES, hr + 1):
        ws.cell(row=ri, column=2, value=f"{aid} — {des}").font = fnt(10)
        ws.merge_cells(start_row=ri, start_column=2, end_row=ri, end_column=4)
        sc = ws.cell(row=ri, column=5, value=st)
        sc.font = fnt(10); sc.alignment = Alignment(horizontal="center")
        avc = ws.cell(row=ri, column=6, value=av)
        avc.alignment = Alignment(horizontal="center"); avc.font = fnt(10)
        ws.merge_cells(start_row=ri, start_column=6, end_row=ri, end_column=7)
        hc = ws.cell(row=ri, column=8, value=conso)
        hc.number_format = "0.0"; hc.font = fnt(10)
        hc.alignment = Alignment(horizontal="right")
        ws.row_dimensions[ri].height = 19
    last = hr + len(ACTIVITES)
    ws.conditional_formatting.add(f"F{hr+1}:F{last}", DataBarRule(
        start_type="num", start_value=0, end_type="num", end_value=100,
        color=BLUE, showValue=True))
    ws.conditional_formatting.add(f"H{hr+1}:H{last}", IconSetRule(
        icon_style="3TrafficLights1", type="percent", values=[0, 33, 67],
        showValue=True))
    statut_cf(ws, f"E{hr+1}:E{last}")

    activites_sheet(wb)
    oil_sheet(wb)
    out = HERE / "Design_A_Studio.xlsx"
    wb.save(out)
    print(f"[OK] {out.name}")


# ══════════════════════════════════════════════════════════════════════════════
# DESIGN B — COCKPIT BANDEAU
# ══════════════════════════════════════════════════════════════════════════════

# Couleurs de bandeau par onglet (bandeau = couleur d'onglet, tres visuel)
TEAL = "0F6E56"; TEAL_CHIP = "085041"; TEAL_TINT = "9FE1CB"
AMB_B = "BA7517"; AMB_CHIP = "854F0B"; AMB_TINT = "FAC775"


def banner_B(ws, subtitle, ncols, bg=NAVY_D, chip=NAVY, tint="B5D4F4"):
    """Bandeau colore lignes 1-4, largeur EXACTE ncols (ne deborde pas).
    Titre UO puis, en dessous, systeme — projet en texte simple (meme fond).
    La couleur du bandeau = la couleur de l'onglet (coherence)."""
    ws.sheet_properties.tabColor = bg
    for rr in range(1, 5):
        for cc in range(1, ncols + 1):
            ws.cell(row=rr, column=cc).fill = fill(bg)
    t = ws.cell(row=1, column=2, value="UO L09U1 — Préparation et passage des IDR")
    t.font = Font(name=F, size=14, bold=True, color=WHITE)
    t.alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=1, start_column=2, end_row=1,
                   end_column=max(ncols - 1, 3))

    # Systeme — Projet : texte simple, meme fond que le bandeau
    sp = ws.cell(row=2, column=2, value="CLIMATISATION  —  PROJET DEMO")
    sp.font = Font(name=F, size=12, bold=True, color=WHITE)
    sp.alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=2, start_column=2, end_row=2,
                   end_column=min(13, max(ncols - 1, 3)))

    nav = [("⌂ Dashboard", "Dashboard"), ("✎ Activités", "Activites"),
           ("⚑ OIL", "OIL")]
    for i, (label, target) in enumerate(nav):
        # navigation adaptative : sur feuille etroite, 1 cellule par pastille
        col = 2 + i * 2 if ncols >= 8 else 2 + i
        c = ws.cell(row=3, column=col)
        c.value = f'=HYPERLINK("#{target}!A1","{label}")'
        c.font = Font(name=F, size=9, bold=True, color=tint)
        if ncols >= 8:
            ws.merge_cells(start_row=3, start_column=col,
                           end_row=3, end_column=col + 1)
    if ncols >= 12:
        s = ws.cell(row=3, column=ncols - 2, value=subtitle + " · J. Dujardin")
        s.font = Font(name=F, size=9, color=tint)
        s.alignment = Alignment(horizontal="right", vertical="center")
        ws.merge_cells(start_row=3, start_column=ncols - 2, end_row=3,
                       end_column=ncols)
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 6


def banner_teal(ws, subtitle, ncols):
    banner_B(ws, subtitle, ncols, bg=TEAL, chip=TEAL_CHIP, tint=TEAL_TINT)


def banner_amber(ws, subtitle, ncols):
    banner_B(ws, subtitle, ncols, bg=AMB_B, chip=AMB_CHIP, tint=AMB_TINT)


def section_box(ws, title, r1, c1, r2, c2):
    """Delimite une zone : bande de titre remplie + cadre fin autour."""
    for cc in range(c1, c2 + 1):
        ws.cell(row=r1, column=cc).fill = fill(BLUE_L)
    tc = ws.cell(row=r1, column=c1, value=title)
    tc.font = fnt(11, bold=True, color=NAVY_D)
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r1].height = 20
    card_border(ws, r1, c1, r2, c2, color=GREY_B)


def make_donut(wb, ws_dash, ws_data, data_row, anchor, label, pct, color):
    """Petit camembert-jauge : pct colore, reste pale. Donnees en feuille cachee."""
    ws_data.cell(row=data_row, column=1, value="fait")
    ws_data.cell(row=data_row, column=2, value=pct)
    ws_data.cell(row=data_row + 1, column=1, value="reste")
    ws_data.cell(row=data_row + 1, column=2, value=100 - pct)
    chart = DoughnutChart()
    data = Reference(ws_data, min_col=2, min_row=data_row, max_row=data_row + 1)
    chart.add_data(data, titles_from_data=False)
    chart.holeSize = 60
    serie = chart.series[0]
    p1 = DataPoint(idx=0); p1.graphicalProperties.solidFill = color
    p2 = DataPoint(idx=1); p2.graphicalProperties.solidFill = "EFEDE7"
    serie.data_points = [p1, p2]
    chart.legend = None
    chart.width = 4.6; chart.height = 4.6
    ws_dash.add_chart(chart, anchor)


def kpi_card_B(ws, col, label, value, sub, border_color, value_color):
    r = 6
    card_border(ws, r, col, r + 3, col + 2, color=border_color)
    for rr in range(r, r + 4):
        for cc in range(col, col + 3):
            ws.cell(row=rr, column=cc).fill = fill(WHITE)
    card_border(ws, r, col, r + 3, col + 2, color=border_color)
    lab = ws.cell(row=r + 1, column=col, value=label)
    lab.font = fnt(9, bold=True, color=GREY_D)
    lab.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=r + 1, start_column=col, end_row=r + 1, end_column=col + 2)
    val = ws.cell(row=r + 2, column=col, value=value)
    val.font = Font(name=F, size=20, bold=True, color=value_color)
    val.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=r + 2, start_column=col, end_row=r + 2, end_column=col + 2)
    s = ws.cell(row=r + 3, column=col, value=sub)
    s.font = fnt(8.5, color=GREY_D)
    s.alignment = Alignment(horizontal="center", vertical="top")
    ws.merge_cells(start_row=r + 3, start_column=col, end_row=r + 3, end_column=col + 2)


def build_design_B():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = NAVY_D
    for i in range(1, 18):
        ws.column_dimensions[get_column_letter(i)].width = 8.5
    ws.column_dimensions["A"].width = 2

    banner_B(ws, "Cockpit de pilotage", 16)

    # Badge sante dans le bandeau (droite, une seule ligne)
    ws.merge_cells(start_row=2, start_column=14, end_row=2, end_column=16)
    b = ws.cell(row=2, column=14, value="SANTÉ ROUGE")
    b.font = Font(name=F, size=11, bold=True, color=WHITE)
    b.fill = fill(RED)
    b.alignment = Alignment(horizontal="center", vertical="center")

    # 4 cartes KPI bordees (lignes 6-9), alignees sur la largeur du bandeau
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 18
    kpi_card_B(ws, 2, "AVANCEMENT UO", "45,0 %", "pondéré par poids", "B5D4F4", NAVY_D)
    kpi_card_B(ws, 6, "CONSOMMÉ", "29,5 %", "59 h / 200 h", GREY_B, "2C2C2A")
    kpi_card_B(ws, 10, "POINTS OUVERTS", "2  ▲1", "dont 1 critique", "FAC775", AMBER_D)
    kpi_card_B(ws, 14, "EAC", "169 h", "dérive −31 h", GREY_B, GREEN_D)

    # Rangee de camemberts-jauges (sans titre — libelles en cellules dessous)
    ws_data = wb.create_sheet("_chart_data")
    ws_data.sheet_state = "hidden"
    donuts = [("B11", 2, "Avancement", 45, BLUE),
              ("F11", 6, "Heures", 30, "888780"),
              ("J11", 10, "Livrables", 33, GREEN),
              ("N11", 14, "Données d'entrée", 67, "EF9F27")]
    for i, (anchor, col, label, pct, color) in enumerate(donuts):
        make_donut(wb, ws, ws_data, 1 + i * 3, anchor, label, pct, color)
        lab = ws.cell(row=20, column=col, value=f"{label} — {pct} %")
        lab.font = fnt(10, bold=True, color=NAVY_D)
        lab.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=20, start_column=col, end_row=20,
                       end_column=col + 2)
    ws.row_dimensions[20].height = 18

    # Prochains livrables — zone delimitee, largeur du bandeau (B..P)
    section_box(ws, "Prochains livrables", 22, 2, 26, 16)
    livs = [("SAA.RS — Requirement Spec", "M3", "EN_COURS"),
            ("RSS.RS — Subsystem Spec", "M3", "A_FAIRE"),
            ("DR.RS — Design Review file", "—", "A_FAIRE")]
    hr = 23
    for ri, (des, mat, st) in enumerate(livs, hr):
        ws.cell(row=ri, column=2, value=des).font = fnt(10)
        ws.merge_cells(start_row=ri, start_column=2, end_row=ri, end_column=11)
        m = ws.cell(row=ri, column=12, value=mat)
        m.font = fnt(10, bold=True, color=NAVY_D)
        m.alignment = Alignment(horizontal="center")
        sc = ws.cell(row=ri, column=13, value=st)
        sc.font = fnt(9); sc.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=ri, start_column=13, end_row=ri, end_column=15)
        for cc in range(2, 16):
            ws.cell(row=ri, column=cc).border = \
                Border(bottom=Side(style="thin", color=GREY_B))
        ws.row_dimensions[ri].height = 19
    statut_cf(ws, f"M{hr}:M{hr+2}")

    # Balle dans quel camp — zone delimitee
    section_box(ws, "La balle est chez…", 28, 2, 30, 16)
    camps = [("FOURNISSEUR", 1, AMBER_L, AMBER_D), ("EXPERT", 1, RED_L, RED_D),
             ("NOUS (SE)", 0, GREEN_L, GREEN_D)]
    for i, (who, n, bg, fg) in enumerate(camps):
        cc = 2 + i * 5
        c = ws.cell(row=29, column=cc, value=f"{who} : {n}")
        c.font = Font(name=F, size=10, bold=True, color=fg)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=29, start_column=cc, end_row=29, end_column=cc + 3)
    ws.row_dimensions[29].height = 22

    # Avancement des activites EN COURS — zone delimitee
    en_cours = [a for a in ACTIVITES if a[2] != "TERMINEE"]
    section_box(ws, "Avancement des activités en cours", 32, 2,
                33 + len(en_cours), 16)
    note = ws.cell(row=32, column=10,
                   value="(les activités terminées ne sont pas affichées)")
    note.font = fnt(8.5, color=GREY_D, italic=True)
    note.alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(start_row=32, start_column=10, end_row=32, end_column=15)
    hr = 33
    for ci, h in [(2, "activité"), (10, "statut"), (12, "avancement"),
                  (15, "heures")]:
        c = ws.cell(row=hr, column=ci, value=h)
        c.font = fnt(9, bold=True, color=GREY_D)
    for cc in range(2, 17):
        ws.cell(row=hr, column=cc).border = \
            Border(bottom=Side(style="thin", color=GREY_B))
    for ri, (aid, des, st, av, load, conso) in enumerate(en_cours, hr + 1):
        ws.cell(row=ri, column=2, value=f"{aid} — {des}").font = fnt(10)
        ws.merge_cells(start_row=ri, start_column=2, end_row=ri, end_column=9)
        sc = ws.cell(row=ri, column=10, value=st)
        sc.font = fnt(10); sc.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=ri, start_column=10, end_row=ri, end_column=11)
        avc = ws.cell(row=ri, column=12, value=av)
        avc.alignment = Alignment(horizontal="center"); avc.font = fnt(10)
        ws.merge_cells(start_row=ri, start_column=12, end_row=ri, end_column=14)
        hc = ws.cell(row=ri, column=15, value=conso)
        hc.number_format = "0.0"; hc.font = fnt(10)
        hc.alignment = Alignment(horizontal="right")
        ws.row_dimensions[ri].height = 19
    last = hr + len(en_cours)
    ws.conditional_formatting.add(f"L{hr+1}:L{last}", DataBarRule(
        start_type="num", start_value=0, end_type="num", end_value=100,
        color=BLUE, showValue=True))
    statut_cf(ws, f"J{hr+1}:J{last}")

    activites_sheet(wb, banner=banner_teal, header_color=TEAL)
    oil_sheet(wb, banner=banner_amber, header_color=AMB_B)
    out = HERE / "Design_B_Cockpit.xlsx"
    wb.save(out)
    print(f"[OK] {out.name}")


if __name__ == "__main__":
    build_design_A()
    build_design_B()
    print("\nOuvre les deux fichiers dans Excel pour comparer :")
    print("  projet_TrainSystem/Design_A_Studio.xlsx")
    print("  projet_TrainSystem/Design_B_Cockpit.xlsx")
