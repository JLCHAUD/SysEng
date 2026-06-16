"""
design_b.py — Bibliothèque de style Design B « Cockpit bandeau ».
=================================================================
Palette, primitives et composants réutilisables par creer_uo.py et design_demo.py.
Ne contient aucune donnée de démo.
"""
from openpyxl.chart import DoughnutChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule, DataBarRule, IconSetRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Palette ───────────────────────────────────────────────────────────────────
NAVY_D = "0C447C"   # marine foncé — titres, bandeaux
NAVY   = "185FA5"   # marine — accents
BLUE   = "378ADD"   # bleu — data bars
BLUE_L = "E6F1FB"   # bleu pâle — fonds doux
GREY_D = "5F5E5A"   # gris foncé — texte secondaire
GREY_L = "F5F4F0"   # gris chaud pâle — fonds de carte
GREY_B = "D3D1C7"   # gris — bordures
GREEN  = "639922";  GREEN_L  = "EAF3DE";  GREEN_D  = "27500A"
AMBER  = "EF9F27";  AMBER_L  = "FAEEDA";  AMBER_D  = "854F0B"
RED    = "E24B4A";  RED_L    = "FCEBEB";  RED_D    = "791F1F"
WHITE  = "FFFFFF"

# Bandeaux d'onglets
TEAL      = "0F6E56"; TEAL_CHIP  = "085041"; TEAL_TINT  = "9FE1CB"
AMB_B     = "BA7517"; AMB_CHIP   = "854F0B"; AMB_TINT   = "FAC775"

# Typographie & bordures globales
F      = "Segoe UI"
THIN_G = Side(style="thin", color=GREY_B)
HAIR   = Border(left=THIN_G, right=THIN_G, top=THIN_G, bottom=THIN_G)

# ── Primitives ────────────────────────────────────────────────────────────────

def fnt(size=11, bold=False, color="2C2C2A", italic=False):
    return Font(name=F, size=size, bold=bold, color=color, italic=italic)


def fill(color):
    return PatternFill("solid", fgColor=color)


def card_border(ws, r1, c1, r2, c2, side=None, color=GREY_B):
    """Encadre une zone rectangulaire d'une bordure fine, option accent gauche."""
    thin = Side(style="thin", color=color)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            b = {}
            if r == r1: b["top"] = thin
            if r == r2: b["bottom"] = thin
            if c == c1: b["left"] = side or thin
            if c == c2: b["right"] = thin
            old = cell.border
            cell.border = Border(
                left=b.get("left", old.left),
                right=b.get("right", old.right),
                top=b.get("top", old.top),
                bottom=b.get("bottom", old.bottom),
            )


def add_table(ws, name, ref):
    """Crée un tableau Excel nommé avec style Light15 et lignes alternées."""
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name="TableStyleLight15", showRowStripes=True)
    ws.add_table(t)


def statut_cf(ws, rng):
    """Badges colorés par valeur de statut (colonne statut des activités)."""
    rules = [
        ("TERMINEE",  GREEN_L,  GREEN_D),
        ("EN_COURS",  BLUE_L,   NAVY_D),
        ("A_FAIRE",   GREY_L,   GREY_D),
        ("STAND_BY",  AMBER_L,  AMBER_D),
    ]
    for val, bg, fg in rules:
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="equal",
            formula=[f'"{val}"'],
            fill=fill(bg),
            font=Font(name=F, size=10, bold=True, color=fg),
        ))


def criticite_cf(ws, rng):
    """Badges colorés par criticité OIL (HAUTE/MOYENNE/BASSE)."""
    rules = [
        ("HAUTE",   RED_L,   RED_D),
        ("MOYENNE", AMBER_L, AMBER_D),
        ("BASSE",   GREEN_L, GREEN_D),
    ]
    for val, bg, fg in rules:
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="equal",
            formula=[f'"{val}"'],
            fill=fill(bg),
            font=Font(name=F, size=10, bold=True, color=fg),
        ))


# ── Bandeaux ──────────────────────────────────────────────────────────────────

def banner_B(ws, subtitle, ncols, bg=NAVY_D, chip=NAVY, tint="B5D4F4",
             title="UO L09U1 — Préparation et passage des IDR",
             project_line="CLIMATISATION  —  PROJET DEMO",
             se="J. Dujardin"):
    """Bandeau coloré lignes 1-4, largeur exacte ncols.
    Couleur de bandeau = couleur d'onglet (cohérence visuelle)."""
    ws.sheet_properties.tabColor = bg
    for rr in range(1, 5):
        for cc in range(1, ncols + 1):
            ws.cell(row=rr, column=cc).fill = fill(bg)
    t = ws.cell(row=1, column=2, value=title)
    t.font = Font(name=F, size=14, bold=True, color=WHITE)
    t.alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=1, start_column=2, end_row=1,
                   end_column=max(ncols - 1, 3))
    sp = ws.cell(row=2, column=2, value=project_line)
    sp.font = Font(name=F, size=12, bold=True, color=WHITE)
    sp.alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=2, start_column=2, end_row=2,
                   end_column=min(13, max(ncols - 1, 3)))
    nav = [("⌂ Dashboard", "Dashboard"), ("✎ Activités", "Activites"),
           ("⚑ OIL", "OIL")]
    for i, (label, target) in enumerate(nav):
        col = 2 + i * 2 if ncols >= 8 else 2 + i
        c = ws.cell(row=3, column=col)
        c.value = f'=HYPERLINK("#{target}!A1","{label}")'
        c.font = Font(name=F, size=9, bold=True, color=tint)
        if ncols >= 8:
            ws.merge_cells(start_row=3, start_column=col,
                           end_row=3, end_column=col + 1)
    if ncols >= 12:
        s = ws.cell(row=3, column=ncols - 2,
                    value=subtitle + (f" · {se}" if se else ""))
        s.font = Font(name=F, size=9, color=tint)
        s.alignment = Alignment(horizontal="right", vertical="center")
        ws.merge_cells(start_row=3, start_column=ncols - 2, end_row=3,
                       end_column=ncols)
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 6


def banner_teal(ws, subtitle, ncols, **kwargs):
    banner_B(ws, subtitle, ncols, bg=TEAL, chip=TEAL_CHIP, tint=TEAL_TINT, **kwargs)


def banner_amber(ws, subtitle, ncols, **kwargs):
    banner_B(ws, subtitle, ncols, bg=AMB_B, chip=AMB_CHIP, tint=AMB_TINT, **kwargs)


# ── Composants de mise en page ────────────────────────────────────────────────

def section_box(ws, title, r1, c1, r2, c2):
    """Zone délimitée : bande de titre fond bleu pâle + cadre fin."""
    for cc in range(c1, c2 + 1):
        ws.cell(row=r1, column=cc).fill = fill(BLUE_L)
    tc = ws.cell(row=r1, column=c1, value=title)
    tc.font = fnt(11, bold=True, color=NAVY_D)
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r1].height = 20
    card_border(ws, r1, c1, r2, c2, color=GREY_B)


def kpi_card_B(ws, col, label, value, sub, border_color, value_color, row=6):
    """Carte KPI 4 colonnes : label + grande valeur + sous-titre + bordure colorée."""
    r = row
    card_border(ws, r, col, r + 3, col + 2, color=border_color)
    for rr in range(r, r + 4):
        for cc in range(col, col + 3):
            ws.cell(row=rr, column=cc).fill = fill(WHITE)
    card_border(ws, r, col, r + 3, col + 2, color=border_color)
    lab = ws.cell(row=r + 1, column=col, value=label)
    lab.font = fnt(9, bold=True, color=GREY_D)
    lab.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=r + 1, start_column=col, end_row=r + 1,
                   end_column=col + 2)
    val = ws.cell(row=r + 2, column=col, value=value)
    val.font = Font(name=F, size=20, bold=True, color=value_color)
    val.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=r + 2, start_column=col, end_row=r + 2,
                   end_column=col + 2)
    s = ws.cell(row=r + 3, column=col, value=sub)
    s.font = fnt(8.5, color=GREY_D)
    s.alignment = Alignment(horizontal="center", vertical="top")
    ws.merge_cells(start_row=r + 3, start_column=col, end_row=r + 3,
                   end_column=col + 2)


def make_donut(wb, ws_dash, ws_data, data_row, anchor, label, pct, color):
    """Graphique anneau (jauge %) — données dans feuille _chart_data cachée."""
    ws_data.cell(row=data_row, column=1, value="fait")
    ws_data.cell(row=data_row, column=2, value=pct)
    ws_data.cell(row=data_row + 1, column=1, value="reste")
    ws_data.cell(row=data_row + 1, column=2, value=100 - pct)
    chart = DoughnutChart()
    data = Reference(ws_data, min_col=2, min_row=data_row, max_row=data_row + 1)
    chart.add_data(data, titles_from_data=False)
    chart.holeSize = 60
    serie = chart.series[0]
    p1 = DataPoint(idx=0); p1.graphicalProperties.solidFill = color
    p2 = DataPoint(idx=1); p2.graphicalProperties.solidFill = "EFEDE7"
    serie.data_points = [p1, p2]
    chart.legend = None
    chart.width = 4.6
    chart.height = 4.6
    ws_dash.add_chart(chart, anchor)
