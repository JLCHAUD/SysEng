"""
Patch des _Manifeste existants dans projet_TrainSystem.

Ajoute pilote_id aux cockpits et met à jour les manifestes dashboard
pour utiliser LIST cockpit_ingenieur + COLLECT tbl_mes_uos.

Usage : python projet_TrainSystem/patcher_manifestes.py
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo

# Mapping ingénieur ->pilote_id
PILOTE_MAP = {
    "Cockpit_Alice_Dubois":   "USR004",
    "Cockpit_Bruno_Lecomte":  "USR004",
    "Cockpit_Camille_Vidal":  "USR004",
    "Cockpit_Denis_Renard":   "USR005",
}

DOSSIER = Path(__file__).parent


def _font_comment():
    return Font(name="Calibri", size=9, color="666666", italic=True)

def _font_bold():
    return Font(name="Calibri", size=9.5, bold=True, color="1F3864")

def _font_normal():
    return Font(name="Calibri", size=9.5)


def patch_cockpit(xlsx_path: Path, file_id: str, pilote_id: str):
    """Ajoute pilote_id: {pilote_id} dans le _Manifeste du cockpit."""
    wb = load_workbook(str(xlsx_path))
    if "_Manifeste" not in wb.sheetnames:
        print(f"  [SKIP] pas de _Manifeste dans {xlsx_path.name}")
        return

    ws = wb["_Manifeste"]

    # Vérifier si pilote_id est déjà présent
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=1):
        cell = row[0]
        if cell.value and str(cell.value).startswith("pilote_id:"):
            print(f"  [OK] pilote_id déjà présent dans {xlsx_path.name}")
            wb.close()
            return

    # Trouver la ligne ingenieur: pour insérer pilote_id juste après
    insert_after = None
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=1):
        cell = row[0]
        if cell.value and str(cell.value).startswith("ingenieur:"):
            insert_after = cell.row
            break

    if insert_after is None:
        print(f"  [WARN] ligne 'ingenieur:' introuvable dans {xlsx_path.name}")
        wb.close()
        return

    # Décaler tout ce qui est après insert_after d'une ligne
    ws.insert_rows(insert_after + 1)
    new_row = insert_after + 1
    ws.cell(row=new_row, column=1, value=f"pilote_id: {pilote_id}").font = _font_normal()
    ws.cell(row=new_row, column=3,
            value="Pilote responsable — permet au dashboard de découvrir ce cockpit"
            ).font = _font_comment()

    wb.save(str(xlsx_path))
    print(f"  [OK] {xlsx_path.name} ->pilote_id: {pilote_id}")


def patch_dashboard(xlsx_path: Path, pilote_id: str):
    """Met à jour le _Manifeste dashboard : LIST cockpit + COLLECT tbl_mes_uos + DEF + PUSH."""
    wb = load_workbook(str(xlsx_path))
    if "_Manifeste" not in wb.sheetnames:
        print(f"  [SKIP] pas de _Manifeste dans {xlsx_path.name}")
        return

    ws = wb["_Manifeste"]

    # Trouver la première ligne LIST ou COLLECT (row >= 7) et effacer jusqu'à fin
    start_clear = None
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=1, max_col=1):
        cell = row[0]
        if cell.value and str(cell.value).strip():
            start_clear = cell.row
            break

    if start_clear:
        for r in range(start_clear, ws.max_row + 1):
            for c in range(1, 4):
                ws.cell(row=r, column=c).value = None

    # Écrire le nouveau manifeste à partir de la ligne 7
    def w(row, instr, comment=""):
        ws.cell(row=row, column=1, value=instr).font = (
            _font_bold() if any(instr.startswith(k) for k in ("DEF ", "PUSH ", "LIST ", "COLLECT "))
            else _font_normal()
        )
        if comment:
            ws.cell(row=row, column=3, value=comment).font = _font_comment()

    w(7, f"LIST mes_cockpits TYPE=cockpit_ingenieur WHERE pilote_id={pilote_id}",
      "Découverte des cockpits ingénieur de l'équipe de ce pilote")
    w(8, "COLLECT tbl_mes_uos FROM mes_cockpits INTO tbl_vue_synthese",
      "Agrégation de toutes les UOs de tous les cockpits de l'équipe")
    w(10, "DEF $synthese = GET_TABLE(Vue Synthèse, tbl_vue_synthese)",
      "Relit la table agrégée après COLLECT")
    w(11, f"PUSH $synthese -> dashboard.{pilote_id}.synthese",
      "Publie la vue consolidée de l'équipe dans le store central")

    # Ajouter la feuille Vue Synthèse si absente
    if "Vue Synthèse" not in wb.sheetnames:
        ws_vs = wb.create_sheet("Vue Synthèse")
        headers = ["_source_file_id", "ingenieur", "UO ID", "Type UO", "Système",
                   "Projet", "Charge (h)", "% Avancement", "H réalisées", "Date fin", "Alerte"]
        for col, h in enumerate(headers, 1):
            ws_vs.cell(row=1, column=col, value=h)
        ws_vs.cell(row=2, column=1, value="(vide — rempli par ExoSync)")
        tbl = Table(displayName="tbl_vue_synthese",
                    ref=f"A1:{chr(64+len(headers))}2")
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws_vs.add_table(tbl)
        print(f"  [+] Feuille 'Vue Synthèse' créée")

    wb.save(str(xlsx_path))
    print(f"  [OK] {xlsx_path.name} ->manifeste mis à jour")


def main():
    print("=== Patch des manifestes TrainSystem ===\n")

    print("Cockpits :")
    for file_id, pilote_id in PILOTE_MAP.items():
        xlsx = DOSSIER / f"{file_id}.xlsx"
        if xlsx.exists():
            patch_cockpit(xlsx, file_id, pilote_id)
        else:
            print(f"  [MISS] {xlsx.name} absent")

    print("\nDashboards :")
    dashboard_map = {
        "Dashboard_Jean-Luc_Bernard.xlsx": "USR004",
        "Dashboard_Denis_Renard.xlsx":     "USR005",
    }
    for fname, pilote_id in dashboard_map.items():
        xlsx = DOSSIER / fname
        if xlsx.exists():
            patch_dashboard(xlsx, pilote_id)
        else:
            print(f"  [MISS] {fname} absent")

    print("\nDone.")


if __name__ == "__main__":
    main()
