"""Generates one cockpit Excel file per engineer."""
from datetime import date, timedelta
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from src.models import UOInstance
from src.styles import (
    BLUE_DARK, BLUE_MID, BLUE_LIGHT, GREEN_LIGHT, ORANGE_LIGHT, RED_LIGHT,
    GREY_LIGHT, WHITE, YELLOW_LIGHT,
    THIN_BORDER, solid_fill, header_font, body_font, center, left,
    style_header_row, style_data_row, set_column_widths, freeze_top_row,
)

OUTPUT_DIR = Path(__file__).parent.parent.parent / "output" / "cockpits"
UO_DIR = Path(__file__).parent.parent.parent / "output" / "UOs"


def generate_cockpit(engineer_name: str, all_uo_instances: List[UOInstance],
                     output_dir: Path = OUTPUT_DIR) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)

    engineer_uo = [uo for uo in all_uo_instances if uo.engineer_name == engineer_name]

    wb = Workbook()
    ws = wb.active
    ws.title = "Cockpit"
    ws.sheet_view.showGridLines = False

    today = date.today()

    # ── Header ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value = f"Cockpit Ingénieur — {engineer_name}   |   {today.strftime('%d/%m/%Y')}"
    t.fill = solid_fill(BLUE_DARK)
    t.font = header_font(size=14)
    t.alignment = center()
    ws.row_dimensions[1].height = 32

    # ── KPI summary row ────────────────────────────────────────────────────────
    ws.merge_cells("A2:C2")
    ws["A2"].value = f"UO en cours : {len(engineer_uo)}"
    ws["A2"].fill = solid_fill(BLUE_LIGHT)
    ws["A2"].font = body_font(bold=True)
    ws["A2"].alignment = center()
    ws["A2"].border = THIN_BORDER

    total_h = sum(uo.total_hours for uo in engineer_uo)
    ws.merge_cells("D2:F2")
    ws["D2"].value = f"Charge totale : {total_h}h"
    ws["D2"].fill = solid_fill(BLUE_LIGHT)
    ws["D2"].font = body_font(bold=True)
    ws["D2"].alignment = center()
    ws["D2"].border = THIN_BORDER

    # ── UO Summary table ──────────────────────────────────────────────────────
    ws.merge_cells("A4:I4")
    sec = ws["A4"]
    sec.value = "Vue d'ensemble de mes UO"
    sec.fill = solid_fill(BLUE_MID)
    sec.font = header_font()
    sec.alignment = center()

    uo_headers = ["UO ID", "Type", "Système", "Projet", "Charge (h)", "% Avancement",
                  "H réalisées", "Date fin", "Alerte"]
    for col, h in enumerate(uo_headers, 1):
        ws.cell(row=5, column=col, value=h)
    style_header_row(ws, 5, 1, 9, color=BLUE_MID)

    for i, uo in enumerate(engineer_uo):
        row = 6 + i
        type_name = uo.uo_type.name if uo.uo_type else uo.uo_type_id
        sys_name = uo.system.name if uo.system else uo.system_id
        proj_name = uo.project.name if uo.project else uo.project_id

        uo_id_cell = ws.cell(row=row, column=1, value=uo.id)
        uo_id_cell.font = body_font(color="0563C1")
        uo_id_cell.alignment = left()
        uo_id_cell.border = THIN_BORDER

        # Hyperlink to UO file
        uo_type_id = uo.uo_type.id if uo.uo_type else uo.uo_type_id
        uo_sys_id = uo.system.id if uo.system else uo.system_id
        uo_filename = f"{uo.id}_{uo_type_id}_{uo_sys_id}.xlsx"
        uo_path = UO_DIR / uo_filename
        uo_id_cell.hyperlink = str(uo_path)

        ws.cell(row=row, column=2, value=type_name)
        ws.cell(row=row, column=3, value=sys_name)
        ws.cell(row=row, column=4, value=proj_name)
        ws.cell(row=row, column=5, value=uo.total_hours)
        ws.cell(row=row, column=6, value=0.0).number_format = "0%"
        ws.cell(row=row, column=7, value=0.0)

        date_cell = ws.cell(row=row, column=8, value=uo.end_date)
        date_cell.number_format = "DD/MM/YYYY"

        # Alert formula
        alert_formula = (
            f'=IF(G{row}>E{row},"⚠ Dérive heures",'
            f'IF(AND(H{row}<>"",H{row}<TODAY()+7),"⏰ Échéance proche","✅ OK"))'
        )
        ws.cell(row=row, column=9, value=alert_formula)

        style_data_row(ws, row, 2, 8, alternate=(i % 2 == 1))
        ws.cell(row=row, column=1).fill = solid_fill(GREY_LIGHT if i % 2 == 1 else WHITE)

    last_uo_row = 5 + len(engineer_uo)

    # Conditional formatting on Alert column
    if engineer_uo:
        alert_range = f"I6:I{last_uo_row}"
        ws.conditional_formatting.add(
            alert_range,
            CellIsRule(operator="equal", formula=['"⚠ Dérive heures"'], fill=solid_fill(RED_LIGHT)),
        )
        ws.conditional_formatting.add(
            alert_range,
            CellIsRule(operator="equal", formula=['"⏰ Échéance proche"'], fill=solid_fill(ORANGE_LIGHT)),
        )
        ws.conditional_formatting.add(
            alert_range,
            CellIsRule(operator="equal", formula=['"✅ OK"'], fill=solid_fill(GREEN_LIGHT)),
        )

    # ── Activités du jour ─────────────────────────────────────────────────────
    act_start_row = last_uo_row + 3

    ws.merge_cells(f"A{act_start_row}:I{act_start_row}")
    sec2 = ws[f"A{act_start_row}"]
    sec2.value = "Activités à traiter — prochaines échéances"
    sec2.fill = solid_fill(BLUE_MID)
    sec2.font = header_font()
    sec2.alignment = center()

    act_header_row = act_start_row + 1
    act_headers = ["UO ID", "Activité", "Heures allouées", "% Avancement", "Date fin prévue", "Action"]
    for col, h in enumerate(act_headers, 1):
        ws.cell(row=act_header_row, column=col, value=h)
    style_header_row(ws, act_header_row, 1, 6, color=BLUE_MID)

    act_row = act_header_row + 1
    for uo in engineer_uo:
        activities = uo.uo_type.activities if uo.uo_type else []
        total_default = sum(a.default_hours for a in activities) or 1
        for act in activities:
            allocated = round(act.default_hours / total_default * uo.total_hours, 1)
            ws.cell(row=act_row, column=1, value=uo.id)
            ws.cell(row=act_row, column=2, value=act.name)
            ws.cell(row=act_row, column=3, value=allocated)
            ws.cell(row=act_row, column=4, value=0.0).number_format = "0%"
            date_cell = ws.cell(row=act_row, column=5, value=uo.end_date)
            date_cell.number_format = "DD/MM/YYYY"
            ws.cell(row=act_row, column=6, value="")
            style_data_row(ws, act_row, 1, 6, alternate=(act_row % 2 == 0))
            act_row += 1

    set_column_widths(ws, {
        "A": 12, "B": 35, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 16, "I": 20,
    })
    freeze_top_row(ws)

    safe_name = engineer_name.replace(" ", "_")
    filename = f"Cockpit_{safe_name}.xlsx"
    filepath = output_dir / filename
    wb.save(filepath)
    return filepath
