"""xl_design — charte graphique Excel centralisée d'ExoSync (classe XD).

Importé par tous les générateurs : aucun style n'est défini inline ailleurs.
Voir docs/superpowers/specs/2026-06-22-design-system-excel-design.md.
"""
from dataclasses import dataclass

from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo


@dataclass(frozen=True)
class SheetStyle:
    banner: str   # ton foncé (bannière, texte blanc)
    header: str   # ton moyen (= tabColor + en-tête de tableau)
    accent: str   # ton clair (lignes alternées, cartes)
    glyph: str    # glyphe monochrome de la bannière


class XD:
    FONT_FAMILY = "Segoe UI"
    DEFAULT_TEXT = "2C2C2A"

    # ── Palette transversale (statuts / neutres) ───────────────
    WHITE = "FFFFFF"
    INPUT = "FFF2CC"
    GREEN_L = "EAF3DE"
    GREEN_D = "27500A"
    BLUE_L = "E6F1FB"
    NAVY_D = "0C447C"
    AMBER_L = "FAEEDA"
    AMBER_D = "854F0B"
    RED_L = "FCEBEB"
    RED_D = "791F1F"
    GREY_L = "F1EFE8"
    GREY_D = "5F5E5A"
    GREY_B = "D3D1C7"

    # ── Couleurs santé de la spine (indépendantes des familles) ─
    SPINE_DONE = "3B6D11"
    SPINE_OK = "0F8A66"
    SPINE_WATCH = "EF9F27"
    SPINE_ALERT = "A32D2D"
    SPINE_TODO = "6B6A64"

    # ── Bordure fine (4 côtés) ─────────────────────────────────
    _SIDE = Side(style="thin", color="D3D1C7")
    HAIR = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)

    # ── Registre des familles d'onglets (palette verrouillée) ──
    SHEETS = {
        "general":        SheetStyle("08335E", "0C447C", "E6F1FB", "⬢"),
        "dashboard":      SheetStyle("0E4474", "1763A8", "E3EFFA", "◉"),
        "description":    SheetStyle("1C5E92", "2E86C8", "E7F2FB", "✎"),
        "planning":       SheetStyle("074E60", "0A6E88", "DEEFF3", "◷"),
        "donnees_entree": SheetStyle("0A6149", "0F8A66", "E1F5EE", "⤓"),
        "activites":      SheetStyle("084434", "0C5E49", "E1F5EE", "✔"),
        "livrables":      SheetStyle("386114", "4F8A1E", "EBF3DE", "▣"),
        "oil":            SheetStyle("791F1F", "A32D2D", "FCEBEB", "⚑"),
        "kpi":            SheetStyle("3C3489", "534AB7", "EEEDFE", "▲"),
        "orga":           SheetStyle("4D4C47", "6B6A64", "F1EFE8", "❖"),
        "manifeste":      SheetStyle("1C1C1A", "2C2C2A", "F1EFE8", "⚙"),
    }

    @classmethod
    def sheet(cls, key):
        return cls.SHEETS[key]

    @classmethod
    def tab_colors(cls):
        return {k: v.header for k, v in cls.SHEETS.items()}

    # ── Primitives ─────────────────────────────────────────────
    @staticmethod
    def fnt(size=10, bold=False, color="2C2C2A", italic=False, underline=False):
        return Font(name=XD.FONT_FAMILY, size=size, bold=bold, color=color,
                    italic=italic, underline="single" if underline else None)

    @staticmethod
    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    @staticmethod
    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    @staticmethod
    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    @classmethod
    def banner(cls, ws, key, title, subtitle="", se="", n_cols=10, height=30):
        """Bannière 1 ligne : glyphe + titre à gauche, sous-titre · SE à droite.
        Pose aussi tabColor = ton moyen de la famille."""
        s = cls.sheet(key)
        ws.sheet_properties.tabColor = s.header
        for c in range(1, n_cols + 1):
            ws.cell(row=1, column=c).fill = cls.fill(s.banner)

        t = ws.cell(row=1, column=1, value=f"{s.glyph}  {title}")
        t.font = cls.fnt(14, bold=True, color=cls.WHITE)
        t.alignment = Alignment(vertical="center", indent=1)
        left_end = max(n_cols - 3, 1)
        if left_end > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1,
                           end_column=left_end)

        right_parts = [p for p in (subtitle, se) if p]
        if right_parts and n_cols > left_end + 1:
            r = ws.cell(row=1, column=left_end + 1,
                        value="   ·   ".join(right_parts))
            r.font = cls.fnt(10, color=cls.WHITE)
            r.alignment = Alignment(horizontal="right", vertical="center",
                                    indent=1)
            ws.merge_cells(start_row=1, start_column=left_end + 1,
                           end_row=1, end_column=n_cols)
        ws.row_dimensions[1].height = height

    @classmethod
    def table_header(cls, ws, row, headers, key, col_start=1):
        """En-tête de tableau coloré au ton moyen de l'onglet, texte blanc."""
        s = cls.sheet(key)
        for i, h in enumerate(headers):
            c = ws.cell(row=row, column=col_start + i, value=h)
            c.fill = cls.fill(s.header)
            c.font = cls.fnt(10, bold=True, color=cls.WHITE)
            c.alignment = cls.center()
            c.border = cls.HAIR
        ws.row_dimensions[row].height = 20

    @classmethod
    def data_row(cls, ws, row, i, col_start, col_end, key):
        """Ligne de données : alternance blanc (i pair) / accent (i impair)."""
        s = cls.sheet(key)
        bg = s.accent if i % 2 else cls.WHITE
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = cls.fill(bg)
            cell.font = cls.fnt(10)
            cell.border = cls.HAIR

    @classmethod
    def named_table(cls, ws, display_name, ref, key):
        """Table Excel nommée (pour GET_TABLE/COLLECT) avec STYLE CLAIR sans
        en-tête imposé + coloration manuelle de l'en-tête au ton de l'onglet.
        L'AutoFilter natif d'Excel est actif automatiquement."""
        s = cls.sheet(key)
        tbl = Table(displayName=display_name, ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight15", showRowStripes=True,
            showFirstColumn=False, showLastColumn=False,
            showColumnStripes=False,
        )
        ws.add_table(tbl)
        min_col, min_row, max_col, _ = range_boundaries(ref)
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=min_row, column=c)
            cell.fill = cls.fill(s.header)
            cell.font = cls.fnt(10, bold=True, color=cls.WHITE)
            cell.alignment = cls.center()

    @classmethod
    def statut_cf(cls, ws, rng):
        """Badges colorés par statut d'activité/livrable."""
        rules = [
            ("TERMINEE", cls.GREEN_L, cls.GREEN_D),
            ("VALIDE",   cls.GREEN_L, cls.GREEN_D),
            ("LIVRE",    cls.BLUE_L,  cls.NAVY_D),
            ("EN_COURS", cls.BLUE_L,  cls.NAVY_D),
            ("A_FAIRE",  cls.GREY_L,  cls.GREY_D),
            ("STAND_BY", cls.AMBER_L, cls.AMBER_D),
        ]
        for val, bg, fg in rules:
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="equal", formula=[f'"{val}"'],
                fill=cls.fill(bg),
                font=cls.fnt(10, bold=True, color=fg)))

    @classmethod
    def criticite_cf(cls, ws, rng):
        """Badges colorés par criticité OIL."""
        rules = [
            ("HAUTE",   cls.RED_L,   cls.RED_D),
            ("MOYENNE", cls.AMBER_L, cls.AMBER_D),
            ("BASSE",   cls.GREEN_L, cls.GREEN_D),
        ]
        for val, bg, fg in rules:
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="equal", formula=[f'"{val}"'],
                fill=cls.fill(bg),
                font=cls.fnt(10, bold=True, color=fg)))

    @classmethod
    def traffic_light(cls, ws, row, col, value, thresholds=(0.5, 0.8)):
        """Cellule au fond rouge/ambre/vert selon value et les seuils."""
        lo, hi = thresholds
        color = cls.RED_L if value < lo else (cls.AMBER_L if value < hi
                                              else cls.GREEN_L)
        cell = ws.cell(row=row, column=col)
        cell.fill = cls.fill(color)
        cell.border = cls.HAIR
        cell.alignment = cls.center()
        return cell

    @classmethod
    def card_border(cls, ws, r1, c1, r2, c2, color=None):
        """Encadre une zone rectangulaire d'une bordure fine."""
        thin = Side(style="thin", color=color or cls.GREY_B)
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cell = ws.cell(row=r, column=c)
                old = cell.border
                cell.border = Border(
                    left=thin if c == c1 else old.left,
                    right=thin if c == c2 else old.right,
                    top=thin if r == r1 else old.top,
                    bottom=thin if r == r2 else old.bottom,
                )

    @classmethod
    def section_box(cls, ws, title, r1, c1, r2, c2, key):
        """Bande de titre (accent de l'onglet) + cadre fin."""
        s = cls.sheet(key)
        for c in range(c1, c2 + 1):
            ws.cell(row=r1, column=c).fill = cls.fill(s.accent)
        tc = ws.cell(row=r1, column=c1, value=title)
        tc.font = cls.fnt(11, bold=True, color=s.banner)
        tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r1].height = 20
        cls.card_border(ws, r1, c1, r2, c2)

    @classmethod
    def health_spine(cls, ws, key, header_row, row_start, row_end,
                     status_col, spine_col=1, pct_col=None):
        """Colonne A fine + en-tête au ton bannière. Pose les règles de mise en
        forme conditionnelle santé, lues sur la colonne statut (recolore live)."""
        s = cls.sheet(key)
        spine_L = get_column_letter(spine_col)
        stat_L = get_column_letter(status_col)
        ws.column_dimensions[spine_L].width = 2.5
        ws.cell(row=header_row, column=spine_col).fill = cls.fill(s.banner)

        rng = f"{spine_L}{row_start}:{spine_L}{row_end}"

        def rule(formula, color):
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[formula], stopIfTrue=True, fill=cls.fill(color)))

        # ordre = priorité (premier vrai gagne)
        rule(f'OR(${stat_L}{row_start}="TERMINEE",${stat_L}{row_start}="VALIDE")',
             cls.SPINE_DONE)
        rule(f'OR(${stat_L}{row_start}="OUVERT",${stat_L}{row_start}="HAUTE")',
             cls.SPINE_ALERT)
        rule(f'${stat_L}{row_start}="STAND_BY"', cls.SPINE_WATCH)
        rule(f'${stat_L}{row_start}="EN_COURS"', cls.SPINE_OK)
        rule(f'OR(${stat_L}{row_start}="A_FAIRE",${stat_L}{row_start}="EN_ATTENTE")',
             cls.SPINE_TODO)
