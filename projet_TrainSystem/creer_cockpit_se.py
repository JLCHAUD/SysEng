"""
creer_cockpit_se.py — Génère un cockpit ingénieur à partir des UO du répertoire.

Scanne tous les fichiers L*.xlsx, extrait l'ingénieur depuis la feuille General,
groupe les UO par ingénieur et génère un cockpit Excel par ingénieur.

Usage :
    python projet_TrainSystem/creer_cockpit_se.py
    python projet_TrainSystem/creer_cockpit_se.py --pilote USR004
"""
import argparse
import sys
from datetime import date, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from src.xl_design import XD

HERE = Path(__file__).parent


# ─── Lecture des UO ──────────────────────────────────────────────────────────

def lire_uo_du_repertoire(dossier: Path) -> list[dict]:
    """Scanne les UO xlsx et retourne la liste de leurs métadonnées."""
    uo_list = []
    for xlsx in sorted(dossier.glob("L*.xlsx")):
        if xlsx.name.startswith("~"):
            continue
        try:
            wb = load_workbook(str(xlsx), data_only=True)
        except Exception:
            continue

        if "General" not in wb.sheetnames or "_Manifeste" not in wb.sheetnames:
            wb.close()
            continue

        ws_g = wb["General"]
        se_name = heures = systeme = projet = None
        for row in ws_g.iter_rows(min_row=1, max_row=30, values_only=True):
            if not row[1]:
                continue
            label = str(row[1])
            if "Ing" in label and "SE" in label:
                se_name = row[2]
            elif "Heures" in label:
                heures = row[2]
            elif "Syst" in label:
                systeme = row[2]
            elif "Projet" in label:
                projet = row[2]

        if se_name:
            uo_list.append({
                "file_id": xlsx.stem,
                "se_name": str(se_name),
                "heures":  heures or 0,
                "systeme": systeme or "",
                "projet":  projet or "",
            })
        wb.close()
    return uo_list


# ─── Onglet Agenda ───────────────────────────────────────────────────────────

def _agenda_section(ws, title: str, start_row: int,
                    uo_list: list[dict], bg_key: str = "planning") -> int:
    """Section agenda (titre + en-têtes + 1 ligne jaune par UO). Retourne la prochaine ligne libre."""
    XD.section_box(ws, title, start_row, 1, start_row, 6, bg_key)
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=6)
    start_row += 1

    headers = ["UO ID", "Activité", "Priorité", "Date échéance", "Statut", "Action"]
    XD.table_header(ws, start_row, headers, bg_key)
    start_row += 1

    for uo in uo_list:
        c1 = ws.cell(row=start_row, column=1, value=uo["file_id"])
        c1.hyperlink = f"{uo['file_id']}.xlsx"
        c1.font = XD.fnt(9.5, color="0563C1", underline=True)
        c1.alignment = XD.left()
        c1.border = XD.HAIR
        for col in range(2, 7):
            c = ws.cell(row=start_row, column=col, value="")
            c.fill = XD.fill(XD.INPUT)
            c.border = XD.HAIR
            c.alignment = XD.center()
        ws.row_dimensions[start_row].height = 18
        start_row += 1

    if not uo_list:
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row, end_column=6)
        c = ws.cell(row=start_row, column=1, value="Aucune UO")
        c.font = XD.fnt(9, color=XD.GREY_D, italic=True)
        c.fill = XD.fill(XD.GREY_L)
        c.alignment = XD.center()
        start_row += 1

    return start_row


def _agenda_po(ws, start_row: int, uo_list: list[dict]) -> int:
    """Section Points ouverts / Actions. Retourne la prochaine ligne libre."""
    XD.section_box(ws, "⚡  Points ouverts / Actions", start_row, 1, start_row, 6, "oil")
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=6)
    start_row += 1

    headers = ["UO ID", "Description action", "Responsable", "Date limite", "Nb points", "Statut"]
    XD.table_header(ws, start_row, headers, "oil")
    start_row += 1

    for uo in uo_list:
        c1 = ws.cell(row=start_row, column=1, value=uo["file_id"])
        c1.hyperlink = f"{uo['file_id']}.xlsx"
        c1.font = XD.fnt(9.5, color="0563C1", underline=True)
        c1.alignment = XD.left()
        c1.border = XD.HAIR
        for col in range(2, 7):
            c = ws.cell(row=start_row, column=col, value="")
            c.fill = XD.fill(XD.INPUT)
            c.border = XD.HAIR
            c.alignment = XD.center()
        ws.row_dimensions[start_row].height = 18
        start_row += 1

    return start_row


def _sheet_agenda(wb: Workbook, se_name: str, uo_list: list[dict]):
    """Onglet Agenda — vue semaine + 30j + points ouverts. Lignes jaunes à remplir par le SE."""
    ws = wb.create_sheet("Agenda")
    ws.sheet_view.showGridLines = False
    XD.banner(ws, "planning", f"Agenda — {se_name}",
              subtitle=date.today().strftime('%d/%m/%Y'), n_cols=6)

    row = 3
    row = _agenda_section(ws, "📅  Semaine en cours", row, uo_list)
    row += 1
    row = _agenda_section(ws, "📋  Prochaines échéances (30 jours)", row, uo_list)
    row += 1
    _agenda_po(ws, row, uo_list)

    for col, w in zip("ABCDEF", [14, 38, 12, 16, 18, 30]):
        ws.column_dimensions[col].width = w


# ─── Génération du cockpit ────────────────────────────────────────────────────

def generer_cockpit(se_name: str, uo_list: list[dict],
                    pilote_id: str, output_dir: Path) -> Path:
    """Génère un fichier cockpit Excel pour un ingénieur."""
    wb = Workbook()
    wb.remove(wb.active)

    _sheet_mes_uos(wb, se_name, uo_list)
    _sheet_agenda(wb, se_name, uo_list)
    _sheet_manifeste(wb, se_name, uo_list, pilote_id)

    safe = se_name.replace(" ", "_")
    out = output_dir / f"Cockpit_{safe}.xlsx"
    wb.save(str(out))
    return out


def _sheet_mes_uos(wb: Workbook, se_name: str, uo_list: list[dict]):
    ws = wb.create_sheet("Mes UOs")
    ws.sheet_view.showGridLines = False

    XD.banner(ws, "general",
              f"Cockpit Ingénieur — {se_name}",
              subtitle=date.today().strftime('%d/%m/%Y'), n_cols=8)

    headers = ["UO ID", "Système", "Projet", "Charge (h)",
               "% Avancement", "H réalisées", "Date fin", "Alerte"]
    row_h = 3
    XD.table_header(ws, row_h, headers, "general")

    for i, uo in enumerate(uo_list):
        row = row_h + 1 + i
        XD.data_row(ws, row, i, 1, 8, "general")
        ws.cell(row=row, column=1).alignment = XD.left()
        for col in range(2, 9):
            ws.cell(row=row, column=col).alignment = XD.center()
        c1 = ws.cell(row=row, column=1, value=uo["file_id"])
        c1.hyperlink = f"{uo['file_id']}.xlsx"
        c1.font = XD.fnt(9.5, color="0563C1", underline=True)
        ws.cell(row=row, column=2, value=uo["systeme"])
        ws.cell(row=row, column=3, value=uo["projet"])
        ws.cell(row=row, column=4, value=uo["heures"])
        for col in (5, 6):
            c = ws.cell(row=row, column=col, value=0)
            c.fill = XD.fill(XD.INPUT)
            c.border = XD.HAIR
            c.alignment = XD.center()
        ws.cell(row=row, column=5).number_format = "0%"

    last_row = row_h + len(uo_list)
    tbl_ref = f"A{row_h}:H{last_row}" if uo_list else f"A{row_h}:H{row_h}"
    XD.named_table(ws, "tbl_mes_uos", tbl_ref, "general")

    for col, w in zip("ABCDEFGH", [20, 18, 22, 12, 16, 14, 14, 22]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = f"A{row_h + 1}"


def _sheet_manifeste(wb: Workbook, se_name: str,
                     uo_list: list[dict], pilote_id: str):
    ws = wb.create_sheet("_Manifeste")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 65
    ws.column_dimensions["C"].width = 60

    def w(row, instr, comment=""):
        c = ws.cell(row=row, column=1, value=instr)
        bold = any(instr.startswith(k) for k in ("DEF ", "PUSH ", "PULL ", "LIST "))
        c.font = XD.fnt(9.5, bold=bold, color="0C447C" if bold else "2C2C2A")
        c.alignment = XD.left()
        if comment:
            ws.cell(row=row, column=3, value=comment).font = XD.fnt(9, color="5F5E5A", italic=True)

    safe = se_name.replace(" ", "_")

    ws["A1"] = "MANIFESTE_V=1"
    ws["A1"].font = XD.fnt(10, bold=True, color="0C447C")
    ws.sheet_properties.tabColor = XD.sheet("manifeste").header

    r = 3
    w(r, "FILE_TYPE: cockpit_ingenieur",  "Type de fichier ExoSync"); r += 1
    w(r, f"FILE_ID: Cockpit_{safe}",      "Identifiant unique du cockpit"); r += 1
    w(r, f"ingenieur: {se_name}",         "Nom de l'ingénieur propriétaire"); r += 1
    if pilote_id:
        w(r, f"pilote_id: {pilote_id}",
          "Pilote responsable — permet au dashboard de découvrir ce cockpit"); r += 1

    r += 1  # séparateur
    w(r, "# ── Lecture de la table des UOs ────────────────────────────"); r += 1
    w(r, "DEF $mes_uos = GET_TABLE(Mes UOs, tbl_mes_uos)",
      "Lit la table des UOs avec les saisies ingénieur (zones jaunes)"); r += 1

    r += 1
    w(r, "# ── Publication vers le store central ──────────────────────"); r += 1
    w(r, f"PUSH $mes_uos -> cockpit.{safe}.mes_uos",
      "Remonte la table UOs (avec avancements) vers le store central ExoSync")


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _cockpit_has_saisies(xlsx_path: Path) -> bool:
    """Retourne True si une cellule des colonnes 5 ou 6 de tbl_mes_uos est non nulle."""
    try:
        wb = load_workbook(str(xlsx_path), data_only=True, read_only=True)
        if "Mes UOs" not in wb.sheetnames:
            return False
        ws = wb["Mes UOs"]
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row,
                                min_col=5, max_col=6, values_only=True):
            if any(v is not None and v != 0 for v in row):
                return True
        return False
    except Exception:
        return False


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description="Génère des cockpits ingénieur depuis les UO du répertoire")
    p.add_argument("--pilote", default="USR004",
                   help="Pilote ID à inscrire dans le manifeste (défaut: USR004)")
    p.add_argument("--ingenieur", default="",
                   help="Filtrer un seul ingénieur (optionnel)")
    p.add_argument("--output", default=str(HERE),
                   help="Répertoire de sortie (défaut: projet_TrainSystem)")
    p.add_argument("--safe", action="store_true",
                   help="Ne régénère que les cockpits sans saisies (cols 5-6 nulles)")
    args = p.parse_args()

    output_dir = Path(args.output)
    uo_list = lire_uo_du_repertoire(HERE)

    if not uo_list:
        sys.exit("[ERR] Aucune UO trouvée dans le répertoire.")

    # Grouper par ingénieur
    par_ingenieur: dict[str, list] = {}
    for uo in uo_list:
        se = uo["se_name"]
        if args.ingenieur and se != args.ingenieur:
            continue
        par_ingenieur.setdefault(se, []).append(uo)

    if not par_ingenieur:
        sys.exit(f"[ERR] Aucun ingénieur trouvé{' pour ' + args.ingenieur if args.ingenieur else ''}.")

    print(f"UOs trouvées : {len(uo_list)}")
    for se_name, uos in sorted(par_ingenieur.items()):
        if args.safe:
            out_path = output_dir / f"Cockpit_{se_name.replace(' ', '_')}.xlsx"
            if out_path.exists() and _cockpit_has_saisies(out_path):
                print(f"  [SKIP --safe] {out_path.name} (saisies détectées)")
                continue
        out = generer_cockpit(se_name, uos, args.pilote, output_dir)
        print(f"  [OK] {out.name}  ({len(uos)} UO, pilote_id={args.pilote})")


if __name__ == "__main__":
    main()
