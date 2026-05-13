"""Génère un fichier Excel squelette pour un Post du registre.

Crée toutes les feuilles définies dans la Classe (required + optional),
avec les headers de colonnes issus des std_tables, et une feuille _Manifeste
pré-remplie avec les DEF/PUSH de la Classe.
"""
import io
from pathlib import Path
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from web.services.config_service import load_registre, load_file_types, load_tables

router = APIRouter()

# ── Styles ────────────────────────────────────────────────────────────────────

_BLUE_DARK  = "1E3A5F"
_BLUE_MID   = "2563EB"
_BLUE_LIGHT = "DBEAFE"
_GREY_LIGHT = "F1F5F9"
_WHITE      = "FFFFFF"

def _hfont(bold=True, size=11, color="FFFFFF"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def _bfont(bold=False, size=10, color="1E293B"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def _fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def _border():
    s = Side(border_style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_row(ws, row, headers, col_widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font   = _hfont()
        c.fill   = _fill(_BLUE_MID)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
    ws.row_dimensions[row].height = 20
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


# ── Générateur MXL embarqué ───────────────────────────────────────────────────

def _build_manifeste_lines(file_id: str, file_rec: dict, ft: dict) -> list[str]:
    lines = [
        f"FILE_TYPE  {file_rec['type_fichier']}",
        f"FILE_ID    {file_id}",
        f"VERSION    1",
        "",
    ]

    push_prefix = ft.get("push_prefix", "")

    # DEF scalaires (min_fields) → DEF SCALAIRES  PUSH
    min_fields = ft.get("min_fields", [])
    if min_fields:
        lines.append("DEF  SCALAIRES  SHEET=_Manifeste")
        for f in min_fields:
            col_line = f"  COL  {f['name']}  TYPE={f.get('field_type','string')}"
            if f.get("label"):
                col_line += f'  HEADER="{f["label"]}"'
            lines.append(col_line)
        lines.append("")

        for f in min_fields:
            if f.get("pushable"):
                key = push_prefix.rstrip(".").replace("{id}", file_id) + f".{f['name']}"
                lines.append(f"PUSH  {f['name']}  TO={key}")
        lines.append("")

    # Tables standard (std_tables depuis tables.json)
    tables_data = load_tables().get("tables", {})
    class_tables = [
        t for t in tables_data.values()
        if t.get("file_id") == f"__class__{file_rec['type_fichier']}"
    ]
    for tbl in class_tables:
        lines.append(f"DEF  {tbl['table_name']}  SHEET={tbl.get('sheet', tbl['table_name'])}")
        for col in tbl.get("columns", []):
            col_line = f"  COL  {col['name']}  TYPE={col.get('col_type','string')}"
            if col.get("header"):
                col_line += f'  HEADER="{col["header"]}"'
            if col.get("is_key"):
                col_line += "  KEY"
            lines.append(col_line)
        lines.append("")

    # PULL namespaces autorisés
    for ns in ft.get("allowed_namespaces", []):
        lines.append(f"# PULL suggéré : PULL  <TABLE>  FROM={ns}<clé>")
    if ft.get("allowed_namespaces"):
        lines.append("")

    return lines


# ── Feuille _Manifeste ────────────────────────────────────────────────────────

def _sheet_manifeste(wb: Workbook, file_id: str, file_rec: dict, ft: dict):
    ws = wb.create_sheet("_Manifeste")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 80

    # Titre
    ws["A1"] = f"_Manifeste — {file_id} ({file_rec['type_fichier']})"
    ws["A1"].font = _hfont(size=12)
    ws["A1"].fill = _fill(_BLUE_DARK)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 24

    lines = _build_manifeste_lines(file_id, file_rec, ft)
    for i, line in enumerate(lines, start=3):
        c = ws.cell(row=i, column=1, value=line)
        c.font = Font(name="Courier New", size=9.5,
                      color="4ADE80" if not line.startswith("#") else "6B7280",
                      bold=line and not line.startswith(" ") and not line.startswith("#"))
        c.fill = _fill("0F172A")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=0)

    ws.freeze_panes = "A3"


# ── Feuille de données générique ──────────────────────────────────────────────

def _sheet_data(wb: Workbook, sheet_name: str, columns: list[dict]):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    # Titre
    title_end = max(len(columns), 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=title_end)
    ws["A1"] = sheet_name
    ws["A1"].font = _hfont(size=11)
    ws["A1"].fill = _fill(_BLUE_DARK)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 22

    if not columns:
        ws.cell(row=3, column=1, value="(feuille vide — définir les colonnes dans Tables & colonnes)")
        ws.cell(row=3, column=1).font = Font(name="Calibri", size=9, color="94A3B8", italic=True)
        ws.column_dimensions["A"].width = 60
        return

    headers  = [c.get("header") or c.get("name", "") for c in columns]
    col_wids = [max(len(h) + 4, 14) for h in headers]
    _header_row(ws, 2, headers, col_wids)

    # 3 lignes vides formatées pour montrer la structure
    for r in range(3, 6):
        for i in range(1, len(columns) + 1):
            c = ws.cell(row=r, column=i, value="")
            c.fill  = _fill(_GREY_LIGHT if r % 2 == 1 else _WHITE)
            c.border = _border()
            c.font  = _bfont()

    ws.freeze_panes = f"A3"


# ── Feuille vide nommée ───────────────────────────────────────────────────────

def _sheet_empty(wb: Workbook, sheet_name: str):
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = sheet_name
    ws["A1"].font  = _hfont(size=11)
    ws["A1"].fill  = _fill(_BLUE_DARK)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.column_dimensions["A"].width = 50
    ws.row_dimensions[1].height = 22


# ── Endpoint ──────────────────────────────────────────────────────────────────

@router.get("/{file_id}")
def generate_xlsx(file_id: str):
    registre  = {f["id"]: f for f in load_registre()}
    if file_id not in registre:
        raise HTTPException(404, f"Post '{file_id}' non trouvé dans le registre")

    file_rec  = registre[file_id]
    ft_all    = load_file_types()
    ft        = ft_all.get(file_rec["type_fichier"])
    if not ft:
        raise HTTPException(422, f"Type de fichier '{file_rec['type_fichier']}' inconnu")

    # Colonnes par feuille depuis std_tables de la Classe
    tables_data  = load_tables().get("tables", {})
    cols_by_sheet: dict[str, list[dict]] = {}
    for t in tables_data.values():
        if t.get("file_id") == f"__class__{file_rec['type_fichier']}":
            sheet = t.get("sheet", t["table_name"])
            cols_by_sheet[sheet] = t.get("columns", [])

    wb = Workbook()
    wb.remove(wb.active)  # enlève le Sheet vide par défaut

    # _Manifeste en premier
    _sheet_manifeste(wb, file_id, file_rec, ft)

    # Feuilles required
    for sheet in ft.get("required_sheets", []):
        if sheet == "_Manifeste":
            continue
        _sheet_data(wb, sheet, cols_by_sheet.get(sheet, []))

    # Feuilles optional
    for sheet in ft.get("optional_sheets", []):
        if sheet in ("_Manifeste", "_Log"):
            _sheet_empty(wb, sheet)
        else:
            _sheet_data(wb, sheet, cols_by_sheet.get(sheet, []))

    # Sérialisation en mémoire
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_id  = file_id.replace("/", "-").replace("\\", "-")
    filename = f"{safe_id}_{file_rec['type_fichier']}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
