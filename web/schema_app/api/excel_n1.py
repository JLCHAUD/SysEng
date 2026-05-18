"""Import Excel → Classe (N1 uniquement).

Workflow :
  1. POST /api/excel/scan-upload  → upload, découverte tables + plages nommées
  2. POST /api/excel/preview-class → mapping proposé tables→TableStd + fields→FieldMin
  3. POST /api/excel/apply-class   → enrichit/crée la Classe dans file_types.yaml

La partie N2 (enregistrement dans registre.json) est dans l'app N2.
"""
import re
import tempfile
import shutil
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, HTTPException, UploadFile, File
from pydantic import BaseModel
import openpyxl

from web.schema_app.services.config_service import load_file_types, save_file_types

router = APIRouter()


# ── Modèles ────────────────────────────────────────────────────────────────────

class ColumnInfo(BaseModel):
    name: str
    suggested_type: str = "string"


class TableInfo(BaseModel):
    name: str
    sheet: str
    ref: str
    columns: list[ColumnInfo]
    row_count: int


class NamedRangeInfo(BaseModel):
    name: str
    sheet: str
    address: str
    kind: str              # "scalar" | "range"
    value_preview: str
    suggested_type: str


class ScanResult(BaseModel):
    filename: str
    sheets: list[str]
    tables: list[TableInfo]
    named_ranges: list[NamedRangeInfo]


class FieldMapping(BaseModel):
    named_range: str
    field_name: str
    field_type: str = "string"
    nature: str = "identitaire"
    source: str = "user_input"
    required: bool = True
    pushable: bool = False
    include: bool = True


class TableMapping(BaseModel):
    excel_name: str
    schema_name: str
    sheet: str
    columns: list[ColumnInfo] = []
    include: bool = True


class PreviewClassRequest(BaseModel):
    scan: ScanResult
    class_id: str = ""


class PreviewClassResponse(BaseModel):
    class_id: str
    suggested_fields: list[FieldMapping]
    suggested_tables: list[TableMapping]


class ApplyClassRequest(BaseModel):
    class_id: str
    label: str = ""
    owner_function: str = ""
    field_mappings: list[FieldMapping] = []
    table_mappings: list[TableMapping] = []
    min_sheets: list[str] = []
    optional_sheets: list[str] = []
    create_if_missing: bool = True


# ── Utilitaires ────────────────────────────────────────────────────────────────

def _guess_type(value) -> str:
    if value is None:
        return "string"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, int):
        return "int"
    if isinstance(value, float):
        return "float"
    import datetime
    if isinstance(value, (datetime.date, datetime.datetime)):
        return "date"
    s = str(value).strip()
    try:
        float(s.replace(",", "."))
        return "float"
    except ValueError:
        pass
    return "string"


def _read_table_columns(ws, tbl) -> list[ColumnInfo]:
    cols = []
    try:
        from openpyxl.utils.cell import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
        data_row = min_row + 1
    except Exception:
        data_row = None

    for tc in tbl.tableColumns:
        suggested = "string"
        if data_row:
            try:
                cell = ws.cell(row=data_row, column=tc.id)
                suggested = _guess_type(cell.value)
            except Exception:
                pass
        cols.append(ColumnInfo(name=tc.name, suggested_type=suggested))
    return cols


def _count_rows(ws, tbl) -> int:
    try:
        from openpyxl.utils.cell import range_boundaries
        _, min_row, _, max_row = range_boundaries(tbl.ref)
        return max(0, max_row - min_row)
    except Exception:
        return 0


def _scan_workbook(wb, filename: str) -> ScanResult:
    sheets = wb.sheetnames

    # Tables Excel
    tables: list[TableInfo] = []
    for ws in wb.worksheets:
        if ws.title.startswith("_"):
            continue
        for tbl in ws.tables.values():
            cols = _read_table_columns(ws, tbl)
            rows = _count_rows(ws, tbl)
            tables.append(TableInfo(name=tbl.name, sheet=ws.title, ref=tbl.ref,
                                    columns=cols, row_count=rows))

    # Plages nommées
    named_ranges: list[NamedRangeInfo] = []
    for dn in wb.defined_names.values():
        name = dn.name
        if name.startswith("_") or name.startswith("xlnm"):
            continue
        try:
            destinations = list(dn.destinations) if hasattr(dn, "destinations") else []
        except Exception:
            continue
        if destinations:
            sheet_name, coord = destinations[0]
            addr = coord.replace("$", "")
            kind = "range" if ":" in addr else "scalar"
        else:
            continue

        value_preview = ""
        if kind == "scalar" and sheet_name in wb.sheetnames:
            try:
                val = wb[sheet_name][addr].value
                value_preview = str(val) if val is not None else ""
            except Exception:
                pass

        named_ranges.append(NamedRangeInfo(
            name=name, sheet=sheet_name, address=addr, kind=kind,
            value_preview=value_preview,
            suggested_type=_guess_type(value_preview) if kind == "scalar" else "string",
        ))

    wb.close()
    return ScanResult(filename=filename, sheets=sheets, tables=tables, named_ranges=named_ranges)


# ── Routes ─────────────────────────────────────────────────────────────────────

@router.post("/scan-upload", response_model=ScanResult)
async def scan_upload(file: UploadFile = File(...)):
    suffix = Path(file.filename).suffix
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = tmp.name
    try:
        wb = openpyxl.load_workbook(tmp_path, read_only=False, data_only=True)
        return _scan_workbook(wb, file.filename)
    except Exception as e:
        raise HTTPException(422, f"Impossible d'ouvrir le fichier : {e}")
    finally:
        Path(tmp_path).unlink(missing_ok=True)


@router.post("/preview-class", response_model=PreviewClassResponse)
def preview_class(body: PreviewClassRequest):
    """Propose un mapping automatique tables → TableStd et plages → FieldMin."""
    scan = body.scan

    # Suggestion champs scalaires depuis plages nommées (scalaires uniquement)
    suggested_fields = [
        FieldMapping(
            named_range=nr.name,
            field_name=re.sub(r"[^a-z0-9_]", "_", nr.name.lower()),
            field_type=nr.suggested_type,
            nature="identitaire",
            source="user_input",
            include=True,
        )
        for nr in scan.named_ranges
        if nr.kind == "scalar"
    ]

    # Suggestion tables (exclure _Manifeste et feuilles système)
    suggested_tables = [
        TableMapping(
            excel_name=t.name,
            schema_name="Tab" + t.name if not t.name.startswith("Tab") else t.name,
            sheet=t.sheet,
            columns=t.columns,
            include=True,
        )
        for t in scan.tables
    ]

    return PreviewClassResponse(
        class_id=body.class_id,
        suggested_fields=suggested_fields,
        suggested_tables=suggested_tables,
    )


@router.post("/apply-class")
def apply_class(body: ApplyClassRequest):
    """Enrichit ou crée la Classe dans file_types.yaml avec le mapping validé."""
    ft_all = load_file_types()

    if body.class_id not in ft_all:
        if not body.create_if_missing:
            raise HTTPException(404, f"Classe '{body.class_id}' introuvable")
        ft_all[body.class_id] = {
            "label": body.label or body.class_id,
            "owner_role": body.owner_function,
            "required_sheets": [],
            "min_sheets": [],
            "optional_sheets": [],
            "allowed_namespaces": [],
            "push_prefix": "",
            "template": "",
            "min_fields": [],
        }

    ft = ft_all[body.class_id]

    # Mise à jour des feuilles (les sélections explicites de l'utilisateur écrasent l'existant)
    if body.min_sheets:
        ft["min_sheets"] = body.min_sheets
        ft["required_sheets"] = body.min_sheets   # alias legacy
    if body.optional_sheets:
        ft["optional_sheets"] = body.optional_sheets

    # Mise à jour des champs scalaires (ajout seulement, pas écrasement)
    existing_field_names = {f["name"] for f in ft.get("min_fields", [])}
    added_fields = 0
    for fm in body.field_mappings:
        if not fm.include or fm.field_name in existing_field_names:
            continue
        ft.setdefault("min_fields", []).append({
            "name": fm.field_name,
            "label": fm.named_range,
            "field_type": fm.field_type,
            "nature": fm.nature,
            "source": fm.source,
            "required": fm.required,
            "pushable": fm.pushable,
            "description": "",
        })
        existing_field_names.add(fm.field_name)
        added_fields += 1

    ft_all[body.class_id] = ft
    save_file_types(ft_all)

    # Tables std → tables.json
    from web.schema_app.services.config_service import load_tables, save_tables
    from web.schema_app.api.classes import ColumnMin, TableStd, _save_std_tables

    included_tables = [t for t in body.table_mappings if t.include]
    if included_tables:
        std_tables = [
            TableStd(
                name=t.schema_name,
                sheet=t.sheet,
                columns=[ColumnMin(name=c.name, col_type=c.suggested_type) for c in t.columns],
            )
            for t in included_tables
        ]
        _save_std_tables(body.class_id, std_tables)

    return {
        "ok": True,
        "class_id": body.class_id,
        "fields_added": added_fields,
        "tables_added": len(included_tables),
        "message": f"Classe '{body.class_id}' mise à jour",
    }
