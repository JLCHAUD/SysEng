# src/generators/cockpit_ingenieur_generator.py
"""Cockpit ingénieur système — Vue Agenda + Mes UOs + _Manifeste."""
from datetime import date, timedelta
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule

from src.models import UOInstance
from src.xl_design import XD

OUTPUT_DIR = Path(__file__).parent.parent.parent / "output" / "cockpits"
UO_DIR = Path(__file__).parent.parent.parent / "output" / "UOs"


def generate_cockpit_ingenieur(
    engineer_name: str,
    all_uo_instances: List[UOInstance],
    output_dir: Path = OUTPUT_DIR,
    uo_dir: Path = UO_DIR,
    pilote_id: str = "",
) -> Path:
    """Génère le cockpit agenda pour un ingénieur système."""
    output_dir.mkdir(parents=True, exist_ok=True)
    uo_list = [u for u in all_uo_instances if u.engineer_name == engineer_name]

    wb = Workbook()
    wb.remove(wb.active)

    _sheet_mes_uos(wb, engineer_name, uo_list, uo_dir)
    _sheet_agenda(wb, engineer_name, uo_list)
    _sheet_manifeste(wb, engineer_name, uo_list, pilote_id=pilote_id)

    safe = engineer_name.replace(" ", "_")
    filepath = output_dir / f"Cockpit_{safe}.xlsx"
    wb.save(filepath)
    return filepath


def _sheet_mes_uos(wb: Workbook, engineer_name: str, uo_list: List[UOInstance], uo_dir: Path = UO_DIR):
    ws = wb.create_sheet("Mes UOs")
    ws.sheet_view.showGridLines = False

    XD.banner(ws, "general", f"Mes UOs — {engineer_name}",
              subtitle=date.today().strftime('%d/%m/%Y'), n_cols=9)

    ws.merge_cells("A2:C2")
    ws["A2"].value = f"UOs actives : {len(uo_list)}"
    ws["A2"].fill = XD.fill(XD.sheet("general").accent)
    ws["A2"].font = XD.fnt(11, bold=True)
    ws["A2"].alignment = XD.center()
    ws["A2"].border = XD.HAIR

    ws.merge_cells("D2:F2")
    ws["D2"].value = f"Charge totale : {sum(u.total_hours for u in uo_list)}h"
    ws["D2"].fill = XD.fill(XD.sheet("general").accent)
    ws["D2"].font = XD.fnt(11, bold=True)
    ws["D2"].alignment = XD.center()
    ws["D2"].border = XD.HAIR

    ws.merge_cells("A4:I4")
    sec = ws["A4"]
    sec.value = "Toutes mes UO"
    sec.fill = XD.fill(XD.sheet("general").header)
    sec.font = XD.fnt(11, bold=True, color=XD.WHITE)
    sec.alignment = XD.center()

    # A=UO ID  B=Type  C=Système  D=Projet  E=Charge(h)  F=% Avancement  G=H réalisées  H=Date fin  I=Alerte
    headers = ["UO ID", "Type UO", "Système", "Projet", "Charge (h)",
               "% Avancement", "H réalisées", "Date fin", "Alerte"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=h)
    XD.table_header(ws, 5, headers, "general")

    for i, uo in enumerate(uo_list):
        row = 6 + i
        type_name = uo.uo_type.name if uo.uo_type else uo.uo_type_id
        sys_name = uo.system.name if uo.system else uo.system_id
        proj_name = uo.project.name if uo.project else uo.project_id

        id_cell = ws.cell(row=row, column=1, value=uo.id)
        id_cell.font = XD.fnt(11, color="0563C1")
        id_cell.alignment = XD.left()
        id_cell.border = XD.HAIR
        uo_type_file = uo.uo_type.id if uo.uo_type else uo.uo_type_id
        uo_sys_file = uo.system.id if uo.system else uo.system_id
        uo_path = uo_dir / f"{uo.id}_{uo_type_file}_{uo_sys_file}.xlsx"
        id_cell.hyperlink = str(uo_path)

        ws.cell(row=row, column=2, value=type_name)
        ws.cell(row=row, column=3, value=sys_name)
        ws.cell(row=row, column=4, value=proj_name)
        ws.cell(row=row, column=5, value=uo.total_hours)

        avanc_cell = ws.cell(row=row, column=6, value=0.0)
        avanc_cell.number_format = "0%"
        avanc_cell.fill = XD.fill(XD.INPUT)
        avanc_cell.border = XD.HAIR
        avanc_cell.alignment = XD.center()

        h_cell = ws.cell(row=row, column=7, value=0.0)
        h_cell.fill = XD.fill(XD.INPUT)
        h_cell.border = XD.HAIR
        h_cell.alignment = XD.center()

        date_cell = ws.cell(row=row, column=8, value=uo.end_date)
        date_cell.number_format = "DD/MM/YYYY"
        date_cell.border = XD.HAIR
        date_cell.alignment = XD.center()

        alert_formula = (
            f'=IF(G{row}>E{row},"⚠ Dérive heures",'
            f'IF(AND(H{row}<>"",H{row}<TODAY()+7),"⏰ Échéance proche","✅ OK"))'
        )
        ws.cell(row=row, column=9, value=alert_formula)

        # style_data_row replacement for cols 2-5 and 8-9
        alternate = (i % 2 == 1)
        bg = XD.sheet("general").accent if alternate else XD.WHITE
        for c in range(2, 6):
            cell = ws.cell(row=row, column=c)
            cell.fill = XD.fill(bg)
            cell.font = XD.fnt(11)
            cell.alignment = XD.left()
            cell.border = XD.HAIR
        for c in range(8, 10):
            cell = ws.cell(row=row, column=c)
            cell.fill = XD.fill(bg)
            cell.font = XD.fnt(11)
            cell.alignment = XD.left()
            cell.border = XD.HAIR

    last_row = 5 + len(uo_list)
    if uo_list:
        alert_range = f"I6:I{last_row}"
        ws.conditional_formatting.add(
            alert_range,
            CellIsRule(operator="equal", formula=['"⚠ Dérive heures"'],
                       fill=XD.fill(XD.RED_L)),
        )
        ws.conditional_formatting.add(
            alert_range,
            CellIsRule(operator="equal", formula=['"⏰ Échéance proche"'],
                       fill=XD.fill(XD.AMBER_L)),
        )
        ws.conditional_formatting.add(
            alert_range,
            CellIsRule(operator="equal", formula=['"✅ OK"'],
                       fill=XD.fill(XD.GREEN_L)),
        )

    # Table nommée pour GET_TABLE(Mes UOs, tbl_mes_uos)
    if uo_list:
        tbl_ref = f"A5:I{last_row}"
        XD.named_table(ws, "tbl_mes_uos", tbl_ref, "general")

    for col, w in {"A": 12, "B": 30, "C": 18, "D": 20, "E": 13,
                   "F": 16, "G": 14, "H": 14, "I": 22}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"


def _sheet_agenda(wb: Workbook, engineer_name: str, uo_list: List[UOInstance]):
    ws = wb.create_sheet("Agenda")
    ws.sheet_view.showGridLines = False
    today = date.today()

    # ── Titre ──────────────────────────────────────────────────────────────────
    XD.banner(ws, "planning",
              f"Agenda — {engineer_name}   |   Semaine du {today.strftime('%d/%m/%Y')}",
              n_cols=6)

    current_row = 3

    # ── Section : Semaine en cours ────────────────────────────────────────────
    current_row = _agenda_section(
        ws, "📅  Semaine en cours", current_row,
        uo_list, today, today + timedelta(days=7),
        color=XD.sheet("general").header,
    )

    current_row += 1

    # ── Section : Prochaines échéances (8-30j) ────────────────────────────────
    current_row = _agenda_section(
        ws, "📋  Prochaines échéances (30 jours)", current_row,
        uo_list, today + timedelta(days=8), today + timedelta(days=30),
        color=XD.sheet("general").accent, header_text_color="1F3864",
    )

    current_row += 1

    # ── Section : Points ouverts ──────────────────────────────────────────────
    current_row = _agenda_points_ouverts(ws, current_row, uo_list)

    for col, w in {"A": 12, "B": 35, "C": 12, "D": 16, "E": 18, "F": 30}.items():
        ws.column_dimensions[col].width = w


def _agenda_section(
    ws, title: str, start_row: int,
    uo_list: List[UOInstance], date_from: date, date_to: date,
    color: str = None, header_text_color: str = "FFFFFF",
) -> int:
    """Affiche une section agenda filtrée par horizon de dates. Retourne la prochaine ligne libre."""
    if color is None:
        color = XD.sheet("general").header
    ws.merge_cells(f"A{start_row}:F{start_row}")
    sec = ws[f"A{start_row}"]
    sec.value = title
    sec.fill = XD.fill(color)
    sec.font = XD.fnt(11, bold=True, color=header_text_color)
    sec.alignment = XD.left()
    start_row += 1

    headers = ["UO ID", "Activité", "Priorité", "Date échéance", "Statut", "Action"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=h)
    XD.table_header(ws, start_row, headers, "general")
    for col in range(1, 7):
        ws.cell(row=start_row, column=col).font = XD.fnt(11, bold=True, color="1F3864")
    start_row += 1

    row = start_row
    for uo in uo_list:
        activities = uo.uo_type.activities if uo.uo_type else []
        for idx, act in enumerate(activities):
            act_end = getattr(act, 'end_date', None) or uo.end_date
            if not act_end or not (date_from <= act_end <= date_to):
                continue

            ws.cell(row=row, column=1, value=uo.id)
            ws.cell(row=row, column=2, value=act.name)

            priorite = "🔴 Haute" if act_end <= date.today() + timedelta(days=3) else "🟡 Normale"
            ws.cell(row=row, column=3, value=priorite)

            date_cell = ws.cell(row=row, column=4, value=act_end)
            date_cell.number_format = "DD/MM/YYYY"

            act_statut = getattr(act, 'statut', None)
            statut_val = act_statut.value if hasattr(act_statut, "value") else str(act_statut) if act_statut else ""
            ws.cell(row=row, column=5, value=statut_val)
            ws.cell(row=row, column=6, value="")

            # style_data_row replacement for cols 1-6
            alternate = (row % 2 == 0)
            bg = XD.sheet("general").accent if alternate else XD.WHITE
            for c in range(1, 7):
                cell = ws.cell(row=row, column=c)
                cell.fill = XD.fill(bg)
                cell.font = XD.fnt(11)
                cell.alignment = XD.left()
                cell.border = XD.HAIR
            row += 1

    if row == start_row:
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"].value = "Aucune activité dans cette période"
        ws[f"A{row}"].fill = XD.fill("F9F9F9")
        ws[f"A{row}"].font = XD.fnt(11, color="999999")
        ws[f"A{row}"].alignment = XD.center()
        row += 1

    return row


def _agenda_points_ouverts(ws, start_row: int, uo_list: List[UOInstance]) -> int:
    """Section Points ouverts en bas de l'Agenda."""
    ws.merge_cells(f"A{start_row}:F{start_row}")
    sec = ws[f"A{start_row}"]
    sec.value = "⚡  Points ouverts / Actions"
    sec.fill = XD.fill(XD.AMBER_L)
    sec.font = XD.fnt(11, bold=True, color="C00000")
    sec.alignment = XD.left()
    start_row += 1

    headers = ["UO ID", "Description action", "Responsable", "Date limite", "Nb points", "Statut"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=h)
    XD.table_header(ws, start_row, headers, "general")
    for col in range(1, 7):
        ws.cell(row=start_row, column=col).font = XD.fnt(11, bold=True, color="1F3864")
    start_row += 1

    for uo in uo_list:
        ws.cell(row=start_row, column=1, value=uo.id)
        for col in range(2, 7):
            c = ws.cell(row=start_row, column=col, value="")
            c.fill = XD.fill(XD.INPUT)
            c.border = XD.HAIR
        ws.cell(row=start_row, column=1).border = XD.HAIR
        ws.cell(row=start_row, column=1).alignment = XD.left()
        start_row += 1

    return start_row


def _sheet_manifeste(wb: Workbook, engineer_name: str, uo_list: List[UOInstance],
                     pilote_id: str = ""):
    """Génère l'onglet _Manifeste au format MXL mono-colonne.
    Col A = instruction MXL, col B = ancre, col C = commentaire français.
    """
    ws = wb.create_sheet("_Manifeste")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = XD.sheet("manifeste").header

    safe_name = engineer_name.replace(" ", "_")

    def _mxl_row(row: int, instr: str, ancre: str = "", comment: str = ""):
        cell_a = ws.cell(row=row, column=1, value=instr)
        cell_a.font = XD.fnt(9.5, bold=instr.startswith(("DEF ", "PUSH ", "PULL ")))
        cell_a.alignment = XD.left()
        if ancre:
            ws.cell(row=row, column=2, value=ancre).font = XD.fnt(9, color="888888")
        if comment:
            c = ws.cell(row=row, column=3, value=comment)
            c.font = XD.fnt(9, color="666666")
            c.fill = XD.fill("F9F9F9")

    # Ligne 1 : version
    ws["A1"] = "MANIFESTE_V=1"
    ws["A1"].font = XD.fnt(11, bold=True, color="1F3864")
    # Ligne 2 intentionnellement vide (skippée par parser.py)

    # Métadonnées
    r = 3
    _mxl_row(r, "FILE_TYPE: cockpit_ingenieur",  comment="Type de fichier ExoSync"); r += 1
    _mxl_row(r, f"FILE_ID: Cockpit_{safe_name}", comment="Identifiant unique du cockpit"); r += 1
    _mxl_row(r, f"ingenieur: {engineer_name}",   comment="Nom de l'ingénieur propriétaire"); r += 1
    if pilote_id:
        _mxl_row(r, f"pilote_id: {pilote_id}",
                 comment="Pilote responsable — permet au dashboard de découvrir ce cockpit")
        r += 1

    r += 1  # ligne vide de séparation
    # Définition de la table
    _mxl_row(r, "DEF $mes_uos = GET_TABLE(Mes UOs, tbl_mes_uos)",
             comment="Référence à la table des UOs de l'ingénieur"); r += 1
    _mxl_row(r, "COL $mes_uos.avancement : WRITE=engineer",
             comment="% avancement saisi par l'ingénieur (zone jaune)"); r += 1
    _mxl_row(r, "COL $mes_uos.heures_realisees : WRITE=engineer",
             comment="Heures réalisées saisies par l'ingénieur (zone jaune)"); r += 1

    r += 1  # ligne vide
    # Export vers store
    _mxl_row(r, f"PUSH $mes_uos -> cockpit.{safe_name}.mes_uos",
             comment="Remonte les saisies ingénieur vers le store central ExoSync")

    for col, w in {"A": 65, "B": 18, "C": 60}.items():
        ws.column_dimensions[col].width = w
