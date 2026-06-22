"""Tests TDD pour le module de design system xl_design."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.utils import get_column_letter

from src.xl_design import XD


class TestPrimitives:
    def test_font_family_segoe(self):
        f = XD.fnt(12, bold=True, color="FFFFFF")
        assert isinstance(f, Font)
        assert f.name == "Segoe UI"
        assert f.size == 12
        assert f.bold is True
        assert f.color.rgb.endswith("FFFFFF")

    def test_font_defaut(self):
        f = XD.fnt()
        assert f.size == 10
        assert f.bold is False
        assert f.color.rgb.endswith("2C2C2A")

    def test_fill_solide(self):
        fill = XD.fill("0C447C")
        assert isinstance(fill, PatternFill)
        assert fill.fgColor.rgb.endswith("0C447C")

    def test_input_jaune_constant(self):
        assert XD.INPUT == "FFF2CC"

    def test_hair_border(self):
        assert isinstance(XD.HAIR, Border)
        assert XD.HAIR.left.style == "thin"

    def test_alignements(self):
        assert XD.center().horizontal == "center"
        assert XD.center().wrap_text is True
        assert XD.left().horizontal == "left"


class TestRegistreOnglets:
    def test_onze_familles(self):
        assert len(XD.SHEETS) == 11

    def test_cles_attendues(self):
        attendues = {"general", "dashboard", "description", "planning",
                     "donnees_entree", "activites", "livrables", "oil",
                     "kpi", "orga", "manifeste"}
        assert set(XD.SHEETS) == attendues

    def test_triple_activites(self):
        s = XD.sheet("activites")
        assert s.banner == "084434"
        assert s.header == "0C5E49"
        assert s.accent == "E1F5EE"
        assert s.glyph == "✔"

    def test_triple_oil_rouge(self):
        assert XD.sheet("oil").header == "A32D2D"

    def test_cle_inconnue_leve(self):
        import pytest
        with pytest.raises(KeyError):
            XD.sheet("inexistant")

    def test_tab_colors_mappe_le_ton_moyen(self):
        tc = XD.tab_colors()
        assert tc["general"] == "0C447C"
        assert tc["kpi"] == "534AB7"


class TestBanner:
    def _ws(self):
        wb = Workbook()
        return wb.active

    def test_tab_color_pose_ton_moyen(self):
        ws = self._ws()
        XD.banner(ws, "activites", "UO L09U1 — Activités", n_cols=10)
        assert ws.sheet_properties.tabColor.rgb.endswith("0C5E49")

    def test_titre_avec_glyphe_en_a1(self):
        ws = self._ws()
        XD.banner(ws, "activites", "UO L09U1 — Activités", n_cols=10)
        assert "✔" in str(ws["A1"].value)
        assert "Activités" in str(ws["A1"].value)

    def test_fond_banniere_fonce(self):
        ws = self._ws()
        XD.banner(ws, "activites", "T", n_cols=10)
        assert ws["A1"].fill.fgColor.rgb.endswith("084434")

    def test_sous_titre_et_se_a_droite(self):
        ws = self._ws()
        XD.banner(ws, "activites", "T", subtitle="Clim", se="J. Dujardin", n_cols=10)
        valeurs = [ws.cell(row=1, column=c).value for c in range(1, 11)]
        joined = " ".join(str(v) for v in valeurs if v)
        assert "Clim" in joined
        assert "J. Dujardin" in joined

    def test_hauteur_ligne1(self):
        ws = self._ws()
        XD.banner(ws, "activites", "T", n_cols=10, height=30)
        assert ws.row_dimensions[1].height == 30


class TestTableHeaderEtDataRow:
    def _ws(self):
        return Workbook().active

    def test_header_au_ton_onglet(self):
        ws = self._ws()
        XD.table_header(ws, 5, ["id", "désignation", "statut"], "activites")
        assert ws.cell(row=5, column=1).fill.fgColor.rgb.endswith("0C5E49")
        assert ws.cell(row=5, column=1).font.color.rgb.endswith("FFFFFF")
        assert ws.cell(row=5, column=2).value == "désignation"

    def test_data_row_paire_blanche(self):
        ws = self._ws()
        XD.data_row(ws, 6, 0, 1, 3, "activites")
        assert ws.cell(row=6, column=1).fill.fgColor.rgb.endswith("FFFFFF")

    def test_data_row_impaire_accent(self):
        ws = self._ws()
        XD.data_row(ws, 7, 1, 1, 3, "activites")
        assert ws.cell(row=7, column=1).fill.fgColor.rgb.endswith("E1F5EE")


class TestNamedTable:
    def _ws_avec_donnees(self):
        ws = Workbook().active
        ws["A5"] = "id"
        ws["B5"] = "désignation"
        ws["C5"] = "statut"
        ws["A6"] = "ACT-1"
        ws["B6"] = "x"
        ws["C6"] = "A_FAIRE"
        return ws

    def test_table_nommee_creee(self):
        ws = self._ws_avec_donnees()
        XD.named_table(ws, "tbl_test", "A5:C6", "activites")
        assert "tbl_test" in ws.tables

    def test_style_clair(self):
        ws = self._ws_avec_donnees()
        XD.named_table(ws, "tbl_test", "A5:C6", "activites")
        assert ws.tables["tbl_test"].tableStyleInfo.name == "TableStyleLight15"

    def test_entete_colore_manuellement(self):
        ws = self._ws_avec_donnees()
        XD.named_table(ws, "tbl_test", "A5:C6", "activites")
        assert ws.cell(row=5, column=1).fill.fgColor.rgb.endswith("0C5E49")
        assert ws.cell(row=5, column=1).font.color.rgb.endswith("FFFFFF")


class TestBadgesConditionnels:
    def test_statut_cf_ajoute_des_regles(self):
        ws = Workbook().active
        XD.statut_cf(ws, "F6:F20")
        assert len(ws.conditional_formatting) >= 1

    def test_criticite_cf_ajoute_des_regles(self):
        ws = Workbook().active
        XD.criticite_cf(ws, "G6:G20")
        assert len(ws.conditional_formatting) >= 1


class TestTrafficEtCadres:
    def test_traffic_rouge_sous_50(self):
        ws = Workbook().active
        XD.traffic_light(ws, 6, 3, 0.3)
        assert ws.cell(row=6, column=3).fill.fgColor.rgb.endswith("FCEBEB")

    def test_traffic_ambre_entre_50_80(self):
        ws = Workbook().active
        XD.traffic_light(ws, 6, 3, 0.65)
        assert ws.cell(row=6, column=3).fill.fgColor.rgb.endswith("FAEEDA")

    def test_traffic_vert_au_dessus_80(self):
        ws = Workbook().active
        XD.traffic_light(ws, 6, 3, 0.9)
        assert ws.cell(row=6, column=3).fill.fgColor.rgb.endswith("EAF3DE")

    def test_card_border_pose_un_cadre(self):
        ws = Workbook().active
        XD.card_border(ws, 2, 2, 4, 4)
        assert ws.cell(row=2, column=2).border.top.style == "thin"

    def test_section_box_titre_et_fond(self):
        ws = Workbook().active
        XD.section_box(ws, "Titre section", 2, 2, 5, 4, "kpi")
        assert ws.cell(row=2, column=2).value == "Titre section"


class TestHealthSpine:
    def test_largeur_colonne_fine(self):
        ws = Workbook().active
        XD.health_spine(ws, "activites", header_row=5, row_start=6,
                        row_end=10, status_col=6, spine_col=1)
        assert ws.column_dimensions[get_column_letter(1)].width <= 3

    def test_entete_spine_au_ton_banniere(self):
        ws = Workbook().active
        XD.health_spine(ws, "activites", header_row=5, row_start=6,
                        row_end=10, status_col=6, spine_col=1)
        assert ws.cell(row=5, column=1).fill.fgColor.rgb.endswith("084434")

    def test_regles_conditionnelles_posees(self):
        ws = Workbook().active
        XD.health_spine(ws, "activites", header_row=5, row_start=6,
                        row_end=10, status_col=6, spine_col=1)
        assert len(ws.conditional_formatting) >= 1
