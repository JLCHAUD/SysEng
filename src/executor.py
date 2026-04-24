"""
ExoSync — executor.py
=====================
Exécute un PasserelleAST sur un fichier Excel en 4 phases séquentielles :

  Phase 1 — PULL   : store central → tables Excel
  Phase 2 — COMPUTE: tables Excel  → variables Python (scalaires + tables)
  Phase 3 — PUSH   : variables     → store central
  Phase 4 — BIND   : variables     → plages nommées Dashboard

Usage :
    from src.executor import execute_ast
    from src.parser   import parse_file
    from src          import store as Store

    ast    = parse_file(Path("UO-001.xlsx"))
    result = execute_ast(ast, Path("UO-001.xlsx"), Store)
    print(result.summary())
"""

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple

from src.parser import PasserelleAST, PullNode, DefNode, BindNode, PushNode


# ─── Résultat d'exécution ────────────────────────────────────────────────────

@dataclass
class ExecutionResult:
    pushed:   List[str] = field(default_factory=list)  # clés store mises à jour
    pulled:   List[str] = field(default_factory=list)  # tables Excel mises à jour
    bound:    List[str] = field(default_factory=list)  # plages nommées écrites
    errors:   List[str] = field(default_factory=list)  # erreurs (SEVERITY=error)
    warnings: List[str] = field(default_factory=list)  # avertissements (SEVERITY=warning)
    skipped:  List[str] = field(default_factory=list)  # instructions ignorées

    @property
    def has_blocking_errors(self) -> bool:
        """True si des erreurs de validation bloquantes ont été détectées."""
        return bool(self.errors)

    def summary(self) -> str:
        lines = [
            f"PULL    : {len(self.pulled)} table(s) mise(s) à jour",
            f"PUSH    : {len(self.pushed)} variable(s) publiée(s)",
            f"BIND    : {len(self.bound)} plage(s) écrite(s)",
            f"Skip    : {len(self.skipped)}",
            f"Warning : {len(self.warnings)}",
            f"Erreur  : {len(self.errors)}",
        ]
        if self.warnings:
            lines += [f"  ⚡ {w}" for w in self.warnings]
        if self.errors:
            lines += [f"  ⚠ {e}" for e in self.errors]
        return "\n".join(lines)


# ─── Palette STATUS → couleur ─────────────────────────────────────────────────

_STATUS_COLORS: Dict[str, Dict[str, str]] = {
    "ROUGE":       {"bg": "C00000", "fg": "FFFFFF"},
    "ORANGE":      {"bg": "FF8C00", "fg": "FFFFFF"},
    "JAUNE":       {"bg": "FFD700", "fg": "000000"},
    "VERT":        {"bg": "00B050", "fg": "FFFFFF"},
    "OK":          {"bg": "00B050", "fg": "FFFFFF"},
    "ATTENTION":   {"bg": "FF8C00", "fg": "FFFFFF"},
    "RETARD":      {"bg": "C00000", "fg": "FFFFFF"},
    "TERMINÉ":     {"bg": "375623", "fg": "FFFFFF"},
    "NON_DÉMARRÉ": {"bg": "D9D9D9", "fg": "595959"},
    "DÉPASSEMENT": {"bg": "7F0000", "fg": "FFFFFF"},
    "BLEU":        {"bg": "0070C0", "fg": "FFFFFF"},
    "GRIS":        {"bg": "808080", "fg": "FFFFFF"},
}


def _apply_status_format(cell, value: Any) -> None:
    """Applique le formatage couleur si la valeur est un STATUS connu."""
    colors = _STATUS_COLORS.get(str(value).upper().strip())
    if not colors:
        return
    cell.fill = PatternFill(
        start_color=colors["bg"],
        end_color=colors["bg"],
        fill_type="solid",
    )
    cell.font = Font(color=colors["fg"], bold=True)
    cell.alignment = Alignment(horizontal="center")


# ─── Helpers openpyxl ─────────────────────────────────────────────────────────

def _read_table_from_ws(ws, table_name: str) -> List[Dict[str, Any]]:
    """
    Lit un tableau Excel nommé (objet Table openpyxl) depuis une feuille.
    Retourne une liste de dicts {colonne: valeur}.
    """
    if table_name not in ws.tables:
        return []

    tbl = ws.tables[table_name]
    cells = list(ws[tbl.ref])
    if not cells:
        return []

    headers = [c.value for c in cells[0]]
    rows = []
    for raw_row in cells[1:]:
        values = [c.value for c in raw_row]
        if all(v is None for v in values):
            continue
        rows.append({
            headers[i]: values[i]
            for i in range(len(headers))
            if headers[i] is not None
        })
    return rows


def _table_ref_bounds(tbl) -> Tuple[int, int, int, int]:
    """Retourne (min_col, min_row, max_col, max_row) depuis tbl.ref."""
    ref = tbl.ref  # ex: "A1:H30"
    start, end = ref.split(":")
    min_col_l = re.sub(r"\d", "", start).replace("$", "")
    min_row   = int(re.sub(r"\D", "", start))
    max_col_l = re.sub(r"\d", "", end).replace("$", "")
    max_row   = int(re.sub(r"\D", "", end))
    return (
        column_index_from_string(min_col_l),
        min_row,
        column_index_from_string(max_col_l),
        max_row,
    )


def _write_cell_safe(ws, row: int, col: int, value: Any) -> None:
    """
    Écrit une valeur dans une cellule, en gérant les cellules fusionnées
    (écrit dans la cellule maître de la plage).
    """
    for merged in ws.merged_cells.ranges:
        if (merged.min_row <= row <= merged.max_row and
                merged.min_col <= col <= merged.max_col):
            ws.cell(merged.min_row, merged.min_col).value = value
            return
    ws.cell(row=row, column=col).value = value


def _resolve_named_range(wb, target_sheet: str, range_name: str
                          ) -> Optional[Tuple[Any, int, int]]:
    """
    Recherche une plage nommée dans le workbook.
    Retourne (worksheet, row, col) ou None.
    Cherche d'abord dans wb.defined_names, puis par correspondance partielle.
    """
    defn = wb.defined_names.get(range_name)
    if defn is not None:
        for title, coord in defn.destinations:
            ws = wb[title] if title in wb.sheetnames else None
            if ws is None:
                continue
            clean = coord.replace("$", "")
            # coord peut être une cellule ("F3") ou une plage ("F3:F3")
            if ":" in clean:
                clean = clean.split(":")[0]
            row, col = coordinate_to_tuple(clean)
            return ws, row, col

    # Fallback : feuille cible explicite
    if target_sheet in wb.sheetnames:
        ws = wb[target_sheet]
        for defn in wb.defined_names.values():
            if defn.name == range_name:
                for title, coord in defn.destinations:
                    if title == target_sheet:
                        clean = coord.replace("$", "").split(":")[0]
                        row, col = coordinate_to_tuple(clean)
                        return ws, row, col
    return None


# ─── Phase 1 — PULL ───────────────────────────────────────────────────────────

def _overwrite_table(ws, table_name: str, data: List[Dict]) -> None:
    """Écrase complètement une table Excel avec les données du store."""
    if not data or table_name not in ws.tables:
        return

    tbl = ws.tables[table_name]
    min_col, min_row, max_col, max_row = _table_ref_bounds(tbl)
    headers = list(data[0].keys())

    # Effacer le contenu existant
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).value = None

    # Écrire les en-têtes
    for ci, h in enumerate(headers, min_col):
        ws.cell(min_row, ci).value = h

    # Écrire les données
    for ri, row_data in enumerate(data, min_row + 1):
        for ci, h in enumerate(headers, min_col):
            ws.cell(ri, ci).value = row_data.get(h)

    # Mettre à jour la ref de la table
    new_max_row = min_row + len(data)
    new_max_col_l = get_column_letter(min_col + len(headers) - 1)
    tbl.ref = f"{get_column_letter(min_col)}{min_row}:{new_max_col_l}{new_max_row}"


def _append_new_rows(ws, table_name: str, data: List[Dict], key: str) -> int:
    """Ajoute les lignes dont la clé n'existe pas encore dans la table. Retourne le nb ajouté."""
    if not data or table_name not in ws.tables:
        return 0

    existing = _read_table_from_ws(ws, table_name)
    existing_keys = {str(row.get(key)) for row in existing}
    new_rows = [r for r in data if str(r.get(key, "")) not in existing_keys]

    if not new_rows:
        return 0

    tbl = ws.tables[table_name]
    min_col, min_row, max_col, max_row = _table_ref_bounds(tbl)
    headers = [ws.cell(min_row, c).value for c in range(min_col, max_col + 1)]

    insert_row = max_row + 1
    for row_data in new_rows:
        for ci, h in enumerate(headers, min_col):
            ws.cell(insert_row, ci).value = row_data.get(h)
        insert_row += 1

    # Étendre la table
    new_max_row = max_row + len(new_rows)
    tbl.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"
    return len(new_rows)


def _update_table(ws, table_name: str, data: List[Dict], key: str) -> int:
    """Met à jour les lignes existantes par clé. Retourne le nb de lignes modifiées."""
    if not data or table_name not in ws.tables:
        return 0

    tbl = ws.tables[table_name]
    min_col, min_row, max_col, max_row = _table_ref_bounds(tbl)
    headers = [ws.cell(min_row, c).value for c in range(min_col, max_col + 1)]

    if key not in headers:
        return 0

    key_col = min_col + headers.index(key)
    store_index = {str(r.get(key)): r for r in data}
    updated = 0

    for ri in range(min_row + 1, max_row + 1):
        cell_key = str(ws.cell(ri, key_col).value)
        if cell_key in store_index:
            row_data = store_index[cell_key]
            for ci, h in enumerate(headers, min_col):
                ws.cell(ri, ci).value = row_data.get(h, ws.cell(ri, ci).value)
            updated += 1

    return updated


def execute_pulls(ast: PasserelleAST, wb, store,
                  result: ExecutionResult) -> None:
    """Phase 1 — Copie les données du store vers les tables Excel."""
    for pull in ast.pulls:
        data = store.get(pull.global_name)
        if data is None:
            result.skipped.append(f"PULL {pull.global_name} — absent du store")
            continue

        if pull.sheet not in wb.sheetnames:
            result.errors.append(
                f"PULL {pull.global_name} — feuille '{pull.sheet}' introuvable"
            )
            continue

        ws = wb[pull.sheet]

        if pull.mode == "OVERWRITE":
            if isinstance(data, list):
                _overwrite_table(ws, pull.table, data)
                result.pulled.append(pull.table)
            else:
                result.errors.append(
                    f"PULL OVERWRITE {pull.global_name} — données non tabulaires"
                )

        elif pull.mode == "APPEND_NEW":
            if isinstance(data, list):
                n = _append_new_rows(ws, pull.table, data, pull.key)
                result.pulled.append(f"{pull.table} (+{n} lignes)")
            else:
                result.errors.append(
                    f"PULL APPEND_NEW {pull.global_name} — données non tabulaires"
                )

        elif pull.mode == "UPDATE":
            if isinstance(data, list):
                n = _update_table(ws, pull.table, data, pull.key)
                result.pulled.append(f"{pull.table} ({n} mises à jour)")
            else:
                result.errors.append(
                    f"PULL UPDATE {pull.global_name} — données non tabulaires"
                )

        elif pull.mode == "READ_ONLY":
            result.skipped.append(f"PULL {pull.global_name} — mode READ_ONLY")


# ─── Phase 2 — COMPUTE ────────────────────────────────────────────────────────

def _resolve_col(col_ref: str, ctx: Dict[str, Any]) -> List[Any]:
    """
    Résout une référence de colonne '$table.col' dans le contexte d'exécution.
    Retourne la liste de valeurs de la colonne (None filtrés par défaut).
    """
    col_ref = col_ref.strip()
    if "." not in col_ref:
        raise ValueError(f"Référence de colonne invalide : '{col_ref}'")
    var, col = col_ref.split(".", 1)
    table = ctx.get(var)
    if table is None:
        raise ValueError(f"Variable '{var}' non définie")
    if not isinstance(table, list):
        raise ValueError(f"'{var}' n'est pas une table")
    return [row.get(col) for row in table]


_MISSING = object()  # sentinelle pour distinguer "absent" de "None"


def _resolve_scalar(ref: str, ctx: Dict[str, Any]) -> Any:
    """Résout un scalaire : variable $x ou littéral numérique."""
    ref = ref.strip()
    if ref.startswith("$"):
        val = ctx.get(ref, _MISSING)
        if val is _MISSING:
            raise ValueError(f"Variable '{ref}' non définie")
        return val  # peut légitimement être None
    try:
        return float(ref) if "." in ref else int(ref)
    except ValueError:
        return ref.strip('"').strip("'")


def _eval_filter(args_str: str, ctx: Dict[str, Any]) -> List[Dict]:
    """
    FILTER($table, col OP "valeur")
    Opérateurs supportés : =, !=, >, >=, <, <=
    Connecteurs supportés : AND, OR
    """
    # Séparer la table de la condition (première virgule hors parenthèses)
    depth = 0
    split_at = -1
    for i, ch in enumerate(args_str):
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
        elif ch == "," and depth == 0:
            split_at = i
            break

    if split_at == -1:
        raise ValueError(f"FILTER : séparateur introuvable dans '{args_str}'")

    table_ref = args_str[:split_at].strip()
    condition  = args_str[split_at + 1:].strip()

    table = ctx.get(table_ref)
    if table is None:
        raise ValueError(f"FILTER : table '{table_ref}' non définie")
    if not isinstance(table, list):
        raise ValueError(f"FILTER : '{table_ref}' n'est pas une table")

    def _matches(row: Dict, cond: str) -> bool:
        # Gérer AND / OR
        if " AND " in cond.upper():
            parts = re.split(r"\bAND\b", cond, flags=re.IGNORECASE)
            return all(_matches(row, p.strip()) for p in parts)
        if " OR " in cond.upper():
            parts = re.split(r"\bOR\b", cond, flags=re.IGNORECASE)
            return any(_matches(row, p.strip()) for p in parts)

        # Condition atomique : col OP valeur
        m = re.match(
            r'^([\w_]+)\s*(!=|>=|<=|=|>|<)\s*(.+)$', cond.strip()
        )
        if not m:
            return True  # condition non parseable → ne filtre pas

        col, op, raw_val = m.group(1), m.group(2), m.group(3).strip()

        # Valeur : chaîne entre guillemets ou numérique
        if raw_val.startswith('"') or raw_val.startswith("'"):
            val = raw_val.strip('"').strip("'")
        else:
            try:
                val = float(raw_val) if "." in raw_val else int(raw_val)
            except ValueError:
                val = raw_val

        cell = row.get(col)
        # Comparer (forcer str si l'un est str et l'autre non)
        if isinstance(val, str) and cell is not None and not isinstance(cell, str):
            cell = str(cell)
        if isinstance(cell, str) and not isinstance(val, str):
            try:
                cell = type(val)(cell)
            except (ValueError, TypeError):
                cell = str(cell)
                val  = str(val)

        if op == "=":
            return cell == val
        if op == "!=":
            return cell != val
        if op == ">":
            return cell is not None and cell > val
        if op == ">=":
            return cell is not None and cell >= val
        if op == "<":
            return cell is not None and cell < val
        if op == "<=":
            return cell is not None and cell <= val
        return True

    return [row for row in table if _matches(row, condition)]


def _eval_formula(formula: str, ctx: Dict[str, Any]) -> Any:
    """
    Évalue une formule COMPUTE et retourne le résultat.
    formula = contenu intérieur de COMPUTE(...).
    """
    formula = formula.strip()

    # ─ FILTER ─────────────────────────────────────────────────────────────────
    if formula.upper().startswith("FILTER("):
        inner = formula[len("FILTER("):-1]
        return _eval_filter(inner, ctx)

    # ─ MEAN_WEIGHTED ──────────────────────────────────────────────────────────
    if formula.upper().startswith("MEAN_WEIGHTED("):
        inner = formula[len("MEAN_WEIGHTED("):-1]
        val_ref, wgt_ref = [s.strip() for s in inner.split(",", 1)]
        values  = [v for v in _resolve_col(val_ref, ctx) if v is not None]
        weights = [w for w in _resolve_col(wgt_ref, ctx) if w is not None]
        pairs   = [(v, w) for v, w in zip(
            _resolve_col(val_ref, ctx),
            _resolve_col(wgt_ref, ctx)
        ) if v is not None and w is not None]
        total_w = sum(w for _, w in pairs)
        if total_w == 0:
            return 0.0
        return sum(v * w for v, w in pairs) / total_w

    # ─ SUM ────────────────────────────────────────────────────────────────────
    if formula.upper().startswith("SUM("):
        inner = formula[len("SUM("):-1].strip()
        return sum(v for v in _resolve_col(inner, ctx) if v is not None)

    # ─ COUNT ──────────────────────────────────────────────────────────────────
    if formula.upper().startswith("COUNT("):
        inner = formula[len("COUNT("):-1].strip()
        return sum(1 for v in _resolve_col(inner, ctx) if v is not None)

    # ─ COUNT_IF ───────────────────────────────────────────────────────────────
    if formula.upper().startswith("COUNT_IF("):
        inner = formula[len("COUNT_IF("):-1]
        col_ref, raw_val = [s.strip() for s in inner.split(",", 1)]
        target = raw_val.strip('"').strip("'")
        return sum(1 for v in _resolve_col(col_ref, ctx) if str(v) == target)

    # ─ AVG ────────────────────────────────────────────────────────────────────
    if formula.upper().startswith("AVG("):
        inner = formula[len("AVG("):-1].strip()
        vals = [v for v in _resolve_col(inner, ctx) if v is not None]
        return sum(vals) / len(vals) if vals else 0.0

    # ─ MIN ────────────────────────────────────────────────────────────────────
    if formula.upper().startswith("MIN("):
        inner = formula[len("MIN("):-1].strip()
        vals = [v for v in _resolve_col(inner, ctx) if v is not None]
        return min(vals) if vals else None

    # ─ MAX ────────────────────────────────────────────────────────────────────
    if formula.upper().startswith("MAX("):
        inner = formula[len("MAX("):-1].strip()
        vals = [v for v in _resolve_col(inner, ctx) if v is not None]
        return max(vals) if vals else None

    # ─ DIV ────────────────────────────────────────────────────────────────────
    if formula.upper().startswith("DIV("):
        inner = formula[len("DIV("):-1]
        a_ref, b_ref = [s.strip() for s in inner.split(",", 1)]
        a = _resolve_scalar(a_ref, ctx)
        b = _resolve_scalar(b_ref, ctx)
        if b == 0 or b is None:
            return 0.0
        return a / b

    # ─ TRAFFIC_LIGHT ──────────────────────────────────────────────────────────
    if formula.upper().startswith("TRAFFIC_LIGHT("):
        inner = formula[len("TRAFFIC_LIGHT("):-1]
        # Ex: $avancement_global, warn=30, ok=70
        parts = [s.strip() for s in inner.split(",")]
        val = _resolve_scalar(parts[0], ctx)
        attrs: Dict[str, float] = {}
        for p in parts[1:]:
            m = re.match(r'(\w+)\s*=\s*([\d.]+)', p.strip())
            if m:
                attrs[m.group(1).lower()] = float(m.group(2))
        warn = attrs.get("warn", 30.0)
        ok   = attrs.get("ok",   70.0)
        if val is None:
            return "ROUGE"
        if val < warn:
            return "ROUGE"
        if val < ok:
            return "ORANGE"
        return "VERT"

    # ─ SWITCH_RANGE ───────────────────────────────────────────────────────────
    if formula.upper().startswith("SWITCH_RANGE("):
        # SWITCH_RANGE($val, [0,25]:"DÉMARRAGE", [26,75]:"EN_COURS", ...)
        inner = formula[len("SWITCH_RANGE("):-1]
        # Premier argument = la valeur
        first_comma = inner.index(",")
        val_ref = inner[:first_comma].strip()
        val = _resolve_scalar(val_ref, ctx)
        rest = inner[first_comma + 1:]
        for m in re.finditer(r'\[(\d+),\s*(\d+)\]\s*:\s*"([^"]+)"', rest):
            lo, hi, label = int(m.group(1)), int(m.group(2)), m.group(3)
            if val is not None and lo <= val <= hi:
                return label
        return None

    # ─ IF ─────────────────────────────────────────────────────────────────────
    if formula.upper().startswith("IF("):
        # IF($var >= 70, "OUI", "NON")
        inner = formula[len("IF("):-1]
        # Trouver les 3 arguments (séparés par virgules hors parenthèses)
        args = _split_args(inner)
        if len(args) < 3:
            raise ValueError(f"IF attend 3 arguments, {len(args)} trouvés")
        cond_str, true_val, false_val = args[0], args[1], args[2]
        if _eval_condition(cond_str.strip(), ctx):
            return _resolve_scalar(true_val.strip(), ctx)
        return _resolve_scalar(false_val.strip(), ctx)

    # ─ IF_NULL ────────────────────────────────────────────────────────────────
    if formula.upper().startswith("IF_NULL("):
        inner = formula[len("IF_NULL("):-1]
        args = [s.strip() for s in inner.split(",", 1)]
        val = _resolve_scalar(args[0], ctx)
        default = _resolve_scalar(args[1], ctx) if len(args) > 1 else None
        return default if val is None else val

    # ─ Littéral string entre guillemets ──────────────────────────────────────
    if (formula.startswith('"') and formula.endswith('"')) or \
       (formula.startswith("'") and formula.endswith("'")):
        return formula[1:-1]

    # ─ Littéral numérique ────────────────────────────────────────────────────
    try:
        return float(formula) if "." in formula else int(formula)
    except ValueError:
        pass

    raise ValueError(f"Fonction COMPUTE inconnue : '{formula}'")


def _split_args(text: str) -> List[str]:
    """Sépare les arguments d'une liste en gérant les parenthèses imbriquées."""
    args, depth, current = [], 0, []
    for ch in text:
        if ch == "(" :
            depth += 1
            current.append(ch)
        elif ch == ")":
            depth -= 1
            current.append(ch)
        elif ch == "," and depth == 0:
            args.append("".join(current).strip())
            current = []
        else:
            current.append(ch)
    if current:
        args.append("".join(current).strip())
    return args


def _eval_condition(cond: str, ctx: Dict[str, Any]) -> bool:
    """Évalue une condition simple (ex: '$var >= 70', '$var != NULL')."""
    m = re.match(r'^(\$[\w]+)\s*(!=|>=|<=|=|>|<)\s*(.+)$', cond)
    if not m:
        return False
    var_ref, op, raw_val = m.group(1), m.group(2), m.group(3).strip()
    val = ctx.get(var_ref)

    if raw_val.upper() == "NULL":
        rhs = None
    elif raw_val.startswith('"') or raw_val.startswith("'"):
        rhs = raw_val.strip('"').strip("'")
    else:
        try:
            rhs = float(raw_val) if "." in raw_val else int(raw_val)
        except ValueError:
            rhs = raw_val

    if op == "=":
        return val == rhs
    if op == "!=":
        return val != rhs
    if op == ">":
        return val is not None and val > rhs
    if op == ">=":
        return val is not None and val >= rhs
    if op == "<":
        return val is not None and val < rhs
    if op == "<=":
        return val is not None and val <= rhs
    return False


def execute_computes(ast: PasserelleAST, wb,
                     result: ExecutionResult) -> Dict[str, Any]:
    """
    Phase 2 — Lit les tables GET_TABLE depuis Excel puis évalue les COMPUTE.
    Retourne le contexte complet {var_name: valeur}.
    """
    ctx: Dict[str, Any] = {}

    for defn in ast.defs:
        try:
            if defn.source_type == "GET_TABLE":
                if defn.sheet not in wb.sheetnames:
                    result.errors.append(
                        f"GET_TABLE {defn.var_name} — feuille '{defn.sheet}' introuvable"
                    )
                    ctx[defn.var_name] = []
                    continue
                ws = wb[defn.sheet]
                ctx[defn.var_name] = _read_table_from_ws(ws, defn.table_name)

            elif defn.source_type == "GET_CELL":
                target = _resolve_named_range(wb, defn.sheet, defn.named_range)
                if target is None:
                    result.errors.append(
                        f"GET_CELL {defn.var_name} — plage '{defn.named_range}' introuvable"
                    )
                    ctx[defn.var_name] = None
                    continue
                ws_t, row, col = target
                ctx[defn.var_name] = ws_t.cell(row, col).value

            elif defn.source_type == "COMPUTE":
                ctx[defn.var_name] = _eval_formula(defn.formula, ctx)

        except Exception as exc:
            result.errors.append(
                f"COMPUTE {defn.var_name} — {exc}"
            )
            ctx[defn.var_name] = None

    return ctx


# ─── Phase 2b — VALIDATE ─────────────────────────────────────────────────────

def _validate_rule(rule: str, values: List[Any]) -> List[str]:
    """
    Applique une règle à une liste de valeurs.
    Retourne la liste des messages de violation (vide = OK).
    """
    rule = rule.strip()
    violations: List[str] = []

    # ── NOT_NULL ────────────────────────────────────────────────────────────
    if rule.upper() == "NOT_NULL":
        nulls = [i for i, v in enumerate(values) if v is None]
        if nulls:
            violations.append(
                f"NOT_NULL : {len(nulls)} valeur(s) nulle(s) (lignes {nulls[:5]}{'…' if len(nulls)>5 else ''})"
            )

    # ── POSITIVE ────────────────────────────────────────────────────────────
    elif rule.upper() == "POSITIVE":
        bad = [v for v in values if v is None or v <= 0]
        if bad:
            violations.append(f"POSITIVE : {len(bad)} valeur(s) <= 0 ({bad[:3]})")

    # ── NON_NEGATIVE ────────────────────────────────────────────────────────
    elif rule.upper() == "NON_NEGATIVE":
        bad = [v for v in values if v is None or v < 0]
        if bad:
            violations.append(f"NON_NEGATIVE : {len(bad)} valeur(s) < 0 ({bad[:3]})")

    # ── UNIQUE ──────────────────────────────────────────────────────────────
    elif rule.upper() == "UNIQUE":
        seen: set = set()
        dupes = []
        for v in values:
            if v is not None:
                key = str(v)
                if key in seen:
                    dupes.append(v)
                seen.add(key)
        if dupes:
            violations.append(f"UNIQUE : {len(dupes)} doublon(s) ({dupes[:3]})")

    # ── RANGE(min, max) ─────────────────────────────────────────────────────
    elif rule.upper().startswith("RANGE("):
        m = re.match(r'RANGE\(\s*([\d.+-]+)\s*,\s*([\d.+-]+)\s*\)', rule, re.IGNORECASE)
        if m:
            lo, hi = float(m.group(1)), float(m.group(2))
            bad = [v for v in values if v is None or not (lo <= v <= hi)]
            if bad:
                violations.append(
                    f"RANGE({lo},{hi}) : {len(bad)} valeur(s) hors plage ({bad[:3]})"
                )

    # ── IN("a", "b", ...) ───────────────────────────────────────────────────
    elif rule.upper().startswith("IN("):
        inner = rule[len("IN("):-1]
        allowed = {s.strip().strip('"').strip("'") for s in inner.split(",")}
        bad = [v for v in values if str(v) not in allowed]
        if bad:
            violations.append(
                f"IN({', '.join(sorted(allowed))}) : {len(bad)} valeur(s) non autorisée(s) ({bad[:3]})"
            )

    else:
        violations.append(f"Règle inconnue : '{rule}'")

    return violations


def execute_validates(ast: "PasserelleAST", ctx: Dict[str, Any],
                      result: ExecutionResult) -> None:
    """
    Phase 2b — Vérifie les règles VALIDATE sur les variables du contexte.

    - SEVERITY=error   → violation ajoutée à result.errors
    - SEVERITY=warning → violation ajoutée à result.warnings
    Le sync continue dans les deux cas ; c'est l'appelant qui décide d'aborter.
    """
    for vnode in ast.validates:
        ref = vnode.var_ref    # ex: "$activites.avancement" ou "$total_heures"

        # Résoudre : colonne de table ou scalaire
        if "." in ref:
            # Référence de colonne : $table.col
            table_var, col_name = ref.split(".", 1)
            table = ctx.get(table_var)
            if table is None:
                result.errors.append(
                    f"VALIDATE {ref} — variable '{table_var}' non définie"
                )
                continue
            if not isinstance(table, list):
                result.errors.append(
                    f"VALIDATE {ref} — '{table_var}' n'est pas une table"
                )
                continue
            values = [row.get(col_name) for row in table]
        else:
            # Scalaire : $total_heures
            val = ctx.get(ref)
            values = [val]

        # Appliquer la règle
        violations = _validate_rule(vnode.rule, values)
        for msg in violations:
            full_msg = f"VALIDATE {ref} : {msg}"
            if vnode.severity == "warning":
                result.warnings.append(full_msg)
            else:
                result.errors.append(full_msg)


# ─── Phase 3 — PUSH ───────────────────────────────────────────────────────────

def execute_pushes(ast: PasserelleAST, ctx: Dict[str, Any],
                   store, result: ExecutionResult) -> None:
    """Phase 3 — Publie les variables calculées dans le store central.

    Si le PushNode a une clause only_if, la condition est évaluée :
    - condition vraie  → PUSH exécuté
    - condition fausse → instruction skippée (valeur non écrasée dans le store)
    """
    to_push: Dict[str, Any] = {}

    for push in ast.pushes:
        # ── Évaluer ONLY_IF si présent ────────────────────────────────────────
        if push.only_if:
            try:
                if not _eval_condition(push.only_if, ctx):
                    result.skipped.append(
                        f"PUSH {push.global_name} — ONLY_IF non satisfait "
                        f"({push.only_if})"
                    )
                    continue
            except Exception as exc:
                result.errors.append(
                    f"PUSH {push.global_name} — erreur ONLY_IF : {exc}"
                )
                continue

        val = ctx.get(push.var_name)
        if val is None:
            result.skipped.append(f"PUSH {push.global_name} — valeur None ignorée")
            continue
        to_push[push.global_name] = val

    if to_push:
        store.set_many(to_push)
        result.pushed.extend(to_push.keys())


# ─── Phase 4 — BIND ───────────────────────────────────────────────────────────

def execute_binds(ast: PasserelleAST, wb, ctx: Dict[str, Any],
                  result: ExecutionResult) -> None:
    """
    Phase 4 — Écrit les valeurs calculées dans les plages nommées du Dashboard.
    Applique le formatage couleur si la valeur est un STATUS connu.
    """
    for bind in ast.binds:
        val = ctx.get(bind.var_name)
        if val is None:
            result.skipped.append(
                f"BIND {bind.var_name} → {bind.target_sheet}.{bind.target_range} — None"
            )
            continue

        target = _resolve_named_range(wb, bind.target_sheet, bind.target_range)
        if target is None:
            result.errors.append(
                f"BIND {bind.var_name} — plage '{bind.target_range}' introuvable "
                f"dans '{bind.target_sheet}'"
            )
            continue

        ws_t, row, col = target
        cell = ws_t.cell(row, col)
        cell.value = val
        _apply_status_format(cell, val)   # no-op si val n'est pas un STATUS
        result.bound.append(bind.target_range)


# ─── Entrée publique ──────────────────────────────────────────────────────────

def execute_ast(
    ast: PasserelleAST,
    filepath: Path,
    store,
) -> ExecutionResult:
    """
    Exécute un PasserelleAST sur le fichier Excel en 4 phases.

    Args:
        ast      : AST produit par parse_file()
        filepath : chemin du fichier Excel (sera lu et sauvegardé)
        store    : module store (doit exposer get(), set_many())

    Returns:
        ExecutionResult avec les listes pushed / pulled / bound / errors / skipped
    """
    result = ExecutionResult()

    if not filepath.exists():
        result.errors.append(f"Fichier introuvable : {filepath}")
        return result

    try:
        wb = load_workbook(str(filepath))
    except Exception as exc:
        result.errors.append(f"Impossible d'ouvrir '{filepath.name}' : {exc}")
        return result

    try:
        # Phase 1 — PULL
        execute_pulls(ast, wb, store, result)

        # Phase 2 — COMPUTE
        ctx = execute_computes(ast, wb, result)

        # Phase 2b — VALIDATE
        execute_validates(ast, ctx, result)

        # Phase 3 — PUSH
        execute_pushes(ast, ctx, store, result)

        # Phase 4 — BIND
        execute_binds(ast, wb, ctx, result)

        # Sauvegarder le fichier modifié
        wb.save(str(filepath))

    except Exception as exc:
        result.errors.append(f"Erreur inattendue : {exc}")
    finally:
        wb.close()

    return result
