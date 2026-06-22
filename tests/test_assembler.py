"""
tests/test_assembler.py — Tests TDD pour assembler.py et les modifications de creer_cockpit_se.py.
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
sys.path.insert(0, str(Path(__file__).parent.parent / "projet_TrainSystem"))

import pytest
from openpyxl import load_workbook

from creer_cockpit_se import generer_cockpit


# ── Tâche 0 : onglet Agenda ───────────────────────────────────────────────────

def test_cockpit_has_agenda_sheet(tmp_path):
    """Le cockpit doit avoir un onglet 'Agenda' avec les 3 sections."""
    out = generer_cockpit("Alice Dubois", [
        {"file_id": "L09U1-CFL2400-CLIM", "systeme": "Clim",
         "projet": "CFL", "heures": 200}
    ], "USR004", tmp_path)

    wb = load_workbook(str(out))
    assert "Agenda" in wb.sheetnames

    ws = wb["Agenda"]
    all_values = [ws.cell(row=r, column=c).value
                  for r in range(1, ws.max_row + 1)
                  for c in range(1, 3)]
    all_str = [str(v) for v in all_values if v]
    assert any("Semaine" in s for s in all_str)
    assert any("30" in s for s in all_str)
    assert any("ouverts" in s.lower() or "action" in s.lower() for s in all_str)


# ── Tâche 1 : table toujours créée ───────────────────────────────────────────

def test_cockpit_vide_has_table(tmp_path):
    """Un cockpit généré sans UO doit quand même avoir tbl_mes_uos."""
    out = generer_cockpit("Test Ingenieur", [], "USR004", tmp_path)
    wb = load_workbook(str(out))
    ws = wb["Mes UOs"]
    assert "tbl_mes_uos" in ws.tables
    tbl = ws.tables["tbl_mes_uos"]
    assert tbl.ref == "A3:H3"   # header seul, 0 lignes de données


# ── Tâche 2 : ajouter_uo_au_cockpit ──────────────────────────────────────────

def _make_cockpit_with_saisies(tmp_path: Path, se_name: str = "Test SE") -> Path:
    """Crée un cockpit avec 1 UO et simule une saisie ingénieur (col 5 = 25%)."""
    out = generer_cockpit(se_name, [
        {"file_id": "L09U1-CFL2400-CLIM", "systeme": "Clim",
         "projet": "CFL", "heures": 200}
    ], "USR004", tmp_path)
    wb = load_workbook(str(out))
    ws = wb["Mes UOs"]
    ws.cell(row=4, column=5, value=0.25)
    wb.save(str(out))
    return out


def test_ajouter_uo_preserve_colonnes_jaunes(tmp_path):
    """ajouter_uo_au_cockpit ne doit pas toucher aux colonnes 5 et 6 existantes."""
    from assembler import ajouter_uo_au_cockpit

    cockpit = _make_cockpit_with_saisies(tmp_path)
    result = ajouter_uo_au_cockpit(cockpit, {
        "file_id": "L11U1-RERNG-FREIN",
        "systeme": "Frein", "projet": "RER NG", "heures": 150
    })

    assert result == "added"
    wb = load_workbook(str(cockpit))
    ws = wb["Mes UOs"]
    assert ws.cell(row=4, column=5).value == pytest.approx(0.25)
    assert ws.cell(row=5, column=1).value == "L11U1-RERNG-FREIN"
    assert ws.cell(row=5, column=5).value is None or ws.cell(row=5, column=5).value == 0
    assert ws.cell(row=5, column=6).value is None or ws.cell(row=5, column=6).value == 0


def test_ajouter_uo_idempotent(tmp_path):
    """Appeler deux fois avec la même UO → 1 seule ligne, résultat 'skipped' au 2e appel."""
    from assembler import ajouter_uo_au_cockpit

    cockpit = generer_cockpit("Test SE", [], "USR004", tmp_path)
    uo = {"file_id": "L09U1-CFL2400-CLIM", "systeme": "Clim", "projet": "CFL", "heures": 200}

    r1 = ajouter_uo_au_cockpit(cockpit, uo)
    r2 = ajouter_uo_au_cockpit(cockpit, uo)

    assert r1 == "added"
    assert r2 == "skipped"
    wb = load_workbook(str(cockpit))
    ws = wb["Mes UOs"]
    count = sum(
        1 for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=1)
        if row[0].value == "L09U1-CFL2400-CLIM"
    )
    assert count == 1


def test_ajouter_uo_etend_table(tmp_path):
    """La ref de tbl_mes_uos doit être étendue d'une ligne après ajout."""
    from assembler import ajouter_uo_au_cockpit

    cockpit = generer_cockpit("Test SE", [], "USR004", tmp_path)
    ajouter_uo_au_cockpit(cockpit, {
        "file_id": "L09U1-CFL2400-CLIM", "systeme": "Clim", "projet": "CFL", "heures": 200
    })

    wb = load_workbook(str(cockpit))
    ws = wb["Mes UOs"]
    assert "tbl_mes_uos" in ws.tables
    assert ws.tables["tbl_mes_uos"].ref == "A3:H4"


def test_ajouter_uo_cree_backup(tmp_path):
    """Un fichier .bak doit exister après modification."""
    from assembler import ajouter_uo_au_cockpit

    cockpit = generer_cockpit("Test SE", [], "USR004", tmp_path)
    ajouter_uo_au_cockpit(cockpit, {
        "file_id": "L09U1-CFL2400-CLIM", "systeme": "Clim", "projet": "CFL", "heures": 200
    })

    assert cockpit.with_suffix(".bak").exists()


# ── Tâche 3 : creer_cockpit_vide ─────────────────────────────────────────────

def test_creer_cockpit_vide(tmp_path):
    """creer_cockpit_vide produit un cockpit avec tbl_mes_uos vide."""
    from assembler import creer_cockpit_vide

    out = creer_cockpit_vide("Alice Dubois", "USR004", tmp_path)

    assert out.exists()
    assert out.name == "Cockpit_Alice_Dubois.xlsx"
    wb = load_workbook(str(out))
    ws = wb["Mes UOs"]
    assert "tbl_mes_uos" in ws.tables
    assert ws.tables["tbl_mes_uos"].ref == "A3:H3"


def test_assembler_cockpit_inexistant_cree_automatiquement(tmp_path):
    """L'assembleur crée automatiquement le cockpit s'il n'existe pas."""
    from assembler import ajouter_uo_au_cockpit, creer_cockpit_vide

    cockpit = creer_cockpit_vide("Marie Dupont", "USR004", tmp_path)
    result = ajouter_uo_au_cockpit(cockpit, {
        "file_id": "L09U1-CFL2400-CLIM",
        "systeme": "Climatisation", "projet": "CFL 2400", "heures": 200
    })

    assert result == "added"
    wb = load_workbook(str(cockpit))
    ws = wb["Mes UOs"]
    assert ws.cell(row=4, column=1).value == "L09U1-CFL2400-CLIM"


# ── Tâche 4 : end-to-end ─────────────────────────────────────────────────────

def test_assembler_end_to_end(tmp_path):
    """L'assembleur crée l'UO, met à jour le cockpit, et est idempotent."""
    from assembler import instancier_uo

    # Cockpit préexistant avec saisie ingénieur (col 5 = 0.5)
    cockpit = generer_cockpit("Alice Dubois", [
        {"file_id": "L09U1-TEST01-CLIM", "systeme": "Clim",
         "projet": "TEST", "heures": 100}
    ], "USR004", tmp_path)
    wb = load_workbook(str(cockpit))
    ws = wb["Mes UOs"]
    ws.cell(row=4, column=5, value=0.5)
    wb.save(str(cockpit))

    # Appel 1 : crée l'UO et ajoute la ligne
    result1 = instancier_uo(
        uo_type="L09U1",
        projet_code="CFL2400", systeme_code="CLIM",
        se_name="Alice Dubois", pilote_id="USR004",
        heures=200, output_dir=tmp_path, sync=False
    )
    assert result1["uo_status"] == "created"
    assert result1["cockpit_status"] == "added"
    assert (tmp_path / "L09U1-CFL2400-CLIM.xlsx").exists()

    # Saisie initiale préservée
    wb = load_workbook(str(cockpit))
    ws = wb["Mes UOs"]
    assert ws.cell(row=4, column=5).value == pytest.approx(0.5)
    assert ws.cell(row=5, column=1).value == "L09U1-CFL2400-CLIM"

    # Appel 2 : idempotent
    result2 = instancier_uo(
        uo_type="L09U1",
        projet_code="CFL2400", systeme_code="CLIM",
        se_name="Alice Dubois", pilote_id="USR004",
        heures=200, output_dir=tmp_path, sync=False
    )
    assert result2["uo_status"] == "skipped"
    assert result2["cockpit_status"] == "skipped"
