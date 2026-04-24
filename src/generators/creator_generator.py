"""
Génère Creator.xlsx — console d'administration centrale de l'écosystème.

10 feuilles :
  1  Ecosystème       — paramètres globaux
  2  Rôles            — définition des rôles et droits
  3  Acteurs          — utilisateurs et profils d'accès
  4  Projets          — projets et acteurs projet
  5  Types UO         — types UO, activités et livrables génériques
  6  Systèmes         — systèmes techniques et REX pré-remplis
  7  Catalogue Tables — tables connues de l'écosystème (auto-enrichi)
  8  Catalogue Vars   — variables connues de l'écosystème (auto-enrichi)
  9  Registre         — fichiers enregistrés et planning de synchro
  10 Créer Fichier    — interface de création d'un nouveau fichier XLS
"""
from pathlib import Path
from typing import Any, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.config_loader import (
    load_acteurs, load_projects, load_registre,
    load_systems, load_uo_instances, load_uo_types,
)
from src.styles import (
    BLUE_DARK, BLUE_MID, BLUE_LIGHT,
    GREEN, GREEN_LIGHT, ORANGE_LIGHT, GREY_LIGHT, GREY_MID, WHITE,
    YELLOW_LIGHT, RED_LIGHT,
    THIN_BORDER, solid_fill, header_font, body_font,
    center, left, set_column_widths, freeze_top_row,
    style_header_row, style_data_row,
)

ROOT = Path(__file__).parent.parent.parent
CREATOR_PATH = ROOT / "Creator.xlsx"

# Couleurs d'onglets
TAB_GREEN  = "70AD47"
TAB_BLUE   = "2F5496"
TAB_ORANGE = "ED7D31"
TAB_GREY   = "808080"
TAB_RED    = "C00000"


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _add_table(ws, ref: str, name: str, style: str = "TableStyleMedium2") -> None:
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True)
    ws.add_table(tbl)


def _section_title(ws, row: int, col_start: int, col_end: int,
                   text: str, color: str = BLUE_MID) -> None:
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row, end_column=col_end,
    )
    c = ws.cell(row=row, column=col_start, value=text)
    c.fill = solid_fill(color)
    c.font = header_font(size=11)
    c.alignment = center()
    ws.row_dimensions[row].height = 22


def _kv_row(ws, row: int, label: str, value: Any = "",
            label_color: str = BLUE_LIGHT) -> None:
    lbl = ws.cell(row=row, column=1, value=label)
    lbl.fill = solid_fill(label_color)
    lbl.font = body_font(bold=True)
    lbl.alignment = left()
    lbl.border = THIN_BORDER

    val = ws.cell(row=row, column=2, value=value)
    val.font = body_font()
    val.alignment = left()
    val.border = THIN_BORDER


def _dv_list(formula: str) -> DataValidation:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Valeur non autorisée"
    dv.errorTitle = "Erreur"
    return dv


# ─── Feuille 1 : Ecosystème ───────────────────────────────────────────────────

def _sheet_ecosysteme(wb: Workbook) -> None:
    ws = wb.create_sheet("Ecosystème")
    ws.sheet_properties.tabColor = TAB_GREEN
    ws.sheet_view.showGridLines = False

    _section_title(ws, 1, 1, 3, "PARAMÈTRES DE L'ÉCOSYSTÈME", BLUE_DARK)

    params = [
        ("Nom écosystème",           "SysEng Ferroviaire"),
        ("Version",                   "1"),
        ("Chemin racine",             str(ROOT)),
        ("SharePoint URL",            ""),
        ("Synchro quotidienne",       "22:00"),
        ("Synchro hebdomadaire",      "Dimanche 22:00"),
        ("Email IT Manager",          "it@corp.fr"),
        ("Mode bootstrap effectué",   "NON"),
    ]
    for i, (label, val) in enumerate(params):
        _kv_row(ws, 3 + i, label, val)

    set_column_widths(ws, {"A": 30, "B": 50, "C": 20})


# ─── Feuille 2 : Rôles ────────────────────────────────────────────────────────

def _sheet_roles(wb: Workbook) -> None:
    ws = wb.create_sheet("Rôles")
    ws.sheet_properties.tabColor = TAB_GREEN
    ws.sheet_view.showGridLines = False

    _section_title(ws, 1, 1, 7, "DÉFINITION DES RÔLES", BLUE_DARK)

    headers = ["ID Rôle", "Nom affiché", "Accès UO", "Accès consolidation",
               "Peut créer UO", "Power BI", "Description"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, 1, 7, BLUE_MID)

    roles = [
        ("ingenieur_sys",      "Ingénieur Système",   "Ses UO",        "Non",     "Non", "Non", "Exécute les UO assignées"),
        ("pilote_tech",        "Pilote Technique",    "Son périmètre", "Partiel", "Oui", "Oui", "Distribue et suit les UO"),
        ("engagement_mgr",     "Engagement Manager",  "Non",           "Complet", "Non", "Oui", "Vision multi-projets"),
        ("it_manager",         "IT Manager",          "Tous",          "Complet", "Oui", "Non", "Administre l'écosystème"),
        ("donneur_ordre",      "Donneur d'Ordre",     "Non",           "Résumé",  "Non", "Oui", "Client externe"),
        ("pilote_tech_client", "Pilote Tech Client",  "Non",           "Filtré",  "Non", "Oui", "Pilote technique côté client"),
        ("resp_projet_client", "Resp. Projet Client", "Non",           "Projet",  "Non", "Oui", "Vision sur un projet"),
        ("expert_client",      "Expert Client",       "Non",           "Non",     "Non", "Non", "Ponctuel sur points ouverts"),
        ("fournisseur",        "Fournisseur",         "Non",           "Non",     "Non", "Non", "Via points ouverts"),
    ]
    for i, row_data in enumerate(roles):
        row = 3 + i
        for col, val in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=val)
        style_data_row(ws, row, 1, 7, i % 2 == 1)

    last = 2 + len(roles)
    _add_table(ws, f"A2:G{last}", "TabRoles", "TableStyleMedium9")
    set_column_widths(ws, {"A": 22, "B": 22, "C": 18, "D": 20, "E": 14, "F": 12, "G": 40})
    freeze_top_row(ws)


# ─── Feuille 3 : Acteurs ──────────────────────────────────────────────────────

def _sheet_acteurs(wb: Workbook) -> None:
    ws = wb.create_sheet("Acteurs")
    ws.sheet_properties.tabColor = TAB_GREEN
    ws.sheet_view.showGridLines = False

    _section_title(ws, 1, 1, 7, "ACTEURS ET PROFILS D'ACCÈS", BLUE_DARK)

    headers = ["ID", "Nom", "Rôle", "Filtre type", "Filtre valeur", "Accès", "Email"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, 1, 7, BLUE_MID)

    acteurs = load_acteurs()
    for i, a in enumerate(acteurs):
        row = 3 + i
        vals = [a.id, a.nom, a.role.value, a.filtre_type.value,
                a.filtre_valeur, a.acces.value, a.email]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v)
        style_data_row(ws, row, 1, 7, i % 2 == 1)

    # Validation rôle
    roles_formula = '"ingenieur_sys,pilote_tech,engagement_mgr,it_manager,donneur_ordre,pilote_tech_client,resp_projet_client,expert_client,fournisseur"'
    dv_role = _dv_list(roles_formula)
    ws.add_data_validation(dv_role)
    last = 2 + len(acteurs)
    for row in range(3, last + 6):
        dv_role.add(ws.cell(row=row, column=3))

    # Validation accès
    acces_formula = '"read/write,read,read_filtered,read_summary,admin"'
    dv_acces = _dv_list(acces_formula)
    ws.add_data_validation(dv_acces)
    for row in range(3, last + 6):
        dv_acces.add(ws.cell(row=row, column=6))

    _add_table(ws, f"A2:G{last}", "TabActeurs", "TableStyleMedium2")
    set_column_widths(ws, {"A": 10, "B": 22, "C": 22, "D": 18, "E": 35, "F": 16, "G": 30})
    freeze_top_row(ws)


# ─── Feuille 4 : Projets ──────────────────────────────────────────────────────

def _sheet_projets(wb: Workbook) -> None:
    ws = wb.create_sheet("Projets")
    ws.sheet_properties.tabColor = TAB_GREEN
    ws.sheet_view.showGridLines = False

    # ── Table projets
    _section_title(ws, 1, 1, 3, "PROJETS", BLUE_DARK)
    for col, h in enumerate(["ID Projet", "Nom", "Description"], 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, 1, 3, BLUE_MID)

    projects = load_projects()
    proj_list = list(projects.values())
    for i, p in enumerate(proj_list):
        row = 3 + i
        for col, v in enumerate([p.id, p.name, ""], 1):
            ws.cell(row=row, column=col, value=v)
        style_data_row(ws, row, 1, 3, i % 2 == 1)

    last_proj = 2 + len(proj_list)
    _add_table(ws, f"A2:C{last_proj}", "TabProjets", "TableStyleMedium2")

    # ── Table acteurs projets
    sep = last_proj + 2
    _section_title(ws, sep, 1, 4, "ACTEURS PAR PROJET", BLUE_MID)
    for col, h in enumerate(["Projet", "Nom acteur", "Rôle dans projet", "Email"], 1):
        ws.cell(row=sep + 1, column=col, value=h)
    style_header_row(ws, sep + 1, 1, 4, BLUE_MID)

    row = sep + 2
    for p in proj_list:
        for j, actor in enumerate(p.actors):
            for col, v in enumerate([p.id, actor.name, actor.role, actor.email], 1):
                ws.cell(row=row, column=col, value=v)
            style_data_row(ws, row, 1, 4, j % 2 == 1)
            row += 1

    last_act = row - 1
    _add_table(ws, f"A{sep+1}:D{last_act}", "TabActeursProjets", "TableStyleMedium9")
    set_column_widths(ws, {"A": 18, "B": 28, "C": 28, "D": 32})


# ─── Feuille 5 : Types UO ─────────────────────────────────────────────────────

def _sheet_types_uo(wb: Workbook) -> None:
    ws = wb.create_sheet("Types UO")
    ws.sheet_properties.tabColor = TAB_BLUE
    ws.sheet_view.showGridLines = False

    uo_types = load_uo_types()

    # ── Table types
    _section_title(ws, 1, 1, 3, "TYPES D'UO", BLUE_DARK)
    for col, h in enumerate(["ID Type", "Nom", "Description"], 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, 1, 3, BLUE_MID)
    for i, t in enumerate(uo_types.values()):
        row = 3 + i
        for col, v in enumerate([t.id, t.name, ""], 1):
            ws.cell(row=row, column=col, value=v)
        style_data_row(ws, row, 1, 3, i % 2 == 1)
    last_type = 2 + len(uo_types)
    _add_table(ws, f"A2:C{last_type}", "TabTypesUO", "TableStyleMedium2")

    # ── Table activités génériques
    sep1 = last_type + 2
    _section_title(ws, sep1, 1, 4, "ACTIVITÉS GÉNÉRIQUES PAR TYPE", BLUE_MID)
    for col, h in enumerate(["ID Type", "ID Activité", "Nom activité", "Heures défaut"], 1):
        ws.cell(row=sep1 + 1, column=col, value=h)
    style_header_row(ws, sep1 + 1, 1, 4, BLUE_MID)
    row = sep1 + 2
    for t in uo_types.values():
        for j, a in enumerate(t.activities):
            for col, v in enumerate([t.id, a.id, a.name, a.default_hours], 1):
                ws.cell(row=row, column=col, value=v)
            style_data_row(ws, row, 1, 4, j % 2 == 1)
            row += 1
    last_act = row - 1
    _add_table(ws, f"A{sep1+1}:D{last_act}", "TabActivitesGeneriques", "TableStyleMedium9")

    # ── Table livrables génériques
    sep2 = last_act + 2
    _section_title(ws, sep2, 1, 3, "LIVRABLES GÉNÉRIQUES PAR TYPE", BLUE_MID)
    for col, h in enumerate(["ID Type", "ID Livrable", "Nom livrable"], 1):
        ws.cell(row=sep2 + 1, column=col, value=h)
    style_header_row(ws, sep2 + 1, 1, 3, BLUE_MID)
    row = sep2 + 2
    for t in uo_types.values():
        for j, d in enumerate(t.deliverables):
            for col, v in enumerate([t.id, d.id, d.name], 1):
                ws.cell(row=row, column=col, value=v)
            style_data_row(ws, row, 1, 3, j % 2 == 1)
            row += 1
    last_del = row - 1
    _add_table(ws, f"A{sep2+1}:C{last_del}", "TabLivrablesGeneriques", "TableStyleLight1")
    set_column_widths(ws, {"A": 22, "B": 28, "C": 40, "D": 16})


# ─── Feuille 6 : Systèmes ─────────────────────────────────────────────────────

def _sheet_systemes(wb: Workbook) -> None:
    ws = wb.create_sheet("Systèmes")
    ws.sheet_properties.tabColor = TAB_BLUE
    ws.sheet_view.showGridLines = False

    systems = load_systems()

    _section_title(ws, 1, 1, 2, "SYSTÈMES TECHNIQUES", BLUE_DARK)
    for col, h in enumerate(["ID Système", "Nom"], 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, 1, 2, BLUE_MID)
    for i, s in enumerate(systems.values()):
        row = 3 + i
        for col, v in enumerate([s.id, s.name], 1):
            ws.cell(row=row, column=col, value=v)
        style_data_row(ws, row, 1, 2, i % 2 == 1)
    last_sys = 2 + len(systems)
    _add_table(ws, f"A2:B{last_sys}", "TabSystemes", "TableStyleMedium2")

    sep = last_sys + 2
    _section_title(ws, sep, 1, 3, "REX PRÉ-REMPLIS PAR SYSTÈME", BLUE_MID)
    for col, h in enumerate(["ID Système", "Catégorie", "Item REX"], 1):
        ws.cell(row=sep + 1, column=col, value=h)
    style_header_row(ws, sep + 1, 1, 3, BLUE_MID)
    row = sep + 2
    for s in systems.values():
        for j, item in enumerate(s.rex_prefill):
            for col, v in enumerate([s.id, "Bonne pratique", item], 1):
                ws.cell(row=row, column=col, value=v)
            style_data_row(ws, row, 1, 3, j % 2 == 1)
            row += 1
    last_rex = row - 1
    _add_table(ws, f"A{sep+1}:C{last_rex}", "TabREX", "TableStyleMedium9")
    set_column_widths(ws, {"A": 18, "B": 20, "C": 70})


# ─── Feuille 7 : Catalogue Tables ────────────────────────────────────────────

def _sheet_catalogue_tables(wb: Workbook) -> None:
    ws = wb.create_sheet("Catalogue Tables")
    ws.sheet_properties.tabColor = TAB_ORANGE
    ws.sheet_view.showGridLines = False

    _section_title(ws, 1, 1, 6, "CATALOGUE DES TABLES DE L'ÉCOSYSTÈME  —  auto-enrichi par le script", BLUE_DARK)

    note = ws.cell(row=2, column=1,
        value="ℹ  Ce catalogue est alimenté automatiquement lors des synchronisations. "
              "Vous pouvez aussi ajouter des tables manuellement.")
    note.fill = solid_fill(YELLOW_LIGHT)
    note.font = body_font(size=10)
    note.alignment = left()
    ws.merge_cells("A2:F2")

    for col, h in enumerate(["ID Table", "Fichier source", "Description",
                               "Colonnes (identifiants)", "Découvert le", "Statut"], 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, 1, 6, BLUE_MID)

    # Lignes d'exemple vides
    exemples = [
        ("uo.activites",     "UO-001", "Activités d'une instance UO",
         "id;nom;heures;avancement;statut;heures_realisees", "", "découvert"),
        ("uo.livrables",     "UO-001", "Livrables d'une UO",
         "id;nom;date_prevue;statut", "", "découvert"),
        ("uo.points_ouverts","UO-001", "Points ouverts d'une UO",
         "id;titre;nature;responsable;statut", "", "découvert"),
        ("projet.acteurs",   "REF-PROJ-MI20", "Acteurs d'un projet",
         "nom;role;email", "", "découvert"),
    ]
    for i, row_data in enumerate(exemples):
        row = 4 + i
        for col, v in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=v)
        style_data_row(ws, row, 1, 6, i % 2 == 1)

    last = 3 + len(exemples)
    _add_table(ws, f"A3:F{last}", "TabCatalogueTables", "TableStyleMedium7")

    # Détail colonnes
    sep = last + 2
    _section_title(ws, sep, 1, 6, "DÉTAIL DES COLONNES PAR TABLE", BLUE_MID)
    for col, h in enumerate(["ID Table", "Colonne", "Type", "Header Excel", "Write", "Description"], 1):
        ws.cell(row=sep + 1, column=col, value=h)
    style_header_row(ws, sep + 1, 1, 6, BLUE_MID)

    colonnes = [
        ("uo.activites", "id",               "KEY",    "ID Activité",     "creation",      ""),
        ("uo.activites", "nom",              "string", "Activité",         "uo_generique",  ""),
        ("uo.activites", "heures",           "float",  "Heures allouées",  "creation",      ""),
        ("uo.activites", "avancement",       "float",  "% Avancement",     "engineer",      ""),
        ("uo.activites", "statut",           "string", "Statut",           "engineer",      ""),
        ("uo.activites", "heures_realisees", "float",  "Heures réalisées", "engineer",      ""),
        ("projet.acteurs","nom",             "string", "Nom",              "creation",      ""),
        ("projet.acteurs","role",            "string", "Rôle",             "creation",      ""),
        ("projet.acteurs","email",           "string", "Email",            "creation",      ""),
    ]
    for i, row_data in enumerate(colonnes):
        row = sep + 2 + i
        for col, v in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=v)
        style_data_row(ws, row, 1, 6, i % 2 == 1)

    last_col = sep + 1 + len(colonnes)
    _add_table(ws, f"A{sep+1}:F{last_col}", "TabCatalogueColonnes", "TableStyleLight9")
    set_column_widths(ws, {"A": 22, "B": 22, "C": 12, "D": 22, "E": 16, "F": 35})
    freeze_top_row(ws)


# ─── Feuille 8 : Catalogue Variables ─────────────────────────────────────────

def _sheet_catalogue_variables(wb: Workbook) -> None:
    ws = wb.create_sheet("Catalogue Variables")
    ws.sheet_properties.tabColor = TAB_ORANGE
    ws.sheet_view.showGridLines = False

    _section_title(ws, 1, 1, 6,
        "CATALOGUE DES VARIABLES DE L'ÉCOSYSTÈME  —  auto-enrichi par le script", BLUE_DARK)

    note = ws.cell(row=2, column=1,
        value="ℹ  Variables scalaires et calculées disponibles pour les PUSH/PULL/BIND dans les passerelles.")
    note.fill = solid_fill(YELLOW_LIGHT)
    note.font = body_font(size=10)
    ws.merge_cells("A2:F2")

    for col, h in enumerate(["ID Variable", "Type", "Fichier source",
                               "Formule / Source", "Découvert le", "Description"], 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, 1, 6, BLUE_MID)

    variables = [
        ("uo.avancement_global",  "COMPUTED", "UO-*",
         "MEAN_WEIGHTED(activites.avancement, activites.heures)", "", "Avancement pondéré UO"),
        ("uo.heures_realisees",   "COMPUTED", "UO-*",
         "SUM(activites.heures_realisees)", "", "Total heures réalisées"),
        ("uo.nb_points_ouverts",  "COMPUTED", "UO-*",
         'COUNT_IF(points_ouverts.statut,"En cours")', "", "Points ouverts actifs"),
        ("uo.statut",             "CELL",     "UO-*",    "",  "", "Statut cycle de vie UO"),
        ("projet.nom",            "CELL",     "REF-PROJ-*", "", "", "Nom du projet"),
    ]
    for i, row_data in enumerate(variables):
        row = 4 + i
        for col, v in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=v)
        style_data_row(ws, row, 1, 6, i % 2 == 1)

    last = 3 + len(variables)
    _add_table(ws, f"A3:F{last}", "TabCatalogueVariables", "TableStyleMedium7")
    set_column_widths(ws, {"A": 28, "B": 12, "C": 18, "D": 45, "E": 16, "F": 35})
    freeze_top_row(ws)


# ─── Feuille 9 : Registre ─────────────────────────────────────────────────────

def _sheet_registre(wb: Workbook) -> None:
    ws = wb.create_sheet("Registre Fichiers")
    ws.sheet_properties.tabColor = TAB_GREY
    ws.sheet_view.showGridLines = False

    _section_title(ws, 1, 1, 7, "REGISTRE DES FICHIERS DE L'ÉCOSYSTÈME", BLUE_DARK)

    headers = ["ID", "Type", "Chemin", "Périodicité",
               "Généré par script", "Dernière synchro", "Statut"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, 1, 7, BLUE_MID)

    entrees = load_registre()
    for i, e in enumerate(entrees):
        row = 3 + i
        vals = [
            e.id, e.type_fichier, e.chemin, e.synchro_periodicite,
            "OUI" if e.genere_par_script else "NON",
            e.derniere_synchro or "",
            e.statut_dernier_synchro or "jamais",
        ]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v)
        style_data_row(ws, row, 1, 7, i % 2 == 1)

        # Couleur statut
        statut_cell = ws.cell(row=row, column=7)
        if e.statut_dernier_synchro == "ok":
            statut_cell.fill = solid_fill(GREEN_LIGHT)
        elif e.statut_dernier_synchro == "erreur":
            statut_cell.fill = solid_fill(RED_LIGHT)
        elif e.statut_dernier_synchro == "skip_verrouille":
            statut_cell.fill = solid_fill(ORANGE_LIGHT)

    # Validation périodicité
    dv_per = _dv_list('"quotidien,hebdomadaire,manuel"')
    ws.add_data_validation(dv_per)
    for row in range(3, 3 + len(entrees) + 10):
        dv_per.add(ws.cell(row=row, column=4))

    last = 2 + len(entrees)
    _add_table(ws, f"A2:G{last}", "TabRegistre", "TableStyleMedium2")
    set_column_widths(ws, {"A": 18, "B": 22, "C": 55, "D": 14, "E": 18, "F": 22, "G": 18})
    freeze_top_row(ws)


# ─── Feuille 10 : Créer Fichier ───────────────────────────────────────────────

def _sheet_creer_fichier(wb: Workbook) -> None:
    ws = wb.create_sheet("Créer Fichier")
    ws.sheet_properties.tabColor = TAB_RED
    ws.sheet_view.showGridLines = False

    # ── Titre principal
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "CRÉER UN NOUVEAU FICHIER EXCEL"
    t.fill = solid_fill(BLUE_DARK)
    t.font = header_font(size=14)
    t.alignment = center()
    ws.row_dimensions[1].height = 30

    # ── Section : Identification
    _section_title(ws, 3, 1, 8, "1.  IDENTIFICATION DU FICHIER", BLUE_MID)

    form_fields = [
        ("Type de fichier",   "cockpit",  "D4",
         '"cockpit,pilote_tech,engagement_mgr,uo_instance,client,referentiel_projet,referentiel_uo,custom"'),
        ("ID du fichier",     "",   None, None),
        ("Destinataire / Ingénieur", "", None, None),
        ("Projet filtré",     "",   None, None),
        ("Chemin de sortie",  "",   None, None),
    ]

    acteurs = load_acteurs()
    acteur_noms = ";".join(a.nom for a in acteurs)
    projets = load_projects()
    projet_ids = ";".join(projets.keys())

    for i, (label, default, _, _) in enumerate(form_fields):
        row = 4 + i
        lbl = ws.cell(row=row, column=1, value=label)
        lbl.fill = solid_fill(BLUE_LIGHT)
        lbl.font = body_font(bold=True)
        lbl.alignment = left()
        lbl.border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

        val = ws.cell(row=row, column=3, value=default)
        val.font = body_font()
        val.border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)

    # Validations
    dv_type = _dv_list('"cockpit,pilote_tech,engagement_mgr,uo_instance,client,referentiel_projet,referentiel_uo,custom"')
    ws.add_data_validation(dv_type)
    dv_type.add(ws["C4"])

    dv_dest = _dv_list(f'"{acteur_noms}"')
    ws.add_data_validation(dv_dest)
    dv_dest.add(ws["C6"])

    dv_proj = _dv_list(f'"{projet_ids}"')
    ws.add_data_validation(dv_proj)
    dv_proj.add(ws["C7"])

    # ── Section : Feuilles
    _section_title(ws, 10, 1, 8, "2.  FEUILLES À INCLURE", BLUE_MID)

    feuille_headers = ["Nom feuille", "Contenu (description)", "Variables incluses",
                       "Tables incluses", "Ordre", "Inclure"]
    for col, h in enumerate(feuille_headers, 1):
        ws.cell(row=11, column=col, value=h)
    style_header_row(ws, 11, 1, 6, BLUE_MID)

    feuilles_defaut = {
        "cockpit":        [("Cockpit",        "Vue synthétique UO", "uo.avancement_global;uo.heures_realisees", "uo.activites", "1")],
        "pilote_tech":    [("Suivi Pilote",   "Vue périmètre",      "uo.avancement_global",                    "uo.activites", "1")],
        "engagement_mgr": [("Vue Globale",    "Multi-projets",      "projet.nom",                              "uo.activites", "1")],
    }
    for i in range(8):
        row = 12 + i
        for col in range(1, 7):
            c = ws.cell(row=row, column=col, value="")
            c.border = THIN_BORDER
            c.fill = solid_fill(GREY_LIGHT if i % 2 == 1 else WHITE)

    dv_incl = _dv_list('"OUI,NON"')
    ws.add_data_validation(dv_incl)
    for row in range(12, 20):
        dv_incl.add(ws.cell(row=row, column=6))

    _add_table(ws, "A11:F19", "TabFeuillees", "TableStyleMedium2")

    # ── Section : Variables disponibles
    _section_title(ws, 21, 1, 8, "3.  VARIABLES DISPONIBLES  (depuis Catalogue Variables)", BLUE_MID)

    var_headers = ["ID Variable", "Type", "Description", "Inclure"]
    for col, h in enumerate(var_headers, 1):
        ws.cell(row=22, column=col, value=h)
    style_header_row(ws, 22, 1, 4, BLUE_MID)

    variables = [
        ("uo.avancement_global",  "COMPUTED", "Avancement pondéré UO"),
        ("uo.heures_realisees",   "COMPUTED", "Total heures réalisées"),
        ("uo.nb_points_ouverts",  "COMPUTED", "Points ouverts actifs"),
        ("uo.statut",             "CELL",     "Statut cycle de vie"),
        ("projet.nom",            "CELL",     "Nom du projet"),
    ]
    for i, (vid, vtype, vdesc) in enumerate(variables):
        row = 23 + i
        for col, v in enumerate([vid, vtype, vdesc, ""], 1):
            ws.cell(row=row, column=col, value=v)
        style_data_row(ws, row, 1, 4, i % 2 == 1)

    dv_var = _dv_list('"OUI,NON"')
    ws.add_data_validation(dv_var)
    last_var = 22 + len(variables)
    for row in range(23, last_var + 1):
        dv_var.add(ws.cell(row=row, column=4))

    _add_table(ws, f"A22:D{last_var}", "TabVariablesDispos", "TableStyleLight9")

    # ── Section : Tables disponibles
    sep_tables = last_var + 2
    _section_title(ws, sep_tables, 1, 8, "4.  TABLES DISPONIBLES  (depuis Catalogue Tables)", BLUE_MID)

    tbl_headers = ["ID Table", "Description", "Colonnes", "Mode PULL", "Inclure"]
    for col, h in enumerate(tbl_headers, 1):
        ws.cell(row=sep_tables + 1, column=col, value=h)
    style_header_row(ws, sep_tables + 1, 1, 5, BLUE_MID)

    tables = [
        ("uo.activites",      "Activités de l'UO",      "id;nom;heures;avancement;statut", "APPEND_NEW"),
        ("uo.livrables",      "Livrables de l'UO",      "id;nom;date_prevue;statut",       "APPEND_NEW"),
        ("uo.points_ouverts", "Points ouverts",          "id;titre;nature;statut",          "READ_ONLY"),
        ("projet.acteurs",    "Acteurs du projet",       "nom;role;email",                  "OVERWRITE"),
    ]
    for i, row_data in enumerate(tables):
        row = sep_tables + 2 + i
        for col, v in enumerate(list(row_data) + [""], 1):
            ws.cell(row=row, column=col, value=v)
        style_data_row(ws, row, 1, 5, i % 2 == 1)

    dv_tbl = _dv_list('"OUI,NON"')
    ws.add_data_validation(dv_tbl)
    last_tbl = sep_tables + 1 + len(tables)
    for row in range(sep_tables + 2, last_tbl + 1):
        dv_tbl.add(ws.cell(row=row, column=5))

    dv_mode = _dv_list('"READ_ONLY,APPEND_NEW,UPDATE,OVERWRITE"')
    ws.add_data_validation(dv_mode)
    for row in range(sep_tables + 2, last_tbl + 1):
        dv_mode.add(ws.cell(row=row, column=4))

    _add_table(ws, f"A{sep_tables+1}:E{last_tbl}", "TabTablesDispos", "TableStyleLight9")

    # ── Section : Génération
    sep_gen = last_tbl + 2
    _section_title(ws, sep_gen, 1, 8, "5.  GÉNÉRATION", BLUE_DARK)

    flag_label = ws.cell(row=sep_gen + 1, column=1, value="Marquer pour génération")
    flag_label.fill = solid_fill(BLUE_LIGHT)
    flag_label.font = body_font(bold=True)
    flag_label.border = THIN_BORDER
    ws.merge_cells(start_row=sep_gen+1, start_column=1, end_row=sep_gen+1, end_column=2)

    flag = ws.cell(row=sep_gen + 1, column=3, value="NON")
    flag.fill = solid_fill(YELLOW_LIGHT)
    flag.font = header_font(size=12, color="000000")
    flag.alignment = center()
    flag.border = THIN_BORDER

    dv_flag = _dv_list('"OUI,NON,GENERE"')
    ws.add_data_validation(dv_flag)
    dv_flag.add(ws.cell(row=sep_gen + 1, column=3))

    statut_label = ws.cell(row=sep_gen + 2, column=1, value="Statut")
    statut_label.fill = solid_fill(BLUE_LIGHT)
    statut_label.font = body_font(bold=True)
    statut_label.border = THIN_BORDER
    ws.merge_cells(start_row=sep_gen+2, start_column=1, end_row=sep_gen+2, end_column=2)

    statut = ws.cell(row=sep_gen + 2, column=3, value="En attente")
    statut.font = body_font()
    statut.border = THIN_BORDER

    set_column_widths(ws, {
        "A": 28, "B": 22, "C": 30, "D": 25, "E": 20, "F": 10, "G": 15, "H": 15,
    })


# ─── Génération principale ────────────────────────────────────────────────────

def generate_creator(path: Path = CREATOR_PATH) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    _sheet_ecosysteme(wb)
    _sheet_roles(wb)
    _sheet_acteurs(wb)
    _sheet_projets(wb)
    _sheet_types_uo(wb)
    _sheet_systemes(wb)
    _sheet_catalogue_tables(wb)
    _sheet_catalogue_variables(wb)
    _sheet_registre(wb)
    _sheet_creer_fichier(wb)

    wb.save(path)
    return path
