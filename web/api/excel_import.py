"""Importe un fichier Excel existant dans l'écosystème ExoSync.

Workflow :
  1. POST /api/excel/scan   → lit le fichier, retourne plages nommées + tables
  2. POST /api/excel/register → enregistre le mapping dans registre + tables
"""
import re
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, HTTPException, UploadFile, File
from pydantic import BaseModel
import openpyxl
from openpyxl.utils import get_column_letter

from web.services.config_service import (
    load_registre, save_registre,
    load_tables, save_tables,
    load_file_types,
    load_hierarchy, save_hierarchy,
)

router = APIRouter()


# ── Modèles de réponse ────────────────────────────────────────────────────────

class NamedRangeInfo(BaseModel):
    name: str
    sheet: str
    address: str
    kind: str           # "scalar" | "range" | "table_ref"
    value_preview: str  # première valeur lue, pour aider l'utilisateur
    suggested_type: str # "string" | "float" | "int" | "date" | "bool" | "ref"


class ColumnInfo(BaseModel):
    name: str
    suggested_type: str  # déduit de la première valeur non-vide


class TableInfo(BaseModel):
    name: str
    sheet: str
    ref: str
    columns: list[ColumnInfo]
    row_count: int


class ScanResult(BaseModel):
    path: str
    sheets: list[str]
    named_ranges: list[NamedRangeInfo]
    tables: list[TableInfo]


class ColumnMapping(BaseModel):
    name: str
    col_type: str = "string"
    header: str = ""
    is_key: bool = False
    required: bool = True


class TableMapping(BaseModel):
    excel_name: str       # nom de la table dans Excel
    schema_name: str      # nom dans le schéma (ex. TabActivites)
    sheet: str
    columns: list[ColumnMapping]
    include: bool = True


class FieldMapping(BaseModel):
    named_range: str      # nom de la plage nommée
    field_name: str       # nom dans min_fields
    field_type: str = "string"
    nature: str = "identitaire"
    source: str = "user_input"
    required: bool = True
    pushable: bool = False
    include: bool = True


class CollectMapping(BaseModel):
    source_class: str
    source_table: str
    target_table: str
    where_clause: str = ""
    cols_filter: str = ""


class RegisterRequest(BaseModel):
    path: str
    file_id: str
    type_fichier: str
    owner_role: str = ""
    synchro_periodicite: str = "quotidien"
    field_mappings: list[FieldMapping] = []
    table_mappings: list[TableMapping] = []
    update_class_fields: bool = False   # si True, met à jour min_fields de la Classe
    collect_mappings: list[CollectMapping] = []


# ── Modèles pour compare-tables ──────────────────────────────────────────────

class CompareTablesRequest(BaseModel):
    scanned_columns: list[str]
    existing_table_id: str


class CompareTablesResponse(BaseModel):
    match_score: float
    matching_columns: list[str]
    missing_in_scan: list[str]
    extra_in_scan: list[str]
    verdict: str  # "identical" | "compatible" | "different"


# ── Utilitaires ───────────────────────────────────────────────────────────────

def _guess_type(value) -> str:
    if value is None:
        return "string"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, int):
        return "int"
    if isinstance(value, float):
        return "float"
    import datetime
    if isinstance(value, (datetime.date, datetime.datetime)):
        return "date"
    s = str(value).strip()
    if re.match(r"^\d{1,3}([.,]\d+)?%$", s):
        return "pct"
    try:
        float(s.replace(",", "."))
        return "float"
    except ValueError:
        pass
    return "string"


def _resolve_named_range(wb, dest_str: str) -> tuple[str, str, str]:
    """Retourne (sheet, address, kind) depuis une destination de plage nommée."""
    # Formats possibles : "Sheet1!$A$1", "'Sheet 1'!$A$1:$B$10", "#REF!", formulas…
    dest_str = dest_str.strip()
    if "!" not in dest_str:
        return ("", dest_str, "formula")

    # Séparer feuille et adresse
    parts = dest_str.rsplit("!", 1)
    sheet_raw = parts[0].strip("'\"")
    addr = parts[1].replace("$", "")

    if ":" in addr:
        kind = "range"
    else:
        kind = "scalar"

    return (sheet_raw, addr, kind)


def _read_scalar_value(wb, sheet_name: str, address: str) -> str:
    """Lit la valeur d'une cellule simple."""
    try:
        ws = wb[sheet_name]
        val = ws[address].value
        return str(val) if val is not None else ""
    except Exception:
        return ""


def _read_table_columns(ws, tbl) -> list[ColumnInfo]:
    """Lit les colonnes d'une Table Excel."""
    cols = []
    # Récupérer la première ligne de données pour deviner le type
    try:
        ref = tbl.ref  # ex. "A1:F20"
        from openpyxl.utils.cell import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        data_row = min_row + 1  # ligne après l'en-tête
    except Exception:
        data_row = None

    for tc in tbl.tableColumns:
        suggested = "string"
        if data_row:
            try:
                cell = ws.cell(row=data_row, column=tc.id)
                suggested = _guess_type(cell.value)
            except Exception:
                pass
        cols.append(ColumnInfo(name=tc.name, suggested_type=suggested))

    return cols


def _count_table_rows(ws, tbl) -> int:
    try:
        from openpyxl.utils.cell import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
        return max(0, max_row - min_row)  # sans l'en-tête
    except Exception:
        return 0


# ── Similarité Jaccard ────────────────────────────────────────────────────────

def _normalize_col(s: str) -> str:
    return re.sub(r"[\s\-]+", "_", s.strip().lower())


def _jaccard(a_cols: list[str], b_cols: list[str]) -> float:
    a = {_normalize_col(c) for c in a_cols}
    b = {_normalize_col(c) for c in b_cols}
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


# ── POST /compare-tables ──────────────────────────────────────────────────────

@router.post("/compare-tables", response_model=CompareTablesResponse)
def compare_tables(body: CompareTablesRequest):
    tables_data = load_tables()
    existing = tables_data.get("tables", {}).get(body.existing_table_id)
    if not existing:
        raise HTTPException(404, f"Table '{body.existing_table_id}' introuvable dans tables.json")

    existing_cols = [c["name"] for c in existing.get("columns", [])]
    scanned_norm = {_normalize_col(c) for c in body.scanned_columns}
    existing_norm = {_normalize_col(c) for c in existing_cols}

    score = _jaccard(body.scanned_columns, existing_cols)
    if score >= 0.90:
        verdict = "identical"
    elif score >= 0.60:
        verdict = "compatible"
    else:
        verdict = "different"

    matching = [c for c in body.scanned_columns if _normalize_col(c) in existing_norm]
    missing = [c for c in existing_cols if _normalize_col(c) not in scanned_norm]
    extra = [c for c in body.scanned_columns if _normalize_col(c) not in existing_norm]

    return CompareTablesResponse(
        match_score=round(score, 4),
        matching_columns=matching,
        missing_in_scan=missing,
        extra_in_scan=extra,
        verdict=verdict,
    )


# ── GET /available-tables ─────────────────────────────────────────────────────

@router.get("/available-tables")
def available_tables():
    tables_data = load_tables()
    registre = load_registre()

    file_type_map = {f["id"]: f.get("type_fichier", "") for f in registre}

    by_class: dict[str, list] = {}
    for tid, tbl in tables_data.get("tables", {}).items():
        file_id = tbl.get("file_id", "")
        type_fichier = file_type_map.get(file_id, "")
        if not type_fichier:
            continue
        entry = {
            "table_id": tid,
            "table_name": tbl.get("table_name", ""),
            "columns": [c["name"] for c in tbl.get("columns", [])],
        }
        by_class.setdefault(type_fichier, []).append(entry)

    return {"by_class": by_class}


# ── POST /scan ────────────────────────────────────────────────────────────────

class ScanRequest(BaseModel):
    path: str


@router.post("/scan", response_model=ScanResult)
def scan_excel(body: ScanRequest):
    p = Path(body.path)
    if not p.exists():
        raise HTTPException(404, f"Fichier introuvable : {body.path}")
    if p.suffix.lower() not in (".xlsx", ".xlsm", ".xlsb", ".xls"):
        raise HTTPException(422, "Le fichier doit être un fichier Excel (.xlsx, .xlsm…)")

    try:
        wb = openpyxl.load_workbook(str(p), read_only=False, data_only=True)
    except Exception as e:
        raise HTTPException(422, f"Impossible d'ouvrir le fichier : {e}")

    sheets = wb.sheetnames

    # ── Plages nommées ─────────────────────────────────────────────────────────
    named_ranges: list[NamedRangeInfo] = []
    for dn in wb.defined_names.values():
        name = dn.name
        # Ignorer les noms internes Excel (_xlnm.*, Print_Area, etc.)
        if name.startswith("_") or name.startswith("xlnm"):
            continue
        try:
            dest = dn.attr_text or ""
            destinations = list(dn.destinations) if hasattr(dn, "destinations") else []
        except Exception:
            continue

        if destinations:
            sheet_name, coord = destinations[0]
            addr = coord.replace("$", "")
            kind = "range" if ":" in addr else "scalar"
        else:
            try:
                sheet_name, addr, kind = _resolve_named_range(wb, dest)
            except Exception:
                continue

        value_preview = ""
        if kind == "scalar" and sheet_name in wb.sheetnames:
            value_preview = _read_scalar_value(wb, sheet_name, addr)

        named_ranges.append(NamedRangeInfo(
            name=name,
            sheet=sheet_name,
            address=addr,
            kind=kind,
            value_preview=value_preview,
            suggested_type=_guess_type(value_preview) if kind == "scalar" else "string",
        ))

    # ── Tables Excel ───────────────────────────────────────────────────────────
    tables: list[TableInfo] = []
    for ws in wb.worksheets:
        for tbl in ws.tables.values():
            cols = _read_table_columns(ws, tbl)
            rows = _count_table_rows(ws, tbl)
            tables.append(TableInfo(
                name=tbl.name,
                sheet=ws.title,
                ref=tbl.ref,
                columns=cols,
                row_count=rows,
            ))

    wb.close()

    return ScanResult(
        path=str(p),
        sheets=sheets,
        named_ranges=named_ranges,
        tables=tables,
    )


# ── POST /scan-upload ─────────────────────────────────────────────────────────

@router.post("/scan-upload", response_model=ScanResult)
async def scan_excel_upload(file: UploadFile = File(...)):
    """Variante upload — lit directement le fichier envoyé par le navigateur."""
    import tempfile, shutil
    suffix = Path(file.filename).suffix
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = tmp.name

    result = scan_excel(ScanRequest(path=tmp_path))
    # Garder le nom original dans le résultat
    result.path = file.filename
    Path(tmp_path).unlink(missing_ok=True)
    return result


# ── POST /register ────────────────────────────────────────────────────────────

@router.post("/register")
def register_excel(body: RegisterRequest):
    # Vérifier que la Classe existe
    ft_all = load_file_types()
    if body.type_fichier not in ft_all:
        raise HTTPException(422, f"Classe '{body.type_fichier}' inconnue")

    # ── Registre ───────────────────────────────────────────────────────────────
    fichiers = load_registre()
    existing = next((f for f in fichiers if f["id"] == body.file_id), None)
    entry = {
        "id": body.file_id,
        "type_fichier": body.type_fichier,
        "chemin": body.path,
        "synchro_periodicite": body.synchro_periodicite,
        "owner_role": body.owner_role,
        "genere_par_script": False,
        "derniere_synchro": None,
        "statut_dernier_synchro": "importé",
    }
    if existing:
        fichiers = [entry if f["id"] == body.file_id else f for f in fichiers]
    else:
        fichiers.append(entry)
    save_registre(fichiers)

    # ── Tables dans tables.json ────────────────────────────────────────────────
    tables_data = load_tables()
    if "tables" not in tables_data:
        tables_data["tables"] = {}

    for tm in body.table_mappings:
        if not tm.include:
            continue
        tid = f"{body.file_id}.{tm.schema_name}"
        tables_data["tables"][tid] = {
            "id": tid,
            "file_id": body.file_id,
            "table_name": tm.schema_name,
            "sheet": tm.sheet,
            "description": f"Importé depuis {tm.excel_name}",
            "columns": [
                {
                    "name": c.name,
                    "col_type": c.col_type,
                    "header": c.header or c.name,
                    "write": "",
                    "is_key": c.is_key,
                    "required": c.required,
                    "description": "",
                }
                for c in tm.columns
            ],
        }

    save_tables(tables_data)

    # ── Mise à jour optionnelle des min_fields de la Classe ───────────────────
    if body.update_class_fields and body.field_mappings:
        import yaml
        from web.services.config_service import FILE_TYPES_PATH
        ft = ft_all[body.type_fichier]
        existing_names = {f["name"] for f in ft.get("min_fields", [])}
        for fm in body.field_mappings:
            if not fm.include or fm.field_name in existing_names:
                continue
            ft.setdefault("min_fields", []).append({
                "name": fm.field_name,
                "label": fm.named_range,
                "field_type": fm.field_type,
                "nature": fm.nature,
                "source": fm.source,
                "required": fm.required,
                "pushable": fm.pushable,
            })
        ft_all[body.type_fichier] = ft
        with open(FILE_TYPES_PATH, encoding="utf-8") as f:
            doc = yaml.safe_load(f) or {}
        doc["file_types"] = ft_all
        with open(FILE_TYPES_PATH, "w", encoding="utf-8") as f:
            yaml.dump(doc, f, allow_unicode=True, default_flow_style=False, sort_keys=False)

    # ── COLLECT dans hierarchy.json ───────────────────────────────────────────
    if body.collect_mappings:
        hierarchy = load_hierarchy()
        for cm in body.collect_mappings:
            hierarchy["collects"].append({
                "source_class": cm.source_class,
                "source_table": cm.source_table,
                "target_file_id": body.file_id,
                "target_table": cm.target_table,
                "where": cm.where_clause,
                "cols": cm.cols_filter,
            })
        save_hierarchy(hierarchy)

    return {
        "ok": True,
        "file_id": body.file_id,
        "tables_created": [t.schema_name for t in body.table_mappings if t.include],
        "collects_added": len(body.collect_mappings),
        "message": f"Post '{body.file_id}' enregistré ({body.type_fichier})",
    }
