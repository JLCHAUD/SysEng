"""
Moteur de synchronisation central.

Flux :
  1. Charger le registre
  2. Pour chaque fichier : tenter l'ouverture → skip si verrouillé
  3. Référentiels d'abord (push → store), puis instances, puis agrégation
  4. Générer le rapport de synchronisation
"""
import json
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from src import store as Store
from src import ecosystem as Ecosystem
from src.config_loader import load_registre, save_registre
from src.executor import execute_ast
from src.models import EntreeRegistre
from src.parser import parse_file
from src.parser import MANIFESTE_SHEET
from src.passerelle import executer_passerelle  # conservé pour rétro-compat (legacy _Passerelle)

ROOT = Path(__file__).parent.parent

ORDRE_TYPES = [
    "referentiel_uo",
    "referentiel_projet",
    "uo_instance",
    "cockpit",
    "pilote",
    "consolidation",
    "client",
]


# ─── Vérification de verrouillage ────────────────────────────────────────────

def _est_verrouille(chemin: Path) -> bool:
    """
    Tente d'ouvrir le fichier en mode exclusif.
    Retourne True si le fichier est verrouillé par Excel ou un autre process.
    """
    if not chemin.exists():
        return False
    try:
        with open(chemin, "r+b"):
            return False
    except (IOError, PermissionError):
        return True


# ─── Rapport de synchronisation ───────────────────────────────────────────────

def _generer_rapport(
    resultats: List[dict],
    debut: datetime,
    fin: datetime,
) -> Path:
    """Génère un rapport JSON de synchronisation."""
    rapport_dir = ROOT / "output" / "rapports"
    rapport_dir.mkdir(parents=True, exist_ok=True)
    nom = f"Rapport_Synchro_{debut.strftime('%Y%m%d_%H%M%S')}.json"
    chemin = rapport_dir / nom

    rapport = {
        "debut": debut.isoformat(timespec="seconds"),
        "fin": fin.isoformat(timespec="seconds"),
        "nb_total": len(resultats),
        "nb_ok": sum(1 for r in resultats if r["statut"] == "ok"),
        "nb_skip": sum(1 for r in resultats if r["statut"] == "skip_verrouille"),
        "nb_erreur": sum(1 for r in resultats if r["statut"] == "erreur"),
        "fichiers": resultats,
    }
    with open(chemin, "w", encoding="utf-8") as f:
        json.dump(rapport, f, ensure_ascii=False, indent=2)
    return chemin


# ─── Synchronisation d'un fichier ────────────────────────────────────────────

def _sync_fichier(entree: EntreeRegistre, force: bool = False) -> dict:
    """
    Synchronise un fichier individuel.
    Retourne un dict de résultat pour le rapport.
    """
    chemin = ROOT / entree.chemin
    log: List[str] = []
    resultat = {
        "id": entree.id,
        "chemin": entree.chemin,
        "type": entree.type_fichier,
        "statut": "ok",
        "log": log,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
    }

    if not chemin.exists():
        resultat["statut"] = "erreur"
        log.append(f"[ERREUR] Fichier introuvable : {chemin}")
        return resultat

    if _est_verrouille(chemin):
        resultat["statut"] = "skip_verrouille"
        log.append(f"[SKIP] Fichier ouvert/verrouillé : {chemin.name}")
        return resultat

    try:
        ast = parse_file(chemin)
        if ast is None:
            # Pas de feuille _Manifeste (ni _Passerelle) → fallback ancien moteur
            pushed = executer_passerelle(chemin, entree.id, log)
            log.append(f"[OK] {len(pushed)} variable(s) synchronisée(s) (legacy)")
        else:
            if ast.errors:
                for err in ast.errors:
                    log.append(f"[WARN] Parse L{err.line_num}: {err.message}")
            exec_result = execute_ast(ast, chemin, Store)
            log.extend(exec_result.errors)
            log.append(
                f"[OK] PULL={len(exec_result.pulled)} "
                f"PUSH={len(exec_result.pushed)} "
                f"BIND={len(exec_result.bound)} "
                f"skip={len(exec_result.skipped)} "
                f"err={len(exec_result.errors)}"
            )
            if exec_result.errors:
                resultat["statut"] = "erreur"

            # Mise à jour Exomap (M03)
            status_eco = "error" if exec_result.errors else "ok"
            Ecosystem.record_file_sync(
                file_id=entree.id,
                path_str=entree.chemin,
                file_type=entree.type_fichier,
                status=status_eco,
                manifest_version=ast.header.version,
            )
            Ecosystem.record_edges_from_ast(ast, file_id=entree.id)
    except Exception as e:
        resultat["statut"] = "erreur"
        log.append(f"[ERREUR] Exception inattendue : {e}")

    return resultat


# ─── Synchronisation complète ─────────────────────────────────────────────────

def synchroniser(
    ids: Optional[List[str]] = None,
    types: Optional[List[str]] = None,
    force: bool = False,
) -> Path:
    """
    Lance une synchronisation.

    Args:
        ids: liste d'IDs de fichiers à synchroniser (None = tous)
        types: liste de types à synchroniser (None = tous)
        force: ignorer la vérification de verrouillage

    Retourne le chemin du rapport généré.
    """
    debut = datetime.now()
    entrees = load_registre()

    # Filtrer par IDs ou types
    if ids:
        entrees = [e for e in entrees if e.id in ids]
    if types:
        entrees = [e for e in entrees if e.type_fichier in types]

    # Trier par ordre de traitement
    def _ordre(e: EntreeRegistre) -> int:
        try:
            return ORDRE_TYPES.index(e.type_fichier)
        except ValueError:
            return 99

    entrees.sort(key=_ordre)

    resultats = []
    for entree in entrees:
        print(f"  Synchro {entree.id} ({entree.type_fichier})...")
        res = _sync_fichier(entree, force=force)
        resultats.append(res)

        # Mettre à jour le registre
        entree.derniere_synchro = res["timestamp"]
        entree.statut_dernier_synchro = res["statut"]
        for line in res["log"]:
            print(f"    {line}")

    save_registre(entrees)

    fin = datetime.now()
    rapport_path = _generer_rapport(resultats, debut, fin)

    nb_ok = sum(1 for r in resultats if r["statut"] == "ok")
    nb_skip = sum(1 for r in resultats if r["statut"] == "skip_verrouille")
    nb_err = sum(1 for r in resultats if r["statut"] == "erreur")
    print(f"\nSynchro terminee : {nb_ok} OK / {nb_skip} skips / {nb_err} erreurs")
    print(f"Rapport : {rapport_path}")

    return rapport_path


# ─── Audit d'onboarding ───────────────────────────────────────────────────────

def auditer_fichier(chemin_str: str) -> dict:
    """
    Audit d'onboarding pour un fichier Excel créé manuellement.
    Vérifie la présence des feuilles requises, des tableaux, du manifeste.
    Génère un template _Manifeste vide si absent.
    """
    from openpyxl import load_workbook
    from src.passerelle import MANIFESTE_SHEET, COLONNES_PASSERELLE
    PASSERELLE_SHEET = MANIFESTE_SHEET  # alias local

    chemin = Path(chemin_str)
    rapport = {"chemin": chemin_str, "ok": True, "alertes": [], "actions": []}

    if not chemin.exists():
        rapport["ok"] = False
        rapport["alertes"].append("Fichier introuvable")
        return rapport

    try:
        wb = load_workbook(chemin)
    except Exception as e:
        rapport["ok"] = False
        rapport["alertes"].append(f"Impossible d'ouvrir : {e}")
        return rapport

    feuilles = wb.sheetnames
    rapport["feuilles_trouvees"] = feuilles

    if PASSERELLE_SHEET not in feuilles:
        rapport["alertes"].append(f"Feuille '{PASSERELLE_SHEET}' absente")
        rapport["actions"].append("Template _Manifeste genere")

        # Générer un template _Manifeste vide
        ws = wb.create_sheet(PASSERELLE_SHEET)
        ws["A1"] = "MANIFESTE_V=1"
        for col, nom in enumerate(COLONNES_PASSERELLE, 1):
            ws.cell(row=2, column=col, value=nom)
        wb.save(chemin)
    else:
        rapport["actions"].append("Feuille _Manifeste existante conservee")

    return rapport
