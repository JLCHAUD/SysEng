"""
creer_uo.py — Assembleur d'instanciation d'UO (brique C4).
===========================================================
Cree un fichier UO complet et synchronisable a partir du Catalogue :

    python projet_TrainSystem/creer_uo.py L09U1-CFL2400-CLIM \\
        --se "Jean Dujardin" --heures 200 \\
        [--projet "CFL 2400"] [--systeme "Climatisation"] [--output DIR]

Le code se decompose : L{lot}U{uo}-{PROJET}-{SYSTEME}
  L09U1  -> entree du Catalogue (activites, livrables, donnees d'entree)
  CFL2400 -> code projet (affiche dans General ; --projet pour le libelle)
  CLIM    -> code systeme (--systeme pour le libelle)

L'assembleur :
  1. lit Catalogue_UO_TrainSystem.xlsx (filtre uo_type)
  2. genere le classeur 11 feuilles avec Design XD (banniere 1 ligne)
  3. pre-remplit Activites / Livrables / Donnees_Entree depuis le catalogue
  4. genere le _Manifeste MXL avec les bonnes cles uo.<code>.*
"""
import argparse
import datetime
import re
import sys
from pathlib import Path

HERE = Path(__file__).parent
ROOT = HERE.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(HERE))

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from src.xl_design import XD

CATALOGUE = HERE / "Catalogue_UO_TrainSystem.xlsx"
RE_CODE = re.compile(r"^(L\d{2}U\d)(?:-([A-Za-z0-9]+))?(?:-([A-Za-z0-9]+))?$")

# Banniere XD occupe la ligne 1 (hauteur=30). Contenu demarre a la ligne T.
T = 2


# ─── Lecture du catalogue ─────────────────────────────────────────────────────

def read_table(ws, table_name):
    ref = ws.tables[table_name].ref
    rows = list(ws[ref])
    headers = [c.value for c in rows[0]]
    return [dict(zip(headers, (c.value for c in r))) for r in rows[1:]
            if any(c.value is not None for c in r)]


def load_catalogue(uo_type):
    if not CATALOGUE.exists():
        sys.exit(f"[ERR] Catalogue introuvable : {CATALOGUE}\n"
                 "      Lancez d'abord : python projet_TrainSystem/build_catalogue.py")
    wb = load_workbook(CATALOGUE, data_only=False)
    index = {r["uo_type"]: r for r in read_table(wb["Index"], "tbl_index")}
    if uo_type not in index:
        sys.exit(f"[ERR] UO type '{uo_type}' absent du catalogue. "
                 f"Disponibles : {', '.join(sorted(index))}")
    acts = [r for r in read_table(wb["Catalogue_Activites"], "tbl_cat_activites")
            if r["uo_type"] == uo_type]
    livs = [r for r in read_table(wb["Catalogue_Livrables"], "tbl_cat_livrables")
            if r["uo_type"] == uo_type]
    des = [r for r in read_table(wb["Catalogue_DonneesEntree"], "tbl_cat_donnees")
           if r["uo_type"] == uo_type]
    return index[uo_type], acts, livs, des


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _write_table(ws, name, headers, rows, widths=None, start_row=T, family="general", col_start=1):
    """Ecrit entetes + donnees via XD, enregistre le tableau Excel nomme."""
    XD.table_header(ws, start_row, headers, family, col_start=col_start)
    ws.row_dimensions[start_row].height = 22
    for ri, rd in enumerate(rows, start_row + 1):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=ri, column=col_start + ci - 1, value=rd.get(h))
            c.border = XD.HAIR
            c.font = XD.fnt(10)
        ws.row_dimensions[ri].height = 20
    last = start_row + max(len(rows), 1)
    start_col_letter = get_column_letter(col_start)
    end_col_letter = get_column_letter(col_start + len(headers) - 1)
    ref = f"{start_col_letter}{start_row}:{end_col_letter}{last}"
    XD.named_table(ws, name, ref, family)
    if widths:
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
    return start_row + 1, last  # (data_start_row, data_end_row)


def named(wb, name, sheet, cell):
    wb.defined_names[name] = DefinedName(
        name, attr_text=f"'{sheet}'!${cell[0]}${cell[1:]}")


def _dv(ws, type_, formula, rng):
    dv = DataValidation(type=type_, formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(rng)


def _kpi_card(ws, col, label, value, sub, border_color, value_color, row):
    """Carte KPI 4 lignes : cadre colore, label, grande valeur, sous-titre."""
    XD.card_border(ws, row, col, row + 3, col + 2, color=border_color)
    for rr in range(row, row + 4):
        for cc in range(col, col + 3):
            ws.cell(row=rr, column=cc).fill = XD.fill(XD.WHITE)
    XD.card_border(ws, row, col, row + 3, col + 2, color=border_color)
    lab = ws.cell(row=row + 1, column=col, value=label)
    lab.font = XD.fnt(9, bold=True, color=XD.GREY_D)
    lab.alignment = XD.center()
    ws.merge_cells(start_row=row + 1, start_column=col,
                   end_row=row + 1, end_column=col + 2)
    val = ws.cell(row=row + 2, column=col, value=value)
    val.font = XD.fnt(20, bold=True, color=value_color)
    val.alignment = XD.center()
    ws.merge_cells(start_row=row + 2, start_column=col,
                   end_row=row + 2, end_column=col + 2)
    s = ws.cell(row=row + 3, column=col, value=sub)
    s.font = XD.fnt(8.5, color=XD.GREY_D)
    s.alignment = Alignment(horizontal="center", vertical="top")
    ws.merge_cells(start_row=row + 3, start_column=col,
                   end_row=row + 3, end_column=col + 2)


# ─── Construction de l'instance ───────────────────────────────────────────────

def build_instance(code, uo_type, projet_code, systeme_code, args):
    info, cat_acts, cat_livs, cat_des = load_catalogue(uo_type)
    wb = Workbook()

    uo_title = f"UO {uo_type} — {info['uo_libelle']}"
    parts = [args.projet or projet_code or "", args.systeme or systeme_code or ""]
    proj_line = "  ·  ".join(p for p in parts if p)
    se = args.se or ""

    # ── General ──────────────────────────────────────────────────────────────
    # Col A = marge etroite ; cols B,C = libelle / valeur
    ws = wb.active
    ws.title = "General"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 52
    XD.banner(ws, "general", uo_title, subtitle=proj_line, se=se, n_cols=8)
    kv = [
        ("Pôle",                   "Train Système"),
        ("Lot",                    f"LOT {info['lot']} — {info['lot_libelle']}"),
        ("UO",                     f"UO{uo_type[-1]} — {info['uo_libelle']}"),
        ("Libellé",                info["lot_libelle"]),
        ("Projet",                 args.projet or projet_code or ""),
        ("Système",                args.systeme or systeme_code or ""),
        ("Ingénieur Système (SE)", se),
        ("Heures vendues",         args.heures),
        ("Date de création",       datetime.date.today().isoformat()),
    ]
    for i, (k, v) in enumerate(kv, T + 1):
        ws.cell(row=i, column=2, value=k).font = XD.fnt(10, bold=True, color=XD.NAVY_D)
        ws.cell(row=i, column=3, value=v).font = XD.fnt(10)
        ws.row_dimensions[i].height = 20
    # "Heures vendues" = index 7 (0-based) → ligne T+1+7 = T+8, colonne C
    named(wb, "heures_vendues", "General", f"C{T + 8}")

    # ── Description_Besoin ───────────────────────────────────────────────────
    # Col A = marge etroite ; col B = texte large
    ws = wb.create_sheet("Description_Besoin")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 108
    XD.banner(ws, "description", uo_title, subtitle=proj_line, se=se, n_cols=6)
    r = T + 2
    for act in cat_acts:
        c = ws.cell(row=r, column=2,
                    value=f"{act['id']} : {act['designation']}")
        c.font = XD.fnt(11, bold=True, color=XD.NAVY_D)
        r += 1
        for line in (act.get("description") or "").split("\n"):
            if line.strip():
                c = ws.cell(row=r, column=2, value="   " + line)
                c.font = XD.fnt(10)
                c.alignment = Alignment(wrap_text=True)
                r += 1
        r += 1
    if info.get("criteres_acceptation"):
        ws.cell(row=r, column=2,
                value="Critères d'acceptation").font = XD.fnt(11, bold=True,
                                                               color=XD.NAVY_D)
        r += 1
        for line in str(info["criteres_acceptation"]).split("\n"):
            ws.cell(row=r, column=2, value="   " + line).font = XD.fnt(10)
            r += 1

    # ── Donnees_Entree ───────────────────────────────────────────────────────
    ws = wb.create_sheet("Donnees_Entree")
    ws.sheet_view.showGridLines = False
    XD.banner(ws, "donnees_entree", uo_title, subtitle=proj_line, se=se, n_cols=10)
    _write_table(ws, "tbl_donnees_entree",
        ["id", "designation", "type", "origine", "pic", "statut",
         "date_reception", "maturite", "date_update", "commentaire"],
        [{"id": f"DE-{i:03d}", "designation": d["designation"],
          "statut": "EN_ATTENTE"}
         for i, d in enumerate(cat_des, 1)],
        widths={"A": 9, "B": 50, "C": 10, "D": 10, "F": 12, "G": 13,
                "H": 9, "I": 12, "J": 28},
        family="donnees_entree")
    _dv(ws, "list", '"NA,EN_ATTENTE,RECUE"', f"F{T+1}:F{T+80}")

    # ── Activites ────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Activites")
    ws.sheet_view.showGridLines = False
    XD.banner(ws, "activites", uo_title, subtitle=proj_line, se=se, n_cols=10)
    act_headers = ["id", "designation", "applicable", "poids", "heures_allouees",
                   "statut", "avancement", "heures_consommees", "reste_a_faire",
                   "commentaire"]
    act_rows = [{"id": a["id"], "designation": a["designation"],
                 "applicable": "OUI", "poids": a.get("poids") or 1,
                 "statut": "A_FAIRE", "avancement": 0, "heures_consommees": 0}
                for a in cat_acts]
    ws.column_dimensions["A"].width = 2.5
    dr_s, dr_e = _write_table(ws, "tbl_activites", act_headers, act_rows,
                               widths={"A": 2.5, "B": 10, "C": 46, "D": 11, "E": 8,
                                       "F": 15, "G": 11, "H": 12, "I": 17, "J": 13,
                                       "K": 32},
                               family="activites",
                               col_start=2)
    n = len(act_rows)
    for r in range(dr_s, dr_e + 1):
        ws.cell(row=r, column=6).value = (
            f'=IF(D{r}="OUI",E{r}/SUMIF($D${dr_s}:$D${dr_e},"OUI"'
            f',$E${dr_s}:$E${dr_e})*heures_vendues,0)')
        ws.cell(row=r, column=10).value = f"=(1-H{r}/100)*F{r}"
        ws.cell(row=r, column=6).number_format = "0.00"
        ws.cell(row=r, column=10).number_format = "0.00"
    _dv(ws, "list", '"OUI,NON"', f"D{T+1}:D{T+60}")
    _dv(ws, "list", '"A_FAIRE,EN_COURS,TERMINEE,STAND_BY"', f"G{T+1}:G{T+60}")
    ws.conditional_formatting.add(
        f"H{T+1}:H{T+n}",
        DataBarRule(start_type="num", start_value=0, end_type="num",
                    end_value=100, color=XD.sheet("activites").header, showValue=True))
    XD.statut_cf(ws, f"G{T+1}:G{T+n}")
    XD.health_spine(ws, "activites", header_row=T, row_start=T+1, row_end=dr_e, status_col=7)
    ws.freeze_panes = f"B{T+1}"

    # ── Livrables ─────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Livrables")
    ws.sheet_view.showGridLines = False
    XD.banner(ws, "livrables", uo_title, subtitle=proj_line, se=se, n_cols=10)
    liv_rows = [{"id": f"LIV-{i:03d}", "designation": l["designation"],
                 "statut": "A_FAIRE"}
                for i, l in enumerate(cat_livs, 1)]
    _write_table(ws, "tbl_livrables",
        ["id", "designation", "type", "maturite_attendue", "date_attendue",
         "pic", "statut", "date_revisee", "date_livraison", "commentaire"],
        liv_rows,
        widths={"A": 9, "B": 50, "C": 8, "D": 16, "E": 14, "F": 10,
                "G": 11, "H": 13, "I": 13, "J": 28},
        family="livrables")
    _dv(ws, "list", '"A_FAIRE,EN_COURS,LIVRE,VALIDE"', f"G{T+1}:G{T+60}")
    XD.statut_cf(ws, f"G{T+1}:G{T+len(liv_rows)}")

    # ── OIL ──────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("OIL")
    ws.sheet_view.showGridLines = False
    XD.banner(ws, "oil", uo_title, subtitle=proj_line, se=se, n_cols=10)
    _write_table(ws, "tbl_oil",
        ["id", "titre", "description", "en_action", "domaine", "criticite",
         "statut", "date_ouverture", "date_besoin", "journal"],
        [{"id": "PO-000", "titre": "(exemple — remplacer par ton premier point)",
          "description": "", "en_action": "SE", "criticite": "BASSE",
          "statut": "CLOS", "journal": ""}],
        widths={"A": 8, "B": 34, "C": 44, "D": 13, "E": 12, "F": 11,
                "G": 9, "H": 13, "I": 12, "J": 50},
        family="oil")
    _dv(ws, "list", '"SE,FOURNISSEUR,EXPERT,AT,CLIENT,AUTRE"', f"D{T+1}:D{T+60}")
    _dv(ws, "list", '"BASSE,MOYENNE,HAUTE"', f"F{T+1}:F{T+60}")
    _dv(ws, "list", '"OUVERT,CLOS"', f"G{T+1}:G{T+60}")
    XD.criticite_cf(ws, f"F{T+1}:F{T+60}")
    ws.conditional_formatting.add(f"G{T+1}:G{T+60}", CellIsRule(
        operator="equal", formula=['"OUVERT"'],
        fill=XD.fill(XD.RED_L), font=XD.fnt(10, bold=True, color=XD.RED_D)))
    ws.conditional_formatting.add(f"G{T+1}:G{T+60}", CellIsRule(
        operator="equal", formula=['"CLOS"'],
        fill=XD.fill(XD.GREEN_L), font=XD.fnt(10, color=XD.GREEN_D)))

    # ── KPI ──────────────────────────────────────────────────────────────────
    # Col A = marge ; cols B,C,D = libelle / valeur / description
    ws = wb.create_sheet("KPI")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 64
    XD.banner(ws, "kpi", uo_title, subtitle=proj_line, se=se, n_cols=8)

    XD.section_box(ws, "KPI calculés par ExoSync (ne pas saisir manuellement)",
                   T, 2, T + 7, 4, "kpi")
    exo_kpis = [
        ("Avancement UO (%)",           "kpi_avancement",     "0.00",
         "Moyenne des avancements pondérée par les poids (activités applicables)"),
        ("Heures consommées",           "kpi_h_conso",        "0.00",
         "Somme des heures consommées (activités applicables)"),
        ("Points ouverts",              "kpi_po_ouverts",     "0",
         "Nombre de points OIL au statut OUVERT"),
        ("Points fermés",               "kpi_po_fermes",      "0",
         "Nombre de points OIL au statut CLOS"),
        ("Points critiques ouverts",    "kpi_po_critiques",   "0",
         "Points OUVERTS avec criticite = HAUTE"),
        ("Dont balle chez fournisseur", "kpi_po_fournisseur", "0",
         "Points OUVERTS avec en_action = FOURNISSEUR"),
        ("Dont balle chez expert",      "kpi_po_expert",      "0",
         "Points OUVERTS avec en_action = EXPERT"),
    ]
    r = T + 1
    for label, nm, fmt, desc in exo_kpis:
        ws.cell(row=r, column=2, value=label).font = XD.fnt(10, bold=True,
                                                             color=XD.NAVY_D)
        named(wb, nm, "KPI", f"C{r}")   # valeur en colonne C
        ws.cell(row=r, column=3).number_format = fmt
        ws.cell(row=r, column=4, value=desc).font = XD.fnt(9, color=XD.GREY_D,
                                                            italic=True)
        ws.row_dimensions[r].height = 20
        r += 1
    # r = T+8 ; ligne T+8 = espaceur ; section Excel demarre a T+9
    r += 1
    XD.section_box(ws,
                   "Indicateurs Excel (formules — s'actualisent à l'ouverture)",
                   r, 2, r + 5, 4, "kpi")
    r += 1
    excel_kpis = [
        ("Heures vendues",              "kpi_h_vendues",
         "=heures_vendues",            "0",
         "Lu depuis General (formule Excel)"),
        ("Reste à faire total (h)",     "kpi_raf",
         f"=SUM(Activites!J{T+1}:J{T+60})", "0.00",
         "Somme colonne reste_a_faire (formule Excel)"),
        ("Heures estimées à terminaison (EAC)", "kpi_eac",
         "=kpi_h_conso+kpi_raf",      "0.00",
         "EAC = consommé + reste à faire (formule Excel)"),
        ("Dérive à terminaison (h)",   "kpi_derive",
         "=kpi_eac-kpi_h_vendues",    "0.00",
         "EAC − heures vendues : >0 = dépassement prévu (formule Excel)"),
        ("Santé",                      "kpi_sante",
         '=IF(kpi_po_critiques>0,"ROUGE",IF(kpi_po_ouverts>0,"ORANGE","VERT"))',
         "General",
         "ROUGE si point critique · ORANGE si point ouvert · VERT sinon"),
    ]
    for label, nm, formula, fmt, desc in excel_kpis:
        ws.cell(row=r, column=2, value=label).font = XD.fnt(10, bold=True,
                                                             color=XD.NAVY_D)
        ws.cell(row=r, column=3, value=formula)
        ws.cell(row=r, column=3).number_format = fmt
        named(wb, nm, "KPI", f"C{r}")   # valeur en colonne C
        ws.cell(row=r, column=4, value=desc).font = XD.fnt(9, color=XD.GREY_D,
                                                            italic=True)
        ws.row_dimensions[r].height = 20
        r += 1

    # ── Dashboard ─────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    for i in range(2, 18):
        ws.column_dimensions[get_column_letter(i)].width = 8.5
    XD.banner(ws, "dashboard", uo_title, subtitle=proj_line, se=se, n_cols=16)

    ROW1, ROW2 = T + 1, T + 6
    for rr in range(ROW1, ROW1 + 4):
        ws.row_dimensions[rr].height = 18
    for rr in range(ROW2, ROW2 + 4):
        ws.row_dimensions[rr].height = 18

    cards_r1 = [
        (2,  "AVANCEMENT UO",  '=ROUND(kpi_avancement,1)&" %"',
         "pondéré par poids",   "B5D4F4", XD.NAVY_D),
        (6,  "HEURES",
         '=ROUND(kpi_h_conso,0)&" / "&heures_vendues&" h"',
         "consommées",          "D3D1C7", "2C2C2A"),
        (10, "POINTS OUVERTS", "=kpi_po_ouverts",
         "au statut OUVERT",    "FAC775", XD.AMBER_D),
        (14, "PTS CRITIQUES",  "=kpi_po_critiques",
         "criticité HAUTE",     "FCEBEB", XD.RED_D),
    ]
    cards_r2 = [
        (2,  "SANTÉ",          "=kpi_sante",
         "",                    "F5F4F0", XD.NAVY_D),
        (6,  "EAC (h)",        "=ROUND(kpi_eac,0)",
         "estimé à terminaison","D3D1C7", "2C2C2A"),
        (10, "DÉRIVE (h)",     "=ROUND(kpi_derive,0)",
         "EAC − heures vendues","F5F4F0", XD.GREY_D),
        (14, "BALLE FOURN.",   "=kpi_po_fournisseur",
         "pts chez fournisseur","FAEEDA", XD.AMBER_D),
    ]
    for col, label, value, sub, bc, vc in cards_r1:
        _kpi_card(ws, col, label, value, sub, bc, vc, row=ROW1)
    for col, label, value, sub, bc, vc in cards_r2:
        _kpi_card(ws, col, label, value, sub, bc, vc, row=ROW2)

    note_row = ROW2 + 5
    note = ws.cell(row=note_row, column=2,
                   value="Cette feuille est libre — l'ingénieur la complète "
                         "(elle ne lit que l'onglet KPI).")
    note.font = XD.fnt(9, color=XD.GREY_D, italic=True)
    ws.merge_cells(start_row=note_row, start_column=2,
                   end_row=note_row, end_column=16)

    # ── Planning ──────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Planning")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    XD.banner(ws, "planning", uo_title, subtitle=proj_line, se=se, n_cols=10)
    ws.cell(row=T + 2, column=2,
            value="Réservé : visualisation planning.").font = \
        XD.fnt(10, color=XD.GREY_D, italic=True)

    # ── Orga ──────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Orga")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    XD.banner(ws, "orga", uo_title, subtitle=proj_line, se=se, n_cols=10)
    ws.cell(row=T + 2, column=2,
            value="Organisation projet (à renseigner depuis le Catalogue Projets).").font = \
        XD.fnt(10, color=XD.GREY_D, italic=True)

    # ── _Manifeste ────────────────────────────────────────────────────────────
    # Col A = instruction MXL (lue par le parser), col C = commentaire francais
    # (ignore par le parser, sert a lire ce que fait chaque instruction).
    ws = wb.create_sheet("_Manifeste")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["C"].width = 60
    ws.sheet_properties.tabColor = XD.sheet("manifeste").header
    ws["A1"] = "MANIFESTE_V=1"
    ws["A1"].font = XD.fnt(11, bold=True, color=XD.NAVY_D)
    ws["C1"] = "Commentaire (non interprété — aide à la lecture)"
    ws["C1"].font = XD.fnt(9, color=XD.GREY_D, italic=True)
    ws["A2"] = "instruction"
    ws["A2"].font = XD.fnt(10, bold=True)

    # (instruction, commentaire). Section/ligne vide → commentaire "".
    se_name = getattr(args, "se", "") if args else ""
    manifeste = [
        ("FILE_TYPE: uo_instance",                 "Type de fichier dans l'écosystème ExoSync"),
        (f"FILE_ID: {code}",                       "Identifiant unique de cette UO (clé du store)"),
        (f"ingenieur: {se_name}" if se_name else "", "Ingénieur Système responsable de cette UO"),
        ("", ""),
        ("# ── Lecture des donnees locales ─────────────────────────────", ""),
        ("DEF $act = GET_TABLE(Activites, tbl_activites)",  "Charge le tableau des activités"),
        ("DEF $oil = GET_TABLE(OIL, tbl_oil)",              "Charge l'Open Items List (points ouverts)"),
        ("DEF $liv = GET_TABLE(Livrables, tbl_livrables)",  "Charge le tableau des livrables"),
        ("", ""),
        ("# ── Sous-ensembles ──────────────────────────────────────────", ""),
        ('DEF $actifs = COMPUTE(FILTER($act, applicable = "OUI"))',
         "Active uniquement les activités applicables (applicable = OUI)"),
        ('DEF $po_ouv = COMPUTE(FILTER($oil, statut = "OUVERT"))',
         "Sous-ensemble des points ouverts encore OUVERTS"),
        ("", ""),
        ("# ── KPI ─────────────────────────────────────────────────────", ""),
        ("DEF $avancement = COMPUTE(MEAN_WEIGHTED($actifs.avancement, $actifs.poids))",
         "Avancement global : moyenne pondérée par les poids des activités applicables"),
        ("DEF $h_conso = COMPUTE(SUM($actifs.heures_consommees))",
         "Total des heures consommées sur les activités applicables"),
        ('DEF $po_ouverts = COMPUTE(COUNT_IF($oil.statut, "OUVERT"))',
         "Nombre de points OIL au statut OUVERT"),
        ('DEF $po_fermes = COMPUTE(COUNT_IF($oil.statut, "CLOS"))',
         "Nombre de points OIL au statut CLOS"),
        ('DEF $po_critiques = COMPUTE(COUNT_IF($po_ouv.criticite, "HAUTE"))',
         "Points ouverts de criticité HAUTE"),
        ('DEF $po_fournisseur = COMPUTE(COUNT_IF($po_ouv.en_action, "FOURNISSEUR"))',
         "Points ouverts dont la balle est chez le fournisseur"),
        ('DEF $po_expert = COMPUTE(COUNT_IF($po_ouv.en_action, "EXPERT"))',
         "Points ouverts dont la balle est chez l'expert"),
        ("", ""),
        ("# ── Regles de qualite ───────────────────────────────────────", ""),
        ("VALIDATE $actifs.avancement : RANGE(0, 100)",
         "Bloque la sync si un avancement sort de l'intervalle 0-100"),
        ('VALIDATE $act.applicable : IN("OUI", "NON")',
         "La colonne applicable ne peut valoir que OUI ou NON"),
        ('VALIDATE $oil.statut : IN("OUVERT", "CLOS")',
         "Le statut OIL ne peut valoir que OUVERT ou CLOS"),
        ("", ""),
        ("# ── Affichage local (onglet KPI) ────────────────────────────", ""),
        ("BIND $avancement -> KPI.kpi_avancement",       "Écrit l'avancement dans l'onglet KPI"),
        ("BIND $h_conso -> KPI.kpi_h_conso",             "Écrit les heures consommées dans KPI"),
        ("BIND $po_ouverts -> KPI.kpi_po_ouverts",       "Écrit le nombre de points ouverts dans KPI"),
        ("BIND $po_fermes -> KPI.kpi_po_fermes",         "Écrit le nombre de points fermés dans KPI"),
        ("BIND $po_critiques -> KPI.kpi_po_critiques",   "Écrit le nombre de points critiques dans KPI"),
        ("BIND $po_fournisseur -> KPI.kpi_po_fournisseur", "Écrit le compteur 'balle fournisseur' dans KPI"),
        ("BIND $po_expert -> KPI.kpi_po_expert",         "Écrit le compteur 'balle expert' dans KPI"),
        ("", ""),
        ("# ── Publication vers l'ecosysteme ───────────────────────────", ""),
        (f"PUSH $avancement -> uo.{code}.avancement",        "Publie l'avancement vers le store central"),
        (f"PUSH $h_conso -> uo.{code}.heures_consommees",    "Publie les heures consommées vers le store"),
        (f"PUSH $po_ouverts -> uo.{code}.po_ouverts",        "Publie le nombre de points ouverts"),
        (f"PUSH $po_fermes -> uo.{code}.po_fermes",          "Publie le nombre de points fermés"),
        (f"PUSH $po_critiques -> uo.{code}.po_critiques",    "Publie le nombre de points critiques"),
        (f"PUSH $actifs -> uo.{code}.activites",             "Publie la table des activités applicables"),
        (f"PUSH $oil -> uo.{code}.points_ouverts",           "Publie la table des points ouverts (OIL)"),
        (f"PUSH $liv -> uo.{code}.livrables",                "Publie la table des livrables"),
    ]
    for i, (instr, comment) in enumerate(manifeste, 3):
        ws.cell(row=i, column=1, value=instr)
        if comment:
            ws.cell(row=i, column=3, value=comment).font = XD.fnt(9, color=XD.GREY_D, italic=True)

    return wb


def main():
    p = argparse.ArgumentParser(description="Cree une UO depuis le Catalogue")
    p.add_argument("code", help="ex: L09U1-CFL2400-CLIM")
    p.add_argument("--se", default="", help="Nom de l'ingenieur systeme")
    p.add_argument("--heures", type=float, default=0, help="Heures vendues")
    p.add_argument("--projet", default="", help="Libelle du projet")
    p.add_argument("--systeme", default="", help="Libelle du systeme")
    p.add_argument("--output", default=str(HERE), help="Repertoire de sortie")
    args = p.parse_args()

    m = RE_CODE.match(args.code)
    if not m:
        sys.exit(f"[ERR] Code invalide '{args.code}' — attendu : L09U1-PROJET-SYSTEME")
    uo_type, projet_code, systeme_code = m.groups()

    wb = build_instance(args.code, uo_type, projet_code, systeme_code, args)
    out = Path(args.output) / f"{args.code}.xlsx"
    wb.save(out)
    print(f"[OK] {out}")
    print(f"     UO type {uo_type} — SE: {args.se or '<a assigner>'} — "
          f"{args.heures:g} h vendues")
    print(f"     Synchroniser : python scripts/valider_un.py "
          f"{out.relative_to(ROOT) if out.is_relative_to(ROOT) else out}")


if __name__ == "__main__":
    main()
