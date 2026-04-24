"""Generates the central consolidation Excel file."""
from datetime import date
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule

from src.models import UOInstance
from src.styles import (
    BLUE_DARK, BLUE_MID, BLUE_LIGHT, GREEN_LIGHT, ORANGE_LIGHT, RED_LIGHT,
    GREY_LIGHT, WHITE, YELLOW_LIGHT,
    THIN_BORDER, solid_fill, header_font, body_font, center, left,
    style_header_row, style_data_row, set_column_widths, freeze_top_row,
)

OUTPUT_DIR = Path(__file__).parent.parent.parent / "output"


def _sheet_vue_globale(wb: Workbook, instances: List[UOInstance]):
    ws = wb.create_sheet("Vue Globale")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = f"Consolidation Globale — {date.today().strftime('%d/%m/%Y')}"
    t.fill = solid_fill(BLUE_DARK)
    t.font = header_font(size=13)
    t.alignment = center()
    ws.row_dimensions[1].height = 28

    headers = ["UO ID", "Ingénieur", "Type UO", "Système", "Projet",
               "Charge (h)", "% Avancement", "H réalisées", "Date fin", "Points ouverts"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, 1, 10, color=BLUE_MID)

    for i, uo in enumerate(instances):
        row = 3 + i
        type_name = uo.uo_type.name if uo.uo_type else uo.uo_type_id
        sys_name = uo.system.name if uo.system else uo.system_id
        proj_name = uo.project.name if uo.project else uo.project_id

        ws.cell(row=row, column=1, value=uo.id)
        ws.cell(row=row, column=2, value=uo.engineer_name)
        ws.cell(row=row, column=3, value=type_name)
        ws.cell(row=row, column=4, value=sys_name)
        ws.cell(row=row, column=5, value=proj_name)
        ws.cell(row=row, column=6, value=uo.total_hours)
        ws.cell(row=row, column=7, value=0.0).number_format = "0%"
        ws.cell(row=row, column=8, value=0.0)
        date_cell = ws.cell(row=row, column=9, value=uo.end_date)
        date_cell.number_format = "DD/MM/YYYY"
        ws.cell(row=row, column=10, value=0)
        style_data_row(ws, row, 1, 10, alternate=(i % 2 == 1))

    # Conditional formatting on % Avancement
    last = 2 + len(instances)
    ws.conditional_formatting.add(
        f"G3:G{last}",
        CellIsRule(operator="greaterThanOrEqual", formula=["1"], fill=solid_fill(GREEN_LIGHT)),
    )
    ws.conditional_formatting.add(
        f"G3:G{last}",
        CellIsRule(operator="between", formula=["0.5", "0.99"], fill=solid_fill(YELLOW_LIGHT)),
    )
    ws.conditional_formatting.add(
        f"G3:G{last}",
        CellIsRule(operator="lessThan", formula=["0.5"], fill=solid_fill(ORANGE_LIGHT)),
    )

    set_column_widths(ws, {
        "A": 12, "B": 22, "C": 28, "D": 20, "E": 25,
        "F": 12, "G": 15, "H": 15, "I": 14, "J": 16,
    })
    freeze_top_row(ws)


def _sheet_par_ingenieur(wb: Workbook, instances: List[UOInstance]):
    ws = wb.create_sheet("Par Ingénieur")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "Synthèse par Ingénieur"
    t.fill = solid_fill(BLUE_DARK)
    t.font = header_font(size=13)
    t.alignment = center()
    ws.row_dimensions[1].height = 28

    engineers = sorted(set(uo.engineer_name for uo in instances))
    current_row = 3

    for eng in engineers:
        eng_uo = [uo for uo in instances if uo.engineer_name == eng]

        # Engineer header
        ws.merge_cells(f"A{current_row}:G{current_row}")
        eng_cell = ws[f"A{current_row}"]
        eng_cell.value = f"▶  {eng}"
        eng_cell.fill = solid_fill(BLUE_MID)
        eng_cell.font = header_font()
        eng_cell.alignment = left()
        current_row += 1

        # Sub-headers
        sub_headers = ["UO ID", "Type", "Système", "Projet", "Charge (h)", "% Avancement", "Date fin"]
        for col, h in enumerate(sub_headers, 1):
            ws.cell(row=current_row, column=col, value=h)
        style_header_row(ws, current_row, 1, 7, color=BLUE_LIGHT)
        for col in range(1, 8):
            ws.cell(row=current_row, column=col).font = body_font(bold=True, color="000000")
        current_row += 1

        for i, uo in enumerate(eng_uo):
            type_name = uo.uo_type.name if uo.uo_type else uo.uo_type_id
            sys_name = uo.system.name if uo.system else uo.system_id
            proj_name = uo.project.name if uo.project else uo.project_id

            ws.cell(row=current_row, column=1, value=uo.id)
            ws.cell(row=current_row, column=2, value=type_name)
            ws.cell(row=current_row, column=3, value=sys_name)
            ws.cell(row=current_row, column=4, value=proj_name)
            ws.cell(row=current_row, column=5, value=uo.total_hours)
            ws.cell(row=current_row, column=6, value=0.0).number_format = "0%"
            date_cell = ws.cell(row=current_row, column=7, value=uo.end_date)
            date_cell.number_format = "DD/MM/YYYY"
            style_data_row(ws, current_row, 1, 7, alternate=(i % 2 == 1))
            current_row += 1

        # Totals row
        total_h = sum(uo.total_hours for uo in eng_uo)
        ws.merge_cells(f"A{current_row}:D{current_row}")
        total_label = ws[f"A{current_row}"]
        total_label.value = f"Total {eng}"
        total_label.fill = solid_fill(BLUE_LIGHT)
        total_label.font = body_font(bold=True)
        total_label.alignment = left()
        total_label.border = THIN_BORDER

        total_h_cell = ws.cell(row=current_row, column=5, value=total_h)
        total_h_cell.fill = solid_fill(BLUE_LIGHT)
        total_h_cell.font = body_font(bold=True)
        total_h_cell.border = THIN_BORDER
        total_h_cell.alignment = center()

        for col in [6, 7]:
            c = ws.cell(row=current_row, column=col)
            c.fill = solid_fill(BLUE_LIGHT)
            c.border = THIN_BORDER

        current_row += 2

    set_column_widths(ws, {"A": 12, "B": 28, "C": 20, "D": 25, "E": 12, "F": 15, "G": 14})


def _sheet_par_projet(wb: Workbook, instances: List[UOInstance]):
    ws = wb.create_sheet("Par Projet")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "Synthèse par Projet"
    t.fill = solid_fill(BLUE_DARK)
    t.font = header_font(size=13)
    t.alignment = center()
    ws.row_dimensions[1].height = 28

    projects = sorted(set(uo.project_id for uo in instances))
    current_row = 3

    for proj_id in projects:
        proj_uo = [uo for uo in instances if uo.project_id == proj_id]
        proj_name = proj_uo[0].project.name if proj_uo[0].project else proj_id

        ws.merge_cells(f"A{current_row}:G{current_row}")
        proj_cell = ws[f"A{current_row}"]
        proj_cell.value = f"▶  {proj_name}"
        proj_cell.fill = solid_fill(BLUE_MID)
        proj_cell.font = header_font()
        proj_cell.alignment = left()
        current_row += 1

        sub_headers = ["UO ID", "Ingénieur", "Type", "Système", "Charge (h)", "% Avancement", "Date fin"]
        for col, h in enumerate(sub_headers, 1):
            ws.cell(row=current_row, column=col, value=h)
        style_header_row(ws, current_row, 1, 7, color=BLUE_LIGHT)
        for col in range(1, 8):
            ws.cell(row=current_row, column=col).font = body_font(bold=True, color="000000")
        current_row += 1

        for i, uo in enumerate(proj_uo):
            type_name = uo.uo_type.name if uo.uo_type else uo.uo_type_id
            sys_name = uo.system.name if uo.system else uo.system_id

            ws.cell(row=current_row, column=1, value=uo.id)
            ws.cell(row=current_row, column=2, value=uo.engineer_name)
            ws.cell(row=current_row, column=3, value=type_name)
            ws.cell(row=current_row, column=4, value=sys_name)
            ws.cell(row=current_row, column=5, value=uo.total_hours)
            ws.cell(row=current_row, column=6, value=0.0).number_format = "0%"
            date_cell = ws.cell(row=current_row, column=7, value=uo.end_date)
            date_cell.number_format = "DD/MM/YYYY"
            style_data_row(ws, current_row, 1, 7, alternate=(i % 2 == 1))
            current_row += 1

        total_h = sum(uo.total_hours for uo in proj_uo)
        ws.merge_cells(f"A{current_row}:D{current_row}")
        total_label = ws[f"A{current_row}"]
        total_label.value = f"Total {proj_name}"
        total_label.fill = solid_fill(BLUE_LIGHT)
        total_label.font = body_font(bold=True)
        total_label.alignment = left()
        total_label.border = THIN_BORDER

        total_h_cell = ws.cell(row=current_row, column=5, value=total_h)
        total_h_cell.fill = solid_fill(BLUE_LIGHT)
        total_h_cell.font = body_font(bold=True)
        total_h_cell.border = THIN_BORDER
        total_h_cell.alignment = center()

        for col in [6, 7]:
            c = ws.cell(row=current_row, column=col)
            c.fill = solid_fill(BLUE_LIGHT)
            c.border = THIN_BORDER

        current_row += 2

    set_column_widths(ws, {"A": 12, "B": 22, "C": 28, "D": 20, "E": 12, "F": 15, "G": 14})


def _sheet_points_ouverts(wb: Workbook, instances: List[UOInstance]):
    ws = wb.create_sheet("Points Ouverts")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "Suivi consolidé des Points Ouverts"
    t.fill = solid_fill(BLUE_DARK)
    t.font = header_font(size=13)
    t.alignment = center()
    ws.row_dimensions[1].height = 28

    headers = ["UO ID", "Ingénieur", "Projet", "ID Point", "Description", "Statut", "Responsable"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, 1, 7, color=BLUE_MID)

    note_row = 3
    ws.merge_cells(f"A{note_row}:G{note_row}")
    note = ws[f"A{note_row}"]
    note.value = "ℹ️  Cette feuille sera alimentée automatiquement lors de la consolidation (phase 2)"
    note.fill = solid_fill(YELLOW_LIGHT)
    note.font = body_font(bold=True)
    note.alignment = center()
    note.border = THIN_BORDER

    set_column_widths(ws, {"A": 12, "B": 22, "C": 25, "D": 12, "E": 45, "F": 14, "G": 22})
    freeze_top_row(ws)


def generate_consolidation(instances: List[UOInstance], output_dir: Path = OUTPUT_DIR) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    _sheet_vue_globale(wb, instances)
    _sheet_par_ingenieur(wb, instances)
    _sheet_par_projet(wb, instances)
    _sheet_points_ouverts(wb, instances)

    filepath = output_dir / "consolidation.xlsx"
    wb.save(filepath)
    return filepath
