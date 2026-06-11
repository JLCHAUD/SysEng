"""Construit les fichiers de test des marches 1-4 EXACTEMENT comme decrits
dans le Parcours de Validation, pour verifier que les resultats annonces
sont corrects. (Script de verification, pas destine a l'utilisateur final.)"""
import sys
from pathlib import Path
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo

VDIR = ROOT / "validation"

def add_table(ws, name, headers, rows, r0=1, c0=1):
    for ci, h in enumerate(headers, c0):
        ws.cell(row=r0, column=ci, value=h)
    for ri, rd in enumerate(rows, r0 + 1):
        for ci, h in enumerate(headers, c0):
            ws.cell(row=ri, column=ci, value=rd.get(h, ""))
    ref = f"{get_column_letter(c0)}{r0}:{get_column_letter(c0+len(headers)-1)}{r0+len(rows)}"
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(t)

def manifeste(ws, lines):
    ws["A1"] = "MANIFESTE_V=1"
    ws["A2"] = "instruction"
    for i, l in enumerate(lines, 3):
        ws.cell(row=i, column=1, value=l)

# m1_source.xlsx
wb = Workbook(); ws = wb.active; ws.title = "Activites"
add_table(ws, "tbl_act", ["id", "libelle", "avancement"],
          [{"id": "A1", "libelle": "Analyse fonctionnelle", "avancement": 80},
           {"id": "A2", "libelle": "Redaction specification", "avancement": 60}])
manifeste(wb.create_sheet("_Manifeste"), [
    "FILE_TYPE: uo_instance", "FILE_ID: TEST-SOURCE", "",
    "DEF $act = GET_TABLE(Activites, tbl_act)", "",
    "PUSH $act -> test.source.activites"])
wb.save(VDIR / "m1_source.xlsx")

# m4_source2.xlsx
wb = Workbook(); ws = wb.active; ws.title = "Activites"
add_table(ws, "tbl_act2", ["id", "libelle", "avancement"],
          [{"id": "A3", "libelle": "Plan de test", "avancement": 50},
           {"id": "A4", "libelle": "Execution recette", "avancement": 30}])
manifeste(wb.create_sheet("_Manifeste"), [
    "FILE_TYPE: uo_instance", "FILE_ID: TEST-SOURCE-2", "",
    "DEF $act = GET_TABLE(Activites, tbl_act2)",
    "PUSH $act -> test.source2.activites"])
wb.save(VDIR / "m4_source2.xlsx")

# m2_recepteur.xlsx  (etat final = marche 4 : 2 PULL + compute + bind)
wb = Workbook(); ws = wb.active; ws.title = "Donnees"
add_table(ws, "tbl_recue", ["id", "libelle", "avancement"],
          [{"id": "-", "libelle": "(en attente de synchro)", "avancement": 0}])
dash = wb.create_sheet("Dashboard")
dash["E3"] = "Avancement moyen"
wb.defined_names["avancement"] = DefinedName("avancement", attr_text="'Dashboard'!$F$3")
manifeste(wb.create_sheet("_Manifeste"), [
    "FILE_TYPE: cockpit", "FILE_ID: TEST-RECEPTEUR", "",
    "PULL test.source.activites  -> FILL_TABLE(Donnees, tbl_recue) MODE=OVERWRITE",
    "PULL test.source2.activites -> FILL_TABLE(Donnees, tbl_recue) MODE=APPEND_NEW KEY=id", "",
    "DEF $recu = GET_TABLE(Donnees, tbl_recue)",
    "DEF $moy = COMPUTE(AVG($recu.avancement))", "",
    "BIND $moy -> Dashboard.avancement"])
wb.save(VDIR / "m2_recepteur.xlsx")
print("OK fichiers construits")
