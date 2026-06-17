"""
Premier test des cockpits ingénieur et dashboard métier.

Ce script orchestre un cycle complet :
  1. Génération des 3 cockpits ingénieurs
  2. Injection de données simulées dans le store (simule la saisie Excel + sync)
  3. Génération du dashboard métier
  4. Vérification des résultats

Usage :
  python scripts/demo_cockpits.py

Sortie attendue :
  output/cockpits/Cockpit_Alice_Dubois.xlsx
  output/cockpits/Cockpit_Bruno_Lecomte.xlsx
  output/cockpits/Cockpit_Camille_Vidal.xlsx
  output/cockpits/Dashboard_Jean-Luc_Bernard.xlsx
"""

import io
import sys
from pathlib import Path

# Force UTF-8 output on Windows consoles
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# S'assure que la racine du projet est dans le path
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from src.config_loader import load_uo_instances, load_acteurs
from src.generators.cockpit_ingenieur_generator import generate_cockpit_ingenieur
from src.generators.dashboard_metier_generator import generate_dashboard_metier
from src.store import JsonStore

OUTPUT_DIR = ROOT / "output" / "cockpits"
STORE_PATH = ROOT / "output" / "store.json"

# ─── Données simulées : ce qu'un ingénieur aurait saisi dans son cockpit ─────

SAISIES_SIMULEES = {
    # Alice Dubois — UO-001 : bien avancée, heures OK
    "uo.UO-001.avancement": 0.65,
    "uo.UO-001.heures_realisees": 20.0,
    # Alice Dubois — UO-002 : en dérive heures
    "uo.UO-002.avancement": 0.40,
    "uo.UO-002.heures_realisees": 52.0,   # > 48h allouées → alerte dérive
    # Bruno Lecomte — UO-003 : en retard d'avancement
    "uo.UO-003.avancement": 0.20,
    "uo.UO-003.heures_realisees": 8.0,
    # Bruno Lecomte — UO-004 : OK
    "uo.UO-004.avancement": 0.50,
    "uo.UO-004.heures_realisees": 12.0,
    # Camille Vidal — UO-005 : OK
    "uo.UO-005.avancement": 0.75,
    "uo.UO-005.heures_realisees": 27.0,
}


def banner(title: str):
    print(f"\n{'-' * 60}")
    print(f"  {title}")
    print(f"{'-' * 60}")


def main():
    banner("Chargement de la configuration")
    all_uos = load_uo_instances()
    acteurs = load_acteurs()
    jlb = next((a for a in acteurs if "Bernard" in a.nom), None)

    if jlb is None:
        print("ERREUR : acteur Jean-Luc Bernard (USR004) introuvable dans config/acteurs.json")
        sys.exit(1)

    engineers = sorted(set(u.engineer_name for u in all_uos))
    print(f"  UOs chargées : {len(all_uos)}")
    print(f"  Ingénieurs   : {', '.join(engineers)}")
    print(f"  Pilote métier: {jlb.nom} (filtre: {jlb.filtre_valeur})")

    # ── Étape 1 : Génération des cockpits ingénieurs ──────────────────────────
    banner("Étape 1 — Génération des cockpits ingénieurs")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    for eng in engineers:
        path = generate_cockpit_ingenieur(eng, all_uos, output_dir=OUTPUT_DIR)
        nb_uos = len([u for u in all_uos if u.engineer_name == eng])
        print(f"  ✅  {path.name}  ({nb_uos} UOs)")

    # ── Étape 2 : Simulation des saisies (remplace ExoSync push) ─────────────
    banner("Étape 2 — Injection des données dans le store (saisies simulées)")
    store = JsonStore(STORE_PATH)
    store.set_many(SAISIES_SIMULEES)
    print(f"  {len(SAISIES_SIMULEES)} valeurs injectées dans {STORE_PATH.name}")
    for k, v in SAISIES_SIMULEES.items():
        uo_id = k.split(".")[1]
        champ = k.split(".")[2]
        print(f"    {uo_id}.{champ} = {v}")

    # ── Étape 3 : Génération du dashboard métier ─────────────────────────────
    banner("Étape 3 — Génération du dashboard métier (Jean-Luc Bernard)")
    dash_path = generate_dashboard_metier(jlb, all_uos, store, output_dir=OUTPUT_DIR)
    print(f"  ✅  {dash_path.name}")

    # ── Étape 4 : Vérification des résultats attendus ─────────────────────────
    banner("Étape 4 — Vérification")
    from openpyxl import load_workbook

    wb = load_workbook(dash_path, data_only=True)

    # 4a. Filtrage : Denis Renard absent (UO-004 est Bruno, pas Denis)
    # Le filtre de JLB exclut Denis Renard (pas dans filtre_valeur)
    ws_synth = wb["Synthèse"]
    all_vals = [ws_synth.cell(row=r, column=c).value
                for r in range(1, 50) for c in range(1, 12)]

    filtrage_ok = "Denis Renard" not in (str(v) for v in all_vals if v)
    uo001_present = "UO-001" in all_vals
    print(f"  Filtrage (Denis Renard absent) : {'✅' if filtrage_ok else '❌'}")
    print(f"  UO-001 (Alice) présente       : {'✅' if uo001_present else '❌'}")

    # 4b. Alerte dérive sur UO-002 (52h > 48h)
    ws_alerts = wb["Alertes"]
    alertes_vals = [ws_alerts.cell(row=r, column=c).value
                    for r in range(1, 30) for c in range(1, 6)]
    alerte_uo002 = "UO-002" in alertes_vals
    print(f"  Alerte dérive UO-002 (52h>48h): {'✅' if alerte_uo002 else '❌'}")

    # 4c. Avancement UO-001 = 0.65 dans Synthèse
    avanc_ok = False
    for row in range(6, 20):
        if ws_synth.cell(row=row, column=1).value == "UO-001":
            v = ws_synth.cell(row=row, column=7).value
            if v is not None and abs(float(v) - 0.65) < 0.01:
                avanc_ok = True
            break
    print(f"  Avancement UO-001 = 65%        : {'✅' if avanc_ok else '❌'}")

    # ── Résumé ────────────────────────────────────────────────────────────────
    banner("Résumé — Fichiers générés")
    for f in sorted(OUTPUT_DIR.glob("*.xlsx")):
        size_kb = f.stat().st_size // 1024
        print(f"  {f.name:<45} {size_kb:>4} Ko")

    print(f"\n  ✅  Cycle push/pull terminé — ouvrez les fichiers dans Excel pour vérification manuelle\n")


if __name__ == "__main__":
    main()
