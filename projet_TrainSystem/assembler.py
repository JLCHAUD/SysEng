"""
assembler.py — Instanciation industrialisée d'une UO (brique C4).

Usage :
    python projet_TrainSystem/assembler.py L09U1 \\
        --projet CFL2400 --systeme CLIM \\
        --se "Alice Dubois" --pilote USR004 --heures 200 [--sync]
"""
import argparse
import json
import re
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = Path(__file__).parent
ROOT = HERE.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(HERE))

RE_CODE = re.compile(r"^(L\d{2}U\d)(?:-([A-Za-z0-9]+))?(?:-([A-Za-z0-9]+))?$")

# ── Helpers style (cohérents avec creer_cockpit_se) ──────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _fnt(size=10, bold=False, color="000000", underline=False) -> Font:
    return Font(name="Segoe UI", size=size, bold=bold, color=color,
                underline="single" if underline else None)

def _thin_border() -> Border:
    s = Side(style="thin", color="D3D1C7")
    return Border(left=s, right=s, top=s, bottom=s)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ── Patch chirurgical du cockpit ─────────────────────────────────────────────

def ajouter_uo_au_cockpit(cockpit_path: Path, uo: dict) -> str:
    """
    Insère une ligne UO dans tbl_mes_uos sans écraser les saisies existantes.

    uo = {"file_id": "L09U1-CFL2400-CLIM", "systeme": "Climatisation",
          "projet": "CFL 2400", "heures": 200}

    Retourne "added" | "skipped" | "error:<msg>".
    Fait un backup .bak avant modification.
    """
    try:
        wb = load_workbook(str(cockpit_path))
    except Exception as e:
        return f"error:{e}"

    if "Mes UOs" not in wb.sheetnames:
        return "error:feuille 'Mes UOs' absente"

    ws = wb["Mes UOs"]

    if "tbl_mes_uos" not in ws.tables:
        return "error:table tbl_mes_uos absente"

    tbl = ws.tables["tbl_mes_uos"]

    # Parser la ref : "A3:H4" → header_row=3, last_row=4
    ref = tbl.ref
    top_ref, bot_ref = ref.split(":")
    header_row = int("".join(c for c in top_ref if c.isdigit()))
    last_row   = int("".join(c for c in bot_ref if c.isdigit()))
    data_start = header_row + 1

    # Idempotence : chercher file_id dans col A des data rows
    file_id = uo["file_id"]
    for r in range(data_start, last_row + 1):
        if ws.cell(row=r, column=1).value == file_id:
            wb.close()
            return "skipped"

    # Backup avant modification
    shutil.copy2(str(cockpit_path), str(cockpit_path.with_suffix(".bak")))

    # Nouvelle ligne
    new_row = last_row + 1 if last_row >= data_start else data_start

    row_idx = new_row - data_start  # 0-based
    bg = "F2F2F2" if row_idx % 2 else "FFFFFF"

    # Écrire cols 1-4 uniquement
    ws.cell(row=new_row, column=1, value=file_id)
    ws.cell(row=new_row, column=2, value=uo.get("systeme", ""))
    ws.cell(row=new_row, column=3, value=uo.get("projet", ""))
    ws.cell(row=new_row, column=4, value=uo.get("heures", 0))

    for col in range(1, 5):
        c = ws.cell(row=new_row, column=col)
        c.fill = _fill(bg)
        c.border = _thin_border()
        c.alignment = _left() if col == 1 else _center()
        c.font = _fnt(9.5, color="0563C1" if col == 1 else "000000",
                      underline=(col == 1))
    ws.cell(row=new_row, column=1).hyperlink = f"{file_id}.xlsx"

    # Cols 5-6 : jaunes, vides (zones ingénieur)
    for col in (5, 6):
        c = ws.cell(row=new_row, column=col)
        c.fill = _fill("FFF2CC")
        c.border = _thin_border()
        c.alignment = _center()
    ws.cell(row=new_row, column=5).number_format = "0%"

    # Cols 7-8 : vides avec fond bg
    for col in (7, 8):
        c = ws.cell(row=new_row, column=col)
        c.fill = _fill(bg)
        c.border = _thin_border()
        c.alignment = _center()

    # Étendre la ref de la table
    tbl.ref = f"{top_ref}:{bot_ref[:1]}{new_row}"

    wb.save(str(cockpit_path))
    wb.close()
    return "added"


# ── Création d'un cockpit vide ────────────────────────────────────────────────

def creer_cockpit_vide(se_name: str, pilote_id: str, output_dir: Path) -> Path:
    """
    Génère Cockpit_{se_name}.xlsx avec tbl_mes_uos vide (0 lignes de données).
    Appelé automatiquement par l'assembleur si le cockpit n'existe pas.
    """
    from creer_cockpit_se import generer_cockpit
    return generer_cockpit(se_name, [], pilote_id, output_dir)


# ── Orchestrateur ────────────────────────────────────────────────────────────

def instancier_uo(
    uo_type: str, projet_code: str, systeme_code: str,
    se_name: str, pilote_id: str, heures: float,
    output_dir: Path, sync: bool = False
) -> dict:
    """
    Orchestre la création d'une UO et la mise à jour du cockpit.

    Retourne {"uo_status": "created"|"skipped",
              "cockpit_status": "added"|"skipped"|"created+added",
              "sync_push": int, "sync_errors": int}
    """
    import types
    import creer_uo

    code = f"{uo_type}-{projet_code}-{systeme_code}"
    uo_file = output_dir / f"{code}.xlsx"
    cockpit_file = output_dir / f"Cockpit_{se_name.replace(' ', '_')}.xlsx"

    # Étape 1 : UO
    uo_status = "skipped"
    if not uo_file.exists():
        args = types.SimpleNamespace(
            se=se_name, heures=heures,
            projet=projet_code, systeme=systeme_code,
            output=str(output_dir)
        )
        wb = creer_uo.build_instance(code, uo_type, projet_code, systeme_code, args)
        wb.save(str(uo_file))
        uo_status = "created"

    # Étape 2 : Cockpit
    cockpit_status_prefix = ""
    if not cockpit_file.exists():
        creer_cockpit_vide(se_name, pilote_id, output_dir)
        cockpit_status_prefix = "created+"

    uo_dict = {
        "file_id": code,
        "systeme": systeme_code,
        "projet":  projet_code,
        "heures":  heures,
    }
    add_result = ajouter_uo_au_cockpit(cockpit_file, uo_dict)
    cockpit_status = cockpit_status_prefix + add_result

    # Étape 3 : Sync optionnelle
    sync_push = sync_errors = 0
    if sync:
        from src.sync import synchroniser_repertoire
        rapport_path = synchroniser_repertoire(output_dir)
        rapport = json.loads(rapport_path.read_text(encoding="utf-8"))
        sync_push = sum(
            len([ln for ln in r.get("log", []) if "PUSH=" in ln])
            for r in rapport.get("fichiers", [])
        )
        sync_errors = rapport.get("nb_erreur", 0)

    return {
        "uo_status": uo_status,
        "cockpit_status": cockpit_status,
        "sync_push": sync_push,
        "sync_errors": sync_errors,
    }


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description="Instancie une UO et met à jour le cockpit de l'ingénieur")
    p.add_argument("uo_type", help="ex: L09U1")
    p.add_argument("--projet",  required=True, help="Code projet ex: CFL2400")
    p.add_argument("--systeme", required=True, help="Code système ex: CLIM")
    p.add_argument("--se",      required=True, help="Nom ingénieur SE")
    p.add_argument("--pilote",  default="USR004", help="Pilote ID")
    p.add_argument("--heures",  type=float, default=0, help="Heures vendues")
    p.add_argument("--output",  default=str(HERE), help="Répertoire de sortie")
    p.add_argument("--sync",    action="store_true", help="Lancer la sync après création")
    args = p.parse_args()

    output_dir = Path(args.output)
    code = f"{args.uo_type}-{args.projet}-{args.systeme}"

    if not RE_CODE.match(code):
        sys.exit(f"[ERR] Code invalide : {code}")

    if (output_dir / f"{code}.xlsx").exists():
        print(f"[SKIP] {code}.xlsx existe déjà — rien à faire.")
        return

    result = instancier_uo(
        uo_type=args.uo_type,
        projet_code=args.projet,
        systeme_code=args.systeme,
        se_name=args.se,
        pilote_id=args.pilote,
        heures=args.heures,
        output_dir=output_dir,
        sync=args.sync,
    )

    print(f"[OK] {code}.xlsx créé dans {output_dir}/")
    cockpit_name = f"Cockpit_{args.se.replace(' ', '_')}.xlsx"
    if "added" in result["cockpit_status"]:
        print(f"[OK] {cockpit_name} mis à jour (1 UO ajoutée)")
    else:
        print(f"[SKIP] {cockpit_name} — UO déjà présente")
    if args.sync:
        print(f"[OK] Sync : {result['sync_push']} PUSH, {result['sync_errors']} erreur(s)")


if __name__ == "__main__":
    main()
